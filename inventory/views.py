from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import (Sum, F, Q, DecimalField, ExpressionWrapper, Count,
                              Max, Min, Avg, Case, When, Value, FloatField,
                              Exists, OuterRef)
from django.db.models.functions import (
    Coalesce, TruncDate, TruncWeek, TruncMonth, ExtractWeekDay,
)
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from django.db import transaction, IntegrityError
from django.http import HttpResponseForbidden, HttpResponse, JsonResponse, Http404
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.views.decorators.cache import never_cache
from django.conf import settings
from django.core.cache import cache
import json as _json
from django.utils import timezone
from datetime import timedelta, datetime, date
import csv
import io
import re

import logging
logger = logging.getLogger(__name__)

import qrcode
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
)

from .models import (
    User, Branch, Product, ProductVariant, BranchStock,
    Category, Group, Intake, Sale, AuditLog, SaleTransaction, Return, Customer, Shift,
    Transfer, TransferLine, StockWriteOff, Stocktake, StocktakeCount, ParkedSale, Promotion,
    PaymentQR, PaymentIntent,
    Supplier, IntakeSession, ProductRequest, CashPayout, CashIn, InvoiceDraft,
    InvoiceImage, QuickSellItem, WebOrder, WebOrderLine, EmployeeDebt,
    EmployeeDebtItem,
    split_breakdown, _norm_pay_method,  # ARCH-6: yagona to'lov-split manbai
    _dec,  # MON-24: pul hisobi FAQAT Decimal'da
    weighted_cost,  # STK-8: o'rtacha-tortilgan tannarx
)
_SIZE_WORDS = {'xs', 's', 'm', 'l', 'xl', 'xxl', 'xxxl', '2xl', '3xl', '4xl'}


def smart_title(raw):
    """Nomlarni bir xil ko'rinishga keltiradi: "head and shoulders" ->
    "Head And Shoulders". Qoidalar:
    - faqat to'liq kichik harfli so'zlar bosh harfga ko'tariladi
      (EDP, XPro kabi aralash/katta yozuvlar tegilmaydi)
    - raqamli so'zlar (400ml, 2/1) o'zgarmaydi
    - o'lcham so'zlari (s, m, l, xl...) to'liq KATTA bo'ladi
    - ortiqcha bo'shliqlar yig'ishtiriladi
    """
    words = (raw or '').split()
    out = []
    for w in words:
        if w.lower() in _SIZE_WORDS:
            out.append(w.upper())
        elif any(ch.isdigit() for ch in w):
            out.append(w)                       # 400ml, 2/1 — tegilmaydi
        elif w.isupper() and len(w) > 3:
            out.append(w[0] + w[1:].lower())    # ELEKTR -> Elektr (CapsLock)
        elif w.islower():
            out.append(w[0].upper() + w[1:])    # elektr -> Elektr
        else:
            out.append(w)                       # EDP, XPro, iPhone — saqlanadi
    return ' '.join(out)


def ean13_check_digit(body12):
    """12 xonali EAN body uchun nazorat raqami."""
    total = 0
    for i, ch in enumerate(body12):
        d = int(ch)
        total += d if (i % 2 == 0) else d * 3
    return (10 - (total % 10)) % 10


def gen_internal_ean13(seed_int):
    """Ichki (do'kon) EAN-13: '2' prefiks + 11 xonali seed + nazorat raqami.
    GS1 '20-29' oralig'i do'kon ichki tovarlar uchun ajratilgan."""
    body = ('2' + str(int(seed_int)).zfill(11))[:12]
    return body + str(ean13_check_digit(body))


def ensure_variant_barcode(variant):
    """Turda shtrix-kod bo'lmasa — beradi (kolliziya bo'lsa boshqasini).

    Kodsiz tur kassada skanerlanmaydi va etiketkasi ham chiqmaydi, shuning
    uchun har bir qabul yo'lida tur yaratilgach shu chaqirilishi kerak.
    """
    from .models import ProductVariant as _PV
    if variant is None or variant.barcode:
        return variant
    code = gen_internal_ean13(variant.pk)
    k = 0
    while _PV.objects.filter(barcode=code).exclude(pk=variant.pk).exists():
        k += 1
        code = gen_internal_ean13(variant.pk + k * 100000)
    variant.barcode = code
    variant.save(update_fields=['barcode'])
    return variant


def parse_dec(raw):
    """Foydalanuvchi kiritgan pul/son matnini Decimal'ga aylantiradi.

    "3 000", "3\u00a0000", "30,000", "30000,50", "30'000" — barchasi ishlaydi.
    Bo'sh bo'lsa 0.
    """
    from decimal import Decimal
    raw = (raw or '').strip()
    for ch in (' ', '\u00a0', '\u202f', "'"):
        raw = raw.replace(ch, '')
    if ',' in raw and '.' in raw:
        raw = raw.replace(',', '')          # 1,234.56 -> 1234.56
    elif ',' in raw:
        head, _, tail = raw.rpartition(',')
        if head and len(tail) in (1, 2) and ',' not in head:
            raw = head + '.' + tail          # 30000,5 -> 30000.5
        else:
            raw = raw.replace(',', '')       # 30,000 -> 30000
    return Decimal(raw) if raw else Decimal('0')


from .forms import (
    LoginForm, BranchForm, ProductForm, CategoryForm,
    IntakeForm, SaleForm, UserCreateForm, UserEditForm, ReportForm,
    PIVOT_METRIC_CHOICES, PIVOT_DIM_CHOICES,
)


def admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not request.user.is_admin():
            return HttpResponseForbidden(
                "<h3>Ruxsat yo'q. Bu sahifa faqat administrator uchun.</h3>"
            )
        return view_func(request, *args, **kwargs)
    return wrapper


def get_user_branch(user):
    """Sotuvchi uchun majburiy filial. Admin uchun None bo'lishi mumkin."""
    return user.branch


def normalize_code(typed):
    """Kiritilgan kodni standart shaklga keltirish: OYO1 → OYO-0001"""
    if not typed:
        return ''
    t = typed.strip().upper().replace(' ', '')
    m = re.match(r'^([A-Z]+)-?(\d+)$', t)
    if m:
        return f'{m.group(1)}-{int(m.group(2)):04d}'
    return t


# ---------- HEALTH ----------

def healthz(request):
    """Liveness probe for Render / monitoring. Cheap, no DB hit."""
    return HttpResponse('ok', content_type='text/plain')


# ---------- AUTH ----------

def csrf_failure(request, reason=""):
    """CSRF xatosini muloyim hal qiladi.

    Ko'pincha sabab — sahifaning ESKIRGANI: bir nechta tab ochiq, brauzer
    tablarni tiklagan yoki "orqaga" tugmasi bosilgan. Login'da Django CSRF
    tokenni yangilaydi, shuning uchun eski formadagi token to'g'ri kelmaydi.
    Qo'rqinchli 403 o'rniga foydalanuvchini yangi token bilan qaytaramiz.
    """
    from django.contrib import messages as _messages
    _messages.warning(
        request, "Sahifa eskirgan edi — iltimos qaytadan urinib ko'ring.")
    if '/login' in request.path or not request.user.is_authenticated:
        return redirect('login')
    return redirect(request.META.get('HTTP_REFERER') or 'home')


@never_cache
def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')

    # S5: rate limit — max 5 failed attempts per IP in 5 minutes.
    # AUTH-5: IP'ni _client_ip orqali (X-Real-IP) olamiz — XFF birinchi
    # elementi soxta qo'yilib budjet yangilanmasin.
    from django.core.cache import cache
    from .middleware import _client_ip
    ip = _client_ip(request) or '0.0.0.0'
    fail_key = f'login_fail:{ip}'
    fail_count = cache.get(fail_key) or 0
    if fail_count >= 5:
        messages.error(request,
            "Juda ko'p urinish. 5 daqiqa kutib qayta urinib ko'ring.")
        return render(request, 'inventory/login.html',
                      {'form': LoginForm(request), 'rate_limited': True})

    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            _u = form.get_user()
            if _u.totp_confirmed and _u.totp_secret:
                # AUTH-2: parol to'g'ri, lekin 2FA hali qolgan — hisoblagichni
                # TOZALAMAYMIZ. Aks holda parolni bilган hujumchi cheksiz TOTP
                # taxmin qilardi. Faqat ikkala bosqich o'tганda tozalanadi.
                request.session['2fa_pending_user'] = _u.id
                return redirect('login_2fa')
            cache.delete(fail_key)  # 2FA yo'q — muvaffaqiyat, hisoblagich tozalanadi
            login(request, _u)
            messages.success(request, f'Xush kelibsiz, {_u.username}!')
            return redirect('home')
        else:
            # Increment fail counter (5-min TTL)
            cache.set(fail_key, fail_count + 1, timeout=300)
            remaining = 5 - (fail_count + 1)
            if remaining > 0:
                messages.warning(request,
                    f"Login muvaffaqiyatsiz. Yana {remaining} ta urinishingiz qoldi.")
    else:
        form = LoginForm(request)
    return render(request, 'inventory/login.html', {'form': form})


def logout_view(request):
    logout(request)
    return redirect('login')


@never_cache
def login_2fa(request):
    """Second step: verify a TOTP or recovery code after a valid password."""
    from .twofa import verify_totp_step, use_recovery_code
    from django.core.cache import cache
    from .middleware import _client_ip
    uid = request.session.get('2fa_pending_user')
    if not uid:
        return redirect('login')
    user = User.objects.filter(id=uid).first()
    if not user:
        request.session.pop('2fa_pending_user', None)
        return redirect('login')

    # AUTH-2: TOTP bosqichi parol bosqichi bilan BIR xil kalitда cheklanadi.
    # Parolni bilган hujumchi ham 5 tadan ko'p taxmin qila olmaydi.
    ip = _client_ip(request) or '0.0.0.0'
    fail_key = f'login_fail:{ip}'
    fail_count = cache.get(fail_key) or 0
    if fail_count >= 5:
        messages.error(request,
            "Juda ko'p urinish. 5 daqiqa kutib qayta urinib ko'ring.")
        return render(request, 'inventory/login_2fa.html', {'rate_limited': True})

    if request.method == 'POST':
        code = request.POST.get('code', '')
        # AUTH-3: mos vaqt-qadamни olamiz; oxirgi qabul qilinganдан katta
        # bo'lса qabul, aks holда replay — rad. Recovery kodlar bir martalik.
        step = verify_totp_step(user.totp_secret, code)
        totp_ok = (step is not None
                   and (user.last_totp_step is None or step > user.last_totp_step))
        if totp_ok or use_recovery_code(user, code):
            if totp_ok:
                user.last_totp_step = step
                user.save(update_fields=['last_totp_step'])
            cache.delete(fail_key)  # ikkala bosqich o'tди — hisoblagич tozalanadi
            request.session.pop('2fa_pending_user', None)
            login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            messages.success(request, f'Xush kelibsiz, {user.username}!')
            return redirect('home')
        cache.set(fail_key, fail_count + 1, timeout=300)
        messages.error(request, "Kod noto'g'ri.")
    return render(request, 'inventory/login_2fa.html', {})


@login_required
def security_2fa(request):
    """Opt-in TOTP 2FA setup for the current user."""
    from .twofa import (gen_secret, verify_totp, otpauth_uri, qr_datauri,
                        gen_recovery_codes, hash_code)
    user = request.user
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'disable':
            # AUTH-4: 2FA'ni o'chirish uchun joriy parol shart. Aks holда
            # o'g'irlangan sessiya himoyani parolsiz olib tashlardi.
            if not user.check_password(request.POST.get('password', '')):
                messages.error(request,
                    "Parol noto'g'ri — 2FA o'chirilmadi.")
                return redirect('security_2fa')
            user.totp_secret = ''
            user.totp_confirmed = False
            user.recovery_codes = []
            user.last_totp_step = None
            user.save()
            messages.success(request, "Ikki bosqichli himoya o'chirildi.")
            return redirect('security_2fa')
        if action == 'enable':
            secret = request.session.get('2fa_setup_secret') or ''
            if secret and verify_totp(secret, request.POST.get('code', '')):
                plain = gen_recovery_codes(8)
                user.totp_secret = secret
                user.totp_confirmed = True
                user.recovery_codes = [hash_code(c) for c in plain]
                user.save()
                request.session.pop('2fa_setup_secret', None)
                messages.success(request, "Ikki bosqichli himoya yoqildi.")
                return render(request, 'inventory/security_2fa.html',
                              {'enabled': True, 'recovery': plain})
            messages.error(request, "Kod noto'g'ri — ilovadagi 6 xonali kodni kiriting.")
    if user.totp_confirmed:
        return render(request, 'inventory/security_2fa.html',
                      {'enabled': True, 'recovery_left': len(user.recovery_codes or [])})
    secret = request.session.get('2fa_setup_secret') or gen_secret()
    request.session['2fa_setup_secret'] = secret
    return render(request, 'inventory/security_2fa.html', {
        'enabled': False, 'secret': secret,
        'qr': qr_datauri(otpauth_uri(secret, user.username)),
    })


@login_required
def home(request):
    # POS tizimiga TEGILMAYDI — sayt butunlay alohida (/shop/).
    if request.user.is_admin():
        return redirect('dashboard')
    return redirect('lookup')


# ---------- LOOKUP (asosiy qidiruv) ----------

@login_required
def lookup(request):
    raw_query = (request.GET.get('code') or '').strip()
    code = normalize_code(raw_query.upper())
    product = None
    branches_data = []
    suggestions = []
    product_stats = None
    popular_products = []
    categories_quick = []

    if raw_query:
        # Try exact code first — either internal code or manufacturer barcode
        product = Product.objects.filter(
            Q(code=code) | Q(external_barcode=raw_query)
        ).first()
        if not product:
            _vm = (ProductVariant.objects.filter(barcode=raw_query)
                   .select_related('product').first())
            if _vm:
                product = _vm.product

        # If no code match, fall back to name / category / description search
        if not product:
            matches = (Product.objects
                .filter(Q(name__icontains=raw_query) |
                        Q(brand__icontains=raw_query) |
                        Q(category__name__icontains=raw_query) |
                        Q(description__icontains=raw_query))
                .select_related('category')[:20])
            if matches.count() == 1:
                product = matches.first()
                code = product.code
            elif matches.exists():
                suggestions = list(matches)
            else:
                messages.warning(request,
                    f"'{raw_query}' bo'yicha mahsulot topilmadi.")

        if product:
            variants = list(product.variants.all())
            sizes = sorted({v.size for v in variants}, key=lambda s: (len(s), s))
            colors = sorted({v.color for v in variants})

            if request.user.is_admin():
                branches = Branch.objects.filter(is_active=True)
            else:
                branches = Branch.objects.filter(pk=request.user.branch_id) \
                    if request.user.branch_id else Branch.objects.none()

            # Aggregate stats across all branches (admin) or single branch
            total_stock = 0
            total_value = 0
            zero_count = 0
            low_count = 0
            for br in branches:
                stocks = BranchStock.objects.filter(
                    variant__product=product, branch=br
                ).select_related('variant')
                matrix = {(s.variant.size, s.variant.color): s for s in stocks}
                br_total = sum(s.stock_count for s in stocks)
                br_value = sum(s.stock_count * float(s.cost_price) for s in stocks)
                total_stock += br_total
                total_value += br_value
                for s in stocks:
                    if s.stock_count == 0:
                        zero_count += 1
                    elif s.stock_count <= 3:
                        low_count += 1
                # Sotuvchi uchun yassi variant ro'yxati: bor bo'lganlar
                # oldinda, ko'p qoldiq oldinda, keyin rang/o'lcham bo'yicha
                var_list = [{
                    'color': st.variant.color,
                    'size': st.variant.size,
                    'barcode': st.variant.barcode or '',
                    'count': st.stock_count,
                    'sale': float(st.sale_price or product.default_sale_price or 0),
                    'wholesale': float(st.wholesale_price or 0),
                    'stock_id': st.id,
                } for st in stocks]
                var_list.sort(key=lambda r: (r['count'] == 0, -r['count'],
                                             r['color'], r['size']))
                branches_data.append({
                    'branch': br, 'matrix': matrix,
                    'sizes': sizes, 'colors': colors,
                    'total': br_total, 'value': br_value,
                    'variants': var_list,
                    'in_stock_n': sum(1 for r in var_list if r['count'] > 0),
                })

            # 30-day sales velocity
            since_30d = timezone.now() - timedelta(days=30)
            velocity_agg = Sale.objects.filter(
                variant__product=product,
                sold_at__gte=since_30d,
            ).aggregate(
                qty=Sum('quantity'),
                txns=Count('transaction', distinct=True),
            )
            sold_30d = velocity_agg['qty'] or 0
            txns_30d = velocity_agg['txns'] or 0
            daily_avg = sold_30d / 30 if sold_30d else 0
            days_left = (total_stock / daily_avg) if daily_avg else None

            product_stats = {
                'total_stock': total_stock,
                'total_value': total_value,
                'variants_count': len(variants),
                'branches_count': len(branches_data),
                'zero_count': zero_count,
                'low_count': low_count,
                'sold_30d': sold_30d,
                'txns_30d': txns_30d,
                'days_left': days_left,
            }
    else:
        # No search — show popular products and category quick-jumps
        since_30d = timezone.now() - timedelta(days=30)
        from django.db.models import Max as _Max
        top_rows = (Sale.objects
                    .filter(sold_at__gte=since_30d)
                    .exclude(variant__product__is_open_price=True)
                    .values('variant__product')
                    .annotate(qty=Sum('quantity'))
                    .order_by('-qty')[:8])
        top_ids = [r['variant__product'] for r in top_rows]
        prods_map = {p.id: p for p in Product.objects
                     .filter(id__in=top_ids).select_related('category')}
        # Haqiqiy sotuv narxi = variantlarning eng katta narxi (default 0 bo'lsa ham)
        price_map = {r['variant__product']: r['pmax'] for r in
                     BranchStock.objects.filter(variant__product_id__in=top_ids)
                     .values('variant__product').annotate(pmax=_Max('sale_price'))}
        for r in top_rows:
            p = prods_map.get(r['variant__product'])
            if p:
                eff = price_map.get(p.id) or p.default_sale_price or 0
                popular_products.append({'p': p, 'qty': r['qty'], 'price': eff})
        categories_quick = list(Category.objects.annotate(
            n=Count('products')
        ).filter(n__gt=0).order_by('-n')[:8])

    return render(request, 'inventory/lookup.html', {
        'code': raw_query, 'product': product,
        'branches_data': branches_data, 'suggestions': suggestions,
        'product_stats': product_stats,
        'popular_products': popular_products,
        'categories_quick': categories_quick,
        'query_looks_like_barcode': bool(raw_query) and raw_query.isdigit() and len(raw_query) >= 6,
    })


# ---------- DASHBOARD ----------

DASHBOARD_CACHE_KEY = 'dashboard:hq:v5'
DASHBOARD_CACHE_TTL = 60  # seconds — heavy aggregates only; recent_sales stays live


def _dashboard_aggregates():
    """Compute the cacheable, expensive part of the dashboard.

    Returned dict is JSON-serializable (Branch model is keyed by id, looked up
    on render). TTL is short (60s) so per-shift drift is invisible to users.
    Invalidated explicitly when a SaleTransaction is saved (see signals.py).
    """
    today = timezone.localdate()
    yesterday = today - timedelta(days=1)

    def _day_range(d):
        tz = timezone.get_current_timezone()
        start = datetime.combine(d, datetime.min.time()).replace(tzinfo=tz)
        return start, start + timedelta(days=1)

    today_start, today_end = _day_range(today)
    yesterday_start, yesterday_end = _day_range(yesterday)

    revenue_expr = ExpressionWrapper(
        F('quantity') * F('sale_price') - F('line_discount'),
        output_field=DecimalField(max_digits=14, decimal_places=2)
    )
    cost_expr = ExpressionWrapper(
        F('quantity') * F('cost_at_sale'),
        output_field=DecimalField(max_digits=14, decimal_places=2)
    )

    def _agg(qs):
        a = qs.aggregate(
            revenue=Sum(revenue_expr),
            cost=Sum(cost_expr),
            qty=Sum('quantity'),
            txns=Count('transaction', distinct=True),
        )
        # SAL-6: KUN — chek darajasidagi o'lchov, chegirma aniq tushadi.
        # Ayirmasak tushum ham, FOYDA ham chegirma miqdorida shishardi.
        _od, _ = _order_discount_share(qs)
        # RET-1: qaytarilgan tovar tushumdan ham, TANNARXdan ham chiqadi.
        _rret, _cret, _ = _returns_adjustment(qs)
        rev = _net_rev(a['revenue'], _od) - float(_rret)
        cost = float(_dec(a['cost'] or 0) - _cret)
        return {
            'revenue': rev,
            'cost': cost,
            'profit': rev - cost,
            'margin': (rev - cost) / rev * 100 if rev else 0,
            'qty': a['qty'] or 0,
            'txns': a['txns'] or 0,
        }

    today_stats = _agg(Sale.objects.filter(sold_at__gte=today_start, sold_at__lt=today_end))
    yesterday_stats = _agg(Sale.objects.filter(sold_at__gte=yesterday_start, sold_at__lt=yesterday_end))

    # Bugungi pul oqimi — to'lov turi bo'yicha. "Aralash" (mixed) cheklar
    # payment_breakdown bo'yicha naqd/karta/o'tkazma qismlarига BO'LINADI —
    # alohida "Aralash" ustuni ko'rsatilmaydi. Har chekning sof summasi
    # qismlar nisbatiga qarab taqsimlanadi (oxirgi qism qoldiqni oladi —
    # yaxlitlash tufayli jami buzilmaydi).
    def _by_method(qs):
        out = {'cash': 0.0, 'card': 0.0, 'transfer': 0.0, 'other': 0.0}
        # SAL-6: pul oqimi KASSAGA TUSHGAN pul bo'lishi kerak — chek chegirmasi
        # ayirilgan (chek bir to'lov turiga tegishli, ulush aniq tushadi).
        _, _od_by_txn = _order_discount_share(qs, 'transaction_id')
        rows = (qs.values('transaction_id', 'transaction__payment_method')
                  .annotate(rev=Sum(revenue_expr)))
        mixed_rev = {}
        for r in rows:
            m = r['transaction__payment_method']
            rev = _net_rev(r['rev'], _od_by_txn.get(r['transaction_id']))
            if m == 'mixed':
                mixed_rev[r['transaction_id']] = mixed_rev.get(r['transaction_id'], 0.0) + rev
            elif m in out:
                out[m] += rev
            else:
                out['other'] += rev
        if mixed_rev:
            txns = (SaleTransaction.objects.filter(id__in=list(mixed_rev.keys()))
                    .values('id', 'payment_breakdown'))
            for t in txns:
                rev = mixed_rev.get(t['id'], 0.0)
                # ARCH-6: yagona split_breakdown() — modeldagi bilan bir xil
                part = split_breakdown(rev, t['payment_breakdown'])
                for k in ('cash', 'card', 'transfer'):
                    out[k] += float(part[k])
        return out

    today_by_method = _by_method(
        Sale.objects.filter(sold_at__gte=today_start, sold_at__lt=today_end))
    yesterday_by_method = _by_method(
        Sale.objects.filter(sold_at__gte=yesterday_start, sold_at__lt=yesterday_end))

    # Bugungi qaytarishlar — ASL to'lov turi bo'yicha (sof savdoni ko'rsatish uchun)
    def _returns_by_method(start, end):
        out = {'cash': 0.0, 'card': 0.0, 'transfer': 0.0}
        rets = (Return.objects.filter(refunded_at__gte=start, refunded_at__lt=end)
                .select_related('sale__transaction'))
        for r in rets:
            amt = float(r.effective_cash_refund)
            if amt <= 0:
                continue
            txn = r.sale.transaction if r.sale_id else None
            pm = txn.payment_method if txn else 'cash'
            if pm != 'mixed' or not txn:
                out[_norm_pay_method(pm)] += amt
                continue
            # ARCH-6: yagona split_breakdown()
            part = split_breakdown(amt, txn.payment_breakdown)
            for k in out:
                out[k] += float(part[k])
        return out

    today_returns_by_method = _returns_by_method(today_start, today_end)

    # 7-day trend — single annotated query
    week_start = today_start - timedelta(days=6)
    trend_qs = (
        Sale.objects.filter(sold_at__gte=week_start, sold_at__lt=today_end)
        .annotate(day=TruncDate('sold_at'))
        .values('day')
        .annotate(rev=Sum(revenue_expr), q=Sum('quantity'))
    )
    day_map = {r['day']: r for r in trend_qs}
    _, _od_by_day = _order_discount_share(
        Sale.objects.filter(sold_at__gte=week_start, sold_at__lt=today_end),
        'sold_at__date')
    trend_labels, trend_revenue, trend_qty = [], [], []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        r = day_map.get(d) or {}
        trend_labels.append(d.strftime('%a %d'))
        trend_revenue.append(_net_rev(r.get('rev'), _od_by_day.get(d)))
        trend_qty.append(r.get('q') or 0)

    # Inventory snapshot — one aggregate covers stock count + value
    inv_totals = BranchStock.objects.aggregate(
        s=Sum('stock_count'),
        v=Sum(ExpressionWrapper(F('stock_count') * F('cost_price'),
                                output_field=DecimalField(max_digits=14, decimal_places=2))),
    )
    total_stock = inv_totals['s'] or 0
    stock_value = float(inv_totals['v'] or 0)

    # Per-branch today — 2 aggregated queries, independent of branch count
    rev_by_branch = {
        r['branch_id']: r for r in
        Sale.objects.filter(sold_at__gte=today_start, sold_at__lt=today_end)
        .values('branch_id')
        .annotate(
            revenue=Sum(revenue_expr),
            qty=Sum('quantity'),
            txns=Count('transaction', distinct=True),
        )
    }
    stock_by_branch = {
        r['branch_id']: r['stock_count__sum']
        for r in BranchStock.objects.values('branch_id').annotate(stock_count__sum=Sum('stock_count'))
    }

    _, _od_by_branch = _order_discount_share(
        Sale.objects.filter(sold_at__gte=today_start, sold_at__lt=today_end),
        'branch_id')
    branch_today_raw = [
        {
            'branch_id': bid,
            # FILIAL — chek darajasidagi o'lchov (SAL-6)
            'revenue': _net_rev((rev_by_branch.get(bid) or {}).get('revenue'),
                                _od_by_branch.get(bid)),
            'qty': (rev_by_branch.get(bid) or {}).get('qty') or 0,
            'txns': (rev_by_branch.get(bid) or {}).get('txns') or 0,
            'stock': stock_by_branch.get(bid, 0),
        }
        for bid in Branch.objects.filter(is_active=True).values_list('id', flat=True)
    ]

    top_today = list(
        Sale.objects.filter(sold_at__gte=today_start, sold_at__lt=today_end)
        .values('variant__product__code', 'variant__product__name')
        .annotate(qty=Sum('quantity'), revenue=Sum(revenue_expr))
        .order_by('-qty')[:5]
    )
    for r in top_today:
        r['revenue'] = float(r['revenue'] or 0)
    # MAHSULOT — qator darajasidagi o'lchov: chegirma alohida tovarga tegishli
    # emas (egasining qoidasi), shuning uchun qatorlar O'Z narxida qoladi va
    # chegirma alohida ko'rsatiladi (SAL-6).
    _od_today, _, _disc_split = _order_discount_share(
        Sale.objects.filter(sold_at__gte=today_start, sold_at__lt=today_end),
        split=True)

    return {
        'today_iso': today.isoformat(),
        'order_discount_today': float(_od_today),
        'discount_split_today': {k: float(v) for k, v in _disc_split.items()},
        'yesterday_iso': yesterday.isoformat(),
        'today_stats': today_stats,
        'yesterday_stats': yesterday_stats,
        'today_by_method': today_by_method,
        'yesterday_by_method': yesterday_by_method,
        'today_returns_by_method': today_returns_by_method,
        'trend_labels': trend_labels,
        'trend_revenue': trend_revenue,
        'trend_qty': trend_qty,
        'total_stock': total_stock,
        'stock_value': stock_value,
        'branch_today_raw': branch_today_raw,
        'top_today': top_today,
    }


@admin_required
def dashboard(request):
    agg = cache.get_or_set(DASHBOARD_CACHE_KEY, _dashboard_aggregates, DASHBOARD_CACHE_TTL)

    today = date.fromisoformat(agg['today_iso'])
    yesterday = date.fromisoformat(agg['yesterday_iso'])
    # Ko'rsatish uchun nusxa — sof foyda tannarx (kirim narxi) kasrли bo'lsa
    # mayda raqamli chiqadi (masalan 5 704 775). Do'kon 1000dan mayda ishlatmaydi,
    # shuning uchun foyda va tushumni eng yaqin 1000 ga yaxlitlaymiz (marja aniq
    # foizда qoladi).
    def _round1k(x):
        try:
            return int(round(float(x) / 1000.0)) * 1000
        except (TypeError, ValueError):
            return x
    today_stats = dict(agg['today_stats'])
    yesterday_stats = dict(agg['yesterday_stats'])
    today_stats['profit'] = _round1k(today_stats.get('profit'))
    today_stats['revenue'] = _round1k(today_stats.get('revenue'))
    yesterday_stats['profit'] = _round1k(yesterday_stats.get('profit'))
    yesterday_stats['revenue'] = _round1k(yesterday_stats.get('revenue'))

    def _delta(now, prev):
        # Kunning boshida (bugun hali 0) -100% ko'rsatmaymiz — bu
        # signal emas, shunchaki kun boshlangani. Delta faqat bugun
        # ham savdo bo'lganda ma'noli.
        if not prev or not now:
            return None
        return (now - prev) / prev * 100

    deltas = {
        'revenue': _delta(today_stats['revenue'], yesterday_stats['revenue']),
        'qty': _delta(today_stats['qty'], yesterday_stats['qty']),
        'txns': _delta(today_stats['txns'], yesterday_stats['txns']),
        'profit': _delta(today_stats['profit'], yesterday_stats['profit']),
    }

    # Hydrate branch_today_raw with live Branch objects (Branch model not cached
    # in case is_active toggles during the 60s window)
    active_branches = {b.id: b for b in Branch.objects.filter(is_active=True)}
    branch_today = []
    for row in agg['branch_today_raw']:
        br = active_branches.get(row['branch_id'])
        if not br:
            continue
        branch_today.append({
            'branch': br,
            'revenue': row['revenue'],
            'qty': row['qty'],
            'txns': row['txns'],
            'stock': row['stock'],
        })

    total_products = Product.objects.count()
    total_branches = len(active_branches)

    # ---- Live (uncached) data: attention widgets + recent activity ----
    low_stock_count = BranchStock.objects.filter(stock_count__lte=3).count()
    low_stock_preview = list(BranchStock.objects.filter(stock_count__lte=3)
                             .select_related('variant__product', 'branch')
                             .order_by('stock_count')[:5])
    out_of_stock_count = BranchStock.objects.filter(stock_count=0).count()
    open_shifts_count = Shift.objects.filter(status=Shift.Status.OPEN).count()
    pending_intents_count = PaymentIntent.objects.filter(
        status=PaymentIntent.Status.PENDING,
        created_at__gte=timezone.now() - timedelta(hours=24),
    ).count()
    in_transit_count = Transfer.objects.filter(status=Transfer.Status.IN_TRANSIT).count()
    open_parked_count = ParkedSale.objects.count()

    recent_sales = (SaleTransaction.objects
                    .select_related('branch', 'sold_by')
                    .prefetch_related('lines')
                    .order_by('-sold_at')[:8])

    # Bugungi pul oqimi — to'lov turi bo'yicha tayyor qatorlar (shablon uchun)
    _tbm = dict(agg['today_by_method'])
    _ybm = dict(agg['yesterday_by_method'])
    # Noma'lum ("other") summa (eski/legacy sotuvlar) — naqdga qo'shamiz,
    # shunда faqat 3 ustun ko'rinadi va jami buzilmaydi.
    _tbm['cash'] = _tbm.get('cash', 0) + _tbm.pop('other', 0)
    _ybm['cash'] = _ybm.get('cash', 0) + _ybm.pop('other', 0)
    flow_total = sum(_tbm.get(k, 0) for k in ('cash', 'card', 'transfer'))
    _flow_defs = [
        ('cash', 'Naqd', 'bi-cash-stack', '#16A34A'),
        ('card', 'Karta', 'bi-credit-card-2-front', '#2563EB'),
        ('transfer', "O'tkazma", 'bi-arrow-left-right', '#D97706'),
    ]
    _trbm = agg.get('today_returns_by_method', {}) or {}
    # Qaytarish — bitta NAQD chiqim (usulларга bo'linmaydi). Karta brutto
    # qoladi (terminal hisoboti bilan mos).
    flow_returns_total = sum(_trbm.get(k, 0) for k in ('cash', 'card', 'transfer'))

    # Do'kon 1000 so'mdan mayda summalar ishlatmaydi — ko'rsatishда eng yaqin
    # 1000 ga yaxlitlaymiz (fraksiya/yuzlik "chiqindi" ko'rinmasin).
    def _r1k(x):
        try:
            return int(round(float(x) / 1000.0)) * 1000
        except (TypeError, ValueError):
            return 0

    flow_returns_total = _r1k(flow_returns_total)
    flow_total = sum(_r1k(_tbm.get(k, 0)) for k in ('cash', 'card', 'transfer'))
    flow_rows = []
    for key, label, icon, color in _flow_defs:
        amt = _r1k(_tbm.get(key, 0))
        flow_rows.append({
            'key': key, 'label': label, 'icon': icon, 'color': color,
            'amount': amt,
            'pct': (amt / flow_total * 100) if flow_total else 0,
            'yesterday': _r1k(_ybm.get(key, 0)),
        })

    return render(request, 'inventory/dashboard.html', {
        'order_discount_today': agg.get('order_discount_today', 0),
        'discount_split_today': agg.get('discount_split_today') or {},
        'today': today, 'yesterday': yesterday,
        'today_stats': today_stats,
        'yesterday_stats': yesterday_stats,
        'today_by_method': agg['today_by_method'],
        'yesterday_by_method': agg['yesterday_by_method'],
        'flow_rows': flow_rows,
        'flow_total': flow_total,
        'flow_returns_total': flow_returns_total,
        'deltas': deltas,
        'trend_labels': agg['trend_labels'],
        'trend_revenue': agg['trend_revenue'],
        'trend_qty': agg['trend_qty'],
        'total_products': total_products,
        'total_branches': total_branches,
        'total_stock': agg['total_stock'],
        'stock_value': agg['stock_value'],
        'branch_today': branch_today,
        'top_today': agg['top_today'],
        'low_stock_count': low_stock_count,
        'low_stock_preview': low_stock_preview,
        'out_of_stock_count': out_of_stock_count,
        'open_shifts_count': open_shifts_count,
        'pending_intents_count': pending_intents_count,
        'in_transit_count': in_transit_count,
        'open_parked_count': open_parked_count,
        'recent_sales': recent_sales,
    })


# ---------- PRODUCTS ----------

@admin_required
@admin_required
def product_bulk_update(request):
    """POST: tanlangan mahsulotlarga bulk amal qo'llash."""
    if request.method != 'POST':
        return redirect('product_list')
    ids = request.POST.getlist('id')
    if not ids:
        messages.warning(request, "Bironta ham mahsulot tanlanmagan.")
        return redirect('product_list')
    op = request.POST.get('op') or ''
    if op == 'merge':
        clean_ids = ','.join(str(int(i)) for i in ids if str(i).isdigit())
        return redirect(f"{reverse('product_merge')}?ids={clean_ids}")
    from decimal import Decimal
    try:
        value = Decimal(request.POST.get('value') or '0')
    except Exception:
        messages.error(request, "Noto'g'ri qiymat.")
        return redirect('product_list')
    if not value.is_finite():      # 'inf'/'nan' — N3/V4
        messages.error(request, "Noto'g'ri qiymat.")
        return redirect('product_list')

    # STK-19: narx amallari uchun qiymat 0 bo'lса — RAD etamiz. Ilgari bo'sh
    # qiymat 0 ga tushib, `multiply_price` butun katalog narxini NOLGA
    # aylantirar (yoki `set_price` hammani 0 qilar) edi.
    if op in ('set_price', 'set_markup', 'multiply_price') and value <= 0:
        messages.error(request, "Qiymat 0 dan katta bo'lishi kerak.")
        return redirect('product_list')

    # N3/V4: YUQORI chegara ham SERVERда — brauzer max faqat maslahat. "×1000"
    # yoki 13-xonali narx butun katalogni buzmasin.
    _caps = {'set_price': Decimal('100000000'),
             'multiply_price': Decimal('1000'),
             'set_markup': Decimal('10000')}
    if op in _caps and value > _caps[op]:
        messages.error(request, "Qiymat juda katta — tekshirib qayta kiriting.")
        return redirect('product_list')

    products = Product.objects.filter(id__in=[int(i) for i in ids if str(i).isdigit()])
    n = 0
    affected_codes = []
    from django.db.models import ProtectedError
    # STK-19: butun amal BITTA tranzaksiyada — yarim qo'llanib qolmasin.
    with transaction.atomic():
        for p in products:
            affected_codes.append(p.code)
            if op == 'set_price':
                p.default_sale_price = value
                p.save(update_fields=['default_sale_price'])
                n += 1
            elif op == 'set_markup':
                p.markup_percent = value
                p.save(update_fields=['markup_percent'])
                n += 1
            elif op == 'multiply_price':
                new_price = (p.default_sale_price * (value / 100)).quantize(Decimal('1'))
                p.default_sale_price = new_price
                p.save(update_fields=['default_sale_price'])
                n += 1
            elif op == 'change_category':
                try:
                    cat_id = int(value)
                    cat = Category.objects.filter(pk=cat_id).first()
                    if cat:
                        p.category = cat
                        p.save(update_fields=['category'])
                        n += 1
                except (ValueError, TypeError):
                    cat = None
            elif op == 'delete':
                try:
                    p.delete()
                    n += 1
                except ProtectedError:
                    messages.warning(
                        request,
                        f"'{p.name}' ({p.code}) o'chirilmadi — sotuv tarixi bor.")

    # M9: audit the bulk operation as a single summary row
    if n > 0:
        OP_LABELS = {
            'set_price': f"narx={value}",
            'set_markup': f"markup={value}%",
            'multiply_price': f"narx×{value}%",
            'change_category': f"category={value}",
            'delete': "DELETE",
        }
        op_label = OP_LABELS.get(op, op)
        codes_preview = ', '.join(affected_codes[:10])
        if len(affected_codes) > 10:
            codes_preview += f", ... (+{len(affected_codes)-10})"
        AuditLog.objects.create(
            user=request.user,
            username_snapshot=request.user.username,
            action=(AuditLog.Action.DELETE if op == 'delete'
                    else AuditLog.Action.UPDATE),
            model_name='ProductBulk',
            object_repr=f"{n}×{op_label}: {codes_preview}"[:200],
        )

    messages.success(request, f"{n} ta mahsulot yangilandi/o'chirildi.")
    if op == 'change_category' and n > 0:
        try:
            _cid = int(value)
            _cat = Category.objects.filter(pk=_cid).first()
            if _cat:
                messages.success(
                    request,
                    f"{n} ta mahsulot \"{_cat.name}\" kategoriyasiga ko'chirildi.")
                return redirect(f"{reverse('product_list')}?category={_cid}")
        except (ValueError, TypeError):
            pass
    return redirect('product_list')


@admin_required
def product_list(request):
    """Mahsulotlar ro'yxati: qidiruv + filtrlar + sortable + 30 kunlik sotilganlik."""
    q = (request.GET.get('q') or '').strip()
    category_id = request.GET.get('category') or ''
    group_slug = (request.GET.get('group') or '').strip()  # men|women|kids|home
    stock_filter = request.GET.get('stock') or ''  # zero|low|in_stock|''
    price_filter = request.GET.get('price') or ''  # 'zero' = narxsiz sotuvda
    sort = request.GET.get('sort') or '-created_at'

    products = (Product.objects
                .select_related('category', 'category__group')
                .exclude(is_open_price=True))

    if group_slug:
        products = products.filter(category__group__slug=group_slug)

    name_conflict = request.GET.get('name_conflict') or ''
    if name_conflict == '1':
        products = products.exclude(pending_name='')

    if q:
        # Subquery orqali — variant JOIN'lari annotatsiya Sum'larini
        # buzmasligi uchun avval mos mahsulot ID'larini topamiz
        _match_ids = Product.objects.filter(
            Q(code__icontains=q) | Q(name__icontains=q) |
            Q(brand__icontains=q) |
            Q(external_barcode__icontains=q) |
            Q(variants__color__icontains=q) |
            Q(variants__size__icontains=q) |
            Q(variants__barcode__icontains=q)
        ).values('id')
        products = products.filter(id__in=_match_ids)
    if category_id:
        try:
            products = products.filter(category_id=int(category_id))
        except ValueError:
            category_id = ''

    # Annotate aggregate stock per product (+ narx diapazoni va real marja
    # uchun qiymatlar — bitta JOIN, qo'shimcha so'rovsiz)
    _val = ExpressionWrapper(
        F('variants__branch_stocks__stock_count') *
        F('variants__branch_stocks__sale_price'),
        output_field=DecimalField(max_digits=16, decimal_places=2))
    _cost = ExpressionWrapper(
        F('variants__branch_stocks__stock_count') *
        F('variants__branch_stocks__cost_price'),
        output_field=DecimalField(max_digits=16, decimal_places=2))
    products = products.annotate(
        total_stock=Coalesce(Sum('variants__branch_stocks__stock_count'), 0),
        variants_count=Count('variants', distinct=True),
        price_min=Min('variants__branch_stocks__sale_price'),
        price_max=Max('variants__branch_stocks__sale_price'),
        sale_val=Sum(_val),
        cost_val=Sum(_cost),
    )

    if stock_filter == 'zero':
        products = products.filter(total_stock=0)
    elif stock_filter == 'low':
        products = products.filter(total_stock__gt=0, total_stock__lte=3)
    elif stock_filter == 'in_stock':
        products = products.filter(total_stock__gt=0)

    if price_filter == 'zero':
        products = products.filter(total_stock__gt=0, default_sale_price=0).filter(
            Q(price_max__isnull=True) | Q(price_max=0))

    # Sort
    allowed_sorts = {
        'name': 'name', '-name': '-name',
        'code': 'code', '-code': '-code',
        'stock': 'total_stock', '-stock': '-total_stock',
        'price': 'default_sale_price', '-price': '-default_sale_price',
        'created': 'created_at', '-created': '-created_at',
        '-created_at': '-created_at',
        'margin': 'margin_calc', '-margin': '-margin_calc',
        'value': 'stock_value', '-value': '-stock_value',
    }
    # Marja va zaxira qiymati bo'yicha saralash uchun annotatsiya
    if sort in ('margin', '-margin', 'value', '-value'):
        products = products.annotate(
            _cv=Sum(ExpressionWrapper(
                F('variants__branch_stocks__stock_count') *
                F('variants__branch_stocks__cost_price'),
                output_field=DecimalField(max_digits=16, decimal_places=2))),
            _sv=Sum(ExpressionWrapper(
                F('variants__branch_stocks__stock_count') *
                F('variants__branch_stocks__sale_price'),
                output_field=DecimalField(max_digits=16, decimal_places=2))),
        ).annotate(
            stock_value=F('_cv'),
            margin_calc=Case(
                When(_cv__gt=0, then=(F('_sv') - F('_cv')) * 100.0 / F('_cv')),
                default=Value(None),
                output_field=FloatField(),
            ),
        )
    products = products.order_by(allowed_sorts.get(sort, '-created_at'))

    # Sahifalash — avval qattiq [:200] chegara bor edi va undan keyingi
    # mahsulotlar UMUMAN ko'rinmasdi (344 tadan 200 tasi).
    from django.core.paginator import Paginator
    try:
        per_page = int(request.GET.get('per') or 100)
    except (TypeError, ValueError):
        per_page = 100
    if per_page not in (50, 100, 200, 500):
        per_page = 100
    show_all = (request.GET.get('per') == 'all')
    if show_all:
        per_page = max(1, products.count())

    # Umumiy yig'indi va KPI — BUTUN ro'yxat bo'yicha (faqat shu sahifa emas).
    # Ilgari bular shablonda `products` ro'yxati ustidan hisoblanardi;
    # sahifalash qo'shilgach faqat joriy sahifani sanab qolardi.
    _all_totals = products.aggregate(
        v=Sum('variants_count'), u=Sum('total_stock'),
        n_in=Count('pk', filter=Q(total_stock__gt=0)),
        n_low=Count('pk', filter=Q(total_stock__gt=0, total_stock__lte=3)),
        n_out=Count('pk', filter=Q(total_stock=0)),
    )
    _pids_all = list(products.values_list('pk', flat=True))
    paginator = Paginator(products, per_page)
    page_obj = paginator.get_page(request.GET.get('page') or 1)
    products = list(page_obj.object_list)
    _pids = [p.id for p in products]
    list_totals = {
        'variants': _all_totals.get('v') or 0,
        'units': _all_totals.get('u') or 0,
    }
    # O'rtacha narx va o'rtacha marja — filtrlangan BUTUN ro'yxat bo'yicha.
    # Marja zaxira bilan tortilgan: qimmat tovar ko'p bo'lsa, o'rtacha ham
    # shuni aks ettiradi (oddiy o'rtacha aldab qo'yardi).
    _stock_qs = BranchStock.objects.filter(variant__product__in=_pids_all)
    _m = _stock_qs.aggregate(
        avg_price=Avg('sale_price', filter=Q(sale_price__gt=0)),
        cost_val=Sum(ExpressionWrapper(
            F('stock_count') * F('cost_price'),
            output_field=DecimalField(max_digits=16, decimal_places=2))),
        sale_val=Sum(ExpressionWrapper(
            F('stock_count') * F('sale_price'),
            output_field=DecimalField(max_digits=16, decimal_places=2))),
    )
    _cv = float(_m.get('cost_val') or 0)
    _sv = float(_m.get('sale_val') or 0)
    kpi = {
        'total': paginator.count,
        'in_stock': _all_totals.get('n_in') or 0,
        'low_stock': _all_totals.get('n_low') or 0,
        'out_stock': _all_totals.get('n_out') or 0,
        'avg_price': float(_m.get('avg_price') or 0),
        'avg_margin': ((_sv - _cv) / _cv * 100) if _cv > 0 else None,
    }

    # 30/365 kunlik sotuvlar — faqat ko'rsatiladigan mahsulotlar bo'yicha
    since_30d = timezone.now() - timedelta(days=30)
    sold_map = {r['variant__product_id']: r['qty'] for r in (
        Sale.objects.filter(sold_at__gte=since_30d,
                            variant__product_id__in=_pids)
        .values('variant__product_id').annotate(qty=Sum('quantity')))}
    since_365d = timezone.now() - timedelta(days=365)
    sold_365_map = {r['variant__product_id']: r['qty'] for r in (
        Sale.objects.filter(sold_at__gte=since_365d,
                            variant__product_id__in=_pids)
        .values('variant__product_id').annotate(qty=Sum('quantity')))}

    # Har mahsulotning tur ranglari (nomi yonida ko'rsatish uchun)
    color_rows = (ProductVariant.objects
                  .filter(product__in=products)
                  .exclude(color='')
                  .order_by('color')
                  .values_list('product_id', 'color')
                  .distinct())
    colors_map = {}
    for pid, color in color_rows:
        colors_map.setdefault(pid, []).append(color)

    # Attach velocity + days_left + turnover
    for p in products:
        p.variant_colors = colors_map.get(p.id, [])
        sold = sold_map.get(p.id, 0)
        p.sold_30d = sold
        daily_avg = sold / 30 if sold else 0
        p.days_left = (p.total_stock / daily_avg) if daily_avg else None
        # Annual turnover ≈ annual_sold / avg_stock. Use current stock as proxy.
        annual_sold = sold_365_map.get(p.id, 0)
        if p.total_stock and p.total_stock > 0:
            p.turnover = annual_sold / p.total_stock
        else:
            p.turnover = None
        # Real marja — ombordagi haqiqiy tannarx/narxdan (og'irlikli):
        # marja% = (sotuv qiymati − tannarx qiymati) / tannarx qiymati
        try:
            cost_val = float(p.cost_val or 0)
            sale_val = float(p.sale_val or 0)
            if cost_val > 0:
                p.unit_profit = sale_val - cost_val
                p.margin_percent = (sale_val - cost_val) / cost_val * 100
            else:
                # fallback: eski markup asosidagi taxmin
                m = float(p.markup_percent or 0)
                if m > 0 and p.default_sale_price:
                    cost = float(p.default_sale_price) / (1 + m / 100)
                    p.unit_profit = float(p.default_sale_price) - cost
                    p.margin_percent = (p.unit_profit /
                                        float(p.default_sale_price) * 100)
                else:
                    p.unit_profit = 0
                    p.margin_percent = 0
        except Exception:
            p.unit_profit = 0
            p.margin_percent = 0

    # Kategoriya ro'yxati: bo'lim tanlangan bo'lsa faqat o'sha bo'limnikini
    categories = Category.objects.select_related('group').order_by('name')
    if group_slug:
        categories = categories.filter(group__slug=group_slug)

    # 4 bo'lim + har biriga nechta mahsulot (filtr tugmalari uchun)
    _grp_counts = dict(
        Product.objects.exclude(is_open_price=True)
        .filter(category__group__isnull=False)
        .values_list('category__group__slug')
        .annotate(n=Count('pk')))
    groups = list(Group.objects.all())
    for g in groups:
        g.n_products = _grp_counts.get(g.slug, 0)

    # Sifat nazorati: sotuvda turgan, lekin narxi 0 bo'lgan tovarlar (0 so'mga ketadi!)
    zero_price_count = (Product.objects.exclude(is_open_price=True)
        .annotate(_stk=Coalesce(Sum('variants__branch_stocks__stock_count'), 0),
                  _pmax=Max('variants__branch_stocks__sale_price'))
        .filter(_stk__gt=0, default_sale_price=0)
        .filter(Q(_pmax__isnull=True) | Q(_pmax=0)).count())

    # Dublikat nomli mahsulotlar (birlashtirish tavsiyasi uchun)
    dup_groups = []
    _name_map = {}
    for pid, pname in Product.objects.exclude(is_open_price=True).values_list('id', 'name'):
        _name_map.setdefault(pname.strip().lower(), []).append(pid)
    for _n, _ids in _name_map.items():
        if len(_ids) > 1:
            dup_groups.append({'name': _n.title(), 'count': len(_ids),
                               'ids': ','.join(map(str, _ids))})
    dup_groups = dup_groups[:5]

    return render(request, 'inventory/product_list.html', {
        'products': products,
        'kpi': kpi,
        'page_obj': page_obj,
        'paginator': paginator,
        'per_page': 'all' if show_all else per_page,
        'dup_groups': dup_groups,
        'list_totals': list_totals,
        'q': q,
        'category_id': category_id,
        'stock_filter': stock_filter,
        'price_filter': price_filter,
        'zero_price_count': zero_price_count,
        'sort': sort,
        'categories': categories,
        'groups': groups,
        'group_slug': group_slug,
        'name_conflict': name_conflict,
        'name_conflict_count': Product.objects.exclude(is_open_price=True)
                                      .exclude(pending_name='').count(),
    })


@admin_required
def employee_debt_list(request):
    """Xodim qarzlari daftari — qo'shish, ro'yxat, 'to'landi' belgilash.

    Sotuvga kiritilmaydi (kassa buzilmaydi). Ojlik kuni to'landi belgilanadi.
    """
    from decimal import Decimal, InvalidOperation
    if request.method == 'POST':
        action = request.POST.get('action') or 'add'
        if action == 'pay':
            d = EmployeeDebt.objects.filter(pk=request.POST.get('pk') or 0,
                                            is_paid=False).first()
            if d:
                d.is_paid = True
                d.paid_at = timezone.now()
                d.paid_by = request.user
                # Naqd kassaga tushadi — ochiq smenga bog'laymiz (kutilgan naqd oshadi)
                open_shift = _open_shift_for(d.branch or getattr(request.user, 'branch', None))
                d.paid_shift = open_shift
                d.save(update_fields=['is_paid', 'paid_at', 'paid_by', 'paid_shift'])
                if open_shift:
                    messages.success(request,
                        f"\"{d.who}\" qarzi to'landi — {d.amount:,.0f} so'm kassaga qo'shildi.")
                else:
                    messages.success(request,
                        f"\"{d.who}\" qarzi to'landi deb belgilandi. (Ochiq smen yo'q — "
                        f"kassaga qo'shilmadi.)")
            return redirect('employee_debt_list')
        if action == 'delete':
            d = (EmployeeDebt.objects.filter(pk=request.POST.get('pk') or 0)
                 .prefetch_related('items__variant__product').first())
            if d:
                # MON-24: TO'LANGAN qarzни o'chirib bo'lmaydi. Aks holda tovar
                # ketgan, pul to'langan bo'lса ham — ombor tiklanib, to'langan
                # naqd smen hisobidan yo'qolib, qarz izsiz g'oyib bo'lardi.
                if d.is_paid:
                    messages.error(request,
                        "To'langan qarzni o'chirib bo'lmaydi (moliyaviy yozuv).")
                    return redirect('employee_debt_list')
                _restored = []
                with transaction.atomic():
                    for it in d.items.all():
                        # MON-25: FAQAT zaxira kamaytirilgan (ochiq narxsiz)
                        # tovar tiklanadi. Ochiq narxli qarzда zaxira
                        # kamaytirilmagan — tiklasak, yo'qdan zaxira "yaratilardi".
                        prod = it.variant.product if it.variant_id else None
                        if (it.variant_id and it.branch_id and prod
                                and not prod.is_open_price):
                            bs = BranchStock.objects.filter(
                                variant_id=it.variant_id, branch_id=it.branch_id).first()
                            if bs:
                                bs.stock_count = F('stock_count') + it.quantity
                                bs.save(update_fields=['stock_count'])
                                _restored.append(f"{prod.code}×{it.quantity}")
                    AuditLog.objects.create(
                        user=request.user, username_snapshot=request.user.username,
                        action=AuditLog.Action.DELETE, model_name='EmployeeDebt',
                        object_id=str(d.pk),
                        object_repr=f"Qarz o'chirildi: {d.who} — {int(d.amount)} so'm"[:300],
                        changes={'amount': float(d.amount), 'who': d.who,
                                 'restored_stock': _restored})
                    d.delete()
                messages.info(request, "Qarz yozuvi o'chirildi (ombor qoldig'i tiklandi).")
            return redirect('employee_debt_list')
        # ---- add ----
        emp_id = request.POST.get('employee') or ''
        emp_name = (request.POST.get('employee_name') or '').strip()[:120]
        note = (request.POST.get('note') or '').strip()[:200]

        # Tovarli qatorlar (skaner/tanlash) — [{stock_id, qty, price}]
        items = []
        raw = request.POST.get('items_json') or ''
        if raw:
            try:
                parsed = _json.loads(raw)
                if isinstance(parsed, list):
                    items = parsed
            except ValueError:
                items = []

        try:
            manual_amount = Decimal(str(request.POST.get('amount') or '0').replace(' ', ''))
        except (InvalidOperation, ValueError):
            manual_amount = Decimal('0')

        if not emp_id and not emp_name:
            messages.error(request, "Xodimni tanlang yoki ismini yozing.")
            return redirect('employee_debt_list')

        if not items and manual_amount <= 0:
            messages.error(request, "Tovar qo'shing yoki summani kiriting.")
            return redirect('employee_debt_list')

        branch = getattr(request.user, 'branch', None)
        try:
            with transaction.atomic():
                debt = EmployeeDebt.objects.create(
                    branch=branch,
                    employee_id=int(emp_id) if emp_id.isdigit() else None,
                    employee_name='' if emp_id.isdigit() else emp_name,
                    amount=Decimal('0'),
                    note=note,
                    created_by=request.user)

                total = Decimal('0')
                # Tovarli qatorlar: qoldiqni bloklab kamaytiramiz (SOTUV emas)
                sids = []
                for ln in items:
                    try:
                        sids.append(int(ln['stock_id']))
                    except (KeyError, ValueError, TypeError):
                        continue
                locked = {}
                if sids:
                    locked = {s.pk: s for s in BranchStock.objects
                              .select_for_update()
                              .select_related('variant__product', 'branch')
                              .filter(pk__in=sids)}
                for ln in items:
                    try:
                        sid = int(ln['stock_id']); qty = int(ln['qty'])
                        price = Decimal(str(ln.get('price') or 0))
                    except (KeyError, ValueError, TypeError, InvalidOperation):
                        raise ValueError("noto'g'ri qator")
                    if qty <= 0:
                        continue
                    stock = locked.get(sid)
                    if not stock:
                        raise ValueError(f"tovar {sid} topilmadi")
                    is_open = stock.variant.product.is_open_price
                    if not is_open and qty > stock.stock_count:
                        raise ValueError(
                            f"{stock.variant.product.code} "
                            f"{stock.variant.size}/{stock.variant.color}: "
                            f"omborda faqat {stock.stock_count} ta bor")
                    if not is_open:
                        stock.stock_count = F('stock_count') - qty
                        stock.save()
                    name = ' '.join(filter(None, [
                        stock.variant.product.name,
                        stock.variant.size, stock.variant.color])) or stock.variant.product.name
                    EmployeeDebtItem.objects.create(
                        debt=debt, variant=stock.variant, branch=stock.branch,
                        product_name=name[:200], quantity=qty, unit_price=price)
                    total += qty * price

                # Qo'lda summa (tovarsiz yoki qo'shimcha)
                if manual_amount > 0:
                    total += manual_amount

                # Decimal(12,2) — 10^10 dan oshsa DB overflow (500) bo'ladi
                if total > Decimal('9999999999.99'):
                    raise ValueError("summa juda katta")

                debt.amount = total
                debt.save(update_fields=['amount'])
        except ValueError as e:
            messages.error(request, f"Xatolik: {e}")
            return redirect('employee_debt_list')

        messages.success(request, "Qarz qo'shildi.")
        return redirect('employee_debt_list')

    debts = list(EmployeeDebt.objects.select_related('employee', 'created_by')
                 .prefetch_related('items')
                 .order_by('is_paid', '-created_at'))
    open_debts = [d for d in debts if not d.is_paid]
    paid_debts = [d for d in debts if d.is_paid][:50]
    # Xodim bo'yicha jami ochiq qarz
    per_emp = {}
    for d in open_debts:
        per_emp[d.who] = per_emp.get(d.who, 0) + float(d.amount)
    per_emp = sorted(per_emp.items(), key=lambda x: -x[1])
    total_open = sum(float(d.amount) for d in open_debts)
    from django.contrib.auth import get_user_model
    employees = get_user_model().objects.filter(is_active=True).order_by('username')
    return render(request, 'inventory/employee_debt_list.html', {
        'open_debts': open_debts, 'paid_debts': paid_debts,
        'per_emp': per_emp, 'total_open': total_open,
        'employees': employees,
    })


@admin_required
def product_resolve_name(request):
    """Nom ziddiyatini hal qilish: bir xil shtrix-kodga 2 xil nom bo'lsa,
    foydalanuvchi qaysi birini qoldirishni tanlaydi.

    choice = 'keep'  -> hozirgi nom qoladi
    choice = 'use'   -> pending_name (nakladnoy nomi) asosiy nom bo'ladi
    Ikkalasida ham pending_name bo'shatiladi (ziddiyat yopiladi).
    """
    if request.method != 'POST':
        return redirect('product_list')
    pk = request.POST.get('pk') or 0
    choice = (request.POST.get('choice') or '').strip()
    product = Product.objects.filter(pk=pk).first()
    if not product or not product.pending_name:
        return redirect(request.META.get('HTTP_REFERER') or 'product_list')
    old, alt = product.name, product.pending_name
    if choice == 'use':
        product.name = alt
    product.pending_name = ''
    product.save(update_fields=['name', 'pending_name'])
    AuditLog.objects.create(
        user=request.user, username_snapshot=request.user.username,
        action=AuditLog.Action.UPDATE, model_name='Product',
        object_id=str(product.pk),
        object_repr=(f"Nom ziddiyati hal qilindi: "
                     f"{'YANGI: ' + alt if choice == 'use' else 'ESKI: ' + old}"))
    messages.success(request, f"Nom tanlandi: \"{product.name}\".")
    return redirect(request.META.get('HTTP_REFERER') or 'product_list')


@admin_required
def product_create(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            _name = smart_title(form.cleaned_data.get('name'))
            product = form.save(commit=False)
            product.name = _name
            product.save()
            # Bir xil nomli boshqa mahsulot bo'lsa — jim birlashtirmaymiz,
            # faqat eslatib qo'yamiz (xohlasa qo'lda birlashtiradi)
            dup = (Product.objects.filter(name__iexact=_name)
                   .exclude(pk=product.pk).first())
            if dup:
                messages.info(
                    request,
                    f"Eslatma: '{_name}' nomli boshqa mahsulot ham bor "
                    f"({dup.code}). Agar bir xil bo'lsa, Mahsulotlar "
                    f"ro'yxatidan birlashtirishingiz mumkin.")
            messages.success(request, f"Mahsulot yaratildi. Kod: {product.code}")
            return redirect('product_detail', code=product.code)
    else:
        # URL prefill: /products/new/?barcode=4607034000234&name=Nike
        initial = {}
        if request.GET.get('barcode'):
            initial['external_barcode'] = request.GET.get('barcode').strip()
        if request.GET.get('name'):
            initial['name'] = request.GET.get('name').strip()
        form = ProductForm(initial=initial)
    return render(request, 'inventory/product_form.html', {
        'form': form, 'title': "Yangi mahsulot qo'shish",
    })


@admin_required
def product_delete(request, code):
    """Mahsulotni o'chirish (POST). Sotuv tarixi bo'lsa DB himoya qiladi
    (PROTECT) — aniq xabar bilan rad etiladi."""
    from django.db.models import ProtectedError
    if request.method != 'POST':
        return redirect('product_list')
    product = get_object_or_404(Product, code=normalize_code(code))
    name, pk = product.name, product.pk
    try:
        product.delete()
    except ProtectedError:
        messages.error(
            request,
            f"'{name}' o'chirilmadi: sotuv/transfer tarixi bor. Tarix "
            f"saqlanishi shart — o'rniga zaxirani 0 ga tushiring yoki "
            f"boshqa mahsulotga birlashtiring.")
        return redirect('product_list')
    AuditLog.objects.create(
        user=request.user,
        username_snapshot=request.user.username,
        action=AuditLog.Action.DELETE,
        model_name='Product',
        object_id=str(pk),
        object_repr=f'{code} — {name}',
    )
    messages.success(request, f"O'chirildi: {name} ({code}).")
    return redirect('product_list')


@admin_required
def variant_delete(request, pk):
    """Turni (rang/o'lcham) o'chirish. Sotuv/transfer/inventarizatsiya
    tarixi bo'lsa DB himoyasi (PROTECT) — aniq xabar bilan rad etiladi."""
    from django.db.models import ProtectedError
    v = get_object_or_404(
        ProductVariant.objects.select_related('product'), pk=pk)
    code = v.product.code
    branch_id = request.POST.get('branch') or ''
    label = f"{v.size or '—'} / {v.color or '—'}"
    back = f"{reverse('product_variants_edit', args=[code])}"
    if branch_id:
        back += f"?branch={branch_id}"
    if request.method != 'POST':
        return redirect(back)
    try:
        v.delete()
    except ProtectedError:
        messages.error(
            request,
            f"{label} o'chirilmadi: sotuv/transfer/inventarizatsiya tarixi "
            f"bor. Tarix saqlanishi shart — o'rniga omborni 0 qiling.")
        return redirect(back)
    AuditLog.objects.create(
        user=request.user,
        username_snapshot=request.user.username,
        action=AuditLog.Action.DELETE,
        model_name='ProductVariant', object_id=str(pk),
        object_repr=f'{code}: {label}')
    messages.success(request, f"O'chirildi: {label}")
    return redirect(back)


@admin_required
def product_variants_move(request, code):
    """Tanlangan turlarni boshqa mahsulotga ko'chirish."""
    product = get_object_or_404(Product, code=normalize_code(code))
    back = reverse('product_variants_edit', args=[product.code])
    if request.method != 'POST':
        return redirect(back)
    target_code = normalize_code(
        (request.POST.get('target_code') or '').strip().upper())
    target = Product.objects.filter(code=target_code).first()
    ids = [int(i) for i in request.POST.getlist('mv') if str(i).isdigit()]
    variants = list(product.variants.filter(pk__in=ids))
    if not variants:
        messages.warning(request, "Ko'chirish uchun tur tanlanmagan.")
        return redirect(back)
    if not target:
        messages.error(request, "Maqsad mahsulot topilmadi — ro'yxatdan tanlang.")
        return redirect(back)
    if target.pk == product.pk:
        messages.warning(request, "Bir xil mahsulot tanlangan.")
        return redirect(back)
    moved, skipped = 0, []
    with transaction.atomic():
        for v in variants:
            if target.variants.filter(size=v.size, color=v.color).exists():
                skipped.append(f"{v.size or '—'}/{v.color or '—'}")
                continue
            v.product = target
            v.save(update_fields=['product'])
            moved += 1
        if moved:
            AuditLog.objects.create(
                user=request.user,
                username_snapshot=request.user.username,
                action=AuditLog.Action.UPDATE,
                model_name='ProductVariant', object_id=str(product.pk),
                object_repr=(f"{moved} ta tur {product.code} -> "
                             f"{target.code}")[:300])
    if moved:
        messages.success(
            request,
            f"{moved} ta tur '{target.name}' ({target.code}) ga ko'chirildi.")
    if skipped:
        messages.warning(
            request,
            f"O'tkazilmadi (maqsadda shu o'lcham/rang bor): {', '.join(skipped)}")
    return redirect(back)


@admin_required
def product_merge(request):
    """Bir nechta mahsulotni bittasiga birlashtirish.

    Tanlanganlarning variantlari, ombor qoldiqlari, sotuv/qabul/transfer
    tarixi asosiy (target) mahsulotga ko'chadi. Manba mahsulotning
    external_barcode'i (agar bitta varianti bo'lsa) o'sha variantning
    shtrix-kodiga aylanadi — skanerlash ishlashda davom etadi.
    """
    ids_raw = (request.GET.get('ids') or request.POST.get('ids') or '')
    ids = [int(i) for i in ids_raw.split(',') if i.strip().isdigit()]
    products = list(Product.objects.filter(id__in=ids)
                    .prefetch_related('variants'))
    if len(products) < 2:
        messages.warning(request, "Birlashtirish uchun kamida 2 ta mahsulot tanlang.")
        return redirect('product_list')

    # Variantlari eng ko'p mahsulot — default target
    products.sort(key=lambda p: (-p.variants.count(), p.created_at))
    default_target = products[0]

    if request.method == 'POST' and request.POST.get('confirm') == '1':
        try:
            target = next(p for p in products
                          if str(p.pk) == request.POST.get('target'))
        except StopIteration:
            messages.error(request, "Asosiy mahsulot noto'g'ri tanlangan.")
            return redirect('product_list')
        sources = [p for p in products if p.pk != target.pk]

        from django.db.models import ProtectedError
        moved_variants = 0
        try:
          with transaction.atomic():
            for src in sources:
                src_variants = list(src.variants.all())
                single = len(src_variants) == 1
                for v in src_variants:
                    tv = target.variants.filter(
                        size=v.size, color=v.color).first()
                    if tv:
                        # Bir xil tur — variant darajasida birlashtiramiz
                        for bs in list(v.branch_stocks.all()):
                            tbs = BranchStock.objects.filter(
                                variant=tv, branch=bs.branch).first()
                            if tbs:
                                tbs.stock_count += bs.stock_count
                                if not tbs.sale_price:
                                    tbs.sale_price = bs.sale_price
                                if not tbs.cost_price:
                                    tbs.cost_price = bs.cost_price
                                if not tbs.wholesale_price:
                                    tbs.wholesale_price = bs.wholesale_price
                                tbs.save()
                                bs.delete()
                            else:
                                bs.variant = tv
                                bs.save(update_fields=['variant'])
                        Sale.objects.filter(variant=v).update(variant=tv)
                        Intake.objects.filter(variant=v).update(variant=tv)
                        TransferLine.objects.filter(variant=v).update(variant=tv)
                        StocktakeCount.objects.filter(variant=v).update(variant=tv)
                        # R5: StockWriteOff.variant PROTECT — qayta bog'lamasak
                        # v.delete() ProtectedError berib butun birlashuv buzilardi.
                        StockWriteOff.objects.filter(variant=v).update(variant=tv)
                        if v.barcode and not tv.barcode:
                            bc = v.barcode
                            v.barcode = None
                            v.save(update_fields=['barcode'])
                            tv.barcode = bc
                            tv.save(update_fields=['barcode'])
                        v.delete()
                    else:
                        v.product = target
                        if (single and src.external_barcode and not v.barcode
                                and not ProductVariant.objects.filter(
                                    barcode=src.external_barcode).exists()):
                            v.barcode = src.external_barcode
                        v.save()
                    moved_variants += 1
                if target.external_barcode is None and src.external_barcode \
                        and not single:
                    # ko'p variantli manba barcode'i variantga bog'lanmadi —
                    # yo'qolmasligi uchun targetga o'tkazamiz
                    bc = src.external_barcode
                    src.external_barcode = None
                    src.save(update_fields=['external_barcode'])
                    target.external_barcode = bc
                    target.save(update_fields=['external_barcode'])
                src.delete()
            AuditLog.objects.create(
                user=request.user,
                username_snapshot=request.user.username,
                action=AuditLog.Action.UPDATE,
                model_name='Product',
                object_id=str(target.pk),
                object_repr=(f"Birlashtirildi -> {target.code}: "
                             + ', '.join(x.code for x in sources))[:300],
            )
        except ProtectedError:
            # R5: kutilmagan PROTECT bog'lanish (masalan onlayn buyurtma qatori)
            # — butun birlashuv orqaga qaytadi, ma'lumot buzilmaydi.
            messages.error(request,
                "Birlashtirib bo'lmadi: bu turlarga bog'liq himoyalangan "
                "yozuvlar bor. Avval ularni ko'rib chiqing.")
            return redirect('product_list')
        messages.success(
            request,
            f"{len(sources)} ta mahsulot '{target.name}' ({target.code}) ga "
            f"birlashtirildi — {moved_variants} ta tur ko'chdi.")
        return redirect('product_detail', code=target.code)

    # Confirm sahifasi
    rows = []
    for p in products:
        rows.append({
            'product': p,
            'variants_count': p.variants.count(),
            'stock': p.total_stock(),
        })
    return render(request, 'inventory/product_merge.html', {
        'rows': rows,
        'ids': ','.join(str(p.pk) for p in products),
        'default_target': default_target,
    })


@admin_required
def product_search_suggest(request):
    """Mahsulotlar sahifasi qidiruvi uchun jonli takliflar:
    mahsulot nomi/kodi, rang, o'lcham, shtrix-kod."""
    q = (request.GET.get('q') or '').strip()
    if len(q) < 2:
        return JsonResponse({'results': []})
    res = []
    for pr in (Product.objects
               .filter(Q(name__icontains=q) | Q(code__icontains=q.upper()) |
                       Q(brand__icontains=q))
               .order_by('name')[:6]):
        res.append({'t': 'mahsulot', 'label': pr.name, 'sub': pr.code,
                    'q': pr.name})
    for br in (Product.objects.filter(brand__icontains=q)
               .exclude(brand='').values_list('brand', flat=True)
               .distinct().order_by('brand')[:4]):
        res.append({'t': 'brend', 'label': br, 'sub': '', 'q': br})
    for cl in (ProductVariant.objects.filter(color__icontains=q)
               .values_list('color', flat=True).distinct()
               .order_by('color')[:4]):
        res.append({'t': 'rang', 'label': cl, 'sub': '', 'q': cl})
    for sz in (ProductVariant.objects.filter(size__icontains=q)
               .values_list('size', flat=True).distinct()
               .order_by('size')[:3]):
        res.append({'t': "o'lcham", 'label': sz, 'sub': '', 'q': sz})
    for bc in (ProductVariant.objects.filter(barcode__istartswith=q)
               .values_list('barcode', flat=True)[:2]):
        res.append({'t': 'shtrix', 'label': bc, 'sub': '', 'q': bc})
    return JsonResponse({'results': res[:12]})


@admin_required
def product_search_for_attach(request):
    """GET /products/search-for-attach/?q=&exclude_with_barcode=1
    Lightweight JSON product search for the "attach EAN to existing
    product" picker. Returns up to 12 matches by name or code."""
    q = (request.GET.get('q') or '').strip()
    if not q:
        return JsonResponse({'results': []})
    # Avval nomi shu harf(lar) bilan BOSHLANADIGANLAR, keyin ichida
    # uchraydiganlar — "F" yozilganda Fructis birinchi chiqadi
    base = Product.objects.all()
    if request.GET.get('exclude_with_barcode') == '1':
        base = base.filter(external_barcode__isnull=True)
    starts = list(base.filter(name__istartswith=q).order_by('name')[:12])
    qs = starts
    if len(qs) < 12 and len(q) >= 2:
        extra = (base.filter(Q(name__icontains=q) | Q(code__icontains=q.upper()))
                 .exclude(pk__in=[x.pk for x in starts])
                 .order_by('name')[:12 - len(qs)])
        qs = qs + list(extra)
    return JsonResponse({'results': [
        {
            'code': p.code,
            'name': p.name,
            'external_barcode': p.external_barcode or '',
            'category': p.category.name if p.category else '',
        } for p in qs
    ]})


@admin_required
def product_attach_barcode(request):
    """POST /products/attach-barcode/  body: code, barcode
    Attach an external_barcode to an existing product. Refuses if the
    barcode is already attached elsewhere (the unique constraint would
    raise IntegrityError; we want a clean JSON error)."""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST required'}, status=405)
    code = normalize_code((request.POST.get('code') or '').upper())
    barcode = (request.POST.get('barcode') or '').strip()
    if not code or not barcode:
        return JsonResponse({'ok': False, 'error': 'code va barcode kerak'},
                            status=400)
    try:
        product = Product.objects.get(code=code)
    except Product.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Mahsulot topilmadi'},
                            status=404)
    clash = (Product.objects
             .filter(external_barcode=barcode)
             .exclude(pk=product.pk)
             .first())
    if clash:
        return JsonResponse({
            'ok': False,
            'error': f"Bu barcode '{clash.name}' ({clash.code}) ga "
                     f"allaqachon biriktirilgan.",
        }, status=409)
    product.external_barcode = barcode
    product.save(update_fields=['external_barcode'])
    return JsonResponse({
        'ok': True,
        'product': {'code': product.code, 'name': product.name,
                    'external_barcode': product.external_barcode},
    })


@admin_required
def product_detail(request, code):
    product = get_object_or_404(Product, code=normalize_code(code))
    # Zaxira yozuvlari bilan birga — har tur uchun alohida so'rov
    # yubormaslik uchun (ilgari sahifada 45 ta so'rov bo'lardi)
    variants = list(product.variants.all()
                    .prefetch_related('branch_stocks__branch'))
    sizes = sorted({v.size for v in variants}, key=lambda s: (len(s), s))
    colors = sorted({v.color for v in variants})

    # Bitta so'rov — barcha filial zaxiralari (filial boshiga alohida
    # so'rov o'rniga; 20 filialda 20 -> 1 query)
    all_stocks = list(BranchStock.objects.filter(variant__product=product)
                      .select_related('variant', 'branch'))
    stocks_by_branch = {}
    for st in all_stocks:
        stocks_by_branch.setdefault(st.branch_id, []).append(st)

    branches_data = []
    for br in Branch.objects.filter(is_active=True):
        stocks = stocks_by_branch.get(br.id, [])
        matrix = {(s.variant.size, s.variant.color): s for s in stocks}
        total = sum(s.stock_count for s in stocks)
        branches_data.append({
            'branch': br, 'matrix': matrix,
            'sizes': sizes, 'colors': colors, 'total': total,
        })

    # Yassi turlar ro'yxati (shtrix-kod + narx + jami ombor)
    stocks_by_variant = {}
    for st in all_stocks:
        stocks_by_variant.setdefault(st.variant_id, []).append(st)
    variant_rows = []
    for v in variants:
        sts = stocks_by_variant.get(v.pk, [])
        prices = [st.sale_price for st in sts if st.sale_price]
        # Tannarx va marja — turlar jadvalida ko'rinsin (ilgari faqat
        # "Turlarni tahrirlash" sahifasida ko'rinardi)
        _costs = [st.cost_price for st in sts if st.cost_price]
        _cv = sum(float(st.stock_count) * float(st.cost_price) for st in sts)
        _sv = sum(float(st.stock_count) * float(st.sale_price) for st in sts)
        if _cv > 0:
            _marja = (_sv - _cv) / _cv * 100
        elif _costs and prices:
            _c = float(min(_costs)); _p = float(min(prices))
            _marja = ((_p - _c) / _c * 100) if _c > 0 else None
        else:
            _marja = None
        variant_rows.append({
            'variant': v,
            'stock_total': sum(st.stock_count for st in sts),
            'price_min': min(prices) if prices else None,
            'price_max': max(prices) if prices else None,
            'cost_min': min(_costs) if _costs else None,
            'cost_max': max(_costs) if _costs else None,
            'marja': _marja,
        })
    _all_prices = [st.sale_price for st in all_stocks if st.sale_price]
    price_min = min(_all_prices) if _all_prices else None
    price_max = max(_all_prices) if _all_prices else None

    recent_intakes = Intake.objects.filter(variant__product=product) \
        .select_related('variant', 'branch', 'received_by').order_by('-received_at')[:20]

    # 30-day sales chart + KPIs
    since_30d = timezone.now() - timedelta(days=30)
    rev_expr = ExpressionWrapper(
        F('quantity') * F('sale_price') - F('line_discount'),
        output_field=DecimalField(max_digits=14, decimal_places=2)
    )
    sales_30d = Sale.objects.filter(variant__product=product, sold_at__gte=since_30d)
    agg = sales_30d.aggregate(
        qty=Sum('quantity'),
        rev=Sum(rev_expr),
        txns=Count('transaction', distinct=True),
    )
    sold_30d = agg['qty'] or 0
    # SAL-6: MAHSULOT — qator darajasidagi o'lchov, chek chegirmasi bu tovarga
    # tegishli emas (egasining qoidasi), shuning uchun rev_30d O'Z narxida
    # qoladi. Chegirma alohida ko'rsatiladi — shunda sahifa "bu mahsulot
    # 300 000 keltirdi" deganда, chekларида 60 000 chegirma borligi ham ko'rinadi.
    rev_30d = float(agg['rev'] or 0)
    _odisc_30d, _ = _order_discount_share(sales_30d)
    _, _, _disc_split = _order_discount_share(sales_30d, split=True)
    order_discount_30d = float(_odisc_30d)
    txns_30d = agg['txns'] or 0
    daily_avg = sold_30d / 30 if sold_30d else 0
    # Single aggregate covers both total stock + total inventory value (was 2 separate queries)
    inv_agg = BranchStock.objects.filter(variant__product=product).aggregate(
        s=Sum('stock_count'),
        v=Sum(ExpressionWrapper(F('stock_count') * F('cost_price'),
                                output_field=DecimalField(max_digits=14, decimal_places=2))),
        sv=Sum(ExpressionWrapper(F('stock_count') * F('sale_price'),
                                 output_field=DecimalField(max_digits=14, decimal_places=2))),
    )
    total_stock = inv_agg['s'] or 0
    total_value = float(inv_agg['v'] or 0)

    # Marja — HAQIQIY tannarx/sotuv bo'yicha hisoblanadi.
    # Ilgari sarlavhada `product.markup_percent` ko'rsatilardi: u saqlanadigan
    # maydon bo'lib, faqat ayrim qabul yo'llarida yozilardi — shuning uchun
    # tannarx 10 200 / sotuv 13 000 bo'lsa ham 0.00% chiqib turardi.
    _cost_val = float(inv_agg['v'] or 0)
    _sale_val = float(inv_agg['sv'] or 0)
    if _cost_val > 0:
        real_markup = (_sale_val - _cost_val) / _cost_val * 100
    else:
        _avg = BranchStock.objects.filter(variant__product=product).aggregate(
            c=Avg('cost_price'), sp=Avg('sale_price'))
        _c = float(_avg['c'] or 0); _sp = float(_avg['sp'] or 0)
        real_markup = ((_sp - _c) / _c * 100) if _c > 0 else None

    # O'rtacha sotuv narxi — zaxira bilan tortilgan (ko'p turgan tur ko'proq
    # ta'sir qiladi). Zaxira bo'lmasa — oddiy o'rtacha.
    if total_stock > 0 and _sale_val > 0:
        avg_price = _sale_val / total_stock
    else:
        avg_price = float(BranchStock.objects
                          .filter(variant__product=product, sale_price__gt=0)
                          .aggregate(a=Avg('sale_price'))['a'] or 0)
    days_left = (total_stock / daily_avg) if daily_avg else None

    # Daily chart data
    from collections import OrderedDict
    chart = OrderedDict()
    today = timezone.localdate()
    for i in range(29, -1, -1):
        d = today - timedelta(days=i)
        chart[d.isoformat()] = 0
    daily_rows = (sales_30d.values_list('sold_at', 'quantity'))
    for sold_at, qty in daily_rows:
        d = sold_at.astimezone(timezone.get_current_timezone()).date()
        key = d.isoformat()
        if key in chart:
            chart[key] += qty
    chart_labels = [d[5:] for d in chart.keys()]  # MM-DD
    chart_qty = list(chart.values())

    # Top branch for this product (by 30-day revenue)
    top_branch_row = (sales_30d.values('branch__name')
                      .annotate(r=Sum(rev_expr))
                      .order_by('-r').first())
    top_branch_name = top_branch_row['branch__name'] if top_branch_row else None
    top_branch_rev = float(top_branch_row['r'] or 0) if top_branch_row else 0

    # Variants summary
    variants_qs = product.variants.all()
    total_variants = variants_qs.count()
    variants_in_stock = (BranchStock.objects
                         .filter(variant__product=product, stock_count__gt=0)
                         .values('variant').distinct().count())

    product_kpis = {
        'total_stock': total_stock,
        'total_value': total_value,
        'sold_30d': sold_30d,
        'rev_30d': rev_30d,
        'order_discount_30d': order_discount_30d,
        'txns_30d': txns_30d,
        'days_left': days_left,
        'top_branch_name': top_branch_name,
        'top_branch_rev': top_branch_rev,
        'total_variants': total_variants,
        'variants_in_stock': variants_in_stock,
    }

    # D5: Cross-sell with confidence + lift (association rules)
    # Confidence P(B|A) = co_count / |A|   — how often B appears when A buys
    # Lift = P(B|A) / P(B)                  — strength of association (1.0 = independent)
    # Sort by lift to surface unusually-strong pairings, not just popular pairs.
    total_txns = SaleTransaction.objects.count() or 1
    txn_ids = list(Sale.objects.filter(variant__product=product)
                   .values_list('transaction_id', flat=True).distinct())
    a_count = len(txn_ids)
    cross_sell = []
    if a_count > 0:
        co_rows = (Sale.objects
                   .filter(transaction_id__in=txn_ids)
                   .exclude(variant__product=product)
                   .values('variant__product_id',
                           'variant__product__code',
                           'variant__product__name')
                   .annotate(co_count=Count('transaction', distinct=True))
                   .order_by('-co_count')[:30])  # candidate pool
        # B-only base counts (how often each candidate B appears overall)
        b_ids = [r['variant__product_id'] for r in co_rows]
        b_counts = {r['variant__product_id']: r['c'] for r in (
            Sale.objects.filter(variant__product_id__in=b_ids)
            .values('variant__product_id')
            .annotate(c=Count('transaction', distinct=True))
        )}
        for r in co_rows:
            pid = r['variant__product_id']
            co = r['co_count']
            b = b_counts.get(pid, 0) or 1
            confidence = co / a_count
            lift = confidence / (b / total_txns) if (b / total_txns) > 0 else 0
            cross_sell.append({
                'code': r['variant__product__code'],
                'name': r['variant__product__name'],
                'co_count': co,
                'confidence': round(confidence * 100, 1),
                'lift': round(lift, 2),
            })
        # Sort by lift (descending), keep min 3 co-occurrences to avoid noise
        cross_sell = [c for c in cross_sell if c['co_count'] >= 2]
        cross_sell.sort(key=lambda c: -c['lift'])
        cross_sell = cross_sell[:5]

    # --- Partiyalar: har tur bo'yicha turli tannarxda kelgan qabullar ---
    # Bir tur bir necha marta, HAR XIL tannarxda kelgan bo'lishi mumkin.
    # BranchStock esa bitta tannarx/narx saqlaydi — ya'ni oxirgi qabul
    # avvalgilarini bosib ketadi. Shu yerda haqiqiy qabul tarixini
    # ko'rsatamiz va kerak bo'lsa alohida shtrix-kodga ajratishni taklif
    # qilamiz.
    _batch_map = {}
    for ink in (Intake.objects
                .filter(variant__product=product, is_return=False)
                .select_related('branch')
                .order_by('variant_id', 'received_at')):
        _batch_map.setdefault((ink.variant_id, ink.branch_id), []).append(ink)

    # Zaxira yozuvlari yuqorida allaqachon o'qilgan (all_stocks) — har tur
    # uchun qaytadan so'rov yubormaymiz (bu yerda 22 ta ortiqcha so'rov edi)
    _stocks_by_variant = {}
    for _st in all_stocks:
        _stocks_by_variant.setdefault(_st.variant_id, []).append(_st)

    batch_groups = []
    for row in variant_rows:
        v = row.get('variant') if isinstance(row, dict) else getattr(row, 'variant', None)
        if v is None:
            continue
        for st in _stocks_by_variant.get(v.pk, []):
            inks = _batch_map.get((v.pk, st.branch_id)) or []
            # AJRATISH SOTUV narxiga qarab bo'ladi (tannarxga emas) — shuning
            # uchun faqat sotuv narxi HAR XIL bo'lgan qabullar ko'rsatiladi.
            # Tannarx tebranib, sotuv narxi bir xil bo'lsa — ajratadigan narsa
            # yo'q, kartochka bezovta qilmaydi.
            sales = {i.sale_price for i in inks if i.sale_price}
            if len(inks) < 2 or len(sales) < 2:
                continue
            batch_groups.append({
                'variant': v, 'branch': st.branch, 'stock': st,
                'intakes': inks,
                'sale_min': min(sales), 'sale_max': max(sales),
            })

    # ---- Harakatlar tarixi (log): har tur bo'yicha kirim/sotuv/qaytarish/qoldiq ----
    _vids = []
    for row in variant_rows:
        v = row.get('variant') if isinstance(row, dict) else getattr(row, 'variant', None)
        if v is not None:
            _vids.append(v.pk)
    _intk = {r['variant']: r['s'] for r in Intake.objects.filter(
        variant_id__in=_vids).values('variant').annotate(s=Sum('quantity'))}
    _sold = {r['variant']: r['s'] for r in Sale.objects.filter(
        variant_id__in=_vids).values('variant').annotate(s=Sum('quantity'))}
    _retn = {r['sale__variant']: r['s'] for r in Return.objects.filter(
        sale__variant_id__in=_vids).values('sale__variant').annotate(s=Sum('quantity'))}
    _cur = {r['variant']: r['s'] for r in BranchStock.objects.filter(
        variant_id__in=_vids).values('variant').annotate(s=Sum('stock_count'))}
    # STK-1: ombor tenglamasiga hisobdan chiqarish, ishchi qarzi va YO'LDAGI
    # ko'chirishlarни ham qo'shamiz. Aks holda ko'chirilgan/yo'qotilgan tovar
    # "farq" bo'lib ko'rinardi. (Ledger kompaniya bo'yicha — qabul qilingan
    # ko'chirish manba−/manzil+ o'zaro yo'qoladi; faqat YO'LDAGISI kamaytiradi.)
    _wo = {r['variant']: r['s'] for r in StockWriteOff.objects.filter(
        variant_id__in=_vids).values('variant').annotate(s=Sum('quantity'))}
    _dbt = {r['variant']: r['s'] for r in EmployeeDebtItem.objects.filter(
        variant_id__in=_vids).values('variant').annotate(s=Sum('quantity'))}
    _tin = {r['variant']: r['s'] for r in TransferLine.objects.filter(
        variant_id__in=_vids, transfer__status=Transfer.Status.IN_TRANSIT
        ).values('variant').annotate(s=Sum('quantity'))}
    _is_open = product.is_open_price
    movement_rows = []
    for row in variant_rows:
        v = row.get('variant') if isinstance(row, dict) else getattr(row, 'variant', None)
        if v is None:
            continue
        ki = _intk.get(v.pk, 0) or 0
        so = _sold.get(v.pk, 0) or 0
        rq = _retn.get(v.pk, 0) or 0
        cu = _cur.get(v.pk, 0) or 0
        wo = _wo.get(v.pk, 0) or 0
        # ochiq narxli tovar sotuv/qarzда zaxira KAMAYTIRMAYDI — bu hadlarni 0
        db = 0 if _is_open else (_dbt.get(v.pk, 0) or 0)
        tin = _tin.get(v.pk, 0) or 0
        if not (ki or so or cu or rq or wo or db or tin):
            continue
        movement_rows.append({
            'variant': v, 'intake': ki, 'sold': so, 'returned': rq, 'current': cu,
            'writeoff': wo, 'debt': db, 'in_transit': tin,
            'balanced': (ki - so + rq - wo - db - tin) == cu,
        })
    movement_rows.sort(key=lambda r: (not r['balanced'], -(r['current'] or 0)))

    return render(request, 'inventory/product_detail.html', {
        'discount_split': _disc_split,   # DISC-1
        'product': product, 'branches_data': branches_data,
        'real_markup': real_markup,
        'avg_price': avg_price,
        'recent_intakes': recent_intakes,
        'movement_rows': movement_rows,
        'product_kpis': product_kpis,
        'variant_rows': variant_rows,
        'batch_groups': batch_groups,
        'variant_ids': ','.join(str(v.pk) for v in variants),
        'price_min': price_min,
        'price_max': price_max,
        'chart_labels': chart_labels,
        'chart_qty': chart_qty,
        'cross_sell': cross_sell,
    })


@admin_required
def product_edit(request, code):
    product = get_object_or_404(Product, code=normalize_code(code))
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, 'Mahsulot yangilandi.')
            return redirect('product_detail', code=product.code)
    else:
        form = ProductForm(instance=product)
    return render(request, 'inventory/product_form.html', {
        'form': form, 'title': f'Tahrirlash — {product.name}', 'product': product,
    })


# ---------- INTAKE ----------

def _split_csv(text):
    """Split a comma- or newline-separated string into trimmed tokens."""
    if not text:
        return []
    parts = []
    for chunk in text.replace('\n', ',').split(','):
        c = chunk.strip()
        if c and c not in parts:
            parts.append(c)
    return parts


@admin_required
def intake_for_product(request, code):
    """Bulk intake: matrix of (size × color) → counts in one form."""
    from decimal import Decimal, InvalidOperation
    product = get_object_or_404(Product, code=normalize_code(code))

    if request.method == 'POST':
        try:
            branch_id = int(request.POST.get('branch') or 0)
            branch = Branch.objects.filter(pk=branch_id, is_active=True).first()
            if not branch:
                messages.error(request, "Filial tanlang.")
                return redirect('intake_for_product', code=product.code)

            # V3: min="0" faqat brauzerда — devtools orqali chetlab o'tilардi va
            # MANFIY cost weighted_cost orqali cost_price'ni manfiyга tortib,
            # o'sha variantning butun marja hisobini buzardi. Shu bois pul
            # maydonlarини SERVERда tekshiramiz (IntakeForm bilan bir xil chegara:
            # min 0, max_digits 12/6). Noto'g'ri bo'lsa ValueError → pastda ushlanib
            # "Maydonlarni tekshiring" xabari chiqadi, hech narsa saqlanmaydi.
            def _money(field, max_int_digits=10):
                raw = (request.POST.get(field) or '').strip()
                for ch in (' ', ' ', ' ', "'"):
                    raw = raw.replace(ch, '')
                raw = raw.replace(',', '.')
                if raw == '':
                    return Decimal('0')
                try:
                    d = Decimal(raw)
                except InvalidOperation:
                    raise ValueError(f"{field} — raqam emas")
                if (not d.is_finite() or d < 0
                        or abs(d) >= (Decimal(10) ** max_int_digits)):
                    raise ValueError(f"{field} — noto'g'ri qiymat")
                return d

            cost = _money('cost_per_unit')
            markup = _money('markup_percent', max_int_digits=4)
            sale_price_raw = (request.POST.get('sale_price') or '').strip()
            if sale_price_raw:
                sale_price = _money('sale_price')
            else:
                sale_price = (cost * (1 + markup / 100)).quantize(Decimal('1'))
            # STK-11: tannarx 0 (bonus/tekin tovar yoki o'qilmagan faktura)
            # bo'lsa, sotuv narxi 0 chiqadi va tovar KASSADA TEKIN o'tardi.
            # 0 bo'lsa katalog narxiga qaytamiz; u ham 0 bo'lsa — pastda
            # stock.sale_price umuman o'zgartirilmaydi (0 yozilmaydi).
            if sale_price <= 0:
                sale_price = product.default_sale_price or Decimal('0')
            wholesale_price = _money('wholesale_price')

            supplier = (request.POST.get('supplier') or '').strip()
            note = (request.POST.get('note') or '').strip()
            update_price = bool(request.POST.get('update_product_price'))

            # Iterate matrix cells. Inputs are named qty[<size>|<color>].
            total_qty = 0
            variants_touched = 0
            with transaction.atomic():
                for key, value in request.POST.items():
                    if not key.startswith('qty[') or not key.endswith(']'):
                        continue
                    payload = key[4:-1]
                    if '|' not in payload:
                        continue
                    size, color = payload.split('|', 1)
                    size = size.strip()
                    color = color.strip()
                    if not size or not color:
                        continue
                    try:
                        qty = int(value)
                    except (ValueError, TypeError):
                        continue
                    if qty <= 0:
                        continue
                    variant, _ = ProductVariant.objects.get_or_create(
                        product=product, size=size, color=color,
                    )
                    # Sotuv narxi boshqacha bo'lsa — alohida tur (o'z kodi bilan)
                    variant = resolve_price_variant(variant, branch, sale_price)
                    stock, _ = BranchStock.objects.get_or_create(
                        variant=variant, branch=branch,
                        defaults={'cost_price': cost, 'sale_price': sale_price,
                                  'wholesale_price': wholesale_price},
                    )
                    stock.cost_price = weighted_cost(  # STK-8 weighted-average
                        stock.stock_count if isinstance(stock.stock_count, int) else 0,
                        stock.cost_price, qty, cost)
                    stock.stock_count = F('stock_count') + qty
                    if sale_price > 0:      # STK-11: 0 narx yozib qo'ymaymiz
                        stock.sale_price = sale_price
                    if wholesale_price > 0:
                        stock.wholesale_price = wholesale_price
                    stock.save()
                    Intake.objects.create(
                        variant=variant, branch=branch,
                        quantity=qty, cost_per_unit=cost, sale_price=sale_price,
                        supplier=supplier, note=note,
                        received_by=request.user,
                    )
                    total_qty += qty
                    variants_touched += 1

                if update_price and sale_price > 0:
                    product.default_sale_price = sale_price
                    product.markup_percent = markup
                    product.save()
                    BranchStock.objects.filter(
                        variant__product=product, sale_price=0
                    ).update(sale_price=sale_price)

            if variants_touched == 0:
                messages.warning(request,
                    "Hech qaysi katakka son kiritilmadi — saqlanmadi.")
            else:
                messages.success(request,
                    f"Qabul saqlandi: {branch.name}ga {variants_touched} ta "
                    f"variant uchun jami {total_qty} dona.")
                return redirect('product_detail', code=product.code)
        except (ValueError, TypeError) as e:
            messages.error(request, f"Maydonlarni tekshiring: {e}")

    # Cost history: last 5 intakes for this product
    last_intakes = list(Intake.objects.filter(variant__product=product)
                        .select_related('variant', 'branch', 'supplier_ref')
                        .order_by('-received_at')[:5])

    # Pre-fill sizes/colors from existing variants if present
    variants = list(product.variants.all())
    existing_sizes = sorted({v.size for v in variants}, key=lambda s: (len(s), s))
    existing_colors = sorted({v.color for v in variants})

    return render(request, 'inventory/intake_form.html', {
        'product': product,
        'branches': Branch.objects.filter(is_active=True),
        'existing_sizes': existing_sizes,
        'existing_colors': existing_colors,
        'last_intakes': last_intakes,
        'suppliers': Supplier.objects.filter(is_active=True).order_by('name'),
    })


@admin_required
def product_variants_edit(request, code):
    """Mahsulot turlarini jadvalda tahrirlash — hamma maydon bir joyda.

    Variant maydonlari (rang, o'lcham, shtrix-kod) filialdan mustaqil;
    narxlar va ombor soni tanlangan filial bo'yicha (BranchStock).
    Ombor soni o'zgartirilsa — farq Intake yozuvi sifatida saqlanadi
    (qo'lda tuzatish) va AuditLog'ga tushadi.
    """
    from decimal import Decimal, InvalidOperation

    product = get_object_or_404(Product, code=normalize_code(code))
    branches = Branch.objects.filter(is_active=True)
    try:
        branch = branches.get(pk=int(request.GET.get('branch')
                                     or request.POST.get('branch') or 0))
    except (Branch.DoesNotExist, ValueError, TypeError):
        branch = branches.first()
    if branch is None:
        messages.error(request, "Faol filial yo'q.")
        return redirect('product_detail', code=product.code)

    _dec = parse_dec

    if request.method == 'POST':
        errors = []
        ids = request.POST.getlist('v_id')
        colors = request.POST.getlist('v_color')
        sizes = request.POST.getlist('v_size')
        barcodes = request.POST.getlist('v_barcode')
        costs = request.POST.getlist('v_cost')
        sales = request.POST.getlist('v_sale')
        wholesales = request.POST.getlist('v_wholesale')
        stocks = request.POST.getlist('v_stock')

        variants = {v.pk: v for v in product.variants.all()}
        # V2: bo'sh katak = "o'zgarmadi". Har bir mavjud tur uchun joriy
        # BranchStock qiymatlarini oldindan olamiz — bo'sh maydon o'shанга
        # tenglashadi, natijaда hech narsa o'zgarmaydi (0 ga tushmaydi).
        cur_stocks = {bs.variant_id: bs for bs in BranchStock.objects.filter(
            branch=branch, variant__product=product)}
        rows = []
        seen_pairs, seen_barcodes = {}, set()   # {(o'lcham, rang): narx}
        for i in range(len(ids)):
            get = lambda lst: (lst[i] if i < len(lst) else '') or ''
            raw_id = (ids[i] or '').strip()
            color = smart_title(get(colors))
            size = smart_title(get(sizes))
            barcode = get(barcodes).strip() or None
            if raw_id:
                try:
                    variant = variants[int(raw_id)]
                except (ValueError, KeyError):
                    continue
            else:
                # Yangi tur qatori — bo'sh bo'lsa jim o'tkazib yuboramiz
                if not (color or size or barcode):
                    continue
                variant = None
            # V2/N2: bo'sh katak = None ("o'zgarmadi"); 'inf'/'nan' rad etiladi.
            def _decn(s):
                s = (s or '').strip()
                if s == '':
                    return None
                d = _dec(s)
                if not d.is_finite():
                    raise InvalidOperation
                return d
            try:
                cost_i = _decn(get(costs))
                sale_i = _decn(get(sales))
                ws_i = _decn(get(wholesales))
                stock_i = _decn(get(stocks))
            except (InvalidOperation, ValueError, TypeError):
                errors.append(f"{i + 1}-qator: raqam maydonlari noto'g'ri.")
                continue
            # Bo'sh maydon: mavjud tur uchun JORIY qiymat, yangi tur uchun 0.
            _cur = cur_stocks.get(variant.pk) if variant is not None else None
            cost = cost_i if cost_i is not None else (_cur.cost_price if _cur else Decimal('0'))
            sale = sale_i if sale_i is not None else (_cur.sale_price if _cur else Decimal('0'))
            wholesale = ws_i if ws_i is not None else (_cur.wholesale_price if _cur else Decimal('0'))
            stock_new = (int(stock_i) if stock_i is not None
                         else (_cur.stock_count if _cur else 0))
            if min(cost, sale, wholesale) < 0 or stock_new < 0:
                errors.append(f"{i + 1}-qator: manfiy qiymat kiritilmaydi.")
                continue
            # Bir xil o'lcham, boshqa narx bo'lsa — narx belgisi bilan ajratamiz.
            # Mavjud turni jimgina qayta nomlamaymiz: unga faqat xato beramiz.
            new_color, is_dup = resolve_row_price_color(
                size, color, sale, seen_pairs)
            if is_dup or (variant is not None and new_color != color):
                errors.append(
                    f"{i + 1}-qator: {size or '—'} / {color or '—'} "
                    f"takrorlangan (o'lchami ham, narxi ham bir xil).")
                continue
            color = new_color
            if barcode:
                if barcode in seen_barcodes:
                    errors.append(f"{i + 1}-qator: shtrix-kod jadvalda takror.")
                    continue
                seen_barcodes.add(barcode)
                v_clash = ProductVariant.objects.filter(barcode=barcode)
                if variant is not None:
                    v_clash = v_clash.exclude(pk=variant.pk)
                v_clash = v_clash.select_related('product').first()
                if v_clash:
                    errors.append(
                        f"Shtrix-kod {barcode} band: {v_clash.product.name} "
                        f"({v_clash.size or '—'}/{v_clash.color or '—'}).")
                p_clash = Product.objects.filter(
                    external_barcode=barcode).first()
                if p_clash:
                    errors.append(
                        f"Shtrix-kod {barcode} '{p_clash.name}' "
                        f"mahsulotiga biriktirilgan.")
            rows.append({'variant': variant, 'color': color, 'size': size,
                         'barcode': barcode, 'cost': cost, 'sale': sale,
                         'wholesale': wholesale, 'stock': stock_new})

        # DB darajasida ham juftlik boshqa variant bilan to'qnashmasin
        _form_pks = {x['variant'].pk for x in rows if x['variant'] is not None}
        for r in rows:
            clash_qs = product.variants.filter(size=r['size'], color=r['color'])
            if r['variant'] is not None:
                clash_qs = clash_qs.exclude(pk=r['variant'].pk)
            clash = clash_qs.first()
            if clash and clash.pk not in _form_pks:
                errors.append(
                    f"{r['size'] or '—'} / {r['color'] or '—'} allaqachon mavjud.")

        if errors:
            for e in errors[:8]:
                messages.error(request, e)
        else:
            changed = []
            with transaction.atomic():
                for r in rows:
                    v = r['variant']
                    if v is None:
                        v = ProductVariant.objects.create(
                            product=product, size=r['size'],
                            color=r['color'], barcode=r['barcode'] or None)
                        # Shtrix-kod kiritilmagan bo'lsa — o'zi beriladi
                        # (boshqa qabul yo'llarida shunday, bu yerda tushib
                        # qolgan edi: kodsiz tur kassada skanerlanmaydi va
                        # etiketka ham chiqmaydi)
                        if not v.barcode:
                            code = gen_internal_ean13(v.pk)
                            k = 0
                            while ProductVariant.objects.filter(
                                    barcode=code).exclude(pk=v.pk).exists():
                                k += 1
                                code = gen_internal_ean13(v.pk + k * 100000)
                            v.barcode = code
                            v.save(update_fields=['barcode'])
                        stock = BranchStock.objects.create(
                            variant=v, branch=branch,
                            cost_price=r['cost'], sale_price=r['sale'],
                            wholesale_price=r['wholesale'],
                            stock_count=r['stock'])
                        if r['stock'] > 0:
                            Intake.objects.create(
                                variant=v, branch=branch,
                                quantity=r['stock'],
                                cost_per_unit=r['cost'],
                                sale_price=r['sale'] or None,
                                note="Yangi tur (tahrirlash sahifasidan)",
                                received_by=request.user)
                        changed.append(
                            f"+ yangi: {v.size or '—'}/{v.color or '—'}")
                        continue
                    v_fields = []
                    if v.color != r['color']:
                        v_fields.append(f"rang {v.color}→{r['color']}")
                        v.color = r['color']
                    if v.size != r['size']:
                        v_fields.append(f"o'lcham {v.size}→{r['size']}")
                        v.size = r['size']
                    if (v.barcode or None) != r['barcode']:
                        v_fields.append(f"shtrix {v.barcode or '—'}→{r['barcode'] or '—'}")
                        v.barcode = r['barcode']
                    if v_fields:
                        v.save()
                    stock, _ = BranchStock.objects.select_for_update().get_or_create(
                        variant=v, branch=branch)
                    s_fields = []
                    if stock.cost_price != r['cost']:
                        s_fields.append(f"tannarx {stock.cost_price}→{r['cost']}")
                        stock.cost_price = r['cost']
                    if stock.sale_price != r['sale']:
                        s_fields.append(f"narx {stock.sale_price}→{r['sale']}")
                        stock.sale_price = r['sale']
                    if stock.wholesale_price != r['wholesale']:
                        s_fields.append(
                            f"ulgurji {stock.wholesale_price}→{r['wholesale']}")
                        stock.wholesale_price = r['wholesale']
                    delta = r['stock'] - stock.stock_count
                    if delta:
                        s_fields.append(
                            f"ombor {stock.stock_count}→{r['stock']}")
                        stock.stock_count = r['stock']
                        Intake.objects.create(
                            variant=v, branch=branch, quantity=delta,
                            cost_per_unit=stock.cost_price,
                            note="Qo'lda tuzatish (turlarni tahrirlash)",
                            received_by=request.user)
                    if s_fields:
                        stock.save()
                    if v_fields or s_fields:
                        changed.append(
                            f"{v.size or '—'}/{v.color or '—'}: "
                            + ', '.join(v_fields + s_fields))
                if changed:
                    AuditLog.objects.create(
                        user=request.user,
                        username_snapshot=request.user.username,
                        action=AuditLog.Action.UPDATE,
                        model_name='ProductVariant',
                        object_id=str(product.pk),
                        object_repr=(f"{product.code} ({branch.name}): "
                                     + ' | '.join(changed))[:300],
                    )
            if changed:
                messages.success(
                    request, f"Saqlandi: {len(changed)} ta tur yangilandi "
                             f"({branch.name}).")
            else:
                messages.info(request, "O'zgarish yo'q.")
            return redirect('product_detail', code=product.code)

    # E1: POST xato bo'lса — submitlaган qatorlarni QAYTA ko'rsatamiz (bitta
    # xato tufayli barcha qatorlar yo'qolmasin; intake_variants'даgi post_back
    # namunasi). Aks holда joriy holatни DB'дан chizamiz.
    def _clean_num(s):
        s = (s or '')
        for ch in (' ', ' ', ' ', "'"):
            s = s.replace(ch, '')
        return s.replace(',', '.')

    if request.method == 'POST' and errors:
        rows = []
        for i in range(len(ids)):
            g = lambda lst: (lst[i] if i < len(lst) else '')
            rows.append({
                'v_id': g(ids), 'color': g(colors), 'size': g(sizes),
                'barcode': g(barcodes),
                'cost': _clean_num(g(costs)), 'sale': _clean_num(g(sales)),
                'wholesale': _clean_num(g(wholesales)), 'stock': _clean_num(g(stocks)),
            })
    else:
        stocks = {st.variant_id: st for st in BranchStock.objects.filter(
            variant__product=product, branch=branch)}
        rows = []
        for v in product.variants.all():
            st = stocks.get(v.pk)
            rows.append({
                'v_id': v.pk, 'color': v.color, 'size': v.size,
                'barcode': v.barcode or '',
                'cost': st.cost_price if st else '',
                'sale': st.sale_price if st else '',
                'wholesale': st.wholesale_price if st else '',
                'stock': st.stock_count if st else 0,
            })
    return render(request, 'inventory/product_variants_edit.html', {
        'product': product, 'branch': branch, 'branches': branches,
        'rows': rows,
        # Qatorlar 'v_id' kalitini ishlatadi ('variant' EMAS). Yangi (saqlanmagan)
        # qatorlarда v_id bo'sh — ularни o'tkazib yuboramiz.
        'variant_ids': ','.join(str(r['v_id']) for r in rows if r['v_id']),
    })


@admin_required
def clothes_intake(request):
    """Kiyim/poyabzal qabul — zavod kodisiz tovarlar. O'lcham x rang
    jadvali, har kombinatsiya uchun avtomatik EAN-13 kod yaratiladi,
    keyin termal etiketka chop etiladi."""
    from decimal import Decimal, InvalidOperation
    categories = Category.objects.order_by('name')
    branches = Branch.objects.filter(is_active=True)

    if request.method == 'POST':
        errors = []
        product = None
        brand = smart_title(request.POST.get('brand'))
        category = Category.objects.filter(
            pk=request.POST.get('category') or 0).first()
        if not brand:
            errors.append("Brend nomini kiriting (masalan: Zara).")
        if not category:
            errors.append("Kategoriya (tur) ni tanlang.")

        branch = Branch.objects.filter(
            pk=request.POST.get('branch') or 0, is_active=True).first()
        if not branch:
            errors.append("Filial tanlang.")

        try:
            price = parse_dec(request.POST.get('price'))
            marja = parse_dec(request.POST.get('marja'))
        except (InvalidOperation, ValueError):
            price = Decimal('0'); marja = Decimal('0')
        if marja < 0:
            marja = Decimal('0')
        # Tannarx (ixtiyoriy) — boshqa qabul sahifalaridagi kabi:
        #   tannarx berilsa -> sotuv = tannarx × (1 + marja/100)
        #   faqat sotuv berilsa -> tannarx = sotuv / (1 + marja/100)
        cost_in = Decimal('0')
        cost_raw = (request.POST.get('cost') or '').strip()
        if cost_raw:
            try:
                cost_in = parse_dec(cost_raw)
            except (InvalidOperation, ValueError):
                cost_in = Decimal('0')
        if cost_in < 0:
            cost_in = Decimal('0')
        if cost_in > 0:
            cost = cost_in.quantize(Decimal('0.01'))
            if price <= 0:
                price = (cost_in * (Decimal('1') + marja / Decimal('100'))
                         ).quantize(Decimal('0.01'))
        elif price > 0:
            cost = (price / (Decimal('1') + marja / Decimal('100'))
                    ).quantize(Decimal('0.01'))
        else:
            cost = Decimal('0')
        if price <= 0:
            errors.append("Sotuv narxini yoki tannarxni kiriting.")

        # Kataklar: qty[<size>]  (kiyim/poyabzalda rang ishlatilmaydi).
        # Eski format qty[<size>|<color>] ham qo'llab-quvvatlanadi.
        cells = []
        for key, val in request.POST.items():
            if not (key.startswith('qty[') and key.endswith(']')):
                continue
            payload = key[4:-1]
            if '|' in payload:
                size, color = payload.split('|', 1)
            else:
                size, color = payload, ''
            size = smart_title(size); color = smart_title(color)
            try:
                qty = int(parse_dec(val))
            except (InvalidOperation, ValueError, TypeError):
                qty = 0
            if qty > 0 and (size or color):
                cells.append((size, color, qty))
        if not cells and not errors:
            errors.append("Kamida bitta o'lcham qatoriga son kiriting.")

        if errors:
            for e in errors[:8]:
                messages.error(request, e)
        else:
            created_ids = []
            total_qty = 0
            split_notes = []      # narxi farq qilgani uchun yangi kod olganlar
            with transaction.atomic():
                # Bir xil brend + kategoriya bo'lsa — o'shanga qo'shamiz
                # (aks holda yangi mahsulot). Nomi = "Brend Kategoriya".
                comp_name = f"{brand} {category.name}".strip()
                product = Product.objects.filter(
                    brand__iexact=brand, category=category).first()
                if product is None:
                    product = Product.objects.create(
                        name=comp_name, brand=brand, category=category,
                        default_sale_price=price)
                session = IntakeSession.objects.create(
                    branch=branch, received_by=request.user,
                    note="Kiyim/poyabzal qabul (avto-kod)")
                for size, color, qty in cells:
                    variant, _ = ProductVariant.objects.get_or_create(
                        product=product, size=size, color=color)
                    if not variant.barcode:
                        code = gen_internal_ean13(variant.pk)
                        # kolliziya bo'lsa (kamdan-kam) — pk+offset
                        n = 0
                        while ProductVariant.objects.filter(
                                barcode=code).exclude(pk=variant.pk).exists():
                            n += 1
                            code = gen_internal_ean13(variant.pk + n * 100000)
                        variant.barcode = code
                        variant.save(update_fields=['barcode'])
                    # Sotuv narxi boshqacha bo'lsa — alohida tur (o'z kodi bilan)
                    before_pk = variant.pk
                    variant = resolve_price_variant(variant, branch, price)
                    if variant.pk != before_pk:
                        # Xodim bilsin: eski kod tegilmadi, bu narxga yangi kod berildi
                        split_notes.append(
                            f"{size or '—'} — {variant.barcode}")
                    stock, _ = BranchStock.objects.get_or_create(
                        variant=variant, branch=branch,
                        defaults={'cost_price': cost, 'sale_price': price})
                    stock.cost_price = weighted_cost(  # STK-8
                        stock.stock_count if isinstance(stock.stock_count, int) else 0,
                        stock.cost_price, qty, cost)
                    stock.sale_price = price
                    stock.stock_count = F('stock_count') + qty
                    stock.save()
                    Intake.objects.create(
                        session=session, variant=variant, branch=branch,
                        quantity=qty, cost_per_unit=cost, sale_price=price,
                        note="Kiyim qabul", received_by=request.user)
                    created_ids.append(variant.pk)
                    total_qty += qty
            from .notifications import notify_intake_session
            notify_intake_session(session)
            messages.success(
                request,
                f"{product.name}: {len(cells)} tur, {total_qty} dona qabul "
                f"qilindi. Endi etiketkalarni chop eting.")
            if split_notes:
                messages.warning(
                    request,
                    "Narxi avvalgisidan farq qilgani uchun yangi shtrix-kod "
                    "berildi (eski koddagi mol o'z narxida qoldi): "
                    + ", ".join(split_notes))
            ids = ','.join(str(i) for i in created_ids)
            return redirect(f"{reverse('variant_labels')}?ids={ids}&price={price}")

    colors_all = (ProductVariant.objects.exclude(color='')
                  .values_list('color', flat=True).distinct().order_by('color')[:500])
    brands_all = (Product.objects.exclude(brand='')
                  .values_list('brand', flat=True).distinct().order_by('brand')[:500])
    return render(request, 'inventory/clothes_intake.html', {
        'categories': categories, 'branches': branches,
        'colors_all': colors_all, 'brands_all': brands_all,
        'post_back': request.POST if request.method == 'POST' else {},
    })



def _svg_scalable(svg):
    """python-barcode SVG'siga viewBox qo'shadi.

    Kutubxona SVG'ni mm birliklarida, viewBox'siz chiqaradi. viewBox
    bo'lmasa CSS'dagi height SVG'ni KICHRAYTIRMAYDI — kesib tashlaydi.
    Shtrix-kod raqami pastda (y≈13.5mm) turgani uchun chop etishda
    (height: 11mm) u qirqilib, ekранda ko'ringani bilan qog'ozga
    tushmasdi.

    Shuning uchun: mm qo'shimchalarini olib tashlab, o'lchamlarni
    foydalanuvchi birligiga aylantiramiz va viewBox qo'yamiz.
    """
    import re as _re
    m = _re.search(r'<svg[^>]*>', svg)
    if not m:
        return svg
    root = m.group(0)
    w = _re.search(r'width="([\d.]+)mm"', root)
    h = _re.search(r'height="([\d.]+)mm"', root)
    if not (w and h):
        return svg
    W, H = w.group(1), h.group(1)

    body = svg[m.end():]
    # pt -> mm (1pt = 25.4/72 mm), aks holda viewBox bilan mos kelmaydi
    def _pt(mm):
        return f'font-size:{float(mm.group(1)) * 25.4 / 72:.3f};'
    body = _re.sub(r'font-size:\s*([\d.]+)pt;', _pt, body)
    # "12.34mm" -> "12.34"
    body = _re.sub(r'="([\d.]+)mm"', r'="\1"', body)

    new_root = (
        f'<svg version="1.1" xmlns="http://www.w3.org/2000/svg" '
        f'width="100%" height="100%" viewBox="0 0 {W} {H}" '
        f'preserveAspectRatio="xMidYMid meet">'
    )
    return new_root + body


@admin_required
def variant_labels(request):
    """Termal etiketka: EAN-13 barcode + QR + kod + narx (variantlar)."""
    import base64
    from barcode import EAN13
    from barcode.writer import SVGWriter

    ids = [int(i) for i in (request.GET.get('ids') or '').split(',')
           if i.strip().isdigit()]
    # Mahsulotlar sahifasidan kelganda MAHSULOT id'lari uzatiladi —
    # ularning hamma turlarini ochamiz.
    # ?products=1,2  ham  ?products=1&products=2  ham ishlasin
    _parts = []
    for chunk in request.GET.getlist('products'):
        _parts.extend(str(chunk).split(','))
    pids = [int(x) for x in _parts if x.strip().isdigit()]
    if pids:
        variants = list(ProductVariant.objects.filter(product_id__in=pids)
                        .select_related('product')
                        .order_by('product__name', 'color', 'size'))
    else:
        variants = list(ProductVariant.objects.filter(pk__in=ids)
                        .select_related('product'))
    # Etiketka nusxalari — har variant uchun nechta (default 1)
    # "copies=stock" -> har tur uchun OMBORDAGI soni qadar yorliq
    copies_raw = (request.GET.get('copies') or '1').strip().lower()
    by_stock = copies_raw in ('stock', 'ombor')
    try:
        copies = 1 if by_stock else max(1, min(200, int(copies_raw)))
    except ValueError:
        copies = 1

    # Narx: variantning filial narxi yoki so'rov paramidan yoki mahsulot default
    price_param = request.GET.get('price')
    # Zaxira yozuvlarini oldindan olamiz — halqa ichida har tur uchun
    # alohida so'rov ketardi (sahifada 51 ta so'rov)
    _stock_by_variant = {}
    for _st in BranchStock.objects.filter(
            variant__in=variants).select_related('branch'):
        _stock_by_variant.setdefault(_st.variant_id, []).append(_st)
    labels = []
    for v in variants:
        bc = v.barcode or gen_internal_ean13(v.pk)
        # EAN-13 barcode SVG (12 body -> lib check qo'shadi)
        try:
            svg_io = io.BytesIO()
            # module_width kattaroq -> shtrix-kod KENGROQ (va skaner uchun
            # ham qulayroq). Balandlik CSS bilan cheklangani uchun, enini
            # oshirishning yagona yo'li — nisbatni kengaytirish.
            # quiet_zone 2.0 saqlanadi: EAN-13 uchun chekka bo'shliq shart.
            EAN13(bc[:12], writer=SVGWriter()).write(
                svg_io, options={'module_height': 9.0, 'module_width': 0.50,
                                 'font_size': 8, 'text_distance': 3.5,
                                 'quiet_zone': 2.0})
            barcode_svg = svg_io.getvalue().decode('utf-8')
            # <?xml ...?> va DOCTYPE'ni olib tashlaymiz (inline uchun)
            i = barcode_svg.find('<svg')
            barcode_svg = barcode_svg[i:] if i >= 0 else barcode_svg
            barcode_svg = _svg_scalable(barcode_svg)
        except Exception:
            barcode_svg = ''
        # QR (kod) -> PNG data URI
        # narx
        _branch = getattr(request.user, 'branch', None)
        _rows = _stock_by_variant.get(v.pk, [])
        _priced = [r for r in _rows if r.sale_price and r.sale_price > 0]
        st = (next((r for r in _priced if _branch and r.branch_id == _branch.pk), None)
              or (max(_priced, key=lambda r: r.sale_price) if _priced else None))
        _stock_row = (next((r for r in _rows if _branch and r.branch_id == _branch.pk), None)
                      or st)
        stock_n = _stock_row.stock_count if _stock_row else 0
        price = (price_param or (st.sale_price if st else v.product.default_sale_price))
        labels.append({
            'stock': stock_n,
            'variant': v, 'code': bc, 'barcode_svg': barcode_svg,
            'price': price,
        })
    # Nusxa: qat'iy son yoki OMBOR soniga qarab (copies=stock).
    # Ombori 0 bo'lgan tur uchun yorliq chiqarilmaydi — bekorga sarf bo'lmasin.
    render_labels = []
    skipped = 0
    for lb in labels:
        n = min(200, lb['stock']) if by_stock else copies
        if by_stock and n <= 0:
            skipped += 1
            continue
        for _ in range(n):
            render_labels.append(lb)
    # "Orqaga" qaytadigan manzil: qaysi sahifadan kelingan bo'lsa o'sha.
    # Yorliq sahifasi yangi oynada ochilgani uchun brauzer tarixi bo'sh —
    # shuning uchun manzilni aniq beramiz. Faqat ichki yo'l ('/...'), tashqi
    # havola emas (ochiq-redirect xavfsizligi).
    back_raw = (request.GET.get('back') or '').strip()
    if back_raw.startswith('/') and not back_raw.startswith('//'):
        back_url = back_raw
    else:
        back_url = reverse('product_list')

    return render(request, 'inventory/variant_labels.html', {
        'store_name': PRICE_LABEL_STORE_NAME,
        'labels': render_labels, 'copies': copies_raw if by_stock else copies,
        'by_stock': by_stock, 'skipped': skipped,
        'uniq_labels': labels,          # o'lcham tanlash paneli uchun
        'variant_count': len(labels),
        'back_url': back_url,
        # products= bilan kelgan bo'lsa ham "Nusxa" formasi ishlashi uchun
        # aniq tur id'larini qaytaramiz
        'ids': (','.join(str(v.pk) for v in variants) if pids
                else request.GET.get('ids', '')),
    })


@admin_required
def intake_variants(request):
    """Jadval usulida qabul — bitta sahifada mahsulot + bir nechta tur.

    Har qator = tur: rang, o'lcham, shtrix-kod (har turga alohida),
    tannarx, sotuv narx, miqdor. Mahsulot yangi yaratiladi yoki
    mavjudlardan tanlanadi (?code= bilan prefill ham mumkin).
    """
    from decimal import Decimal, InvalidOperation

    categories = Category.objects.order_by('name')
    branches = Branch.objects.filter(is_active=True)
    suppliers = Supplier.objects.filter(is_active=True).order_by('name')

    prefill_product = None
    if request.method == 'GET' and request.GET.get('code'):
        prefill_product = Product.objects.filter(
            code=normalize_code(request.GET['code'].upper())).first()

    post_back = None
    if request.method == 'POST':
        errors = []

        # ---------- mahsulot: mavjud yoki yangi ----------
        product = None
        product_code = (request.POST.get('product_code') or '').strip()
        new_name = smart_title(request.POST.get('new_name'))
        if product_code:
            product = Product.objects.filter(
                code=normalize_code(product_code.upper())).first()
            if not product:
                errors.append(f"Mahsulot topilmadi: {product_code}")
        elif new_name:
            # Yangi nom yozilgan -> YANGI mahsulot (jim birlashtirmaymiz).
            # Mavjud mahsulotga qo'shmoqchi bo'lsa, "Mavjud mahsulot" rejimida
            # ro'yxatdan tanlaydi (product_code). Bir brend ostida (masalan
            # Ezo) turli mahsulotlar bo'lishi mumkin.
            product = None  # keyin yangi yaratiladi
        else:
            errors.append("Mahsulot tanlang yoki yangi mahsulot nomini kiriting.")

        branch = Branch.objects.filter(
            pk=request.POST.get('branch') or 0, is_active=True).first()
        if not branch:
            errors.append("Filial tanlang.")

        supplier_text = (request.POST.get('supplier') or '').strip()
        note = (request.POST.get('note') or '').strip()

        # ---------- qatorlar ----------
        colors = request.POST.getlist('row_color')
        sizes = request.POST.getlist('row_size')
        barcodes = request.POST.getlist('row_barcode')
        costs = request.POST.getlist('row_cost')
        marjas = request.POST.getlist('row_marja')
        prices = request.POST.getlist('row_price')
        qtys = request.POST.getlist('row_qty')

        _dec = parse_dec

        # Standart marja — row_marja bo'sh bo'lsa fallback sifatida ishlatiladi.
        # Tannarx berilsa: sotuv = tannarx × (1 + marja/100).
        # Tannarx bo'sh, faqat sotuv berilsa: tannarx = sotuv / (1 + marja/100).
        try:
            marja = _dec(request.POST.get('marja'))
        except (InvalidOperation, ValueError):
            marja = Decimal('0')
        if marja < 0:
            marja = Decimal('0')

        rows, raw_rows = [], []
        seen_pairs, seen_barcodes = {}, set()   # {(o'lcham, rang): narx}
        for i in range(len(colors)):
            get = lambda lst: (lst[i] if i < len(lst) else '') or ''
            color = smart_title(get(colors))
            size = smart_title(get(sizes))
            barcode = get(barcodes).strip() or None
            cost_raw, marja_raw = get(costs), get(marjas)
            price_raw, qty_raw = get(prices), get(qtys)
            if not (color or size or barcode or qty_raw.strip()
                    or price_raw.strip() or cost_raw.strip()):
                continue  # butunlay bo'sh qator
            raw_rows.append({'color': color, 'size': size,
                             'barcode': barcode or '',
                             'cost': cost_raw, 'marja': marja_raw,
                             'price': price_raw, 'qty': qty_raw})
            try:
                price = _dec(price_raw)
                qty = int(_dec(qty_raw))
            except (InvalidOperation, ValueError, TypeError):
                errors.append(f"{i + 1}-qator: narx yoki miqdor noto'g'ri.")
                continue
            # Tannarx (ixtiyoriy) va shu qatorning marjasi
            cost_in = Decimal('0')
            if cost_raw.strip():
                try:
                    cost_in = _dec(cost_raw)
                except (InvalidOperation, ValueError):
                    cost_in = Decimal('0')
            row_marja = None
            if marja_raw.strip():
                try:
                    row_marja = _dec(marja_raw)
                except (InvalidOperation, ValueError):
                    row_marja = None
            if qty < 0 or price < 0 or cost_in < 0:
                errors.append(f"{i + 1}-qator: manfiy qiymat kiritilmaydi.")
                continue
            eff_marja = row_marja if row_marja is not None else marja
            if cost_in > 0:
                # Tannarx to'g'ridan-to'g'ri kiritilgan
                cost = cost_in.quantize(Decimal('0.01'))
                if price <= 0:
                    price = (cost_in * (Decimal('1') + eff_marja / Decimal('100'))
                             ).quantize(Decimal('0.01'))
            elif price > 0:
                # Faqat sotuv narx berilgan — tannarxni marja orqali chiqaramiz
                cost = (price / (Decimal('1') + eff_marja / Decimal('100'))
                        ).quantize(Decimal('0.01'))
            else:
                cost = Decimal('0')
            # Bir xil o'lcham, boshqa narx — narx belgisi bilan alohida tur
            color, is_dup = resolve_row_price_color(size, color, price, seen_pairs)
            if is_dup:
                errors.append(
                    f"{i + 1}-qator: {size or '—'} / {color or '—'} "
                    f"takrorlangan (o'lchami ham, narxi ham bir xil).")
                continue
            if barcode:
                if barcode in seen_barcodes:
                    errors.append(f"{i + 1}-qator: shtrix-kod jadvalda takror.")
                    continue
                seen_barcodes.add(barcode)
            rows.append({'color': color, 'size': size, 'barcode': barcode,
                         'cost': cost, 'price': price, 'qty': qty})

        if not rows and not errors:
            errors.append("Kamida bitta tur qatorini kiriting.")

        # Shtrix-kod bazadagi boshqa yozuvlar bilan to'qnashmasin
        for r in rows:
            if not r['barcode']:
                continue
            v_clash = ProductVariant.objects.filter(barcode=r['barcode'])
            if product:
                v_clash = v_clash.exclude(
                    product=product, size=r['size'], color=r['color'])
            v_clash = v_clash.select_related('product').first()
            if v_clash:
                errors.append(
                    f"Shtrix-kod {r['barcode']} band: {v_clash.product.name} "
                    f"({v_clash.size or '—'}/{v_clash.color or '—'}).")
            p_clash = Product.objects.filter(
                external_barcode=r['barcode']).first()
            if p_clash:
                errors.append(
                    f"Shtrix-kod {r['barcode']} '{p_clash.name}' "
                    f"mahsulotiga biriktirilgan.")

        if errors:
            for e in errors[:8]:
                messages.error(request, e)
            post_back = {
                'marja': request.POST.get('marja') or '0',
                'product_code': product_code,
                'product_name': product.name if product else '',
                'new_name': new_name,
                'category': request.POST.get('category') or '',
                'branch': request.POST.get('branch') or '',
                'supplier': supplier_text,
                'note': note,
                'rows': raw_rows,
            }
        else:
            total_qty = 0
            with transaction.atomic():
                if product is None:
                    category = Category.objects.filter(
                        pk=request.POST.get('category') or 0).first()
                    product = Product.objects.create(
                        name=new_name, category=category)
                supplier_obj = None
                if supplier_text:
                    supplier_obj = Supplier.objects.filter(
                        name__iexact=supplier_text).first()
                session = None
                if any(r['qty'] > 0 for r in rows):
                    session = IntakeSession.objects.create(
                        branch=branch, supplier=supplier_obj,
                        supplier_text='' if supplier_obj else supplier_text,
                        received_by=request.user, note=note)
                for r in rows:
                    # HAQIQAT MANBAI = SHTRIX-KOD. Qatorда kod bo'lsa, qoldiq
                    # AYNAN o'sha kodli turга tushadi va narx bo'yicha "birodar"га
                    # KO'CHIRILMAYDI (narx-split yo'q). Aks holда skanerlanган
                    # kod bir turда, qoldiq boshqa turда qolib, kassada "omborda
                    # yo'q" chiqardi. Faqat kodsiz turlар uchun narx-split ishlaydi.
                    if r['barcode']:
                        variant = (ProductVariant.objects
                                   .filter(barcode=r['barcode']).first())
                        if variant is None:
                            variant, _ = ProductVariant.objects.get_or_create(
                                product=product, size=r['size'], color=r['color'])
                            if variant.barcode != r['barcode']:
                                variant.barcode = r['barcode']
                                variant.save(update_fields=['barcode'])
                    else:
                        variant, _ = ProductVariant.objects.get_or_create(
                            product=product, size=r['size'], color=r['color'])
                        ensure_variant_barcode(variant)
                        variant = resolve_price_variant(variant, branch, r['price'])
                    ensure_variant_barcode(variant)
                    stock, _ = BranchStock.objects.get_or_create(
                        variant=variant, branch=branch,
                        defaults={'cost_price': r['cost'],
                                  'sale_price': r['price']})
                    stock.cost_price = weighted_cost(  # STK-8
                        stock.stock_count if isinstance(stock.stock_count, int) else 0,
                        stock.cost_price, r['qty'], r['cost'])
                    if r['price'] > 0:
                        stock.sale_price = r['price']
                    if r['qty'] > 0:
                        stock.stock_count = F('stock_count') + r['qty']
                    stock.save()
                    if r['qty'] > 0:
                        Intake.objects.create(
                            session=session, supplier_ref=supplier_obj,
                            variant=variant, branch=branch,
                            quantity=r['qty'], cost_per_unit=r['cost'],
                            sale_price=r['price'] or None,
                            supplier=supplier_text, note=note,
                            received_by=request.user)
                        total_qty += r['qty']
                if product.default_sale_price == 0:
                    first_price = next(
                        (r['price'] for r in rows if r['price'] > 0), None)
                    if first_price:
                        product.default_sale_price = first_price
                        product.save(update_fields=['default_sale_price'])
            from .notifications import notify_intake_session
            notify_intake_session(session)
            messages.success(
                request,
                f"Saqlandi: {product.name} — {len(rows)} ta tur, "
                f"jami {total_qty} dona ({branch.name}).")
            return redirect('product_detail', code=product.code)

    # Rang nomlaridagi YAKKA so'zlar — so'zma-so'z taklif uchun
    # ("Cleaning Anti Micro" -> Cleaning, Anti, Micro)
    _color_words = set()
    for _c in (ProductVariant.objects.exclude(color='')
               .values_list('color', flat=True).distinct()):
        for _w in _c.split():
            if len(_w) > 1:
                _color_words.add(_w)
    color_options = sorted(_color_words)[:800]
    return render(request, 'inventory/intake_variants.html', {
        'categories': categories,
        'branches': branches,
        'suppliers': suppliers,
        'prefill_product': prefill_product,
        'post_back': post_back,
        'color_options': color_options,
    })


PARFUM_BRANDS = [
    # Parfyumeriya
    'Chanel', 'Dior', 'Lancome', 'Guerlain', 'Yves Saint Laurent',
    'Giorgio Armani', 'Versace', 'Dolce & Gabbana', 'Gucci', 'Burberry',
    'Hugo Boss', 'Calvin Klein', 'Paco Rabanne', 'Carolina Herrera',
    'Givenchy', 'Hermes', 'Tom Ford', 'Montale', 'Mancera', 'Ajmal',
    'Lattafa', 'Armaf', 'Rasasi', 'Swiss Arabian', 'Al Haramain',
    # Soch/tana parvarishi
    'Head and Shoulders', 'Pantene', 'Gliss Kur', 'Schwarzkopf', 'Syoss',
    "L'Oreal", 'Garnier', 'Nivea', 'Dove', 'Rexona', 'Old Spice', 'AXE',
    'Palmolive', 'Le Petit Marseillais', 'Fa', 'Camay', 'Safeguard',
    # Koreys kosmetikasi
    'The Face Shop', 'Missha', 'Etude House', 'Innisfree', 'Laneige',
    'COSRX', 'Some By Mi', 'Holika Holika', 'Tony Moly', "It's Skin",
    '3W Clinic', 'FarmStay', 'Ekel', 'Jigott', 'Deoproce', 'Lebelage',
    'Enough', 'Esfolio', 'Eunyul', 'Mizon', 'SNP', 'Dr.Jart+',
    'Beauty of Joseon', 'Round Lab', 'Anua', 'Medicube', 'Skin1004',
    # Maishiy
    'Colgate', 'Oral-B', 'Sensodyne', 'Fairy', 'Ariel', 'Tide', 'Persil',
]

IMPORT_HEADERS = ['Mahsulot', 'Rang/Tur', "O'lcham", 'Shtrix-kod',
                  'Sotuv narx', 'Miqdor']


@admin_required
def intake_import_template(request):
    """yurit-import.xlsx shablonini yaratib beradi.

    1-varaq: import shabloni (+2 namuna qator)
    2-varaq: joriy katalog (faqat ma'lumot uchun — import o'qimaydi)
    3-varaq: mashhur brendlar ro'yxati (ma'lumot uchun)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    head_font = Font(bold=True, color='FFFFFF')
    head_fill = PatternFill('solid', fgColor='4472C4')

    ws = wb.active
    ws.title = 'Import'
    ws.append(IMPORT_HEADERS)
    for c in ws[1]:
        c.font = head_font
        c.fill = head_fill
    ws.append(['NAMUNA: Chanel Bleu de Chanel', 'EDP', '100ml',
               '3145891073607', '1500000', '2'])
    ws.append(['NAMUNA: Head and Shoulders', 'Mentol', '400',
               '5000174896190', '45000', '10'])
    note = ("Eslatma: NAMUNA qatorlarini o'chirib, o'z tovarlaringizni "
            "yozing. Bir mahsulotning har bir turi alohida qator. "
            "Import faqat shu varaqni o'qiydi.")
    ws.append([])
    ws.append([note])
    widths = [34, 22, 12, 18, 14, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet('Katalogim')
    ws2.append(['Mahsulot', 'Rang/Tur', "O'lcham", 'Shtrix-kod'])
    for c in ws2[1]:
        c.font = head_font
        c.fill = head_fill
    for v in (ProductVariant.objects.select_related('product')
              .order_by('product__name', 'size', 'color')[:2000]):
        ws2.append([_csv_safe(x) for x in   # SEC-20
                    (v.product.name, v.color, v.size, v.barcode or
                     v.product.external_barcode or '')])
    for i, w in enumerate([34, 26, 12, 18], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws3 = wb.create_sheet('Brendlar')
    ws3.append(['Mashhur brendlar (ma\'lumot uchun)'])
    ws3[1][0].font = head_font
    ws3[1][0].fill = head_fill
    for b in PARFUM_BRANDS:
        ws3.append([b])
    ws3.column_dimensions['A'].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.'
                     'spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="yurit-import.xlsx"'
    return resp


@admin_required
def intake_import(request):
    """Excel (xlsx) orqali ommaviy qabul: mahsulot + turlar + shtrix-kodlar.

    Hammasi-yoki-hech-narsa: bironta qatorda xato bo'lsa, hech nima
    saqlanmaydi (qayta yuklashda ikki marta qo'shilib ketmasligi uchun).
    """
    from decimal import Decimal

    branches = Branch.objects.filter(is_active=True)

    if request.method == 'POST' and request.FILES.get('file'):
        from openpyxl import load_workbook

        branch = Branch.objects.filter(
            pk=request.POST.get('branch') or 0, is_active=True).first()
        if not branch:
            messages.error(request, "Filial tanlang.")
            return redirect('intake_import')
        try:
            marja = parse_dec(request.POST.get('marja'))
        except Exception:
            marja = Decimal('0')
        if marja < 0:
            marja = Decimal('0')

        try:
            wb = load_workbook(request.FILES['file'], read_only=True,
                               data_only=True)
        except Exception:
            messages.error(request, "Fayl o'qilmadi — .xlsx formatda yuklang.")
            return redirect('intake_import')
        ws = wb['Import'] if 'Import' in wb.sheetnames else wb.worksheets[0]

        errors, rows = [], []
        seen_barcodes = {}
        for idx, raw in enumerate(ws.iter_rows(min_row=2, values_only=True),
                                  start=2):
            vals = [(str(x).strip() if x is not None else '')
                    for x in (list(raw) + [''] * 6)[:6]]
            name, color, size, barcode, price_raw, qty_raw = vals
            if not any(vals):
                continue
            if name.upper().startswith('NAMUNA'):
                continue
            name = smart_title(name)
            color = smart_title(color)
            size = smart_title(size)
            if not name:
                errors.append(f"{idx}-qator: mahsulot nomi bo'sh.")
                continue
            if not (color or size or barcode):
                errors.append(f"{idx}-qator: tur ma'lumoti yo'q.")
                continue
            try:
                price = parse_dec(price_raw)
                qty = int(parse_dec(qty_raw))
            except Exception:
                errors.append(f"{idx}-qator: narx/miqdor noto'g'ri "
                              f"({price_raw!r}, {qty_raw!r}).")
                continue
            if price < 0 or qty < 0:
                errors.append(f"{idx}-qator: manfiy qiymat.")
                continue
            barcode = barcode or None
            if barcode:
                if barcode in seen_barcodes:
                    errors.append(
                        f"{idx}-qator: shtrix-kod {barcode} faylda takror "
                        f"({seen_barcodes[barcode]}-qatorda ham).")
                    continue
                seen_barcodes[barcode] = idx
            rows.append({'line': idx, 'name': name, 'color': color,
                         'size': size, 'barcode': barcode,
                         'price': price, 'qty': qty})

        if not rows and not errors:
            errors.append("Faylda import qilinadigan qator topilmadi.")

        # nom -> mavjud mahsulot (bir nechta bo'lsa xato)
        products_cache = {}
        for r in rows:
            key = r['name'].lower()
            if key in products_cache:
                continue
            qs = list(Product.objects.filter(name__iexact=r['name'])[:2])
            if len(qs) > 1:
                errors.append(
                    f"'{r['name']}' nomida bir nechta mahsulot bor — avval "
                    f"birlashtiring (Mahsulotlar → Birlashtirish).")
                products_cache[key] = None
            else:
                products_cache[key] = qs[0] if qs else None

        # shtrix-kod bazada bandmi?
        for r in rows:
            if not r['barcode']:
                continue
            existing_product = products_cache.get(r['name'].lower())
            v_clash = ProductVariant.objects.filter(barcode=r['barcode'])
            if existing_product:
                v_clash = v_clash.exclude(product=existing_product,
                                          size=r['size'], color=r['color'])
            v_clash = v_clash.select_related('product').first()
            if v_clash:
                errors.append(
                    f"{r['line']}-qator: shtrix-kod {r['barcode']} band — "
                    f"{v_clash.product.name} "
                    f"({v_clash.size or '—'}/{v_clash.color or '—'}).")
            p_clash = Product.objects.filter(
                external_barcode=r['barcode']).first()
            if p_clash and (not existing_product
                            or p_clash.pk != existing_product.pk):
                errors.append(
                    f"{r['line']}-qator: shtrix-kod {r['barcode']} "
                    f"'{p_clash.name}' mahsulotida.")

        if errors:
            for e in errors[:12]:
                messages.error(request, e)
            if len(errors) > 12:
                messages.warning(request,
                                 f"... yana {len(errors) - 12} ta xato.")
            messages.warning(request,
                             "Hech narsa saqlanmadi — faylni tuzatib "
                             "qayta yuklang.")
        else:
            created_products = 0
            total_qty = 0
            with transaction.atomic():
                session = None
                if any(r['qty'] > 0 for r in rows):
                    session = IntakeSession.objects.create(
                        branch=branch, received_by=request.user,
                        note="Excel import")
                for r in rows:
                    key = r['name'].lower()
                    product = products_cache.get(key)
                    if product is None:
                        product = Product.objects.create(name=r['name'])
                        products_cache[key] = product
                        created_products += 1
                    variant, _ = ProductVariant.objects.get_or_create(
                        product=product, size=r['size'], color=r['color'])
                    if r['barcode'] and variant.barcode != r['barcode']:
                        variant.barcode = r['barcode']
                        variant.save(update_fields=['barcode'])
                    cost = (r['price'] / (Decimal('1') + marja /
                            Decimal('100'))).quantize(Decimal('0.01')) \
                        if r['price'] > 0 else Decimal('0')
                    # Sotuv narxi boshqacha bo'lsa — alohida tur (o'z kodi bilan)
                    variant = resolve_price_variant(variant, branch, r['price'])
                    stock, _ = BranchStock.objects.get_or_create(
                        variant=variant, branch=branch,
                        defaults={'cost_price': cost,
                                  'sale_price': r['price']})
                    stock.cost_price = weighted_cost(  # STK-8
                        stock.stock_count if isinstance(stock.stock_count, int) else 0,
                        stock.cost_price, r['qty'], cost)
                    if r['price'] > 0:
                        stock.sale_price = r['price']
                    if r['qty'] > 0:
                        stock.stock_count = F('stock_count') + r['qty']
                    stock.save()
                    if r['qty'] > 0:
                        Intake.objects.create(
                            session=session, variant=variant, branch=branch,
                            quantity=r['qty'], cost_per_unit=cost,
                            sale_price=r['price'] or None,
                            note="Excel import", received_by=request.user)
                        total_qty += r['qty']
                    if product.default_sale_price == 0 and r['price'] > 0:
                        product.default_sale_price = r['price']
                        product.save(update_fields=['default_sale_price'])
                AuditLog.objects.create(
                    user=request.user,
                    username_snapshot=request.user.username,
                    action=AuditLog.Action.CREATE,
                    model_name='Product',
                    object_id='',
                    object_repr=(f"Excel import: {len(rows)} qator, "
                                 f"{created_products} yangi mahsulot, "
                                 f"{total_qty} dona ({branch.name})")[:300],
                )
            from .notifications import notify_intake_session
            notify_intake_session(session)
            messages.success(
                request,
                f"Import tayyor: {len(rows)} qator — {created_products} ta "
                f"yangi mahsulot, jami {total_qty} dona ({branch.name}).")
            return redirect('product_list')

    return render(request, 'inventory/intake_import.html', {
        'branches': branches,
    })


@admin_required
def intake_new(request):
    """Qabul dashboard'i: 3 ta kirish nuqtasi + so'nggi sessiyalar + low-stock."""
    products = Product.objects.order_by('-created_at')[:50]
    recent_sessions = (IntakeSession.objects
                       .select_related('branch', 'supplier', 'received_by')
                       .prefetch_related('intakes')
                       .order_by('-received_at')[:10])
    # Low-stock: o'rtacha kunlik sotuvga nisbatan kam qolgan mahsulotlar
    low_stock = (BranchStock.objects
                 .filter(stock_count__lte=3)
                 .select_related('variant__product', 'branch')
                 .order_by('stock_count')[:10])
    return render(request, 'inventory/intake_choose.html', {
        'products': products,
        'recent_sessions': recent_sessions,
        'low_stock': low_stock,
        'draft_count': InvoiceDraft.objects.count(),
    })


# ---------- QUICK INTAKE (scanner-driven, multi-product session) ----------

@admin_required
def intake_lookup(request):
    """GET /intake/lookup/?q=... — qabul uchun mahsulot qidirish.
    POS'dagidan farqi: stock cheklov yo'q (yangi mahsulot ham qabul qilinadi)."""
    q = (request.GET.get('q') or '').strip()
    if not q:
        return JsonResponse({'found': False})
    code = normalize_code(q.upper())
    product = Product.objects.filter(
        Q(code=code) | Q(external_barcode=q)
    ).first()
    if not product:
        _vm = ProductVariant.objects.filter(barcode=q).select_related('product').first()
        if _vm:
            product = _vm.product
    if not product:
        matches = list(Product.objects.filter(
            Q(name__icontains=q) | Q(brand__icontains=q) |
            Q(category__name__icontains=q)
        )[:8])
        if len(matches) == 1:
            product = matches[0]
        elif matches:
            return JsonResponse({
                'found': False,
                'suggestions': [{'code': p.code, 'name': p.name} for p in matches],
            })
        else:
            return JsonResponse({'found': False, 'suggestions': []})

    # Existing variants — for matrix products show them, allow new ones too
    variants = list(product.variants.values_list('size', 'color').distinct())
    # Last 5 intakes for cost-history hint
    last_intakes = (Intake.objects.filter(variant__product=product)
                    .select_related('variant', 'branch')
                    .order_by('-received_at')[:5])
    last_intakes_data = [{
        'date': i.received_at.strftime('%d.%m.%Y'),
        'branch': i.branch.name,
        'qty': i.quantity,
        'cost': float(i.cost_per_unit),
        'supplier': i.supplier_ref.name if i.supplier_ref else (i.supplier or ''),
    } for i in last_intakes]

    return JsonResponse({
        'found': True,
        'product': {
            'id': product.id,
            'code': product.code,
            'external_barcode': product.external_barcode or '',
            'name': product.name,
            'default_sale_price': float(product.default_sale_price),
            'markup_percent': float(product.markup_percent),
            'has_variants': bool(variants),
        },
        'variants': [{'size': s, 'color': c} for (s, c) in variants],
        'last_intakes': last_intakes_data,
    })


@admin_required
def intake_supplier_search(request):
    """GET /intake/supplier-search/?q=... — autocomplete."""
    q = (request.GET.get('q') or '').strip()
    qs = Supplier.objects.filter(is_active=True)
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(phone__icontains=q))
    return JsonResponse({
        'results': [{
            'id': s.id, 'name': s.name, 'phone': s.phone,
            'contact': s.contact_person,
        } for s in qs.order_by('name')[:10]],
    })


@admin_required
def send_daily_summary_now(request):
    """POST /dashboard/send-daily/ — Telegram'ga kunlik xulosa yuborish."""
    if request.method != 'POST':
        return redirect('dashboard')
    from .notifications import daily_summary_text, send_telegram, _enabled
    if not _enabled():
        messages.error(request,
            "Telegram sozlanmagan: TELEGRAM_BOT_TOKEN va TELEGRAM_CHAT_IDS env'ni belgilang.")
        return redirect('dashboard')
    try:
        text = daily_summary_text()
        ok = send_telegram(text)
        if ok:
            messages.success(request, "✓ Kunlik xulosa Telegram'ga yuborildi.")
        else:
            messages.error(request, "Yuborilmadi — Telegram API javob bermadi.")
    except Exception as e:
        messages.error(request, f"Xatolik: {e}")
    return redirect('dashboard')


def _csv_safe(v):
    """SEC-20: elektron jadval FORMULA IN'EKTSIYASINI zararsizlantiradi.

    Excel/Sheets `=`, `+`, `-`, `@` (yoki tab/CR) bilan boshlangan katakni
    FORMULA deb bajaradi — masalan `=HYPERLINK(...)` fayl ochilganда ishga
    tushardi. Bunday matn oldiga apostrof qo'yamiz (ko'rinishга ta'sir qilmaydi).
    """
    if isinstance(v, str) and v and v[0] in ('=', '+', '-', '@', '\t', '\r'):
        return "'" + v
    return v


_MAX_IMG_BYTES = 20 * 1024 * 1024  # SEC-21: yuklanadigan rasm uchun 20 MB shift


def _oversize_image(files):
    """SEC-21: ro'yxatdagi biror rasm 20 MB'dan katta bo'lsa nomini qaytaradi.

    Ko'p-varaqli yuklashда (getlist) har bir faylни tekshiradi — aks holда
    ulkan fayl xotira/diskни to'ldirib DoS qilardi (dekompressiya bombасидан
    tashqari, oddiy hajm ham).
    """
    for f in (files or []):
        if getattr(f, 'size', 0) and f.size > _MAX_IMG_BYTES:
            return getattr(f, 'name', 'rasm')
    return None


def _clean_text(s, max_len=None):
    """Strip HTML tags and normalize whitespace. Prevents stored XSS from CSV."""
    if not s:
        return ''
    import re
    # Remove anything that looks like an HTML tag
    s = re.sub(r'<[^>]*>', '', str(s))
    # Decode HTML entities
    import html
    s = html.unescape(s)
    s = s.strip()
    if max_len:
        s = s[:max_len]
    return s


@admin_required
def csv_import(request):
    """CSV import: mahsulot, mijoz yoki zaxira.

    Format va template'lar pastda. Birinchi qatordagi headerlar bizning
    columns'larga mos kelishi kerak.
    """
    entity = request.GET.get('entity') or request.POST.get('entity') or 'products'
    if entity not in ('products', 'customers', 'stock'):
        entity = 'products'

    if request.method == 'POST' and 'csv_file' in request.FILES:
        f = request.FILES['csv_file']
        # S6: 5MB hard limit on CSV uploads (DATA_UPLOAD_MAX_MEMORY_SIZE
        # is 10MB but a CSV that big = ~50K rows which Django can't process
        # in one request anyway — block early with a useful message)
        if f.size > 5 * 1024 * 1024:
            messages.error(request,
                f"Fayl juda katta ({f.size//1024} KB). "
                "Maksimum 5 MB. Faylni qismlarga bo'ling.")
            return redirect(f'/import/?entity={entity}')
        try:
            data = f.read().decode('utf-8-sig')  # handle BOM
        except UnicodeDecodeError:
            data = f.read().decode('cp1251', errors='ignore')
        reader = csv.DictReader(io.StringIO(data))
        rows = list(reader)
        if not rows:
            messages.error(request, "Bo'sh fayl yoki noto'g'ri format.")
            return redirect(f'/import/?entity={entity}')

        created = 0
        updated = 0
        failed = []

        try:
            with transaction.atomic():
                if entity == 'products':
                    for i, row in enumerate(rows, 2):  # row 2 = first data row
                        try:
                            name = _clean_text(row.get('name'), 200)
                            if not name:
                                raise ValueError("nom (name) bo'sh")
                            code = _clean_text(row.get('code'), 20).upper()
                            cat_name = _clean_text(row.get('category'), 120)
                            category = None
                            if cat_name:
                                category, _ = Category.objects.get_or_create(name=cat_name)
                            from decimal import Decimal
                            price = Decimal(row.get('default_sale_price') or '0')
                            markup = Decimal(row.get('markup_percent') or '40')
                            desc = _clean_text(row.get('description'), 1000)
                            ext_barcode = (row.get('external_barcode') or '').strip() or None
                            if ext_barcode:
                                clash = Product.objects.filter(
                                    external_barcode=ext_barcode
                                )
                                if code:
                                    clash = clash.exclude(code=code)
                                if clash.exists():
                                    raise ValueError(
                                        f"barcode {ext_barcode} boshqa mahsulotda mavjud"
                                    )
                            if code:
                                # STK-17: FAQAT CSV'да haqiqatan berilgan
                                # ustunlarni yangilaymiz. Ilgari to'liq defaults
                                # bloki tufayli, ikki ustunli (code+qty) eksport
                                # mavjud mahsulotning nomi/kategoriya/narxini
                                # bo'shatib, narxни 0 ga tushirardi.
                                _defaults = {'name': name}
                                if category is not None:
                                    _defaults['category'] = category
                                if ext_barcode:
                                    _defaults['external_barcode'] = ext_barcode
                                if (row.get('default_sale_price') or '').strip():
                                    _defaults['default_sale_price'] = price
                                if (row.get('markup_percent') or '').strip():
                                    _defaults['markup_percent'] = markup
                                if (row.get('description') or '').strip():
                                    _defaults['description'] = desc
                                p, was_created = Product.objects.update_or_create(
                                    code=code, defaults=_defaults,
                                )
                            else:
                                p = Product.objects.create(
                                    name=name, category=category,
                                    external_barcode=ext_barcode,
                                    default_sale_price=price,
                                    markup_percent=markup,
                                    description=desc,
                                )
                                was_created = True
                            if was_created:
                                created += 1
                            else:
                                updated += 1
                        except Exception as e:
                            failed.append({'row': i, 'reason': str(e)})

                elif entity == 'customers':
                    for i, row in enumerate(rows, 2):
                        try:
                            phone = _clean_text(row.get('phone'), 40)
                            name = _clean_text(row.get('name'), 120)
                            if not phone and not name:
                                raise ValueError("phone va name ikkalasi bo'sh")
                            tags = _clean_text(row.get('tags'), 200)
                            inn = _clean_text(row.get('inn'), 14)
                            note = _clean_text(row.get('note'), 500)
                            if phone:
                                c, was_created = Customer.objects.update_or_create(
                                    phone=phone,
                                    defaults={'name': name, 'tags': tags,
                                              'inn': inn, 'note': note},
                                )
                            else:
                                c = Customer.objects.create(
                                    name=name, tags=tags, inn=inn, note=note,
                                )
                                was_created = True
                            if was_created:
                                created += 1
                            else:
                                updated += 1
                        except Exception as e:
                            failed.append({'row': i, 'reason': str(e)})

                elif entity == 'stock':
                    for i, row in enumerate(rows, 2):
                        try:
                            code = (row.get('product_code') or '').strip().upper()
                            branch_name = (row.get('branch') or '').strip()
                            size = (row.get('size') or '—').strip() or '—'
                            color = (row.get('color') or '—').strip() or '—'
                            qty = int(row.get('quantity') or 0)
                            from decimal import Decimal
                            cost = Decimal(row.get('cost_price') or '0')
                            sale = Decimal(row.get('sale_price') or '0')

                            if not code or not branch_name or qty <= 0:
                                raise ValueError("code, branch, va quantity>0 kerak")
                            product = Product.objects.filter(code=code).first()
                            if not product:
                                raise ValueError(f"mahsulot {code} topilmadi")
                            branch = Branch.objects.filter(name=branch_name).first()
                            if not branch:
                                raise ValueError(f"filial '{branch_name}' topilmadi")

                            variant, _ = ProductVariant.objects.get_or_create(
                                product=product, size=size, color=color,
                            )
                            stock, was_created = BranchStock.objects.get_or_create(
                                variant=variant, branch=branch,
                                defaults={'cost_price': cost, 'sale_price': sale,
                                          'stock_count': qty},
                            )
                            if not was_created:
                                if cost > 0:  # STK-8 weighted-average
                                    stock.cost_price = weighted_cost(
                                        stock.stock_count if isinstance(stock.stock_count, int) else 0,
                                        stock.cost_price, qty, cost)
                                stock.stock_count = F('stock_count') + qty
                                if sale > 0: stock.sale_price = sale
                                stock.save()
                            # Create intake record
                            Intake.objects.create(
                                variant=variant, branch=branch,
                                quantity=qty, cost_per_unit=cost,
                                sale_price=sale or None,
                                supplier='CSV import',
                                received_by=request.user,
                            )
                            if was_created:
                                created += 1
                            else:
                                updated += 1
                        except Exception as e:
                            failed.append({'row': i, 'reason': str(e)})

                if failed and len(failed) == len(rows):
                    # Hammasi xato — rollback
                    raise ValueError(f"Hammasi xato: birinchi sabab — {failed[0]['reason']}")
        except Exception as e:
            messages.error(request, f"Import to'xtatildi: {e}")
            return redirect(f'/import/?entity={entity}')

        if created or updated:
            messages.success(request,
                f"Qo'shildi: {created}, yangilandi: {updated}, xato: {len(failed)}")
        if failed:
            preview = '; '.join(f"qator {f['row']}: {f['reason']}" for f in failed[:5])
            messages.warning(request, f"Xato qatorlar: {preview}"
                             + (f" (jami {len(failed)} ta)" if len(failed) > 5 else ''))
        return redirect(f'/import/?entity={entity}')

    return render(request, 'inventory/csv_import.html', {
        'entity': entity,
    })


@admin_required
def cashier_stats(request, user_id):
    """Sotuvchi performans dashboard'i: 30-kun kpi, kunlik trend,
    soatlar bo'yicha aktivlik, top mahsulotlar, komissiya."""
    seller = get_object_or_404(User, pk=user_id)

    since_30d = timezone.now() - timedelta(days=30)
    rev_expr = ExpressionWrapper(
        F('quantity') * F('sale_price') - F('line_discount'),
        output_field=DecimalField(max_digits=14, decimal_places=2)
    )
    cost_expr = ExpressionWrapper(
        F('quantity') * F('cost_at_sale'),
        output_field=DecimalField(max_digits=14, decimal_places=2)
    )

    sales_30d = Sale.objects.filter(sold_by=seller, sold_at__gte=since_30d)
    agg = sales_30d.aggregate(
        revenue=Sum(rev_expr),
        cost=Sum(cost_expr),
        qty=Sum('quantity'),
        txns=Count('transaction', distinct=True),
    )
    # SAL-6: chek chegirmasi sotuvchiga ANIQ tegishli (bitta chek — bitta
    # sotuvchi), shuning uchun tushumdan ayiriladi. Bu yerda ayirmaslik ikki
    # marta yanglishtirardi: FOYDA chegirma miqdorida oshib ko'rinardi
    # (300 000 − 180 000 = 120 000, aslida 240 000 − 180 000 = 60 000) va
    # KOMISSIYA kassir chegirma qilib bergan puldan ham hisoblanardi.
    _odisc_seller, _ = _order_discount_share(sales_30d)
    _, _, _disc_split = _order_discount_share(sales_30d, split=True)
    # RET-1: qaytarish tushumdan ham, tannarxdan ham chiqadi — aks holda
    # kassirning foydasi ham, komissiyasi ham qaytgan tovardan hisoblanardi.
    _ret_rev, _ret_cost, _ = _returns_adjustment(sales_30d)
    revenue = _net_rev(agg['revenue'], _odisc_seller) - float(_ret_rev)
    cost = float(_dec(agg['cost'] or 0) - _ret_cost)
    qty = agg['qty'] or 0
    txns = agg['txns'] or 0
    avg_ticket = revenue / txns if txns else 0
    profit = revenue - cost
    commission_pct = float(seller.commission_percent or 0)
    commission = revenue * commission_pct / 100 if commission_pct else 0

    # Refunds initiated by this user
    refunds_30d = Return.objects.filter(refunded_by=seller, refunded_at__gte=since_30d)
    refund_count = refunds_30d.count()
    refund_qty = refunds_30d.aggregate(s=Sum('quantity'))['s'] or 0
    refund_rate = (refund_qty / qty * 100) if qty else 0

    # Daily trend (last 30 days)
    from collections import OrderedDict
    daily = OrderedDict()
    today = timezone.localdate()
    for i in range(29, -1, -1):
        d = today - timedelta(days=i)
        daily[d.isoformat()] = 0
    _, _odisc_by_day = _order_discount_share(sales_30d, 'sold_at__date')
    for row in (sales_30d.annotate(d=F('sold_at__date')).values('d')
                .annotate(r=Sum(rev_expr))):
        key = row['d'].isoformat() if row['d'] else None
        if key and key in daily:
            # kun ham CHEK darajasidagi o'lchov — ulush aniq tushadi
            daily[key] = _net_rev(row['r'], _odisc_by_day.get(row['d']))
    daily_labels = [k[5:] for k in daily.keys()]
    daily_values = list(daily.values())

    # Hourly heatmap
    hour_qty = [0] * 24
    for sold_at, in sales_30d.values_list('sold_at'):
        h = sold_at.astimezone(timezone.get_current_timezone()).hour
        hour_qty[h] += 1
    hour_labels = [f'{h:02d}' for h in range(24)]

    # Top products — MAHSULOT qator darajasidagi o'lchov: egasining qoidasi
    # bo'yicha chek chegirmasi alohida tovarga tegishli emas, shuning uchun
    # qatorlar O'Z narxida qoladi va chegirma alohida ko'rsatiladi (SAL-6).
    top_products = list(sales_30d.values('variant__product__code', 'variant__product__name')
                        .annotate(qty=Sum('quantity'), revenue=Sum(rev_expr))
                        .order_by('-revenue')[:5])
    order_discount_total = float(_odisc_seller)

    # Sales per active day
    active_days = sales_30d.values('sold_at__date').distinct().count() or 1
    sales_per_day = txns / active_days

    # D3: Anomaly detection — compare this seller to peer average
    _peer_qs = Sale.objects.filter(sold_at__gte=since_30d).exclude(sold_by=seller)
    # D3 taqqoslash bir xil asosда bo'lsin — aks holда bu kassir sof, peer'lar
    # brutto bo'lib, chegirма ko'p qilgan kassir "yomon" ko'rinardi.
    _, _odisc_peer = _order_discount_share(_peer_qs, 'sold_by_id')
    peer_stats = (_peer_qs
                  .values('sold_by_id')
                  .annotate(rev=Sum(rev_expr), qty=Sum('quantity'),
                            txns=Count('transaction', distinct=True)))
    peers = [dict(p, rev=_net_rev(p['rev'], _odisc_peer.get(p['sold_by_id'])))
             for p in peer_stats]
    peer_count = len(peers)
    anomalies = []
    if peer_count >= 2:  # need enough peers for comparison
        peer_avg_tickets = [float(p['rev'] or 0) / p['txns'] if p['txns'] else 0
                            for p in peers]
        peer_avg_ticket = sum(peer_avg_tickets) / len(peer_avg_tickets) or 1
        peer_avg_qty_per_txn = sum((p['qty'] or 0) / p['txns'] if p['txns'] else 0
                                   for p in peers) / len(peers) or 1

        my_avg_ticket = float(avg_ticket or 0)
        my_qty_per_txn = (qty / txns) if txns else 0

        # Refund rate (peer avg)
        peer_refund_qty = Return.objects.filter(
            refunded_at__gte=since_30d
        ).exclude(refunded_by=seller).aggregate(s=Sum('quantity'))['s'] or 0
        peer_sold_qty = sum((p['qty'] or 0) for p in peers)
        peer_refund_rate = (peer_refund_qty / peer_sold_qty * 100) if peer_sold_qty else 0

        # 1) Average ticket significantly different
        if my_avg_ticket > peer_avg_ticket * 2:
            anomalies.append({
                'level': 'warning',
                'icon': 'arrow-up-circle',
                'title': "O'rtacha chek juda yuqori",
                'detail': f"Sizning: {my_avg_ticket:,.0f}, hamkasblar o'rtacha: {peer_avg_ticket:,.0f} so'm",
            })
        elif my_avg_ticket > 0 and my_avg_ticket < peer_avg_ticket * 0.4:
            anomalies.append({
                'level': 'warning',
                'icon': 'arrow-down-circle',
                'title': "O'rtacha chek juda past",
                'detail': f"Sizning: {my_avg_ticket:,.0f}, hamkasblar o'rtacha: {peer_avg_ticket:,.0f} so'm",
            })

        # 2) Refund rate much higher than peers
        if refund_rate > peer_refund_rate * 2 and refund_rate > 5:
            anomalies.append({
                'level': 'danger',
                'icon': 'exclamation-triangle',
                'title': "Qaytarish foizi juda yuqori",
                'detail': f"Sizning: {refund_rate:.1f}%, hamkasblar o'rtacha: {peer_refund_rate:.1f}%",
            })

        # 3) Very high txn count (productive) or very low
        if peers:
            peer_txns_avg = sum(p['txns'] for p in peers) / len(peers)
            if txns > peer_txns_avg * 3 and txns > 50:
                anomalies.append({
                    'level': 'info',
                    'icon': 'star-fill',
                    'title': "Yuqori sotuv hajmi",
                    'detail': f"Sizning: {txns} chek, hamkasblar o'rtacha: {peer_txns_avg:.0f} — top performer!",
                })

    return render(request, 'inventory/cashier_stats.html', {
        'order_discount_total': order_discount_total,
        'discount_split': _disc_split,   # DISC-1
        'seller': seller,
        'revenue': revenue, 'cost': cost, 'profit': profit,
        'qty': qty, 'txns': txns,
        'avg_ticket': avg_ticket,
        'commission_pct': commission_pct,
        'commission': commission,
        'refund_count': refund_count,
        'refund_qty': refund_qty,
        'refund_rate': refund_rate,
        'daily_labels': daily_labels,
        'daily_values': daily_values,
        'hour_labels': hour_labels,
        'hour_qty': hour_qty,
        'top_products': top_products,
        'active_days': active_days,
        'sales_per_day': sales_per_day,
        'anomalies': anomalies,
        'peer_count': peer_count,
    })


@admin_required
def reorder_page(request):
    """Smart reorder: past zaxiradagi mahsulotlar + tavsiya etilgan miqdor.

    Formula: 30-kunlik sotuvni 30 ga bo'lib kunlik o'rtacha velocity, undan
    keyin (BUFFER_DAYS = 14) ga ko'paytirib taklif miqdorini chiqaramiz.
    """
    BUFFER_DAYS = int(request.GET.get('buffer') or 14)
    LOW_THRESHOLD = int(request.GET.get('threshold') or 5)
    branch_id = request.GET.get('branch') or ''

    since_30d = timezone.now() - timedelta(days=30)

    stocks_qs = (BranchStock.objects
                 .filter(stock_count__lte=LOW_THRESHOLD)
                 .select_related('variant__product__category', 'branch'))
    if branch_id:
        try:
            stocks_qs = stocks_qs.filter(branch_id=int(branch_id))
        except ValueError:
            branch_id = ''

    # 30-day sales per (variant, branch) pair
    sales_map = {}
    for row in (Sale.objects.filter(sold_at__gte=since_30d)
                .values('variant_id', 'branch_id')
                .annotate(qty=Sum('quantity'))):
        sales_map[(row['variant_id'], row['branch_id'])] = row['qty'] or 0

    # D4: trend velocity — last 7 days vs previous 7 days. If accelerating,
    # adjust suggested order qty upward.
    last_7 = timezone.now() - timedelta(days=7)
    prev_7_start = timezone.now() - timedelta(days=14)
    last7_map = {}
    for row in (Sale.objects.filter(sold_at__gte=last_7)
                .values('variant_id', 'branch_id')
                .annotate(qty=Sum('quantity'))):
        last7_map[(row['variant_id'], row['branch_id'])] = row['qty'] or 0
    prev7_map = {}
    for row in (Sale.objects.filter(sold_at__gte=prev_7_start, sold_at__lt=last_7)
                .values('variant_id', 'branch_id')
                .annotate(qty=Sum('quantity'))):
        prev7_map[(row['variant_id'], row['branch_id'])] = row['qty'] or 0

    # Most recent supplier per product (from last intake)
    supplier_map = {}
    for row in (Intake.objects.filter(quantity__gt=0)
                .order_by('-received_at')
                .values('variant__product_id', 'supplier_ref__name', 'supplier')[:1000]):
        pid = row['variant__product_id']
        if pid not in supplier_map:
            name = row['supplier_ref__name'] or row['supplier'] or ''
            if name:
                supplier_map[pid] = name

    items = []
    for s in stocks_qs:
        sold30 = sales_map.get((s.variant_id, s.branch_id), 0)
        daily = sold30 / 30 if sold30 else 0
        # D4: trend factor — if last 7 days >> previous 7 days, demand is rising;
        # multiply daily by trend factor (capped at 2.0 to avoid wild over-orders)
        l7 = last7_map.get((s.variant_id, s.branch_id), 0)
        p7 = prev7_map.get((s.variant_id, s.branch_id), 0)
        if p7 > 0 and l7 > 0:
            trend_ratio = l7 / p7
            trend_factor = max(0.5, min(2.0, trend_ratio))
        elif l7 > 0 and p7 == 0:
            trend_factor = 1.5  # new product gaining traction
        else:
            trend_factor = 1.0
        adjusted_daily = daily * trend_factor

        # Target = buffer-days of stock; subtract what's already there
        target = adjusted_daily * BUFFER_DAYS
        suggested = max(0, int(round(target - s.stock_count)))
        if suggested == 0 and sold30 == 0 and s.stock_count > 0:
            # Not moving, has some stock — skip
            continue

        trend_label = ''
        if trend_factor > 1.2:
            trend_label = 'up'
        elif trend_factor < 0.8:
            trend_label = 'down'

        items.append({
            'stock': s,
            'product': s.variant.product,
            'variant': s.variant,
            'branch': s.branch,
            'sold_30d': sold30,
            'daily_avg': daily,
            'adjusted_daily': adjusted_daily,
            'trend_factor': trend_factor,
            'trend_label': trend_label,
            'last_7': l7,
            'prev_7': p7,
            'days_left': (s.stock_count / adjusted_daily) if adjusted_daily else None,
            'suggested': suggested,
            'cost': float(s.cost_price),
            'estimated_cost': suggested * float(s.cost_price),
            'last_supplier': supplier_map.get(s.variant.product_id, ''),
        })

    # Group by supplier
    by_supplier = {}
    for it in items:
        key = it['last_supplier'] or '(noma\'lum)'
        by_supplier.setdefault(key, []).append(it)
    suppliers_grouped = sorted(by_supplier.items(),
                               key=lambda kv: -sum(x['estimated_cost'] for x in kv[1]))

    # M5: aggregate per product (across branches) — single row per product
    # showing total stock across all branches, total suggested qty, etc.
    by_product = {}
    for it in items:
        pid = it['product'].id
        if pid not in by_product:
            by_product[pid] = {
                'product': it['product'],
                'last_supplier': it['last_supplier'],
                'stock_total': 0,
                'sold_30d_total': 0,
                'suggested_total': 0,
                'estimated_cost_total': 0.0,
                'branches': set(),
            }
        agg = by_product[pid]
        agg['stock_total'] += it['stock'].stock_count
        agg['sold_30d_total'] += it['sold_30d']
        agg['suggested_total'] += it['suggested']
        agg['estimated_cost_total'] += it['estimated_cost']
        agg['branches'].add(it['branch'].name)
    products_grouped = sorted(by_product.values(),
                              key=lambda x: -x['estimated_cost_total'])
    # Convert branches set to sorted list for template
    for p in products_grouped:
        p['branches'] = sorted(p['branches'])

    total_suggested_value = sum(it['estimated_cost'] for it in items)
    total_items = len(items)
    total_unique_products = len(by_product)

    view_mode = request.GET.get('view') or 'supplier'  # supplier | product

    return render(request, 'inventory/reorder.html', {
        'items': items,
        'by_supplier': suppliers_grouped,
        'products_grouped': products_grouped,
        'view_mode': view_mode,
        'total_items': total_items,
        'total_unique_products': total_unique_products,
        'total_suggested_value': total_suggested_value,
        'buffer_days': BUFFER_DAYS,
        'low_threshold': LOW_THRESHOLD,
        'branch_id': branch_id,
        'branches': Branch.objects.filter(is_active=True).order_by('name'),
    })


@admin_required
def supplier_list(request):
    if request.method == 'POST':
        action = request.POST.get('action') or 'create'
        if action == 'delete':
            try:
                pk = int(request.POST.get('pk') or 0)
                Supplier.objects.filter(pk=pk).delete()
                messages.success(request, "Yetkazib beruvchi o'chirildi.")
            except (TypeError, ValueError):
                pass
            return redirect('supplier_list')
        # create or edit
        try:
            pk = int(request.POST.get('pk') or 0)
        except (TypeError, ValueError):
            pk = 0
        instance = Supplier.objects.filter(pk=pk).first() if pk else None
        s = instance or Supplier()
        s.name = (request.POST.get('name') or '').strip()[:200]
        s.phone = (request.POST.get('phone') or '').strip()[:40]
        s.address = (request.POST.get('address') or '').strip()[:255]
        s.inn = (request.POST.get('inn') or '').strip()[:14]
        s.contact_person = (request.POST.get('contact_person') or '').strip()[:120]
        s.notes = (request.POST.get('notes') or '').strip()
        s.is_active = bool(request.POST.get('is_active'))
        if not s.name:
            messages.error(request, "Nom kerak.")
            return redirect('supplier_list')
        try:
            s.save()
            verb = 'Yangilandi' if instance else "Qo'shildi"
            messages.success(request, f"{verb}: {s.name}")
        except Exception as e:
            messages.error(request, f"Xato: {e}")
        return redirect('supplier_list')

    suppliers = Supplier.objects.order_by('-is_active', 'name')
    # Per-supplier aggregate: total intakes, last delivery date
    for s in suppliers:
        s.total_qty_received = (Intake.objects.filter(supplier_ref=s)
                                .aggregate(s=Sum('quantity'))['s'] or 0)
        s.last_intake = (Intake.objects.filter(supplier_ref=s)
                         .order_by('-received_at').first())
    return render(request, 'inventory/supplier_list.html',
                  {'suppliers': suppliers})


# ---------- INTAKE SESSION DETAIL ----------

# ---------- FAKTURA RASMIDAN QABUL (AI) ----------

@admin_required
def intake_photo(request):
    """Yetkazib beruvchi fakturasini suratga olib qabul qilish.

    AI faqat O'QIYDI — natija tahrirlanadigan jadvalga tushadi va
    foydalanuvchi tasdiqlagandan keyingina omborga yoziladi.
    """
    from .invoice_ai import is_enabled
    draft = None
    if request.GET.get('draft'):
        # SEC-22: faqat O'Z qoralamang — boshqa adminning qoralamasini o'qib/
        # tahrirlab/o'chirib bo'lmaydi.
        draft = InvoiceDraft.objects.filter(
            pk=request.GET['draft'], created_by=request.user).first()
    draft_data = None
    if draft:
        draft_data = dict(draft.payload or {})
        draft_data['id'] = draft.pk
        draft_data['image_url'] = draft.image.url if draft.image else ''
        # nakladnoy bir necha varaq bo'lishi mumkin
        draft_data['pages'] = [p.image.url for p in draft.pages.all() if p.image]
        if not draft_data['pages'] and draft_data['image_url']:
            draft_data['pages'] = [draft_data['image_url']]
    return render(request, 'inventory/intake_photo.html', {
        'branches': Branch.objects.filter(is_active=True).order_by('name'),
        'categories': Category.objects.order_by('name'),
        'suppliers': Supplier.objects.filter(is_active=True).order_by('name'),
        'product_names': list(Product.objects.order_by('name')
                              .values_list('name', flat=True)),
        'ai_enabled': is_enabled(),
        'drafts': (InvoiceDraft.objects.filter(created_by=request.user)  # SEC-22
                   .select_related('branch', 'created_by')
                   .exclude(pk=draft.pk if draft else 0)[:10]),
        'draft': draft,
        'draft_data': draft_data,
    })


def _draft_payload(payload):
    """Qoralama uchun tozalangan payload (hech qanday hisob-kitobsiz)."""
    rows = []
    for r in (payload.get('rows') or [])[:400]:
        rows.append({
            'name': str(r.get('name') or '')[:200],
            'product': str(r.get('product') or '')[:200],
            'type': str(r.get('type') or '')[:120],
            'size': str(r.get('size') or '')[:60],
            'barcode': str(r.get('barcode') or '')[:64],
            'qty': str(r.get('qty') or '')[:20],
            'cost': str(r.get('cost') or '')[:20],
            'marja': str(r.get('marja') or '')[:20],
            'sale': str(r.get('sale') or '')[:20],
            'category': str(r.get('category') or '')[:120],
            'line_sum': str(r.get('line_sum') or '')[:24],
            'qty_note': str(r.get('qty_note') or '')[:80],
            'hint': str(r.get('hint') or '')[:200],
        })
    return {
        'branch': str(payload.get('branch') or '')[:20],
        'supplier': str(payload.get('supplier') or '')[:200],
        'agent': str(payload.get('agent') or '')[:120],
        'agent_phone': str(payload.get('agent_phone') or '')[:40],
        'invoice_no': str(payload.get('invoice_no') or '')[:80],
        'date': str(payload.get('date') or '')[:20],
        'marja': str(payload.get('marja') or '')[:20],
        'total': str(payload.get('total') or '')[:24],
        'rows': rows,
    }


@admin_required
def intake_photo_draft(request):
    """Yarim ishni saqlash — telefonda boshlab, kompyuterda davom ettirish."""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST only'}, status=405)
    try:
        payload = _json.loads(request.POST.get('payload') or '{}')
    except ValueError:
        return JsonResponse({'ok': False, 'error': 'bad JSON'}, status=400)
    data = _draft_payload(payload)
    if not data['rows']:
        return JsonResponse({'ok': False, 'error': "Saqlash uchun qator yo'q."},
                            status=400)

    draft = None
    if payload.get('draft_id'):
        draft = InvoiceDraft.objects.filter(  # SEC-22: faqat o'z qoralamang
            pk=payload['draft_id'], created_by=request.user).first()
    if draft is None:
        draft = InvoiceDraft(created_by=request.user)
    draft.branch = Branch.objects.filter(
        pk=data['branch'] or 0, is_active=True).first()
    draft.supplier_text = data['supplier']
    draft.invoice_number = data['invoice_no']
    draft.payload = data
    draft.save()
    # Yangi yuborilgan sahifalarni qo'shamiz (eskilari saqlanib qoladi).
    files = request.FILES.getlist('image')
    _big = _oversize_image(files)   # SEC-21
    if _big:
        return JsonResponse({'ok': False,
                             'error': f"Rasm juda katta ({_big}) — 20 MB dan kichik bo'lsin."},
                            status=400)
    if files:
        start = draft.pages.count()
        for i, f in enumerate(files, start=1):
            InvoiceImage.objects.create(draft=draft, image=f, order=start + i)
        if not draft.image:
            draft.image = draft.pages.first().image
            draft.save(update_fields=['image'])
    return JsonResponse({'ok': True, 'id': draft.pk,
                         'pages': draft.pages.count(),
                         'rows': len(data['rows']),
                         'resume_url': reverse('intake_photo') + f'?draft={draft.pk}'})


@admin_required
def intake_photo_draft_delete(request, pk):
    if request.method != 'POST':
        return redirect('intake_photo')
    InvoiceDraft.objects.filter(pk=pk, created_by=request.user).delete()  # SEC-22
    messages.success(request, "Qoralama o'chirildi.")
    return redirect('intake_photo')


def _norm_name(s):
    return re.sub(r'\s+', ' ', (s or '').strip()).casefold()


_SIZE_RE = re.compile(
    r'\b\d+[.,]?\d*\s*(?:ml|мл|l|л|g|gr|г|гр|kg|кг|шт|sht|dona|pcs|pc|%)\.?\b',
    re.I)


def _match_catalog(rows):
    """AI ajratgan nomni do'kondagi mavjud mahsulotlarga moslaymiz.

    Maqsad: bir xil tovar har qabulda BIR XIL mahsulot ostiga tushsin.
    Mavjud mahsulot nomi qator nomining boshida tursa — o'sha nom olinadi
    (eng uzun moslik ustun), qolgani tur/hajmga bo'linadi.
    """
    catalog = sorted(
        ((_norm_name(p.name), p.name) for p in Product.objects.only('name')),
        key=lambda t: -len(t[0]))
    by_exact = {n: real for n, real in catalog}

    for r in rows:
        full = _norm_name(r.get('name'))
        ai_product = (r.get('product') or '').strip()
        r['matched'] = False

        # 1) AI ajratgan nom aynan bazadagi mahsulot bo'lsa — o'shani olamiz
        hit = by_exact.get(_norm_name(ai_product))
        if hit:
            r['product'] = hit
            r['matched'] = True
            continue

        # 2) Butun qator nomi bazadagi mahsulot bilan boshlansa — eng uzunini
        for norm, real in catalog:
            if not norm or len(norm) < 3:
                continue
            if full == norm or full.startswith(norm + ' '):
                rest = (r.get('name') or '')[len(norm):].strip(' -,.')
                r['product'] = real
                r['matched'] = True
                # qolganini tur/hajmga ajratamiz (AI bermagan bo'lsa)
                if rest and not (r.get('type') or r.get('size')):
                    m = _SIZE_RE.search(rest)
                    if m:
                        r['size'] = m.group(0).strip()
                        r['type'] = (rest[:m.start()] + ' '
                                     + rest[m.end():]).strip(' -,.')
                    else:
                        r['type'] = rest
                break
    return rows


@admin_required
def intake_photo_extract(request):
    """POST rasm -> AI o'qigan qatorlar (JSON). Hech narsa saqlanmaydi."""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST only'}, status=405)
    f = request.FILES.get('image')
    if not f:
        return JsonResponse({'ok': False, 'error': 'Rasm yuborilmadi.'}, status=400)
    if f.size > 20 * 1024 * 1024:
        return JsonResponse({'ok': False,
                             'error': "Rasm juda katta (20 MB dan kichik bo'lsin)."},
                            status=400)
    from .invoice_ai import extract_invoice, InvoiceAIError
    try:
        data = extract_invoice(f)
    except InvoiceAIError as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)
    except Exception:
        logger.exception('intake_photo_extract failed')
        return JsonResponse({'ok': False, 'error': 'Kutilmagan xato.'}, status=500)
    data['rows'] = _match_catalog(data.get('rows') or [])
    data['ok'] = True
    return JsonResponse(data)


@admin_required
def intake_photo_save(request):
    """Tekshirilgan qatorlarni omborga yozamiz (bitta qabul sessiyasi)."""
    from decimal import Decimal, InvalidOperation
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST only'}, status=405)
    try:
        payload = _json.loads(request.POST.get('payload') or '{}')
    except ValueError:
        return JsonResponse({'ok': False, 'error': 'bad JSON'}, status=400)

    branch = Branch.objects.filter(
        pk=payload.get('branch') or 0, is_active=True).first()
    if not branch:
        return JsonResponse({'ok': False, 'error': 'Filial tanlang.'}, status=400)

    # STK-15: takroriy saqlash (telefonда ikki marta bosish yoki sekin AI
    # javobida qayta yuborish) ombor sonini IKKI marta oshirardi. Bir xil
    # payload'ni 5 daqiqa ichiga bir marta qabul qilamiz (atomik cache.add).
    import hashlib as _hl
    from django.core.cache import cache as _cache
    _idem = (payload.get('idempotency_key') or '').strip()[:64]
    if not _idem:
        _idem = _hl.sha256((request.POST.get('payload') or '').encode('utf-8')).hexdigest()[:32]
    _idem_key = f'intake_save:{_idem}'
    _dup_resp = JsonResponse({'ok': True, 'duplicate': True, 'saved': 0,
                              'error': 'Bu qabul allaqachon saqlangan (takror bosildi).'})
    # Tez yo'l: bir workerда takror bosish (LocMemCache).
    if not _cache.add(_idem_key, True, 300):
        return _dup_resp
    # S4: ASOSIY himoya — DB unikал kaliti. LocMemCache har Gunicorn workerда
    # alohида, shu bois turli workerlarга tushган ikki tegish ikkalasи ham
    # cache'дан o'tar edi. IntakeSession.idempotency_key esa BUTUN bazada yagona.
    if IntakeSession.objects.filter(idempotency_key=_idem).exists():
        return _dup_resp

    # Har qatorning o'z kategoriyasi bo'lishi mumkin — bitta yetkazib beruvchi
    # turli toifadagi mahsulot keltiradi. Ro'yxatda yo'q nom yozilsa, yangi
    # kategoriya yaratiladi (prefiks avtomatik chiqadi).
    cat_cache = {c.name.casefold(): c for c in Category.objects.all()}

    def resolve_category(text):
        text = (text or '').strip()[:120]
        if not text:
            return None
        key = text.casefold()
        if key not in cat_cache:
            cat_cache[key] = Category.objects.create(name=text)
        return cat_cache[key]

    clean, errors, seen_barcodes, seen_keys = [], [], set(), set()
    for idx, r in enumerate(payload.get('rows') or [], 1):
        # Mahsulot = guruh (Colgate), tur+hajm = variant (Fresh / 100ml).
        # Bir xil "product" nomli qatorlar BITTA mahsulot ostiga yig'iladi.
        name = smart_title((r.get('product') or r.get('name') or '').strip())
        if not name:
            continue
        vtype = smart_title((r.get('type') or '').strip())[:120]
        vsize = smart_title((r.get('size') or '').strip())[:60]
        key = (_norm_name(name), _norm_name(vtype), _norm_name(vsize))
        if key in seen_keys:
            errors.append(
                f"{idx}-qator: {name} / {vtype or '—'} / {vsize or '—'} takrorlangan.")
            continue
        seen_keys.add(key)
        try:
            qty = int(float(r.get('qty') or 0))
        except (TypeError, ValueError):
            qty = 0
        if qty <= 0:
            # STK-14: jimgina tashlamaymiz — xodim qatorни ko'rib chiqsin.
            errors.append(f"{idx}-qator: {name} — soni 0/noto'g'ri, o'tkazilmadi.")
            continue
        try:
            cost = Decimal(str(r.get('cost') or 0)).quantize(Decimal('0.01'))
        except (InvalidOperation, TypeError, ValueError):
            cost = Decimal('0')
        try:
            sale = Decimal(str(r.get('sale') or 0)).quantize(Decimal('0.01'))
        except (InvalidOperation, TypeError, ValueError):
            sale = Decimal('0')
        if cost < 0 or sale < 0:
            errors.append(f"{idx}-qator: {name} — manfiy narx/tannarx, o'tkazilmadi.")
            continue
        barcode = (r.get('barcode') or '').strip()[:64]
        if barcode:
            if barcode in seen_barcodes:
                errors.append(f"{idx}-qator: shtrix-kod jadvalda takror ({barcode}).")
                continue
            seen_barcodes.add(barcode)
        clean.append({'name': name[:200], 'type': vtype, 'size': vsize,
                      'qty': qty, 'cost': cost, 'sale': sale,
                      'barcode': barcode, 'variant': None,
                      'category': (r.get('category') or '').strip()[:120]})

    if not clean and not errors:
        _cache.delete(_idem_key)   # STK-15: xato — qayta urinishga ruxsat
        return JsonResponse({'ok': False,
                             'error': "Saqlash uchun qator yo'q (miqdor > 0 bo'lsin)."},
                            status=400)

    # Shtrix-kod bazada bo'lsa — o'sha turga qabul qilamiz (kod = haqiqat manbai).
    # Boshqa mahsulotga biriktirilgan bo'lsa — jim o'g'irlamay, xato qaytaramiz.
    for r in clean:
        if not r['barcode']:
            continue
        v = (ProductVariant.objects.filter(barcode=r['barcode'])
             .select_related('product').first())
        if v:
            r['variant'] = v
            # Foydalanuvchi so'rovi: kod bir xil, lekin nakladnoydagi nom
            # boshqacha bo'lsa — MAHSULOT NOMI AI/nakladnoy nomiga YANGILANADI
            # (r['name'] o'zgarmaydi; saqlashda mahsulot qayta nomlanadi).
            r['old_name'] = v.product.name
            continue
        p_clash = Product.objects.filter(external_barcode=r['barcode']).first()
        if p_clash:
            errors.append(f"Shtrix-kod {r['barcode']} "
                          f"'{p_clash.name}' mahsulotiga biriktirilgan.")

    if errors:
        _cache.delete(_idem_key)   # STK-15: xato — qayta urinishga ruxsat
        return JsonResponse({'ok': False, 'error': ' '.join(errors[:5])}, status=400)

    supplier_text = (payload.get('supplier') or '').strip()[:200]
    supplier_obj = (Supplier.objects.filter(name__iexact=supplier_text).first()
                    if supplier_text else None)
    agent_name = (payload.get('agent') or '').strip()[:120]
    agent_phone = (payload.get('agent_phone') or '').strip()[:40]

    # SEC-21: rasm hajmini atomik blokdan OLDIN tekshiramiz — aks holда
    # blok ichida return commit qilib, ulkan fayl bilan sessiya saqlanardi.
    _big = _oversize_image(request.FILES.getlist('image'))
    if _big:
        return JsonResponse({'ok': False,
                             'error': f"Rasm juda katta ({_big}) — 20 MB dan kichik bo'lsin."},
                            status=400)

    with transaction.atomic():
        session = IntakeSession.objects.create(
            branch=branch,
            supplier=supplier_obj,
            supplier_text='' if supplier_obj else supplier_text,
            received_by=request.user,
            invoice_number=(payload.get('invoice_no') or '').strip()[:80],
            agent_name=agent_name,
            agent_phone=agent_phone,
            note=(payload.get('note') or '').strip()[:500] or 'Faktura rasmidan (AI)',
            idempotency_key=_idem,   # S4: DB unikал kaliti (cross-worker himoya)
        )
        # Yetkazib beruvchi kartochkasi bo'sh bo'lsa — agent ma'lumoti bilan
        # to'ldiramiz. Mavjud qiymatlar hech qachon ustidan yozilmaydi.
        if supplier_obj:
            fill = []
            if agent_name and not supplier_obj.contact_person:
                supplier_obj.contact_person = agent_name
                fill.append('contact_person')
            if agent_phone and not supplier_obj.phone:
                supplier_obj.phone = agent_phone
                fill.append('phone')
            if fill:
                supplier_obj.save(update_fields=fill)
        # Faktura sanasi berilgan bo'lsa — sessiya sanasini o'shanga qo'yamiz
        d = (payload.get('date') or '').strip()
        if d:
            try:
                dt = datetime.strptime(d, '%Y-%m-%d')
                session.received_at = timezone.make_aware(
                    dt.replace(hour=12), timezone.get_current_timezone())
                session.save(update_fields=['received_at'])
            except (ValueError, TypeError):
                pass
        draft = (InvoiceDraft.objects.filter(  # SEC-22: faqat o'z qoralamang
                    pk=payload.get('draft_id'), created_by=request.user).first()
                 if payload.get('draft_id') else None)
        # --- faktura sahifalari (nakladnoy bir necha varaq bo'lishi mumkin) ---
        files = request.FILES.getlist('image')
        order = 0
        for f in files:
            order += 1
            InvoiceImage.objects.create(session=session, image=f, order=order)
        if draft:
            # Qoralamadagi sahifalar serverda turibdi — faylni ko'chirmasdan
            # shunchaki sessiyaga biriktiramiz.
            for pg in draft.pages.all():
                order += 1
                pg.draft = None
                pg.session = session
                pg.order = order
                pg.save(update_fields=['draft', 'session', 'order'])
        first = session.pages.first()
        if first:
            session.invoice_image.name = first.image.name
            session.save(update_fields=['invoice_image'])
        elif draft and draft.image:
            session.invoice_image.name = draft.image.name
            session.save(update_fields=['invoice_image'])

        # Bir xil nomli qatorlar bitta mahsulot ostiga — har biri alohida tur.
        # Kategoriya mahsulotga tegishli: guruhdagi birinchi to'ldirilgani.
        products = {}

        def group_product(row):
            key = _norm_name(row['name'])
            if key in products:
                return products[key]
            product = Product.objects.filter(name__iexact=row['name']).first()
            if product is None:
                cat = next((resolve_category(x['category']) for x in clean
                            if _norm_name(x['name']) == key and x['category']),
                           None)
                product = Product.objects.create(
                    name=row['name'], category=cat,
                    default_sale_price=row['sale'])
            elif row['sale'] > 0 and not product.default_sale_price:
                product.default_sale_price = row['sale']
                product.save(update_fields=['default_sale_price'])
            products[key] = product
            return product

        created = 0
        for r in clean:
            variant = r['variant']
            if variant is not None:
                # Shtrix-kod orqali topildi — mavjud turga qo'shamiz
                product = variant.product
                # Nakladnoy (AI) nomi boshqacha bo'lsa — AVTOMATIK qayta
                # nomlamaymiz. Muqobil nomni pending_name'ga yozamiz; Mahsulotlar
                # sahifasida "2 xil nom" ogohlantirishi chiqadi va foydalanuvchi
                # qaysi nomni qoldirishni o'zi tanlaydi.
                ai_name = (r.get('name') or '').strip()[:200]
                if (ai_name and _norm_name(ai_name) != _norm_name(product.name)
                        and _norm_name(ai_name) != _norm_name(product.pending_name)):
                    product.pending_name = ai_name
                    product.save(update_fields=['pending_name'])
                    AuditLog.objects.create(
                        user=request.user,
                        username_snapshot=request.user.username,
                        action=AuditLog.Action.UPDATE,
                        model_name='Product', object_id=str(product.pk),
                        object_repr=(f"Nom ziddiyati (nakladnoy): "
                                     f"'{product.name}' ↔ '{ai_name}'"))
                products.setdefault(_norm_name(product.name), product)
                if r['sale'] > 0 and not product.default_sale_price:
                    product.default_sale_price = r['sale']
                    product.save(update_fields=['default_sale_price'])
            else:
                product = group_product(r)
                variant, _ = ProductVariant.objects.get_or_create(
                    product=product, size=r['size'], color=r['type'])
                if r['barcode'] and not variant.barcode:
                    variant.barcode = r['barcode']
                    variant.save(update_fields=['barcode'])
            # Sotuv narxi boshqacha bo'lsa — alohida tur (o'z kodi bilan)
            variant = resolve_price_variant(variant, branch, r['sale'])
            stock, _ = BranchStock.objects.get_or_create(
                variant=variant, branch=branch,
                defaults={'cost_price': r['cost'], 'sale_price': r['sale']})
            stock.cost_price = weighted_cost(  # STK-8
                stock.stock_count if isinstance(stock.stock_count, int) else 0,
                stock.cost_price, r['qty'], r['cost'])
            if r['sale'] > 0:
                stock.sale_price = r['sale']
            stock.stock_count = F('stock_count') + r['qty']
            stock.save()
            Intake.objects.create(
                session=session, variant=variant, branch=branch,
                quantity=r['qty'], cost_per_unit=r['cost'],
                sale_price=r['sale'] or None,
                supplier=supplier_text, note='Faktura rasmidan (AI)',
                received_by=request.user)
            created += 1

        if draft:
            # Rasm(lar) sessiyaga o'tdi — qoralamani o'chirganda fayl o'chmasin
            draft.image = None
            draft.save(update_fields=['image'])
            draft.delete()

    from .notifications import notify_intake_session
    notify_intake_session(session)
    return JsonResponse({
        'ok': True, 'session_id': session.pk, 'lines': created,
        'products': len(products),
        'redirect': reverse('intake_session_detail', args=[session.pk]),
    })


@admin_required
def intake_session_detail(request, pk):
    session = get_object_or_404(
        IntakeSession.objects.select_related('branch', 'supplier', 'received_by')
        .prefetch_related('intakes__variant__product'), pk=pk
    )
    if request.method == 'POST' and 'invoice_image' in request.FILES:
        session.invoice_image = request.FILES['invoice_image']
        session.save(update_fields=['invoice_image'])
        messages.success(request, "Faktura rasmi yuklandi.")
        return redirect('intake_session_detail', pk=session.pk)
    return render(request, 'inventory/intake_session_detail.html', {'session': session})


# ---------- STOCKTAKE (physical count vs system) ----------

@admin_required
def stocktake_list(request):
    """Inventarizatsiyalar + filtrlar + ochiq sessiyalar banner."""
    status = request.GET.get('status') or ''
    branch_id = request.GET.get('branch') or ''

    qs = (Stocktake.objects.select_related('branch', 'started_by', 'applied_by'))
    if status:
        qs = qs.filter(status=status)
    if branch_id:
        try:
            qs = qs.filter(branch_id=int(branch_id))
        except ValueError:
            branch_id = ''
    sessions = list(qs.order_by('-started_at')[:50])

    # Variance summary per session
    for s in sessions:
        agg = s.counts.aggregate(
            variance=Sum(F('counted_qty') - F('system_qty')),
            cnt=Count('id'),
        )
        s.variance_total = agg['variance'] or 0
        s.count_lines = agg['cnt'] or 0

    open_sessions = Stocktake.objects.filter(status=Stocktake.Status.OPEN) \
        .select_related('branch', 'started_by').order_by('-started_at')

    return render(request, 'inventory/stocktake_list.html', {
        'sessions': sessions,
        'open_sessions': open_sessions,
        'status': status,
        'branch_id': branch_id,
        'branches': Branch.objects.filter(is_active=True).order_by('name'),
        'status_choices': Stocktake.Status.choices,
    })


@admin_required
def stocktake_create(request):
    branches = Branch.objects.filter(is_active=True)
    if request.method == 'POST':
        try:
            branch_id = int(request.POST.get('branch') or 0)
        except ValueError:
            branch_id = 0
        branch = Branch.objects.filter(pk=branch_id, is_active=True).first()
        if not branch:
            messages.error(request, "Filial tanlang.")
            return redirect('stocktake_create')
        # STK-9: bitta filialда bir vaqtда faqat BITTA ochiq inventarizatsiya.
        # Aks holда ikkita ochiq sessiya bir-birining tuzatishini bekor qilardi.
        existing = Stocktake.objects.filter(
            branch=branch, status=Stocktake.Status.OPEN).first()
        if existing:
            messages.warning(request,
                f"{branch.name}да allaqachon ochiq inventarizatsiya bor "
                f"(#{existing.pk}). Avval uni yakunlang.")
            return redirect('stocktake_detail', pk=existing.pk)
        # Snapshot every variant with stock>0 in this branch
        with transaction.atomic():
            session = Stocktake.objects.create(
                branch=branch, started_by=request.user,
                note=(request.POST.get('note') or '').strip()[:200],
            )
            stocks = BranchStock.objects.filter(branch=branch).select_related('variant')
            counts = [
                StocktakeCount(session=session, variant=s.variant,
                               system_qty=s.stock_count, counted_qty=s.stock_count)
                for s in stocks
            ]
            StocktakeCount.objects.bulk_create(counts)
        messages.success(request, f"Inventarizatsiya boshlandi: {len(counts)} ta variant.")
        return redirect('stocktake_detail', pk=session.pk)
    return render(request, 'inventory/stocktake_create.html', {'branches': branches})


@admin_required
def stocktake_detail(request, pk):
    session = get_object_or_404(
        Stocktake.objects.select_related('branch', 'started_by', 'applied_by'),
        pk=pk,
    )
    counts = (session.counts.select_related('variant__product')
              .order_by('variant__product__code', 'variant__size'))

    if request.method == 'POST' and session.status == Stocktake.Status.OPEN:
        action = request.POST.get('action')
        if action == 'save':
            # Update counted_qty for each row
            with transaction.atomic():
                for c in counts:
                    val = request.POST.get(f'count[{c.pk}]')
                    if val is None or val == '':
                        continue
                    try:
                        c.counted_qty = max(0, int(val))
                        c.save(update_fields=['counted_qty'])
                    except ValueError:
                        pass
            messages.success(request, "Hisoblangan miqdorlar saqlandi.")
            return redirect('stocktake_detail', pk=session.pk)

        if action == 'apply':
            # H6 fix: lock the session row so concurrent apply attempts can't
            # both succeed. select_for_update inside atomic() blocks the second
            # transaction until the first commits.
            with transaction.atomic():
                locked = (Stocktake.objects
                          .select_for_update()
                          .filter(pk=session.pk, status=Stocktake.Status.OPEN)
                          .first())
                if not locked:
                    messages.error(request,
                        "Bu inventarizatsiya allaqachon tasdiqlangan yoki bekor qilingan.")
                    return redirect('stocktake_detail', pk=session.pk)
                # Re-fetch counts under the lock window
                fresh_counts = list(locked.counts.select_related('variant__product'))
                # D2: "Tasdiqlash" bosilганда ham formadagi HOZIRGI kataklarни
                # saqlaymiz. Aks holда "Saqlash" bosilmagan bo'lса counted_qty ==
                # system_qty bo'lib, barcha delta 0 chiqar, sanoq yo'qolib
                # "tasdiqlandi" degan yolg'on xabar chiqardi.
                for c in fresh_counts:
                    _v = request.POST.get(f'count[{c.pk}]')
                    if _v is None or _v == '':
                        continue
                    try:
                        _iv = max(0, int(_v))
                    except (ValueError, TypeError):
                        continue
                    if _iv != c.counted_qty:
                        c.counted_qty = _iv
                        c.save(update_fields=['counted_qty'])
                adjustments = []
                for c in fresh_counts:
                    bs = BranchStock.objects.filter(
                        variant=c.variant, branch=locked.branch
                    ).select_for_update().first()
                    if not bs:
                        continue
                    # STK-9: MUTLAQ yozish EMAS — sanashда topilgan FARQ (counted −
                    # system_qty) HOZIRGI qoldiqqa qo'shiladi. Ilgari counted_qty
                    # to'g'ridan-to'g'ri yozilib, sanash bilan tasdiq orasidagi
                    # BARCHA sotuvlar jimgina bekor bo'lardi (dushanba sanog'ini
                    # jumada qo'llasa — bir haftalik savdo yo'qolardi).
                    delta = int(c.counted_qty) - int(c.system_qty)
                    if delta != 0:
                        bs.stock_count = F('stock_count') + delta
                        bs.save(update_fields=['stock_count'])
                        adjustments.append({
                            'code': c.variant.product.code,
                            'variant': f"{c.variant.size or ''}/{c.variant.color or ''}",
                            'system': int(c.system_qty), 'counted': int(c.counted_qty),
                            'delta': delta})
                locked.status = Stocktake.Status.APPLIED
                locked.applied_by = request.user
                locked.applied_at = timezone.now()
                locked.save()
                # STK-9: har tasdiqni AUDIT LOGGA yozamiz (ombor tuzatishlari
                # ilgari umuman iz qoldirmasdi).
                AuditLog.objects.create(
                    user=request.user, username_snapshot=request.user.username,
                    action=AuditLog.Action.UPDATE, model_name='Stocktake',
                    object_id=str(locked.pk),
                    object_repr=(f"Inventarizatsiya #{locked.pk} tasdiqlandi "
                                 f"({locked.branch.name}): {len(adjustments)} tuzatish")[:300],
                    changes={'branch': locked.branch.name,
                             'adjustments': adjustments[:500]})
            messages.success(request, "Inventarizatsiya tasdiqlandi va ombor yangilandi.")
            return redirect('stocktake_detail', pk=session.pk)

    diffs = sum(1 for c in counts if c.diff != 0)
    return render(request, 'inventory/stocktake_detail.html', {
        'session': session, 'counts': counts, 'diff_count': diffs,
    })


# ---------- TRANSFERS (inter-branch stock moves) ----------

@admin_required
def transfer_list(request):
    """Ko'chirishlar ro'yxati + filtrlar + overdue alert + status counts."""
    status = request.GET.get('status') or ''
    branch_id = request.GET.get('branch') or ''

    qs = (Transfer.objects.select_related('from_branch', 'to_branch',
                                          'created_by', 'received_by')
          .prefetch_related('lines'))

    if status:
        qs = qs.filter(status=status)
    if branch_id:
        try:
            bid = int(branch_id)
            qs = qs.filter(Q(from_branch_id=bid) | Q(to_branch_id=bid))
        except ValueError:
            branch_id = ''

    transfers = list(qs.order_by('-created_at')[:100])

    # Per-status counts (all transfers, not filtered)
    counts = {s: 0 for s, _ in Transfer.Status.choices}
    for s, n in (Transfer.objects.values_list('status')
                 .annotate(n=Count('id')).values_list('status', 'n')):
        counts[s] = n

    # Overdue alert: in_transit older than 3 days
    overdue_cutoff = timezone.now() - timedelta(days=3)
    overdue_transfers = list(Transfer.objects
        .filter(status=Transfer.Status.IN_TRANSIT, dispatched_at__lt=overdue_cutoff)
        .select_related('from_branch', 'to_branch')
        .order_by('dispatched_at')[:5])

    return render(request, 'inventory/transfer_list.html', {
        'transfers': transfers,
        'status': status,
        'branch_id': branch_id,
        'branches': Branch.objects.filter(is_active=True).order_by('name'),
        'status_choices': Transfer.Status.choices,
        'counts': counts,
        'overdue_transfers': overdue_transfers,
    })


@admin_required
def transfer_create(request):
    branches = Branch.objects.filter(is_active=True)
    if request.method == 'POST':
        try:
            from_id = int(request.POST.get('from_branch') or 0)
            to_id = int(request.POST.get('to_branch') or 0)
        except ValueError:
            messages.error(request, "Filial tanlanmagan."); return redirect('transfer_create')
        if from_id == to_id:
            messages.error(request, "Bir xil filialga ko'chirib bo'lmaydi.")
            return redirect('transfer_create')
        from_branch = Branch.objects.filter(pk=from_id).first()
        to_branch = Branch.objects.filter(pk=to_id).first()
        if not from_branch or not to_branch:
            messages.error(request, "Filiallar topilmadi."); return redirect('transfer_create')

        # Parse line items: qty[<variant_id>]
        lines_data = []
        for key, value in request.POST.items():
            if not key.startswith('qty[') or not key.endswith(']'):
                continue
            try:
                variant_id = int(key[4:-1])
                qty = int(value)
            except ValueError:
                continue
            if qty <= 0:
                continue
            lines_data.append((variant_id, qty))

        if not lines_data:
            messages.error(request, "Birorta tovar miqdori kiritilmagan.")
            return redirect('transfer_create')

        # Validate enough stock in source branch
        problems = []
        for variant_id, qty in lines_data:
            stock = BranchStock.objects.filter(variant_id=variant_id, branch=from_branch).first()
            available = stock.stock_count if stock else 0
            if qty > available:
                v = ProductVariant.objects.get(pk=variant_id)
                problems.append(f"{v.product.code} {v.size}/{v.color}: {from_branch.name}da {available} bor, so'rov {qty}")
        if problems:
            for p in problems:
                messages.error(request, p)
            return redirect('transfer_create')

        with transaction.atomic():
            t = Transfer.objects.create(
                from_branch=from_branch, to_branch=to_branch,
                created_by=request.user,
                status=Transfer.Status.IN_TRANSIT,
                dispatched_at=timezone.now(),
                note=(request.POST.get('note') or '').strip()[:200],
            )
            for variant_id, qty in lines_data:
                TransferLine.objects.create(transfer=t, variant_id=variant_id, quantity=qty)
                # Decrement source branch immediately on dispatch
                stock = BranchStock.objects.filter(
                    variant_id=variant_id, branch=from_branch).first()
                stock.stock_count = F('stock_count') - qty
                stock.save()

        messages.success(request, f"Ko'chirish #{t.pk} yo'lga chiqarildi.")
        return redirect('transfer_detail', pk=t.pk)

    return render(request, 'inventory/transfer_create.html', {'branches': branches})


@admin_required
def transfer_detail(request, pk):
    t = get_object_or_404(
        Transfer.objects.select_related('from_branch', 'to_branch',
                                         'created_by', 'received_by')
                         .prefetch_related('lines__variant__product'),
        pk=pk,
    )
    return render(request, 'inventory/transfer_detail.html', {'transfer': t})


@admin_required
def transfer_receive(request, pk):
    t = get_object_or_404(Transfer, pk=pk)
    if t.status != Transfer.Status.IN_TRANSIT:
        messages.warning(request, "Bu ko'chirish allaqachon yopilgan.")
        return redirect('transfer_detail', pk=t.pk)

    if request.method == 'POST':
        with transaction.atomic():
            for line in t.lines.all():
                # R8: ko'chirilayotgan tovarning TANNARXI yo'qolmasin. Manba
                # filialdagi tannarxni olib, qabul qiluvchi filial zaxirasiga
                # o'rtacha-tortilgan (STK-8) qo'shamiz. Ilgari yangi qator
                # cost_price=0 bilan yaratilib, tovar "tekin" ko'rinar edi.
                src = (BranchStock.objects.filter(
                    variant=line.variant, branch=t.from_branch).first())
                src_cost = src.cost_price if src else 0
                dst = (BranchStock.objects.select_for_update().filter(
                    variant=line.variant, branch=t.to_branch).first())
                if dst is None:
                    BranchStock.objects.create(
                        variant=line.variant, branch=t.to_branch,
                        stock_count=line.quantity,
                        cost_price=src_cost,
                        sale_price=(src.sale_price if src else 0),
                    )
                else:
                    dst.cost_price = weighted_cost(   # STK-8
                        dst.stock_count, dst.cost_price,
                        line.quantity, src_cost)
                    dst.stock_count = F('stock_count') + line.quantity
                    dst.save(update_fields=['stock_count', 'cost_price'])
            t.status = Transfer.Status.RECEIVED
            t.received_by = request.user
            t.received_at = timezone.now()
            t.save()
        messages.success(request, f"Ko'chirish #{t.pk} qabul qilindi.")
        return redirect('transfer_detail', pk=t.pk)

    return render(request, 'inventory/transfer_receive.html', {'transfer': t})


# ---------- STOCK WRITE-OFF (STK-1) ----------

@admin_required
def writeoff_list(request):
    """Tovarni hisobdan chiqarish (shikast/yo'qolish/...) — faqat admin.

    GET: forma (filial, shtrix-kod, miqdor, sabab, izoh) + so'nggi yozuvlar.
    POST: variantni shtrix-kod bo'yicha topadi, zaxirani atomik kamaytiradi,
    StockWriteOff + AuditLog yozadi. Bu yozuv ombor tenglamasiga kiradi.
    """
    branches = Branch.objects.filter(is_active=True).order_by('name')

    if request.method == 'POST':
        raw = (request.POST.get('code') or '').strip()
        try:
            branch_id = int(request.POST.get('branch') or 0)
        except ValueError:
            branch_id = 0
        try:
            qty = int(request.POST.get('quantity') or 0)
        except ValueError:
            qty = 0
        reason = (request.POST.get('reason') or '').strip()
        note = (request.POST.get('note') or '').strip()[:200]

        valid_reasons = {r for r, _ in StockWriteOff.Reason.choices}
        branch = Branch.objects.filter(pk=branch_id).first()
        # Variantni shtrix-kod bo'yicha aniqlaymiz (jismoniy tovardagi kod)
        variant = (ProductVariant.objects.filter(barcode=raw)
                   .select_related('product').first())
        if not variant:
            variant = (ProductVariant.objects.filter(product__external_barcode=raw)
                       .select_related('product').first())

        if not branch:
            messages.error(request, "Filialni tanlang.")
        elif not variant:
            messages.error(request, f"Shtrix-kod topilmadi: {raw}. Tovar kodini skanerlang.")
        elif reason not in valid_reasons:
            messages.error(request, "Sababni tanlang.")
        elif qty <= 0:
            messages.error(request, "Miqdor 0 dan katta bo'lsin.")
        else:
            with transaction.atomic():
                stock = (BranchStock.objects.select_for_update()
                         .filter(variant=variant, branch=branch).first())
                have = stock.stock_count if stock else 0
                if qty > have:
                    messages.error(request,
                        f"Omborда faqat {have} ta bor, so'rov {qty}.")
                    return redirect('writeoff_list')
                stock.stock_count = F('stock_count') - qty
                stock.save(update_fields=['stock_count'])
                wo = StockWriteOff.objects.create(
                    variant=variant, branch=branch, quantity=qty,
                    reason=reason, note=note,
                    cost_at_writeoff=stock.cost_price or 0,
                    created_by=request.user,
                )
                AuditLog.objects.create(
                    user=request.user,
                    username_snapshot=request.user.username,
                    action=AuditLog.Action.DELETE,
                    model_name='StockWriteOff',
                    object_id=str(wo.pk),
                    object_repr=(f"{variant.product.code} "
                                 f"{variant.size or ''}/{variant.color or ''} × {qty} "
                                 f"— {dict(StockWriteOff.Reason.choices).get(reason, reason)}")[:300],
                    changes={'branch': branch.name, 'qty': qty, 'reason': reason,
                             'note': note, 'loss_value': float(wo.loss_value)},
                )
            messages.success(request,
                f"Hisobdan chiqarildi: {variant.product.code} × {qty} "
                f"({dict(StockWriteOff.Reason.choices).get(reason, reason)}).")
            return redirect('writeoff_list')

    recent = (StockWriteOff.objects
              .select_related('variant__product', 'branch', 'created_by')
              .order_by('-created_at')[:100])
    # 30 kunlik jami yo'qotish qiymati
    since = timezone.now() - timedelta(days=30)
    loss_30d = sum((w.loss_value for w in StockWriteOff.objects
                    .filter(created_at__gte=since)), 0)
    return render(request, 'inventory/writeoff.html', {
        'branches': branches,
        'reasons': StockWriteOff.Reason.choices,
        'recent': recent,
        'loss_30d': loss_30d,
    })


# ---------- SHIFTS ----------

def _open_shift_for(branch):
    """Returns the single open shift for a branch, or None."""
    return Shift.objects.filter(branch=branch, status=Shift.Status.OPEN).first()


@login_required
def shift_open(request):
    branch = _user_branch_or_403(request)
    if branch is None:
        messages.error(request, "Filial biriktirilmagan.")
        return redirect('lookup')

    existing = _open_shift_for(branch)
    if existing:
        messages.info(request, f"Ochiq smen allaqachon bor (#{existing.pk}).")
        return redirect('pos_terminal')

    if request.method == 'POST':
        try:
            opening_cash = max(0, float(request.POST.get('opening_cash') or 0))
        except ValueError:
            opening_cash = 0
        Shift.objects.create(
            branch=branch, opened_by=request.user,
            opening_cash=opening_cash,
            note=(request.POST.get('note') or '').strip()[:200],
        )
        messages.success(request, f"Smen ochildi (boshlang'ich naqd: {opening_cash:,.0f} so'm).")
        return redirect('pos_terminal')

    # Hint: oxirgi yopilgan smen kassasi
    last_closed = (Shift.objects.filter(branch=branch, status=Shift.Status.CLOSED)
                   .order_by('-closed_at').first())
    return render(request, 'inventory/shift_open.html', {
        'branch': branch,
        'last_closed': last_closed,
    })


@login_required
def shift_close(request):
    branch = _user_branch_or_403(request)
    if branch is None:
        return redirect('lookup')

    shift = _open_shift_for(branch)
    if not shift:
        messages.warning(request, "Ochiq smen yo'q.")
        return redirect('pos_terminal')

    if request.method == 'POST':
        # N1: sanab chiqilган naqд MAJBURIY. Bo'sh bo'lса 0 yozib, kunlik
        # tushumга teng SOXTA kamomad qotirib qo'yardi. Bo'shликни serverда rad
        # etamiz va bo'shликка sabab bo'ladigan "1 200 000" (probel) formatни ham
        # qabul qilamiz (type=number bo'sh qaytaradi).
        _cc = (request.POST.get('counted_cash') or '').strip()
        for ch in (' ', ' ', ' ', "'"):
            _cc = _cc.replace(ch, '')
        _cc = _cc.replace(',', '.')
        if _cc == '':
            messages.error(request, "Sanab chiqilgan naqd summasini kiriting.")
            return redirect('shift_close')
        try:
            counted = max(0.0, float(_cc))
        except ValueError:
            messages.error(request, "Naqd summasi noto'g'ri — faqat raqam kiriting.")
            return redirect('shift_close')
        if counted > 1e12:
            messages.error(request, "Summa juda katta — tekshirib qayta kiriting.")
            return redirect('shift_close')
        with transaction.atomic():
            shift.counted_cash = counted
            shift.closed_by = request.user
            shift.closed_at = timezone.now()
            # MON-22: kutilgan naqdни YOPILISH paytida qotiramiz (keyingi
            # tahrirlar tarixiy farqni qayta yozmasin).
            shift.closing_expected_cash = shift.compute_expected_cash()
            shift.status = Shift.Status.CLOSED
            shift.note = (
                (shift.note + ' | ' if shift.note else '')
                + (request.POST.get('note') or '').strip()
            )[:200]
            shift.save()
        var = shift.variance()
        if var is not None and abs(var) >= 1:
            messages.warning(request,
                f"Smen yopildi. Kassa farqi: {var:+,.0f} so'm "
                f"({'ortiq' if var > 0 else 'kam'}).")
        else:
            messages.success(request, "Smen yopildi. Kassa to'g'ri.")
        # Yopilgach — chek printerida Z-hisobotni chop etish (avtomatik)
        return redirect(f"{reverse('shift_receipt', args=[shift.pk])}?autoprint=1")

    # Aralash cheklar naqd/karta/o'tkazmaga bo'linadi (alohida "Aralash" yo'q).
    # BRUTTO ko'rsatiladi — karta terminal hisoboti bilan solishtirish uchun.
    # Qaytarish alohida NAQD chiqim sifatida (usulларга bo'linmaydi).
    by_method = shift.sales_by_method_split()
    from decimal import Decimal as _D
    _refund = shift.refunds_total()
    payouts = list(shift.payouts.select_related('created_by').order_by('created_at'))
    _defs = [
        ('cash', 'Naqd', 'bi-cash', 'text-success'),
        ('card', 'Karta', 'bi-credit-card', 'text-primary'),
        ('transfer', "O'tkazma", 'bi-arrow-left-right', 'text-info'),
    ]
    method_rows = [{'label': label, 'icon': icon, 'cls': cls,
                    'gross': by_method.get(k, _D('0'))}
                   for k, label, icon, cls in _defs]
    return render(request, 'inventory/shift_close.html', {
        'shift': shift,
        'expected': shift.expected_cash(),
        'cash_sales': shift.cash_sales(),
        'debt_payments': shift.debt_payments_total(),
        'payouts': payouts,
        'payouts_total': shift.payouts_total(),
        'refund_total': _refund,
        'method_rows': method_rows,
        # Jami savdo = SOF (qaytarilgan pul ayirilgan).
        'sales_total': sum(by_method.values(), _D('0')) - _D(str(_refund or 0)),
    })


@login_required
def shift_detail(request, pk):
    shift = get_object_or_404(Shift, pk=pk)
    if not request.user.is_admin() and request.user.branch_id != shift.branch_id:
        return HttpResponseForbidden()
    txns = list(shift.transactions.select_related('sold_by')
                .prefetch_related('lines')
                .order_by('-sold_at')[:200])

    # Payment method breakdown
    pm_breakdown = {}
    for t in txns:
        pm = t.get_payment_method_display()
        if pm not in pm_breakdown:
            pm_breakdown[pm] = {'count': 0, 'total': 0}
        pm_breakdown[pm]['count'] += 1
        pm_breakdown[pm]['total'] += float(t.total)
    pm_list = sorted(pm_breakdown.items(), key=lambda kv: -kv[1]['total'])

    # Hourly activity (24-hour)
    hour_qty = [0] * 24
    hour_revenue = [0] * 24
    for t in txns:
        h = t.sold_at.astimezone(timezone.get_current_timezone()).hour
        hour_qty[h] += 1
        hour_revenue[h] += float(t.total)
    # Trim to active range
    active_hours = [(h, q, r) for h, (q, r) in enumerate(zip(hour_qty, hour_revenue)) if q]

    # Top products in this shift
    top_prod = {}
    for t in txns:
        for ln in t.lines.all():
            key = ln.variant.product.code
            if key not in top_prod:
                top_prod[key] = {
                    'code': key,
                    'name': ln.variant.product.name,
                    'qty': 0,
                    'revenue': 0,
                }
            top_prod[key]['qty'] += ln.quantity
            top_prod[key]['revenue'] += float(ln.total)
    top_products = sorted(top_prod.values(), key=lambda x: -x['qty'])[:5]

    # Variance value
    variance_value = shift.variance() if callable(getattr(shift, 'variance', None)) else None

    # Cash taken out of the till during the shift (payouts)
    payouts = list(shift.payouts.select_related('created_by').order_by('-created_at'))
    payouts_total = shift.payouts_total()
    # Cash added to the till (MON-4)
    cash_ins = list(shift.cash_ins.select_related('created_by').order_by('-created_at'))
    cash_ins_total = shift.cash_ins_total()

    return render(request, 'inventory/shift_detail.html', {
        'shift': shift,
        'txns': txns,
        'cash_sales': shift.cash_sales(),
        'expected': shift.expected_cash(),
        'payouts': payouts,
        'payouts_total': payouts_total,
        'cash_ins': cash_ins,
        'cash_ins_total': cash_ins_total,
        'pm_list': pm_list,
        'hour_labels': [f'{h:02d}:00' for h, _, _ in active_hours],
        'hour_qty': [q for _, q, _ in active_hours],
        'hour_revenue': [r for _, _, r in active_hours],
        'top_products': top_products,
        'variance_value': variance_value,
    })


@login_required
def shift_receipt(request, pk):
    """Smena yopilish cheki (Z-hisobot) — 80mm chek printerida chop etish uchun.

    Ichida: smen ma'lumoti, sotuvlar xulosasi + to'lov turlari, kassa
    hisob-kitobi (ochilish + naqd sotuv − kassadan olingan = kutilgan),
    kassadan olingan pullar ro'yxati va qaytarishlar.
    """
    shift = get_object_or_404(Shift.objects.select_related(
        'branch', 'opened_by', 'closed_by'), pk=pk)
    if not request.user.is_admin() and request.user.branch_id != shift.branch_id:
        return HttpResponseForbidden()

    # ARCH-5/MON-26: SOTUV bloki ham KASSA bloki bilan AYNAN bir xil cheklar
    # to'plamini ko'rishi SHART. Ilgari bu yerda `shift.transactions` (faqat FK)
    # ishlatilardi, `cash_sales()`/`expected_cash()` esa `_txn_qs()` — u FK'siz
    # eski cheklarni ham vaqt oynasi bo'yicha oladi. Natijada bitta chekda
    # "Naqd 310 000" (SOTUV) va "+ Naqd savdo 340 000" (KASSA) — bir xil
    # yorliq, ikki xil son. Kassir esa KASSA raqami bo'yicha javob beradi.
    # Yagona manba: modelning o'z ta'rifi.
    txns = list(shift._txn_qs().select_related('sold_by').prefetch_related('lines'))

    # MON-24: bu view'дagi HAMMA pul Decimal. Float yig'indi (0.1+0.2) `som`
    # yaxlitlashida boshqa tomonga ketib, chop etilgan qatorlarни JAMIдan
    # ajratib yuborardi.
    def _amt(x):
        try:
            return _dec(x or 0)
        except (TypeError, ValueError, InvalidOperation):
            return Decimal('0')

    total_rev = Decimal('0')
    item_qty = 0
    # DIQQAT (ARCH-6): bu Z-hisobot ATAYLAB split_breakdown()'dan farq qiladi.
    # Bu yerda maqsad — karta terminal hisoboti bilan solishtirish uchun har
    # to'lov turining BRUTTO (kiritilgan) summasini ko'rsatish; sof (net)
    # taqsimlash EMAS. Shuning uchun kiritilgan summalar to'g'ridan-to'g'ri
    # olinadi va qoldiq "Boshqa"ga tushadi. Usul nomlari _norm_pay_method
    # orqali normallashtiriladi (payme/click → transfer), jimgina cash'ga emas.
    PM = SaleTransaction.PaymentMethod
    labels = dict(PM.choices)
    money = {PM.CASH: Decimal('0'), PM.CARD: Decimal('0'), PM.TRANSFER: Decimal('0')}
    # DIQQAT: count = SHU usul YAKKA to'lov bo'lgan cheklar soni. Aralash chek
    # bittagina — u ikki bucket'ni oshirmasin (aks holda badge'lar yig'indisi
    # Cheklar'dan oshib, to'g'ri hisobot buzuq ko'rinardi). Pul (money) esa
    # BRUTTO qoladi: aralash chekning karta ulushi Karta summasiga kiradi
    # (terminal hisoboti bilan solishtirish uchun — ARCH-6).
    counts = {PM.CASH: 0, PM.CARD: 0, PM.TRANSFER: 0}
    other_money = Decimal('0')
    other_count = 0
    mixed_count = 0
    mixed_money = Decimal('0')
    # DISC-5: smen davomida BERILGAN CHEGIRMA. Z-hisobot smenning yagona
    # rasmiy hujjati va kassir aynan shu bo'yicha baholanadi — lekin chegirma
    # unда umuman ko'rinmasdi. JAMI SAVDO allaqachon chegirma AYIRILGAN
    # summa, ya'ni kassir bir smenда million so'm qo'lда chegirma bersa ham
    # chekда hech qanday iz qolmasdi. Taqsimlash kerak emas: chek butunlay
    # shu smenga tegishli, shuning uchun to'g'ridan-to'g'ri yig'amiz.
    disc_line = Decimal('0')
    disc_promo = Decimal('0')
    disc_exch = Decimal('0')
    disc_order = Decimal('0')
    for t in txns:
        total_rev += _dec(t.total)
        disc_order += _dec(t.order_discount)
        disc_promo += _dec(t.promo_discount)
        disc_exch += _dec(t.exchange_credit)
        for ln in t.lines.all():
            item_qty += ln.quantity
            disc_line += _dec(ln.line_discount)
        if t.payment_method == PM.MIXED:
            used = set()
            covered = Decimal('0')
            for part in (t.payment_breakdown or []):
                m = _norm_pay_method(part.get('method'))
                amt = _amt(part.get('amount'))
                if amt <= 0:
                    continue
                covered += amt
                if m in money:
                    money[m] += amt
                    used.add(m)
                else:
                    other_money += amt
            # Breakdown to'liq bo'lmasa (eski/chala yozuvlar) — qolgan summa
            # "Boshqa"ga tushadi, shunda qismlar yig'indisi JAMI bilan teng bo'ladi.
            rem = _dec(t.total) - covered
            if abs(rem) > Decimal('0.005'):
                other_money += rem
            # Chekni FAQAT bir marta — aralash sifatida — sanaymiz.
            mixed_count += 1
            mixed_money += _dec(t.total)
        elif t.payment_method in money:
            money[t.payment_method] += _dec(t.total)
            counts[t.payment_method] += 1
        else:
            other_money += _dec(t.total)
            other_count += 1
    # BRUTTO — karta terminal hisoboti bilan solishtirish uchun. Qaytarish
    # alohida (QAYTARISHLAR bo'limida, bitta naqd chiqim sifatida).
    pay_rows = [{'label': labels.get(k, k), 'amount': money[k], 'count': counts[k]}
                for k in (PM.CASH, PM.CARD, PM.TRANSFER)]

    # Qo'lда = chek chegirmasi − aksiya − almashtirish krediti (DISC-1).
    _disc_manual = disc_order - disc_promo - disc_exch
    if _disc_manual < 0:
        _disc_manual = Decimal('0')
    # Bu blok MA'LUMOT uchun — kassa balansiga kirmaydi, shu bois yaxlitlash
    # `som` filtriga qoldiriladi (u yarmini YUQORIGA yaxlitlaydi — MON-24).
    shift_discount = {
        'line': disc_line,
        'promo': disc_promo,
        'manual': _disc_manual,
        'exchange': disc_exch,
        'total': disc_line + disc_order,
    }

    payouts = list(shift.payouts.select_related('created_by').order_by('created_at'))
    # R7: kassaga qo'shilган naqd (cash_in) ham Z-hisobotда ko'rinsin — u
    # kutilган naqdни oshiradi; ko'rsatilmasa reconciliation shaffof bo'lmaydi.
    cash_ins = list(shift.cash_ins.select_related('created_by').order_by('created_at'))
    cash_ins_total = shift.cash_ins_total()

    refunds = list(Return.objects.filter(shift=shift)
                   .select_related('sale__variant__product', 'refunded_by')
                   .order_by('refunded_at'))
    # MON-24: qaytarilgan naqdни DECIMAL'да yig'amiz (float EMAS). Ilgari bu
    # float yig'indi edi, expected_cash() esa Decimal yo'lда ayirar edi — REF-2
    # fraksiyali refundни kiritgach ikki yo'l 1 so'mга farq qilib, chek qatorlari
    # JAMI bilan yopilmay qolardi. Endi bitta Decimal manba.
    refund_total = sum((r.effective_cash_refund for r in refunds), Decimal('0'))
    refund_qty = sum(r.quantity for r in refunds)

    # ---- KASSA bloki: chop etilgan qatorlar YIG'INDISI = chop etilgan JAMI.
    # Har money qatori `|som` bilan butun so'mга yaxlitlanadi; JAMIni (= HOZIRGI
    # QOLDIQ) AYNI shu yaxlitlangan qiymatlardан chiqaramiz — kassir qo'lда
    # qayta hisoblaganда aynan bir xil son chiqsin (§4.4). expected_cash()/
    # variance() modelда audit uchun o'zgarmasдан qoladi; bu faqat CHOP uchun.
    def _somi(v):
        # `som` filtri bilan AYNAN bir xil: butun so'm, YARMI YUQORIGA.
        # (`round()` bank yaxlitlashi qiladi — round(2.5)=2 — kassir esa 3
        # kutadi; ikki joyда ikki xil yaxlitlash aynan 1 so'm farqni tug'dirardi.)
        return int(_amt(v).quantize(Decimal('1'), rounding=ROUND_HALF_UP))
    _opening_i = _somi(shift.opening_cash)
    _cash_i = _somi(shift.cash_sales())
    _debt_i = _somi(shift.debt_payments_total())
    _cashins_i = _somi(cash_ins_total)
    _payouts_i = _somi(shift.payouts_total())
    _refund_i = _somi(refund_total)
    _live_i = (_opening_i + _cash_i + _debt_i + _cashins_i
               - _payouts_i - _refund_i)

    # MON-22 (chop yo'lida ham): YOPILGAN smen uchun JAMI — yopilishда
    # QOTIRILGAN qiymat. Komponentlar esa HOZIR qayta hisoblanadi, shuning
    # uchun smen yopilgandan keyin qaytarish/o'chirish bo'lса ikkisi ajraladi.
    # Farqni JAMIга singdirib yubormaymiz (aks holda oylar oldingi smenning
    # FARQi jimgina qayta yozilib, aybni o'sha kassirга ag'darardi) — alohida
    # qatorда ko'rsatamiz, shunda qatorlar yig'indisi baribir JAMIga teng.
    expected_display = _somi(shift.expected_cash())
    _post_close_delta = _somi(_dec(shift.expected_cash())
                              - _dec(shift.compute_expected_cash()))
    # Qolgani — sof yaxlitlash: har qator ALOHIDA butun so'mга keltirilgani
    # uchun yig'indi 1-2 so'mга siljishi mumkin. Uni ham ochiq ko'rsatamiz,
    # chunki chekда "yig'ilmayotgan" raqamdan yomoni yo'q.
    _rounding_delta = expected_display - _live_i - _post_close_delta
    if shift.counted_cash is not None:
        variance_display = _somi(shift.counted_cash) - expected_display
    else:
        variance_display = shift.variance()

    return render(request, 'inventory/shift_receipt.html', {
        'shift': shift,
        'txn_count': len(txns),
        'shift_discount': shift_discount,   # DISC-5
        'item_qty': item_qty,
        'total_rev': total_rev,
        'pay_rows': pay_rows,
        'other_money': other_money,
        'other_count': other_count,
        'mixed_count': mixed_count,
        'mixed_money': mixed_money,
        'cash_sales': shift.cash_sales(),
        'debt_payments': shift.debt_payments_total(),
        'payouts': payouts,
        'payouts_total': shift.payouts_total(),
        'cash_ins': cash_ins,
        'cash_ins_total': cash_ins_total,
        'expected': expected_display,
        'post_close_delta': _post_close_delta,
        'rounding_delta': _rounding_delta,
        'variance_value': variance_display,
        'refunds': refunds,
        'refund_total': refund_total,
        'refund_qty': refund_qty,
        # SOF SAVDO = JAMI − Qaytarilgan, chop etilgan qiymatlardan (yopiladi).
        'net_sales': _somi(total_rev) - _refund_i,
        'printed_at': timezone.now(),
    })


@login_required
def shift_returns(request, pk):
    """Smen davomidagi QAYTARISHLAR tafsiloti — har qator: chek #, sotilган
    narx, chegirma (va %), qaytarilган summa. Jami qaytarilган = Z-hisobotдаgi
    'Qaytarilgan (naqd)' bilan bir xil. 'Fraksiya qayerdan?' savoliga aniq javob.
    """
    from decimal import Decimal
    shift = get_object_or_404(Shift.objects.select_related('branch'), pk=pk)
    if not request.user.is_admin() and request.user.branch_id != shift.branch_id:
        return HttpResponseForbidden()

    rows = []
    total_sticker = Decimal('0')
    total_discount = Decimal('0')
    total_refund = Decimal('0')
    returns = (Return.objects.filter(shift=shift)
               .select_related('sale__variant__product', 'sale__transaction',
                               'refunded_by')
               .order_by('refunded_at'))
    for r in returns:
        sale = r.sale
        # sale_price / refund_amount / effective_cash_refund allaqачон Decimal.
        sticker = Decimal(r.quantity) * sale.sale_price   # yorliq narx × dona
        refund = r.refund_amount            # qatorning o'z narxi (order_discount taqsimlanmaydi)
        cash = r.effective_cash_refund      # kassaдан HAQIQIY chiqqan
        disc = sticker - refund
        pct = (disc / sticker * 100) if sticker > 0 else Decimal('0')
        rows.append({
            'refunded_at': r.refunded_at,
            'check_id': sale.transaction_id,
            'public_id': sale.transaction.public_id if sale.transaction else None,
            'sold_at': sale.sold_at,
            'product': sale.variant.product.name,
            'code': sale.variant.product.code,
            'qty': r.quantity,
            'unit_price': sale.sale_price,
            'sticker': sticker,
            'discount': disc,
            'pct': pct,
            'refund': refund,
            'cash': cash,
            'is_exchange': r.is_exchange,
        })
        total_sticker += sticker
        total_discount += disc
        total_refund += cash

    return render(request, 'inventory/shift_returns.html', {
        'shift': shift,
        'rows': rows,
        'count': len(rows),
        'qty_total': sum(x['qty'] for x in rows),
        'total_sticker': total_sticker,
        'total_discount': total_discount,
        'total_refund': total_refund,
    })


# MON-17: bitta kassa harakati uchun aqlли yuqori chegara (1 milliard so'm).
# Bundan katta summa — deyarli har doim kiritishда xato yoki suiiste'mol.
_MAX_CASH_MOVE = 1_000_000_000


@login_required
def cash_payout(request):
    """Kassadan pul olishni qayd etish (naqd chiqim).

    Ikki joydan chaqiriladi:
      - POS (fetch/JSON) → JSON javob, sahifa yangilanmaydi
      - Smen sahifasi (form POST) → xabar + redirect
    Har doim joriy ochiq smenga bog'lanadi.
    """
    if request.method != 'POST':
        return redirect('pos_terminal')

    ctype = request.content_type or ''
    wants_json = 'application/json' in ctype or \
        request.headers.get('x-requested-with') == 'XMLHttpRequest'
    if 'application/json' in ctype:
        try:
            data = _json.loads(request.body.decode('utf-8'))
        except ValueError:
            return JsonResponse({'ok': False, 'error': 'bad JSON'}, status=400)
    else:
        data = request.POST

    def _fail(msg, status=400):
        if wants_json:
            return JsonResponse({'ok': False, 'error': msg}, status=status)
        messages.warning(request, msg)
        back = request.POST.get('next')
        return redirect(back if back and back.startswith('/') else 'pos_terminal')

    branch = _user_branch_or_403(request)
    if branch is None:
        return _fail("Filial biriktirilmagan.", 403)

    shift = _open_shift_for(branch)
    if not shift:
        return _fail("Ochiq smen yo'q — avval smen oching.", 400)

    # R6: 'inf'/'1e400' kabi qiymat round()да OverflowError berardi — ushlaymiz.
    try:
        amount = round(float(data.get('amount') or 0))
    except (ValueError, TypeError, OverflowError):
        amount = 0
    if amount <= 0:
        return _fail("Summa 0 dan katta bo'lishi kerak.", 400)
    if amount > _MAX_CASH_MOVE:   # MON-17
        return _fail("Summa juda katta — tekshirib qayta kiriting.", 400)

    category = (data.get('category') or 'other').strip()
    if category not in CashPayout.VALID_CATEGORIES:
        category = 'other'
    note = (data.get('note') or '').strip()[:200]

    # S5: takroriy yuborishни (double-tap / timeout) bloklaymiz — bir martalik
    # kalit bilan. Mavjud kalit bo'lsa yangi yozuv YARATMAYMIZ.
    _idem = (data.get('idempotency_key') or '').strip()[:64] or None

    def _payout_result(p, dup=False):
        if wants_json:
            return JsonResponse({'ok': True, 'id': p.id, 'amount': int(p.amount),
                'category': p.get_category_display(),
                'payouts_total': float(shift.payouts_total()),
                'expected': float(shift.expected_cash()), 'duplicate': dup})
        if dup:
            messages.info(request, "Bu chiqim allaqachon yozilgan.")
        else:
            messages.success(request,
                f"Kassadan olindi: {int(p.amount):,} so'm — {p.get_category_display()}.")
        return redirect('shift_detail', pk=shift.pk)

    if _idem:
        _dupe = CashPayout.objects.filter(idempotency_key=_idem).first()
        if _dupe:
            return _payout_result(_dupe, dup=True)
    try:
        payout = CashPayout.objects.create(
            shift=shift, branch=branch, amount=amount,
            category=category, note=note, created_by=request.user,
            idempotency_key=_idem,
        )
    except IntegrityError:
        _dupe = CashPayout.objects.filter(idempotency_key=_idem).first()
        if _dupe:
            return _payout_result(_dupe, dup=True)
        raise

    # MON-17: har bir kassa chiqimini AUDIT LOGGA yozamiz + katta summада
    # Telegram ogohlantirish. Ilgari chiqim hech qayerда qayd etilmasdi —
    # kassir 500 000 olib, "chiqim" yozib, smenни nol farq bilan yopardi.
    try:
        AuditLog.objects.create(
            user=request.user, username_snapshot=request.user.username,
            action=AuditLog.Action.CREATE, model_name='CashPayout',
            object_id=str(payout.id),
            object_repr=(f"Kassa chiqimi {int(amount)} so'm — "
                         f"{payout.get_category_display()} ({branch.name})")[:300],
            changes={'amount': float(amount), 'category': category,
                     'note': note, 'shift_id': shift.id},
        )
        _threshold = getattr(settings, 'CASH_PAYOUT_ALERT_SOM', 500000)
        if amount >= _threshold:
            try:
                from .notifications import send_telegram
                send_telegram(
                    f"💸 <b>Katta kassa chiqimi</b>\n"
                    f"Filial: {branch.name} · Kassir: {request.user.username}\n"
                    f"Summa: <b>{int(amount)}</b> so'm — {payout.get_category_display()}\n"
                    f"Izoh: {note or '—'} · Smen #{shift.id}")
            except Exception:
                logger.exception('cash-payout telegram alert failed')
    except Exception:
        logger.exception('cash-payout audit failed (payout %s)', payout.id)

    if wants_json:
        return JsonResponse({
            'ok': True,
            'id': payout.id,
            'amount': amount,
            'category': payout.get_category_display(),
            'payouts_total': float(shift.payouts_total()),
            'expected': float(shift.expected_cash()),
        })
    messages.success(
        request,
        f"Kassadan olindi: {amount:,.0f} so'm — {payout.get_category_display()}."
    )
    return redirect('shift_detail', pk=shift.pk)


@login_required
def cash_in(request):
    """Kassaga naqd qo'shishни qayd etish (MON-4) — cash_payout'ning aksi.
    POS (JSON) yoki smen sahifasi (form) dan chaqiriladi."""
    if request.method != 'POST':
        return redirect('pos_terminal')

    ctype = request.content_type or ''
    wants_json = 'application/json' in ctype or \
        request.headers.get('x-requested-with') == 'XMLHttpRequest'
    if 'application/json' in ctype:
        try:
            data = _json.loads(request.body.decode('utf-8'))
        except ValueError:
            return JsonResponse({'ok': False, 'error': 'bad JSON'}, status=400)
    else:
        data = request.POST

    def _fail(msg, status=400):
        if wants_json:
            return JsonResponse({'ok': False, 'error': msg}, status=status)
        messages.warning(request, msg)
        back = request.POST.get('next')
        return redirect(back if back and back.startswith('/') else 'pos_terminal')

    branch = _user_branch_or_403(request)
    if branch is None:
        return _fail("Filial biriktirilmagan.", 403)

    shift = _open_shift_for(branch)
    if not shift:
        return _fail("Ochiq smen yo'q — avval smen oching.", 400)

    # R6: OverflowError'ни ham ushlaymiz (inf/1e400).
    try:
        amount = round(float(data.get('amount') or 0))
    except (ValueError, TypeError, OverflowError):
        amount = 0
    if amount <= 0:
        return _fail("Summa 0 dan katta bo'lishi kerak.", 400)
    if amount > _MAX_CASH_MOVE:   # MON-17
        return _fail("Summa juda katta — tekshirib qayta kiriting.", 400)

    category = (data.get('category') or 'other').strip()
    if category not in CashIn.VALID_CATEGORIES:
        category = 'other'
    note = (data.get('note') or '').strip()[:200]

    # S5: takroriy yuborish (double-tap/timeout) bir martalik kalit bilan bloklanadi.
    _idem = (data.get('idempotency_key') or '').strip()[:64] or None

    def _cashin_result(c, dup=False):
        if wants_json:
            return JsonResponse({'ok': True, 'id': c.id, 'amount': int(c.amount),
                'category': c.get_category_display(),
                'cash_ins_total': float(shift.cash_ins_total()),
                'expected': float(shift.expected_cash()), 'duplicate': dup})
        if dup:
            messages.info(request, "Bu qo'shimcha allaqachon yozilgan.")
        else:
            messages.success(request,
                f"Kassaga qo'shildi: {int(c.amount):,} so'm — {c.get_category_display()}.")
        return redirect('shift_detail', pk=shift.pk)

    if _idem:
        _dupe = CashIn.objects.filter(idempotency_key=_idem).first()
        if _dupe:
            return _cashin_result(_dupe, dup=True)
    try:
        ci = CashIn.objects.create(
            shift=shift, branch=branch, amount=amount,
            category=category, note=note, created_by=request.user,
            idempotency_key=_idem,
        )
    except IntegrityError:
        _dupe = CashIn.objects.filter(idempotency_key=_idem).first()
        if _dupe:
            return _cashin_result(_dupe, dup=True)
        raise

    # MON-4/17: kassaga naqd qo'shishни ham AUDIT LOGGA yozamiz + katta summада
    # ogohlantirish (chiqim bilan bir xil nazorat — aks holда qo'shimchalar
    # hech qayerда qayd etilmай, kutilган naqdни shishirar edi).
    try:
        AuditLog.objects.create(
            user=request.user, username_snapshot=request.user.username,
            action=AuditLog.Action.CREATE, model_name='CashIn',
            object_id=str(ci.id),
            object_repr=(f"Kassaga qo'shildi {int(amount)} so'm — "
                         f"{ci.get_category_display()} ({branch.name})")[:300],
            changes={'amount': float(amount), 'category': category,
                     'note': note, 'shift_id': shift.id},
        )
        _threshold = getattr(settings, 'CASH_PAYOUT_ALERT_SOM', 500000)
        if amount >= _threshold:
            try:
                from .notifications import send_telegram
                send_telegram(
                    f"💰 <b>Katta kassa qo'shimchasi</b>\n"
                    f"Filial: {branch.name} · Kassir: {request.user.username}\n"
                    f"Summa: <b>{int(amount)}</b> so'm — {ci.get_category_display()}\n"
                    f"Izoh: {note or '—'} · Smen #{shift.id}")
            except Exception:
                logger.exception('cash-in telegram alert failed')
    except Exception:
        logger.exception('cash-in audit failed (cash_in %s)', ci.id)

    if wants_json:
        return JsonResponse({
            'ok': True,
            'id': ci.id,
            'amount': amount,
            'category': ci.get_category_display(),
            'cash_ins_total': float(shift.cash_ins_total()),
            'expected': float(shift.expected_cash()),
        })
    messages.success(
        request,
        f"Kassaga qo'shildi: {amount:,.0f} so'm — {ci.get_category_display()}."
    )
    return redirect('shift_detail', pk=shift.pk)


@admin_required
def shift_list(request):
    """Smenlar ro'yxati + filtrlar + ochiq smenlar banner."""
    branch_id = request.GET.get('branch') or ''
    status = request.GET.get('status') or ''
    seller_id = request.GET.get('seller') or ''
    date_from = request.GET.get('date_from') or ''
    date_to = request.GET.get('date_to') or ''

    qs = Shift.objects.select_related('branch', 'opened_by', 'closed_by')

    if branch_id:
        try:
            qs = qs.filter(branch_id=int(branch_id))
        except ValueError:
            branch_id = ''
    if status:
        qs = qs.filter(status=status)
    if seller_id:
        try:
            qs = qs.filter(opened_by_id=int(seller_id))
        except ValueError:
            seller_id = ''
    try:
        if date_from:
            qs = qs.filter(opened_at__date__gte=datetime.strptime(date_from, '%Y-%m-%d').date())
    except ValueError:
        date_from = ''
    try:
        if date_to:
            qs = qs.filter(opened_at__date__lte=datetime.strptime(date_to, '%Y-%m-%d').date())
    except ValueError:
        date_to = ''

    shifts = list(qs.order_by('-opened_at')[:100])

    # Ochiq smenlar — har doim alohida ko'rsatamiz
    open_shifts = (Shift.objects.filter(status=Shift.Status.OPEN)
                   .select_related('branch', 'opened_by')
                   .order_by('-opened_at'))

    # Variance stats — total positive vs negative variance from closed shifts.
    # Shift.variance is a method, so call it (cache per row).
    closed = []
    for s in shifts:
        if s.status == Shift.Status.CLOSED:
            v = s.variance() if callable(s.variance) else s.variance
            s._variance_value = v
            closed.append(s)
    variance_positive = sum(float(s._variance_value) for s in closed
                            if s._variance_value is not None and s._variance_value > 0)
    variance_negative = sum(float(s._variance_value) for s in closed
                            if s._variance_value is not None and s._variance_value < 0)
    avg_variance = (sum(float(s._variance_value) for s in closed
                        if s._variance_value is not None) / len(closed)) if closed else 0

    return render(request, 'inventory/shift_list.html', {
        'shifts': shifts,
        'open_shifts': open_shifts,
        'variance_positive': variance_positive,
        'variance_negative': variance_negative,
        'avg_variance': avg_variance,
        'branch_id': branch_id,
        'status': status,
        'seller_id': seller_id,
        'date_from': date_from,
        'date_to': date_to,
        'branches': Branch.objects.filter(is_active=True).order_by('name'),
        'sellers': User.objects.filter(is_active=True).order_by('username'),
    })


# ---------- POS TERMINAL ----------

POS_BRANCH_SESSION_KEY = 'pos_branch_id'


def _user_branch_or_403(request):
    """Sellers are locked to their assigned branch. Admins can pick
    which branch the POS operates on via a dropdown — the choice is
    stored in the session under POS_BRANCH_SESSION_KEY.

    Fallback order for admins:
      1) session-selected branch
      2) ?branch_id= query (used once by pos_terminal to set the session)
      3) admin's own assigned branch (if set)
      4) branch of admin's own open shift
      5) any open shift's branch
      6) first active branch
    """
    if not request.user.is_admin():
        return request.user.branch

    sb_id = request.session.get(POS_BRANCH_SESSION_KEY)
    if sb_id:
        b = Branch.objects.filter(pk=sb_id, is_active=True).first()
        if b:
            return b

    q_id = request.GET.get('branch_id')
    if q_id:
        try:
            b = Branch.objects.filter(pk=int(q_id), is_active=True).first()
            if b:
                return b
        except ValueError:
            pass

    if request.user.branch_id:
        return request.user.branch

    own_shift = Shift.objects.filter(
        opened_by=request.user, status=Shift.Status.OPEN
    ).select_related('branch').order_by('-opened_at').first()
    if own_shift:
        return own_shift.branch

    any_shift = Shift.objects.filter(
        status=Shift.Status.OPEN
    ).select_related('branch').order_by('-opened_at').first()
    if any_shift:
        return any_shift.branch

    return Branch.objects.filter(is_active=True).first()


@login_required
def pos_customer_display(request):
    """Mijoz tomonidagi ekran — POS bilan BroadcastChannel orqali sinxron.
    Faqat o'qish (interaktiv emas). Iste'mol uchun: 2-monitor yoki ikkinchi tab."""
    branch = _user_branch_or_403(request)
    return render(request, 'inventory/pos_display.html', {
        'branch': branch,
        'shop_name': 'yurit',  # TODO: shop name from settings/branding
    })


def _open_price_product():
    """Yashirin 'ochiq narx' mahsuloti (kodsiz kiyim/poyabzal sotuvi)."""
    product = Product.objects.filter(is_open_price=True).first()
    if product is None:
        product = Product.objects.create(
            name='Kiyim / Poyabzal', is_open_price=True,
            default_sale_price=0, markup_percent=0)
    return product


def _open_price_stock(branch):
    """Qo'lda summa uchun BranchStock. Ombor tekshirilmaydi.

    DIQQAT: tezkor sotuv toifalari ham shu mahsulotning turlari bo'lgani uchun
    bo'sh turni ANIQ tanlaymiz (.first() toifa turini qaytarib yuborishi mumkin).
    """
    product = _open_price_product()
    variant, _ = ProductVariant.objects.get_or_create(
        product=product, size='', color='')
    stock, _ = BranchStock.objects.get_or_create(
        variant=variant, branch=branch,
        defaults={'stock_count': 0, 'sale_price': 0, 'cost_price': 0})
    return stock


QUICK_SELL_CATEGORY = 'Kiyim Kechak'


def _quick_sell_product(item):
    """Toifaning ombor yuritiladigan mahsuloti (Kiyim Kechak kategoriyasida)."""
    if item.product_id:
        return item.product
    cat, _ = Category.objects.get_or_create(
        name=QUICK_SELL_CATEGORY, defaults={'prefix': 'KIY'})
    product = (Product.objects.filter(name__iexact=item.name, category=cat).first()
               or Product.objects.create(name=item.name, category=cat))
    item.product = product
    item.save(update_fields=['product'])
    return product


def _quick_sell_sync(item, branch=None):
    """Har NARX uchun alohida tur + zaxira yozuvi bo'lishini ta'minlaymiz.

    2500 so'mlik paypoq va 5000 so'mlik paypoq — boshqa-boshqa tovar, shuning
    uchun har narx alohida tur. Shunda qabul (miqdor/tannarx/marja) va ombor
    oddiy mahsulotdek ishlaydi.
    """
    product = _quick_sell_product(item)
    branches = ([branch] if branch is not None
                else list(Branch.objects.filter(is_active=True)))
    for price in item.price_list:
        variant, _ = ProductVariant.objects.get_or_create(
            product=product, size=str(price), color='')
        for b in branches:
            BranchStock.objects.get_or_create(
                variant=variant, branch=b,
                defaults={'stock_count': 0, 'sale_price': price,
                          'cost_price': 0, 'wholesale_price': 0})
    return product


def _quick_sell_items(branch):
    """Faol toifalar + har narx uchun tur, zaxira va stock id.

    Toifalar bazadan o'qiladi (Narxlar -> Tezkor sotuv sahifasida
    tahrirlanadi), narxlar esa endi haqiqiy tur bo'lgani uchun ombor
    hisobga olinadi.
    """
    out = []
    for item in (QuickSellItem.objects.filter(is_active=True)
                 .select_related('product')):
        product = _quick_sell_sync(item, branch)
        # Tugmalar HAQIQIY turlardan quriladi — sozlamadagi ro'yxatdan emas.
        # Shunda mahsulot sahifasida yangi narx (tur) qo'shilsa, u shu yerda
        # ham darrov ko'rinadi. _quick_sell_sync sozlamadagi narxlar uchun
        # turlarni yaratib qo'yadi, qolganini ombor hal qiladi.
        by_price = {}
        for st in (BranchStock.objects
                   .filter(variant__product=product, branch=branch)
                   .select_related('variant')):
            price = st.sale_price or 0
            if price <= 0:
                continue
            key = int(price)
            prev = by_price.get(key)
            # Bir narxda bir nechta tur bo'lsa — zaxirasi ko'pini olamiz
            if prev is None or st.stock_count > prev['stock']:
                by_price[key] = {'price': key, 'sid': st.id,
                                 'stock': st.stock_count}
        rows = sorted(by_price.values(), key=lambda r: r['price'])
        out.append({
            'key': f'qs{item.pk}',
            'name': item.name,
            'icon': item.icon or 'bi-bag',
            'code': product.code,
            'prices': rows,
        })
    return out


@login_required
def pos_terminal(request):
    """Single-page POS UI. Browser maintains the cart, posts via AJAX."""
    # If admin clicked the branch dropdown, persist the choice
    if request.user.is_admin():
        q_id = request.GET.get('branch_id')
        if q_id:
            try:
                b = Branch.objects.filter(pk=int(q_id), is_active=True).first()
                if b:
                    request.session[POS_BRANCH_SESSION_KEY] = b.pk
            except ValueError:
                pass

    branch = _user_branch_or_403(request)
    if branch is None:
        messages.error(request, "Filial biriktirilmagan. Administrator bilan bog'laning.")
        return redirect('lookup')

    open_shift = _open_shift_for(branch)
    if not open_shift:
        return redirect('shift_open')

    # POS pastida — FAQAT bugungi sotuvlar (kalendar sana bo'yicha, smenadan
    # qat'i nazar). Kassir bugun sotgan har qanday chekni qaytarish/almashtirish
    # qilishi mumkin.
    # Patch 1e: yangi qatorlar jonli kelib turgani uchun boshlang'ich ro'yxat
    # qisqa (12 ta) bo'lsa yetadi — 200 ta chek + har chekning barcha `lines`
    # prefetch'i (item_count hisoblash uchun) POS yuklanishini sekinlashtirardi.
    # To'liq tarix "Hammasi →" (sales_list) orqali bir bosishда ochiladi.
    recent_txns = (SaleTransaction.objects.filter(
                       branch=branch, sold_at__date=timezone.localdate())
                   .select_related('sold_by')
                   .prefetch_related('lines')
                   .order_by('-sold_at')[:12])

    # Top 12 most-sold products in this branch over the last 30 days.
    # Used to populate the "Tez sotiluvchilar" quick-tap grid on the POS.
    since = timezone.now() - timedelta(days=30)
    top_rows = (Sale.objects
                .filter(branch=branch, sold_at__gte=since)
                .values('variant__product')
                .annotate(sold_qty=Sum('quantity'))
                .order_by('-sold_qty')[:12])
    top_ids = [r['variant__product'] for r in top_rows]
    qty_by_id = {r['variant__product']: r['sold_qty'] for r in top_rows}
    products_by_id = {p.id: p for p in Product.objects.filter(id__in=top_ids)}
    favorites = []
    for pid in top_ids:
        p = products_by_id.get(pid)
        if not p:
            continue
        favorites.append({
            'code': p.code,
            'name': p.name,
            'price': float(p.default_sale_price),
            'image': p.image.url if p.image else '',
            'sold_qty': qty_by_id.get(pid, 0),
        })

    branches_list = []
    if request.user.is_admin():
        branches_list = list(Branch.objects.filter(is_active=True).order_by('name'))

    parked_sales = list(ParkedSale.objects
                        .filter(branch=branch)
                        .select_related('parked_by')
                        .order_by('-created_at')[:20])

    # Static QR codes for this branch — user uploads once per provider.
    # Mijoz QR'ni telefon ilovasidan scan qiladi, summa kiritadi, to'laydi.
    # Kassir manual ravishda chekni yakunlaydi.
    static_qrs = list(PaymentQR.objects
                      .filter(branch=branch, is_active=True)
                      .order_by('provider'))
    # Map provider -> QR for fast lookup in template
    qr_by_provider = {}
    for q in static_qrs:
        qr_by_provider.setdefault(q.provider, q)

    # Build the provider list shown on the POS. Combine: providers that
    # have a static QR uploaded (manual confirm flow) + the in-code
    # provider abstraction (currently stubbed, future merchant API).
    from .payments import available_providers, _REGISTRY
    payment_providers = []
    for p in available_providers():
        static = qr_by_provider.get(p.name)
        payment_providers.append({
            'name': p.name,
            'display_name': p.display_name,
            'icon': p.icon,
            'is_installment': p.is_installment,
            'has_static_qr': bool(static),
            'static_qr_id': static.id if static else None,
        })
    # Also include providers that have a static QR but aren't in the
    # in-code registry (e.g. "humo" or "other")
    in_registry = {p['name'] for p in payment_providers}
    for q in static_qrs:
        if q.provider in in_registry:
            continue
        payment_providers.append({
            'name': q.provider,
            'display_name': q.get_provider_display(),
            'icon': 'bi-qr-code',
            'is_installment': q.provider in ('anor', 'alif', 'iman', 'zoodpay'),
            'has_static_qr': True,
            'static_qr_id': q.id,
        })

    return render(request, 'inventory/pos.html', {
        'branch': branch,
        'open_price_sid': _open_price_stock(branch).id,
        'quick_sell': _quick_sell_items(branch),
        'shift': open_shift,
        'recent_txns': recent_txns,
        'favorites': favorites,
        'parked_sales': parked_sales,
        'payment_methods': SaleTransaction.PaymentMethod.choices,
        'payment_providers': payment_providers,
        'branches_list': branches_list,
        'discount_reasons': DISCOUNT_REASONS,      # DISC-7
    })


@login_required
def pos_lookup(request):
    """GET /pos/lookup/?q=OYO-0001 [&branch=ID]
    Admins may pass &branch= to query a different branch (used by transfers).
    Returns JSON:
      { found: bool,
        product: {code, name, default_sale_price},
        variants: [{variant_id, stock_id, size, color, stock_count, sale_price, cost_price}],
        suggestions: [{code, name}] }
    """
    # Admins may override the branch context (e.g. transfer create form)
    branch_override = request.GET.get('branch')
    if branch_override and request.user.is_admin():
        try:
            branch = Branch.objects.get(pk=int(branch_override))
        except (Branch.DoesNotExist, ValueError):
            branch = _user_branch_or_403(request)
    else:
        branch = _user_branch_or_403(request)
    if branch is None:
        return JsonResponse({'error': 'no branch'}, status=403)

    q = (request.GET.get('q') or '').strip()
    if not q:
        return JsonResponse({'found': False})

    # Exact code first — match either internal code or manufacturer barcode
    code = normalize_code(q.upper())
    product = Product.objects.filter(
        Q(code=code) | Q(external_barcode=q.strip())
    ).first()

    # Tur (variant) shtrix-kodi skanerlangan bo'lishi mumkin — aniq turga moslaymiz
    matched_variant_id = None
    _matched_variant = None
    if not product:
        _vm = (ProductVariant.objects.filter(barcode=q.strip())
               .select_related('product').first())
        if _vm:
            product = _vm.product
            matched_variant_id = _vm.id
            _matched_variant = _vm

    if not product:
        # "Paypoq", "Ich kiyim" kabi tezkor sotuv toifasi yozilgan bo'lishi
        # mumkin — u oddiy mahsulot emas, shuning uchun panelga yo'naltiramiz.
        # DIQQAT: icontains juda ochko'z — "Oq" rangi "Payp[oq]"ga tushib
        # ketardi. Shuning uchun aniq moslik yoki (3+ harfda) boshlanishi.
        qs_q = Q(name__iexact=q)
        if len(q) >= 3:
            qs_q |= Q(name__istartswith=q)
        qs_item = (QuickSellItem.objects.filter(is_active=True)
                   .filter(qs_q).first())
        if qs_item:
            variant, _ = ProductVariant.objects.get_or_create(
                product=_open_price_product(), size='', color=qs_item.name)
            stock, _ = BranchStock.objects.get_or_create(
                variant=variant, branch=branch,
                defaults={'stock_count': 0, 'sale_price': 0, 'cost_price': 0})
            return JsonResponse({
                'found': False,
                'quick_sell': {
                    'name': qs_item.name,
                    'prices': qs_item.price_list,
                    'sid': stock.id,
                },
            })

        # Nom bo'yicha qidiruv — tur (rang/o'lcham) matni ham hisobga olinadi,
        # masalan "Fresh" yoki "1.8 L". Ochiq narxli yashirin mahsulot chiqmaydi.
        matches = list(Product.objects.filter(
            Q(name__icontains=q) | Q(brand__icontains=q) |
            Q(category__name__icontains=q) |
            Q(variants__color__icontains=q) | Q(variants__size__icontains=q)
        ).exclude(is_open_price=True).distinct()[:8])
        if len(matches) == 1:
            product = matches[0]
        elif matches:
            return JsonResponse({
                'found': False,
                'suggestions': [{'code': p.code, 'name': p.name} for p in matches],
            })
        else:
            return JsonResponse({'found': False, 'suggestions': []})

    stocks = (BranchStock.objects.filter(variant__product=product, branch=branch)
              .select_related('variant')
              .order_by('variant__size', 'variant__color'))

    variants = [{
        'stock_id': s.id,
        'variant_id': s.variant_id,
        'size': s.variant.size,
        'color': s.variant.color,
        'barcode': s.variant.barcode or '',
        'stock_count': s.stock_count,
        'sale_price': float(s.sale_price or product.default_sale_price),
        'wholesale_price': float(s.wholesale_price or 0),
        'cost_price': float(s.cost_price),
    } for s in stocks]

    # Skanerlangan barcode turi BO'SH (0), lekin AYNAN SHU RANG+O'LCHAMDAGI
    # (ya'ni bir xil tovar, boshqa barcode bilan qayta qabul qilingan) birodарда
    # qoldiq bor bo'lsa — o'shanga yo'naltiramiz. Shunda skaner "omborda yo'q"
    # deб dead-end bo'lmaydi. Faqat rang BO'SH BO'LMAGANда (aniq bir xil kod).
    if matched_variant_id and _matched_variant is not None:
        mv_row = next((v for v in variants if v['variant_id'] == matched_variant_id), None)
        if (mv_row is None or mv_row['stock_count'] <= 0):
            mc = (_matched_variant.color or '').strip().lower()
            ms = (_matched_variant.size or '').strip().lower()
            if mc:  # faqat aniq (bo'sh bo'lmagan) rang bo'yicha — xavfsiz
                cands = [v for v in variants if v['stock_count'] > 0
                         and (v['color'] or '').strip().lower() == mc
                         and (v['size'] or '').strip().lower() == ms]
                if cands:
                    cands.sort(key=lambda v: -v['stock_count'])
                    matched_variant_id = cands[0]['variant_id']

    # If the product exists but has no variants in this branch, tell the
    # user explicitly — the silent "topilmadi" is misleading.
    other_branches_with_stock = []
    if not variants:
        other_ids = (BranchStock.objects
                     .filter(variant__product=product, stock_count__gt=0)
                     .exclude(branch=branch)
                     .values_list('branch_id', flat=True).distinct())
        for b in Branch.objects.filter(pk__in=other_ids):
            other_branches_with_stock.append({'name': b.name, 'id': b.id})

    return JsonResponse({
        'found': True,
        'product': {
            'code': product.code,
            'external_barcode': product.external_barcode or '',
            'name': product.name,
            'default_sale_price': float(product.default_sale_price),
        },
        'branch_name': branch.name,
        'variants': variants,
        'matched_variant_id': matched_variant_id,
        'other_branches': other_branches_with_stock,
    })


# MON-8 flag-and-audit: qo'lda kiritilgan narx katalog narxidan shu chegaralardan
# ko'proq farq qilsa — audit logga yoziladi (sotuv baribir o'tadi). Ikki shart
# ham bajarilishi kerak: kamida N% VA kamida M so'm — rounding shovqinini kesadi.
PRICE_OVERRIDE_PCT = Decimal('5')       # 5%
PRICE_OVERRIDE_MIN_ABS = Decimal('1000')  # 1000 so'm


@login_required
def pos_checkout(request):
    """POST /pos/checkout/ with JSON body:
      { lines: [{stock_id, qty, sale_price}],
        payment_method: 'cash'|'card'|'transfer'|'mixed',
        customer_name, customer_phone, note }
    Creates SaleTransaction + Sales atomically. Returns {ok, txn_id, receipt_url}.
    """
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST only'}, status=405)

    branch = _user_branch_or_403(request)
    if branch is None:
        return JsonResponse({'ok': False, 'error': 'no branch'}, status=403)

    try:
        data = _json.loads(request.body.decode('utf-8'))
    except ValueError:
        return JsonResponse({'ok': False, 'error': 'bad JSON'}, status=400)

    lines = data.get('lines') or []
    if not lines:
        return JsonResponse({'ok': False, 'error': 'savat boʻsh'}, status=400)

    payment_method = (data.get('payment_method') or 'cash').strip()
    # Mixed payment: client sends [{method, amount}]. Validate sum after we
    # know the total. Keep as list of dicts for storage.
    payment_breakdown = data.get('payment_breakdown') or []
    if not isinstance(payment_breakdown, list):
        payment_breakdown = []

    # PAY-2 / MON-11: QR provider nomlari ('payme','click'...) 'transfer'ga
    # moslashtiriladi; NOMA'LUM usul RAD etiladi (jimgina 'cash'ga aylantirilmaydi
    # — aks holda karta/QR pul kassaga tushган kabi ko'rinib, kamomad bo'lardi).
    _VALID_METHODS = {'cash', 'card', 'transfer'}
    _QR_PROVIDERS = {'payme', 'click', 'uzum', 'humo', 'anor', 'alif',
                     'iman', 'zoodpay', 'other', 'noop'}

    def _norm_method(m):
        m = (m or '').strip().lower()
        return 'transfer' if m in _QR_PROVIDERS else m

    if payment_method != 'mixed':
        _pm = _norm_method(payment_method)
        if _pm not in _VALID_METHODS:
            return JsonResponse({'ok': False,
                'error': f"Noma'lum to'lov turi: {payment_method}"}, status=400)
        payment_method = _pm

    customer_name = (data.get('customer_name') or '').strip()[:120]
    customer_phone = (data.get('customer_phone') or '').strip()[:40]
    note = (data.get('note') or '').strip()[:200]
    # OFF-2: offline replay bir sotuvni ikki marta yozmasin. Mijoz bir martalik
    # kalit yuboradi; o'sha kalit bilan chek allaqachon bo'lsa — yangi
    # yaratmaymiz, mavjudini qaytaramiz.
    idem_key = (data.get('idempotency_key') or '').strip()[:64] or None
    # Pul qiymatlari Decimal bo'lishi SHART: modeldagi maydonlar DecimalField,
    # float bilan aralashsa `Decimal + float` -> TypeError (chek yig'indisi
    # hisoblanganda 500 xato). Shuning uchun hammasini Decimal'ga o'giramiz.
    from decimal import Decimal, InvalidOperation

    # Pul maydonlari DB'da Decimal(12,2) — mutlaq qiymati 10^10 dan kichik
    # bo'lishi SHART, aks holda "numeric field overflow" (500). Xato terilgan
    # ulkan raqam (masalan numpad'дa qo'shimcha nol) 500 emas, tushunarli
    # ogohlantirish berishi kerak.
    MAX_MONEY = Decimal('9999999999.99')

    def _money(v, allow_zero=True):
        try:
            d = Decimal(str(v if v not in (None, '') else 0))
        except (InvalidOperation, ValueError, TypeError):
            d = Decimal('0')
        if d < 0:
            d = Decimal('0')
        return d

    order_discount = _money(data.get('order_discount'))
    if order_discount > MAX_MONEY:
        return JsonResponse({'ok': False,
            'error': "Chegirma juda katta. Iltimos, to'g'ri summa kiriting."}, status=400)
    discount_reason = (data.get('discount_reason') or '').strip()[:200]

    # Parse line shape without touching the DB (input validation only).
    # The stock check that *can* race is done inside the atomic block below
    # with select_for_update, so two concurrent kassirs cannot both pass
    # the qty check against the same row.
    parsed_lines = []
    for ln in lines:
        try:
            sid = int(ln['stock_id'])
            qty = int(ln['qty'])
            line_discount = _money(ln.get('line_discount'))
            # MON-13: narx QAT'IY tekshiriladi — "abc" kabi axlat jimgina 0 ga
            # aylanib, bepul tovarga aylanmasin. Noto'g'ri narx = xato.
            _raw_price = ln['sale_price']
            price = Decimal(str(_raw_price if _raw_price not in (None, '') else 0))
        except (KeyError, ValueError, TypeError, InvalidOperation):
            return JsonResponse({'ok': False, 'error': "noto'g'ri narx yoki qator"}, status=400)
        if qty <= 0 or price < 0:
            return JsonResponse({'ok': False, 'error': 'qty/narx noto\'g\'ri'}, status=400)
        if price > MAX_MONEY or line_discount > MAX_MONEY:
            return JsonResponse({'ok': False,
                'error': "Narx juda katta. Iltimos, to'g'ri narx kiriting."}, status=400)
        parsed_lines.append({'sid': sid, 'qty': qty, 'price': price, 'ld': line_discount})

    # STK-16 / R1: bir xil stock_id li qatorlarni BIRLASHTIRMAYMIZ (ochiq
    # narxli tovarlar bitta stock_id ostida, ammo HAR XIL qo'lда kiritilgan
    # narxда bo'ladi — birlashtirish oxirgi narxда IKKALASINI yozib, kam pul
    # olardi). Buning o'rniga QOLDIQ tekshiruvini stock_id bo'yicha JAMLAB
    # bajaramiz (quyida), narx esa har qatorда o'ziniki bo'lib qoladi.
    _qty_by_sid = {}
    for _l in parsed_lines:
        _qty_by_sid[_l['sid']] = _qty_by_sid.get(_l['sid'], 0) + _l['qty']

    # MON-18: chegirma savat summasidan OSHMASLIGI kerak. Ilgari chegirma faqat
    # MAX_MONEY ga tekshirilardi — 1 000 000 chegirma 1 000 so'mlik sotuvга
    # qo'yilsa, jami −999 000 bo'lib, cash_sales()/expected_cash() minusга
    # tushib, kassadan yashirin pul olish yo'li ochilardi. Endi cheklaymiz.
    _subtotal = Decimal('0')
    for _l in parsed_lines:
        _line_gross = Decimal(_l['qty']) * _l['price']
        if _l['ld'] > _line_gross:      # qator chegirmasi qatordan oshmasin
            _l['ld'] = _line_gross
        _subtotal += _line_gross - _l['ld']
    if order_discount > _subtotal:      # chek chegirmasi savatdan oshmasin
        order_discount = _subtotal

    # V5/S2: aksiya chegirmasini SERVERда qayta hisoblaymiz — mijoz yuborган
    # applied_promos ga ishonmaymiz (aks holда soxta aksiya nomi bilan butun
    # savatni 0 ga tushirish mumkin edi). Aksiyадан tashqari qo'lда chegirма
    # esa SABAB talab qiladi (auditда ko'rinsin).
    _promo_stocks = {s.id: s for s in BranchStock.objects
                     .filter(id__in=list(_qty_by_sid.keys()))
                     .select_related('variant__product')}
    _cart_lines = []
    for _l in parsed_lines:
        _s = _promo_stocks.get(_l['sid'])
        if not _s:
            continue
        _cart_lines.append({
            'stock_id': _l['sid'], 'qty': _l['qty'], 'price': float(_l['price']),
            'product_id': _s.variant.product_id,
            'category_id': _s.variant.product.category_id,
        })
    _server_applied, _server_promo_f = _evaluate_promotions(_cart_lines)
    _server_promo = Decimal(str(_server_promo_f))
    _claimed_promo = Decimal('0')
    for _p in (data.get('applied_promos') or []):
        try:
            _claimed_promo += Decimal(str(_p.get('discount') or 0))
        except (InvalidOperation, TypeError):
            pass
    if _claimed_promo > _server_promo + Decimal('1'):
        return JsonResponse({'ok': False,
            'error': "Aksiya chegirmasi tekshiruvdan o'tmadi — savatni yangilang."},
            status=400)
    _manual_disc = order_discount - _claimed_promo
    if _manual_disc < Decimal('-1'):
        return JsonResponse({'ok': False,
            'error': "Chegirma summasi mos emas — qayta urinib ko'ring."}, status=400)
    if _manual_disc < 0:
        _manual_disc = Decimal('0')
    if _manual_disc > Decimal('0.5'):
        if not discount_reason:
            return JsonResponse({'ok': False,
                'error': "Qo'lda chegirma uchun sabab kiriting."}, status=400)
        # DISC-7: sabab bor, lekin "s" yoki "3000" bo'lsa — audit uchun foydasiz.
        if not _valid_discount_reason(discount_reason):
            return JsonResponse({'ok': False,
                'error': "Sababni ro'yxatdan tanlang yoki qisqacha yozing "
                         "(kamida 3 harf)."}, status=400)
    # Klient sonига emas, SERVER hisoblаган aksiyaga + qo'lда qismga ishonamiz.
    order_discount = _server_promo + _manual_disc
    # DISC-1: aksiya ulushini ALOHIDA saqlaymiz. Ilgari ikkalasi bitta songa
    # qo'shilib ketardi va hisobotlarда kassir bergan chegirma bilan egasi
    # sozlagan aksiya farqlanmasdi. Kassir ixtiyoridagi qism =
    # order_discount − promo.
    # DISC-2 tuzatishi: bu yerda ilgari "sababsiz 289 chek — aslida aksiya"
    # deb yozilgan edi. NOTO'G'RI: Promotion jadvali bo'sh, sabab majburiyati
    # esa 2026-08-26 da qo'shilgan. O'sha cheklar QO'LDA berilган chegirma —
    # shunчaki sababi yozilmagan davrdan qolgan (0061 migratsiyasi tuzatdi).
    _promo_part = _server_promo
    if order_discount > _subtotal:
        order_discount = _subtotal
    if _promo_part > order_discount:      # savat cheklovi aksiyani ham qirqsa
        _promo_part = order_discount

    # Resolve / auto-create Customer by phone (most reliable key)
    customer = None
    if customer_phone:
        cleaned_phone = ''.join(c for c in customer_phone if c.isdigit() or c == '+')
        _digits = sum(c.isdigit() for c in cleaned_phone)
        if cleaned_phone:
            customer = Customer.objects.filter(phone=cleaned_phone).first()
            # SEC-8: juda qisqa/soxta raqamдан YANGI mijoz yaratmaymiz (baza
            # ifloslanishi va enumeratsiya oldini oladi). O'zbek raqami kamida
            # 9 raqam. Mavjud mijozni topsak bog'laymiz; aks holda mijozsiz sotuv.
            if not customer and _digits >= 9:
                customer = Customer.objects.create(
                    phone=cleaned_phone, name=customer_name,
                )
            elif customer and customer_name and not customer.name:
                customer.name = customer_name
                customer.save(update_fields=['name'])

    open_shift = _open_shift_for(branch)

    # OFF-7: sotuvning ASL vaqti (client_ts). Offline navbatдan kech kelsa ham
    # chek shu vaqtни ko'rsatadi va smen shu vaqt bo'yicha aniqlanadi — sinxron
    # vaqti bo'yicha emas (aks holda ertalabki savdo kechki smenга tushardi).
    target_sold_at = None
    target_shift = open_shift
    # R4: client_ts'ga FAQAT offline replayда ishonamiz. Oddiy onlayn sotuvда
    # pos.html client_ts yuborса ham, sotuv vaqti = SERVER vaqti bo'lsin (qurilma
    # soati adashса ertalabki savdo kechki smenга tushib ketmasin). Backdating
    # faqat haqiqiy navbat-replay uchun.
    _client_ts = (data.get('client_ts') or '').strip()
    if _client_ts and data.get('is_offline_replay'):
        from django.utils.dateparse import parse_datetime
        _dt = parse_datetime(_client_ts)
        if _dt is not None:
            if timezone.is_naive(_dt):
                _dt = timezone.make_aware(_dt, timezone.get_current_timezone())
            _now = timezone.now()
            # kelajak yoki >7 kun eski vaqt — qurilma soati xato, e'tiborsiz
            if _now - timedelta(days=7) <= _dt <= _now + timedelta(minutes=5):
                target_sold_at = _dt
                # O'sha payt ochiq bo'lgan smenni topamiz (hozir yopilgan bo'lsa
                # ham). ARCH-5 tufayli FK bo'yicha o'sha smen hisobotiga tushadi.
                _hist = (Shift.objects.filter(branch=branch, opened_at__lte=_dt)
                         .filter(Q(closed_at__isnull=True) | Q(closed_at__gte=_dt))
                         .order_by('-opened_at').first())
                if _hist:
                    target_shift = _hist
                    # R3: agar bu smen ALLAQACHON yopilган bo'lса, uning
                    # closing_expected_cash'i (MON-22) QOTIRILган — kech kelган
                    # bu sotuv o'sha raqamга kirmaydi. Jimgina buzмaslik uchun
                    # egaга ko'rinadigan qilib belgilaymiz (audit + xabar).
                    if _hist.closed_at is not None:
                        try:
                            AuditLog.objects.create(
                                user=request.user,
                                username_snapshot=request.user.username,
                                action=AuditLog.Action.CREATE,
                                model_name='SaleTransaction',
                                object_id='',
                                object_repr=(f"Kech offline sotuv YOPILGAN smen #{_hist.id} "
                                             f"ga tushdi ({branch.name}, {_dt:%Y-%m-%d %H:%M}) "
                                             f"— yopilish naqd hisobiga kirmaydi")[:300],
                                changes={'shift_id': _hist.id, 'client_ts': _client_ts},
                            )
                            from .notifications import send_telegram
                            send_telegram(
                                f"⚠️ <b>Kech offline sotuv</b>\n"
                                f"Filial: {branch.name} · Smen #{_hist.id} allaqachon yopilgan.\n"
                                f"Sotuv vaqti: {_dt:%Y-%m-%d %H:%M}. Yopilish naqd hisobiga "
                                f"kirmaydi — Z-hisobotни tekshiring.")
                        except Exception:
                            logger.exception('late-replay flag failed (shift %s)', _hist.id)

    # MON-9: smensiz sotuv YO'Q. Aks holda shift=None bilan yozilib, hech qaysi
    # Z-hisobotга tushmaydi — kassa "ortiq" chiqadi. (Offline replay o'zining
    # tarixiy smenига tushishi mumkin; oddiy sotuv hozirgi ochiq smenга.)
    if not target_shift:
        return JsonResponse({'ok': False,
            'error': "Smen ochilmagan. Sotuv faqat ochiq smen davomida amalga oshiriladi."},
            status=400)

    def _receipt_response(txn):
        return JsonResponse({
            'ok': True,
            'txn_id': txn.pk,
            'receipt_url': f'/transaction/{txn.public_id}/?autoprint=1',
            'total': float(txn.total),
            'item_count': txn.item_count,
            'sms': None,
            'duplicate': True,
        })

    # OFF-2: bu kalit bilan chek allaqachon yozilgan bo'lsa — takror EMAS,
    # o'sha chekni qaytaramiz (ombor ikki marta kamaymaydi, tushum ikki
    # barobar bo'lmaydi).
    if idem_key:
        dup = SaleTransaction.objects.filter(idempotency_key=idem_key).first()
        if dup:
            return _receipt_response(dup)

    # Sanitize breakdown: QR provider -> 'transfer'; NOMA'LUM usul RAD etiladi.
    clean_breakdown = []
    for entry in payment_breakdown:
        try:
            m = _norm_method(entry.get('method'))
            a = float(entry.get('amount') or 0)
        except (AttributeError, TypeError, ValueError):
            continue
        if not m or a <= 0:
            continue
        if m not in _VALID_METHODS:
            return JsonResponse({'ok': False,
                'error': f"Noma'lum to'lov turi (aralash): {entry.get('method')}"}, status=400)
        clean_breakdown.append({'method': m, 'amount': round(a, 2)})

    # If mixed but breakdown empty, fallback to single-method
    if payment_method == 'mixed' and not clean_breakdown:
        payment_method = 'cash'

    # MON-10: ARALASH to'lov summasi chekdan KAM bo'lmasin. Brauzer buni
    # tekshiradi, lekin devtools yoki offline-replay chetlab o'tishi mumkin —
    # shuning uchun serverда ham tekshiramiz.
    if payment_method == 'mixed':
        _order_total = Decimal('0')
        for _l in parsed_lines:
            _order_total += Decimal(_l['qty']) * _l['price'] - _l['ld']
        _order_total -= order_discount
        if _order_total < 0:
            _order_total = Decimal('0')
        _paid = sum((Decimal(str(e['amount'])) for e in clean_breakdown), Decimal('0'))
        # MON-16: IKKI TOMONLAMA tekshiruv. Ilgari faqat KAM to'lov rad etilardi;
        # ORTIQ to'lov ham xuddi shu nolinchi-farq ekspluatatsiyasига olib
        # kelardi (split_breakdown qoldiqni oxirgi usulga tashlab, kutilgan
        # naqdни shishiradi). Aralashда summa chekка AYNAN teng bo'lishi kerak
        # (ortiqcha naqd = qaytim, u alohida maydon — to'lov legi emas).
        if abs(_paid - _order_total) > Decimal('0.5'):
            _diff = _order_total - _paid
            _msg = (f"Aralash to'lov yetishmaydi: {int(_diff)} so'm kam."
                    if _diff > 0 else
                    f"Aralash to'lov ortiqcha: {int(-_diff)} so'm. Summa chekка teng bo'lsin.")
            return JsonResponse({'ok': False, 'error': _msg}, status=400)

    class _CheckoutAbort(Exception):
        def __init__(self, payload, status=400):
            self.payload = payload
            self.status = status

    try:
        with transaction.atomic():
            # Lock all referenced stocks first (sort by pk to avoid deadlocks
            # when two checkouts overlap on the same items in different order).
            sids = sorted({l['sid'] for l in parsed_lines})
            locked = {
                s.pk: s
                for s in BranchStock.objects
                    .select_for_update()
                    .select_related('variant__product', 'branch')
                    .filter(pk__in=sids, branch=branch)
            }
            for sid in sids:
                if sid not in locked:
                    raise _CheckoutAbort({'ok': False, 'error': f'stock {sid} topilmadi'})

            # Re-validate quantities against the FRESHLY locked stock_count.
            # No concurrent checkout can change it between this check and the
            # F() deduction below — they queue on the row lock instead.
            for _sid in sids:
                stock = locked[_sid]
                # R1/STK-16: shu stock_id bo'yicha JAMI soni (bir nechta qator
                # bir tovarga tegishli bo'lса) qoldiqдан oshmasin.
                _need = _qty_by_sid.get(_sid, 0)
                if (not stock.variant.product.is_open_price
                        and _need > stock.stock_count):
                    err = (f"{stock.variant.product.code} {stock.variant.size}/{stock.variant.color}: "
                           f"omborda faqat {stock.stock_count} ta bor, soʻrov {_need}")
                    # C2: if this is an offline-queue replay, alert admin via Telegram
                    # and log to AuditLog — kassir's offline sale was rejected at sync time.
                    if data.get('is_offline_replay'):
                        try:
                            from .notifications import send_telegram
                            # OFF-6: TO'LIQ payload — sotuv yo'qolib ketmasin.
                            # Barcha qatorlar, narxlar, to'lov turi, mijoz.
                            _pl_lines = []
                            for _l in parsed_lines:
                                _s = locked.get(_l['sid'])
                                _nm = (f"{_s.variant.product.code} "
                                       f"{_s.variant.size or ''}/{_s.variant.color or ''}"
                                       if _s else f"stock {_l['sid']}")
                                _pl_lines.append(
                                    f"  • {_nm} — {_l['qty']} × {int(_l['price'])}")
                            _payload = {
                                'lines': [{'sid': _l['sid'], 'qty': _l['qty'],
                                           'price': float(_l['price'])}
                                          for _l in parsed_lines],
                                'payment_method': payment_method,
                                'payment_breakdown': clean_breakdown,
                                'customer_phone': customer_phone,
                                'idempotency_key': idem_key,
                            }
                            _cust = f"\nMijoz: {customer_phone}" if customer_phone else ""
                            send_telegram(
                                f"⚠️ <b>Offline sotuv RAD ETILDI</b>\n"
                                f"Filial: {branch.name}\n"
                                f"Kassir: {request.user.username}\n"
                                f"Sabab: {err}\n"
                                f"<b>Sotuv tarkibi:</b>\n" + "\n".join(_pl_lines) +
                                f"\nTo'lov: {payment_method}{_cust}\n"
                                f"⚠️ Pul allaqachon kassada bo'lishi mumkin — qo'lda tekshiring."
                            )
                            AuditLog.objects.create(
                                user=request.user,
                                username_snapshot=request.user.username,
                                action=AuditLog.Action.UPDATE,
                                model_name='OfflineConflict',
                                object_repr=err[:200],
                                changes={'rejected_sale': _payload, 'reason': err[:200]},
                            )
                        except Exception:
                            logger.exception('offline-conflict alert failed')
                    raise _CheckoutAbort({'ok': False, 'error': err})

            # MON-5: smen checkout davomida yopilib qolmasin (poyga —
            # _open_shift_for atomic blokдан oldin chaqirilgan). Offline replay
            # tarixiy yopiq smenга ATAYLAB tushadi, uni tekshirmaymiz.
            if not data.get('is_offline_replay'):
                _st = (Shift.objects.filter(pk=target_shift.pk)
                       .values_list('status', flat=True).first())
                if _st != Shift.Status.OPEN:
                    raise _CheckoutAbort({'ok': False,
                        'error': "Smen yopildi. Sotuvni qayta boshlang."})

            txn = SaleTransaction.objects.create(
                branch=branch,
                sold_by=request.user,
                payment_method=payment_method,
                payment_breakdown=clean_breakdown,
                customer=customer,
                customer_name=customer_name,
                customer_phone=customer_phone,
                note=note,
                order_discount=order_discount,
                promo_discount=_promo_part,   # DISC-1
                discount_reason=discount_reason,
                shift=target_shift,                      # OFF-7 / ARCH-5
                sold_at=target_sold_at or timezone.now(),  # OFF-7
                idempotency_key=idem_key,
            )
            price_overrides = []  # MON-8: qo'lda kiritilgan narx auditи
            for ln in parsed_lines:
                stock = locked[ln['sid']]
                prod = stock.variant.product
                if not prod.is_open_price:
                    stock.stock_count = F('stock_count') - ln['qty']
                    stock.save(update_fields=['stock_count'])  # STK-7
                Sale.objects.create(
                    transaction=txn,
                    variant=stock.variant, branch=stock.branch,
                    quantity=ln['qty'],
                    sale_price=ln['price'],
                    cost_at_sale=stock.cost_price,
                    line_discount=ln['ld'],
                    sold_by=request.user,
                )
                # MON-8 (flag-and-audit): qo'lda kiritilgan narxni QABUL qilamiz
                # (kelishilgan/ulgurji narx — ish jarayoni buzilmaydi), LEKIN
                # katalog narxidan sezilarli farq qilsa yoki TANNARXDAN past
                # bo'lsa — audit uchun belgilaymiz (kim, qachon, qancha).
                if not prod.is_open_price:
                    catalog = stock.sale_price or Decimal('0')
                    entered = ln['price']
                    cost = stock.cost_price or Decimal('0')
                    below_cost = bool(cost > 0 and entered < cost)
                    material = False
                    pct = Decimal('0')
                    if catalog > 0:
                        pct = abs(entered - catalog) / catalog * 100
                        # rounding shovqinidan qochish: kamida 5% VA 1000 so'm farq
                        material = (pct >= PRICE_OVERRIDE_PCT
                                    and abs(entered - catalog) >= PRICE_OVERRIDE_MIN_ABS)
                    if material or below_cost:
                        price_overrides.append({
                            'code': prod.code, 'name': prod.name,
                            'variant': ('/'.join(p for p in
                                        (stock.variant.size, stock.variant.color) if p)
                                        or (stock.variant.barcode or '')),
                            'catalog': float(catalog), 'entered': float(entered),
                            'cost': float(cost), 'qty': ln['qty'],
                            'diff_pct': round(float(pct), 1),
                            'below_cost': below_cost,
                        })
    except _CheckoutAbort as e:
        return JsonResponse(e.payload, status=e.status)
    except IntegrityError:
        # OFF-2 poyga: bir vaqtning o'zida ikki replay bir xil kalit bilan
        # kelsa, ikkinchi create unique cheklovга urилadi. Mavjud chekни
        # qaytaramiz — sotuv baribir bir marta yozilgan.
        if idem_key:
            dup = SaleTransaction.objects.filter(idempotency_key=idem_key).first()
            if dup:
                return _receipt_response(dup)
        raise

    # OFF-2: chek ALLAQACHON saqlangan (atomic commit bo'ldi). Bundan keyingi
    # "best-effort" chaqiruvlar (fiskal, SMS) XATO bersa ham 500 qaytmasligi
    # kerak — aks holda offline navbat chekни ikkinchi marta yuborib, sotuv
    # ikki marta yozilardi (ikki barobar tushum, ombor ikki marta kamayadi).
    # MON-8: narx o'zgartirishlarini audit logga yozamiz (best-effort).
    if price_overrides:
        try:
            for ov in price_overrides:
                flag = ', TANNARXDAN PAST!' if ov['below_cost'] else ''
                AuditLog.objects.create(
                    user=request.user,
                    username_snapshot=request.user.username,
                    action=AuditLog.Action.UPDATE,
                    model_name='PriceOverride',
                    object_id=str(txn.pk),
                    object_repr=(f"{ov['code']} {ov['variant']}: katalog "
                                 f"{int(ov['catalog'])} → kiritilgan "
                                 f"{int(ov['entered'])} ({ov['diff_pct']}%{flag})")[:300],
                    changes={'price_override': ov, 'txn_id': txn.pk},
                )
            # Faqat TANNARXDAN PAST (zararli) sotuvга Telegram ogohlantirish —
            # bu eng o'tkir "pul oqib ketmoqda" signali, kamdan-kam bo'ladi.
            loss = [o for o in price_overrides if o['below_cost']]
            if loss:
                try:
                    from .notifications import send_telegram
                    rows = '\n'.join(
                        f"• {o['code']} {o['variant']}: {int(o['entered'])} so'm "
                        f"(tannarx {int(o['cost'])})" for o in loss)
                    send_telegram(
                        f"🔴 <b>Tannarxdan past sotuv</b>\n"
                        f"Kassir: {request.user.username} · Chek #{txn.pk}\n{rows}")
                except Exception:
                    logger.exception('below-cost telegram alert failed (txn %s)', txn.pk)
        except Exception:
            logger.exception('price-override audit failed for txn %s', txn.pk)

    try:
        from .fiscal import submit_for_transaction
        submit_for_transaction(txn)
    except Exception:
        logger.exception('fiscal submit failed for txn %s (sotuv saqlangan)', txn.pk)

    sms_result = None
    if data.get('send_sms') and customer_phone:
        try:
            from .sms import send_receipt
            sms_result = send_receipt(txn, customer_phone)
        except Exception:
            logger.exception('SMS receipt failed for txn %s', txn.pk)

    return JsonResponse({
        'ok': True,
        'txn_id': txn.pk,
        'receipt_url': f'/transaction/{txn.public_id}/?autoprint=1',
        'total': float(txn.total),
        'item_count': txn.item_count,
        'sms': sms_result,
    })


# ---------- POS PARK / HOLD ----------

@login_required
def pos_park(request):
    """POST /pos/park/ — savatni vaqtincha saqlash.
    Body JSON: {label, lines, customer_name, customer_phone, order_discount, discount_reason}
    """
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST only'}, status=405)
    branch = _user_branch_or_403(request)
    if branch is None:
        return JsonResponse({'ok': False, 'error': 'no branch'}, status=403)
    try:
        data = _json.loads(request.body.decode('utf-8'))
    except ValueError:
        return JsonResponse({'ok': False, 'error': 'bad JSON'}, status=400)

    label = (data.get('label') or '').strip()[:80]
    if not label:
        return JsonResponse({'ok': False, 'error': 'yorliq kerak'}, status=400)
    lines = data.get('lines') or []
    if not lines:
        return JsonResponse({'ok': False, 'error': 'savat bo\'sh'}, status=400)

    try:
        order_discount = max(0, float(data.get('order_discount') or 0))
    except (ValueError, TypeError):
        order_discount = 0

    parked = ParkedSale.objects.create(
        branch=branch,
        parked_by=request.user,
        label=label,
        cart_json=_json.dumps(lines),
        customer_name=(data.get('customer_name') or '').strip()[:120],
        customer_phone=(data.get('customer_phone') or '').strip()[:30],
        order_discount=order_discount,
        discount_reason=(data.get('discount_reason') or '').strip()[:200],
    )
    return JsonResponse({'ok': True, 'id': parked.pk, 'label': parked.label})


@login_required
def pos_parked_resume(request, pk):
    """POST /pos/parked/<pk>/resume/ — saqlangan savatni qaytarish va o'chirish."""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST only'}, status=405)
    branch = _user_branch_or_403(request)
    if branch is None:
        return JsonResponse({'ok': False, 'error': 'no branch'}, status=403)
    parked = ParkedSale.objects.filter(pk=pk, branch=branch).first()
    if not parked:
        return JsonResponse({'ok': False, 'error': 'topilmadi'}, status=404)
    try:
        lines = _json.loads(parked.cart_json)
    except ValueError:
        lines = []
    payload = {
        'ok': True,
        'lines': lines,
        'customer_name': parked.customer_name,
        'customer_phone': parked.customer_phone,
        'order_discount': float(parked.order_discount),
        'discount_reason': parked.discount_reason,
    }
    parked.delete()
    return JsonResponse(payload)


@admin_required
def payment_qr_list(request):
    """Admin: barcha filiallardagi static QR'lar ro'yxati + yangi qo'shish formasi."""
    if request.method == 'POST':
        action = request.POST.get('action') or 'create'
        if action == 'delete':
            try:
                pk = int(request.POST.get('pk') or 0)
            except ValueError:
                pk = 0
            PaymentQR.objects.filter(pk=pk).delete()
            messages.success(request, "QR o'chirildi.")
            return redirect('payment_qr_list')
        # create
        try:
            branch_id = int(request.POST.get('branch') or 0)
            branch_obj = Branch.objects.filter(pk=branch_id, is_active=True).first()
            provider = (request.POST.get('provider') or '').strip()
            if not branch_obj or not provider:
                messages.error(request, "Filial va provider tanlang.")
                return redirect('payment_qr_list')
            qr = PaymentQR(
                branch=branch_obj,
                provider=provider,
                label=(request.POST.get('label') or '').strip()[:120],
                qr_payload=(request.POST.get('qr_payload') or '').strip()[:500],
                instructions=(request.POST.get('instructions') or '').strip()[:200],
                is_active=True,
            )
            if 'qr_image' in request.FILES:
                qr.qr_image = request.FILES['qr_image']
            if not qr.qr_image and not qr.qr_payload:
                messages.error(request,
                    "QR rasm yuklang yoki QR ichidagi matn/URL kiriting.")
                return redirect('payment_qr_list')
            qr.save()
            messages.success(request,
                f"{branch_obj.name} uchun {qr.get_provider_display()} QR saqlandi.")
        except (ValueError, TypeError) as e:
            messages.error(request, f"Xatolik: {e}")
        return redirect('payment_qr_list')

    qrs = (PaymentQR.objects.select_related('branch')
           .order_by('branch__name', 'provider'))
    branches = Branch.objects.filter(is_active=True).order_by('name')
    return render(request, 'inventory/payment_qr_list.html', {
        'qrs': qrs, 'branches': branches,
        'providers': PaymentQR.Provider.choices,
    })


@login_required
def pos_static_qr(request, pk):
    """GET /pos/qr/<pk>/ — joriy filialdagi static QR ma'lumotini qaytaradi.
    Image URL (yoki qr_payload'dan generatsiya qilingan data URL) + ko'rsatma."""
    branch = _user_branch_or_403(request)
    if branch is None:
        return JsonResponse({'ok': False, 'error': 'no branch'}, status=403)
    qr = PaymentQR.objects.filter(pk=pk, branch=branch, is_active=True).first()
    if not qr:
        return JsonResponse({'ok': False, 'error': 'topilmadi'}, status=404)

    image_url = ''
    if qr.qr_image:
        image_url = qr.qr_image.url
    elif qr.qr_payload:
        # Generate QR on the fly into a data: URL
        import io, base64
        img = qrcode.make(qr.qr_payload)
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        b64 = base64.b64encode(buf.getvalue()).decode('ascii')
        image_url = f'data:image/png;base64,{b64}'

    return JsonResponse({
        'ok': True,
        'provider': qr.provider,
        'provider_display': qr.get_provider_display(),
        'label': qr.label,
        'image_url': image_url,
        'instructions': qr.instructions,
    })


@login_required
def pos_payment_create(request):
    """POST /pos/payment/create/ JSON: {provider, amount, lines}
    PaymentIntent yaratadi va ref_code qaytaradi. POS modal'i shu ref_code'ni
    mijozga ko'rsatadi (mijoz to'lov izohi sifatida kiritadi)."""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST only'}, status=405)
    branch = _user_branch_or_403(request)
    if branch is None:
        return JsonResponse({'ok': False, 'error': 'no branch'}, status=403)
    try:
        data = _json.loads(request.body.decode('utf-8'))
    except ValueError:
        return JsonResponse({'ok': False, 'error': 'bad JSON'}, status=400)
    provider = (data.get('provider') or '').strip()
    try:
        amount = float(data.get('amount') or 0)
    except (TypeError, ValueError):
        amount = 0
    if not provider or amount <= 0:
        return JsonResponse({'ok': False, 'error': 'parametr yetishmaydi'}, status=400)
    import secrets as _secrets
    ref_code = _secrets.token_hex(3).upper()  # 6 belgili kod
    intent = PaymentIntent.objects.create(
        branch=branch, initiated_by=request.user,
        provider=provider, amount=amount, ref_code=ref_code,
        cart_snapshot=_json.dumps(data.get('lines') or []),
    )
    return JsonResponse({
        'ok': True,
        'intent_id': intent.id,
        'ref_code': ref_code,
        'status': intent.status,
    })


@login_required
def pos_payment_check(request, pk):
    """GET /pos/payment/check/<id>/ — intent holatini qaytaradi.

    Real merchant API yo'q paytda DEMO_AUTO_PAY_SECONDS o'tgan bo'lsa
    avtomatik 'paid' deb belgilanadi (sinov uchun)."""
    branch = _user_branch_or_403(request)
    if branch is None:
        return JsonResponse({'ok': False, 'error': 'no branch'}, status=403)
    intent = PaymentIntent.objects.filter(pk=pk, branch=branch).first()
    if not intent:
        return JsonResponse({'ok': False, 'error': 'topilmadi'}, status=404)

    # Demo auto-pay: ma'lum vaqt o'tgach 'paid' bo'lib turadi
    demo_seconds = getattr(settings, 'DEMO_AUTO_PAY_SECONDS', 0)
    if (intent.status == PaymentIntent.Status.PENDING and demo_seconds > 0):
        elapsed = (timezone.now() - intent.created_at).total_seconds()
        if elapsed >= demo_seconds:
            intent.status = PaymentIntent.Status.PAID
            intent.paid_at = timezone.now()
            intent.provider_txn_id = f'demo-{intent.id}'
            intent.save(update_fields=['status', 'paid_at', 'provider_txn_id'])

    return JsonResponse({
        'ok': True,
        'intent_id': intent.id,
        'status': intent.status,
        'ref_code': intent.ref_code,
        'provider_txn_id': intent.provider_txn_id,
        'paid_at': intent.paid_at.isoformat() if intent.paid_at else None,
    })


@csrf_exempt
def payments_webhook(request, provider):
    """POST /payments/webhook/<provider>/ — real merchant API callback.

    PAY-1 XAVFSIZLIK: bu endpoint IMZO (signature) bilan himoyalangan.
      1. PAYMENTS_WEBHOOK_SECRET sozlanmagan bo'lsa — HAMMA chaqiruv RAD etiladi
         (hech qanday to'lov 'paid' bo'lmaydi). Hozir shунday — QR real ulanмaган.
      2. Sozlanган bo'lsa — 'X-Signature' sarlavhasi tananing HMAC-SHA256 imzosiga
         AYNAN mos kelishi shart (hmac.compare_digest).
      3. Faqat ANIQ ref_code bo'yicha topiladi — 'summa + oxirgi 30 daqiqa'
         fallback OLIB TASHLANDI (mijoz summani bilса, boshqa chekни paid
         qilиб qo'ymasin).
    """
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST only'}, status=405)

    body_raw = request.body

    # 1) Imzo tekshiruvi — sirsiz endpoint hech narsani paid qilмaydi
    import hmac as _hmac, hashlib as _hashlib
    sig = (request.headers.get('X-Signature') or request.headers.get('X-Signature-Sha256') or '').strip()
    secret = getattr(settings, 'PAYMENTS_WEBHOOK_SECRET', '') or ''
    # Imzosiz chaqiruv — HAR DOIM rad (403). Mijoz kassada turib telefonidan
    # "to'ladim" deb yuborgan xabar hech qachon o'tmasligi SHART.
    if not sig:
        logger.warning('[payments webhook %s] REJECTED — missing signature', provider)
        return JsonResponse({'ok': False, 'error': 'signature required'}, status=403)
    # Imzo bor, lekin server siri sozlanmagan — tekshirib bo'lmaydi, RAD (503).
    if not secret:
        logger.warning('[payments webhook %s] REJECTED — no PAYMENTS_WEBHOOK_SECRET configured', provider)
        return JsonResponse({'ok': False, 'error': 'webhook not configured'}, status=503)
    expected = _hmac.new(secret.encode('utf-8'), body_raw, _hashlib.sha256).hexdigest()
    if not _hmac.compare_digest(sig, expected):
        logger.warning('[payments webhook %s] REJECTED — bad signature', provider)
        return JsonResponse({'ok': False, 'error': 'invalid signature'}, status=403)

    try:
        payload = _json.loads(body_raw.decode('utf-8')) if body_raw else {}
    except (ValueError, UnicodeDecodeError):
        payload = {}

    logger.info('[payments webhook %s] %s', provider, payload)

    # 2) FAQAT aniq ref_code bo'yicha (amount-window fallback yo'q)
    ref_code = (payload.get('ref_code') or payload.get('comment')
                or payload.get('memo') or '').strip().upper()
    intent = None
    if ref_code:
        intent = (PaymentIntent.objects
                  .filter(provider=provider, ref_code=ref_code,
                          status=PaymentIntent.Status.PENDING)
                  .order_by('-created_at').first())
    if not intent:
        return JsonResponse({'ok': False, 'error': 'mos intent topilmadi'},
                            status=404)

    intent.status = PaymentIntent.Status.PAID
    intent.paid_at = timezone.now()
    intent.provider_txn_id = str(payload.get('txn_id') or payload.get('id') or '')[:120]
    intent.save(update_fields=['status', 'paid_at', 'provider_txn_id'])
    return JsonResponse({'ok': True, 'intent_id': intent.id})


@login_required
def pos_promo_eval(request):
    """POST /pos/promo-eval/ {lines: [{stock_id, qty, sale_price}]}
    Aktiv aksiyalarni tekshiradi va qo'llab bo'ladigan chegirmalar
    ro'yxatini qaytaradi. Hech narsa o'zgartirmaydi — faqat ko'rsatish."""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST only'}, status=405)
    try:
        data = _json.loads(request.body.decode('utf-8'))
    except ValueError:
        return JsonResponse({'ok': False, 'error': 'bad JSON'}, status=400)

    lines_raw = data.get('lines') or []
    if not lines_raw:
        return JsonResponse({'ok': True, 'promotions': [], 'total_discount': 0})

    # Stock IDs -> product/category lookup
    stock_ids = [int(l['stock_id']) for l in lines_raw if 'stock_id' in l]
    stocks = {s.id: s for s in BranchStock.objects
              .filter(id__in=stock_ids)
              .select_related('variant__product__category')}

    # Build a per-line view: stock_id, qty, price, product_id, category_id
    cart_lines = []
    for ln in lines_raw:
        try:
            sid = int(ln['stock_id'])
            qty = int(ln['qty'])
            price = float(ln['sale_price'])
        except (KeyError, ValueError, TypeError):
            continue
        s = stocks.get(sid)
        if not s:
            continue
        cart_lines.append({
            'stock_id': sid,
            'qty': qty,
            'price': price,
            'product_id': s.variant.product_id,
            'category_id': s.variant.product.category_id,
        })

    applied, total_discount = _evaluate_promotions(cart_lines)
    return JsonResponse({
        'ok': True,
        'promotions': applied,
        'total_discount': round(total_discount, 2),
    })


def _evaluate_promotions(cart_lines):
    """V5: aktiv aksiyalardan kelib chiqadigan chegirmani SERVERда hisoblaydi.

    cart_lines: [{stock_id, qty, price, product_id, category_id}]. Bu yagona
    manba — pos_promo_eval (ko'rsatish) ham, pos_checkout (tekshirish) ham shuni
    chaqiradi, shunда mijoz o'ylab topган aksiya chegirmasi qabul qilinmaydi.
    (applied_list, total_discount_float) qaytaradi.
    """
    now = timezone.now()
    promos = (Promotion.objects
              .filter(is_active=True, valid_from__lte=now)
              .filter(Q(valid_until__isnull=True) | Q(valid_until__gte=now))
              .prefetch_related('target_products'))

    applied = []
    total_discount = 0.0
    for promo in promos:
        # Find lines that match this promo's target
        target_product_ids = set(promo.target_products.values_list('id', flat=True))
        def matches(line):
            if target_product_ids and line['product_id'] not in target_product_ids:
                return False
            if promo.category_id and line['category_id'] != promo.category_id:
                return False
            return True
        matched = [l for l in cart_lines if matches(l)]
        if not matched:
            continue
        total_matched_qty = sum(l['qty'] for l in matched)
        discount = 0.0
        detail = ''
        if promo.promo_type == Promotion.Type.PERCENT_OFF:
            if total_matched_qty < promo.qty_required:
                continue
            matched_subtotal = sum(l['qty'] * l['price'] for l in matched)
            discount = matched_subtotal * float(promo.percent) / 100
            detail = f"{promo.percent}% × {total_matched_qty} dona"
        elif promo.promo_type == Promotion.Type.BUY_X_GET_Y:
            x = promo.qty_required
            y = promo.qty_free
            if x == 0 or total_matched_qty < x + y:
                continue
            # Number of full groups
            groups = total_matched_qty // (x + y)
            free_qty = groups * y
            # Free the cheapest items
            unit_prices = []
            for l in matched:
                unit_prices.extend([l['price']] * l['qty'])
            unit_prices.sort()
            discount = sum(unit_prices[:free_qty])
            detail = f"{x}+{y} bepul × {groups} marta = {free_qty} dona bepul"
        elif promo.promo_type == Promotion.Type.NTH_PERCENT:
            n = promo.qty_required
            if n == 0 or total_matched_qty < n:
                continue
            count = total_matched_qty // n
            # Apply to cheapest items
            unit_prices = []
            for l in matched:
                unit_prices.extend([l['price']] * l['qty'])
            unit_prices.sort()
            target_prices = unit_prices[:count]
            discount = sum(p * float(promo.percent) / 100 for p in target_prices)
            detail = f"{count} ta mahsulotga −{promo.percent}%"
        if discount > 0:
            applied.append({
                'id': promo.id,
                'name': promo.name,
                'type': promo.promo_type,
                'discount': round(discount, 2),
                'detail': detail,
            })
            total_discount += discount

    return applied, total_discount


@login_required
def pos_payment_intent(request):
    """POST /pos/payment/intent/ {provider, amount, txn_ref}
    Provider'dan QR/deeplink oladi. Sotuv hali yakunlanmagan — bu faqat
    intent yaratish. Mijoz to'laganidan keyin POS pos_checkout'ni chaqiradi."""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST only'}, status=405)
    try:
        data = _json.loads(request.body.decode('utf-8'))
    except ValueError:
        return JsonResponse({'ok': False, 'error': 'bad JSON'}, status=400)
    from .payments import get_provider
    provider_name = (data.get('provider') or '').strip()
    provider = get_provider(provider_name)
    if not provider:
        return JsonResponse({'ok': False, 'error': "noma'lum provider"}, status=400)
    try:
        amount = float(data.get('amount') or 0)
    except (TypeError, ValueError):
        amount = 0
    if amount <= 0:
        return JsonResponse({'ok': False, 'error': "summa noto'g'ri"}, status=400)
    txn_ref = (data.get('txn_ref') or f'tmp-{request.user.id}-{int(timezone.now().timestamp())}')[:60]
    try:
        intent = provider.create_intent(amount, txn_ref)
    except Exception as e:
        logger.exception('Payment intent failed: %s', e)
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)
    return JsonResponse({
        'ok': True,
        'provider': provider.name,
        'provider_display': provider.display_name,
        'is_installment': provider.is_installment,
        **intent,
    })


@login_required
def pos_payment_status(request):
    """GET /pos/payment/status/?provider=click&intent_id=..."""
    provider_name = request.GET.get('provider') or ''
    intent_id = request.GET.get('intent_id') or ''
    from .payments import get_provider, available_providers
    # SEC-19: mijoz tanlagan provider'ga ISHONMAYMIZ. Faqat ANIQ yoqilgan
    # provider'lar. Aks holда ?provider=noop&intent_id=har-narsa → 'paid'
    # qaytarib, istalgan kirgan foydalanuvchи to'lovni "tasdiqlab" olardi.
    _enabled = {p.name for p in available_providers()}
    if provider_name not in _enabled:
        return JsonResponse({'ok': False, 'error': 'provider yoqilmagan'}, status=400)
    provider = get_provider(provider_name)
    if not provider or not intent_id:
        return JsonResponse({'ok': False, 'error': 'parametr yetishmaydi'}, status=400)
    try:
        status = provider.check_status(intent_id)
    except Exception as e:
        logger.exception('Payment status check failed: %s', e)
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)
    return JsonResponse({'ok': True, **status})


@login_required
def pos_unlock(request):
    """POST /pos/unlock/ {password} — joriy foydalanuvchining paroli to'g'ri
    bo'lsa, idle lock ochiladi. AJAX'dan chaqiriladi."""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST only'}, status=405)
    try:
        data = _json.loads(request.body.decode('utf-8'))
    except ValueError:
        return JsonResponse({'ok': False, 'error': 'bad JSON'}, status=400)
    password = (data.get('password') or '').strip()
    if not password:
        return JsonResponse({'ok': False, 'error': 'parol kerak'}, status=400)
    if request.user.check_password(password):
        return JsonResponse({'ok': True})
    return JsonResponse({'ok': False, 'error': "noto'g'ri parol"}, status=401)


@login_required
def pos_txn_refundable(request, pk):
    """GET /pos/txn/<pk>/refundable/ — returns lines that can still be refunded."""
    branch = _user_branch_or_403(request)
    if branch is None:
        return JsonResponse({'ok': False, 'error': 'no branch'}, status=403)
    txn = (SaleTransaction.objects.filter(pk=pk, branch=branch)
           .prefetch_related('lines__variant__product', 'lines__returns')
           .first())
    if not txn:
        return JsonResponse({'ok': False, 'error': 'topilmadi'}, status=404)
    lines = []
    for s in txn.lines.all():
        already = sum(r.quantity for r in s.returns.all())
        remaining = s.quantity - already
        if remaining <= 0:
            continue
        lines.append({
            'sale_id': s.id,
            'product_code': s.variant.product.code,
            'product_name': s.variant.product.name,
            'size': s.variant.size,
            'color': s.variant.color,
            'sold_qty': s.quantity,
            'already_returned': already,
            'remaining': remaining,
            'sale_price': float(s.sale_price),
        })
    return JsonResponse({'ok': True, 'txn_id': txn.pk, 'lines': lines})


@login_required
def pos_refund(request):
    """POST /pos/refund/ JSON body: {lines: [{sale_id, qty, reason}]}.
    Processes refunds atomically; restores stock; returns total refunded.

    Requires an OPEN shift: refund returns cash from the drawer, so it must
    be attributed to the current shift for cash reconciliation."""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST only'}, status=405)
    branch = _user_branch_or_403(request)
    if branch is None:
        return JsonResponse({'ok': False, 'error': 'no branch'}, status=403)

    # C1 fix: require open shift
    open_shift = _open_shift_for(branch)
    if not open_shift:
        return JsonResponse({
            'ok': False,
            'error': "Smen ochilmagan. Qaytarish faqat ochiq smen davomida amalga oshiriladi.",
        }, status=400)

    try:
        data = _json.loads(request.body.decode('utf-8'))
    except ValueError:
        return JsonResponse({'ok': False, 'error': 'bad JSON'}, status=400)

    items = data.get('lines') or []
    if not items:
        return JsonResponse({'ok': False, 'error': 'qator yo\'q'}, status=400)

    # S3: takroriy qaytarishни (timeout/qayta bosish) bloklaymiz. Bir martalik
    # kalit butun qaytarish partiyasига tegishli — birinchi Return qatoriga
    # yoziladi. O'sha kalit bilan qaytarish bo'lган bo'lса — takror emas.
    _idem = (data.get('idempotency_key') or '').strip()[:64] or None
    if _idem and Return.objects.filter(idempotency_key=_idem).exists():
        return JsonResponse({'ok': True, 'refunded_qty': 0,
                             'refunded_total': 0, 'duplicate': True})

    # REF-1: HAMMA qatorni atomic blokдан OLDIN tekshiramiz. Ilgari tekshiruv
    # atomic ichida `return` qilardi — bitta qator o'tib, keyingisi yiqilса,
    # o'tgani COMMIT bo'lib, 400 qaytardi. Kassir tuzatib qayta yuborsa —
    # birinchi qator IKKINCHI marta qaytarilardi (ikki barobar naqd/ombor).
    parsed = []
    for it in items:
        try:
            sid = int(it['sale_id'])
            qty = int(it['qty'])
        except (KeyError, ValueError, TypeError):
            return JsonResponse({'ok': False, 'error': 'noto\'g\'ri qator'}, status=400)
        if qty <= 0:
            continue
        reason = (it.get('reason') or '').strip()[:200]
        sale = (Sale.objects.select_related('variant', 'branch')
                .filter(pk=sid, branch=branch).first())
        if not sale:
            return JsonResponse({'ok': False, 'error': f'sale {sid} topilmadi'}, status=400)
        parsed.append({'sid': sale.pk, 'qty': qty, 'reason': reason,
                       'code': sale.variant.product.code})
    if not parsed:
        return JsonResponse({'ok': False, 'error': 'qator yo\'q'}, status=400)

    class _RefundAbort(Exception):
        def __init__(self, payload, status=400):
            self.payload = payload
            self.status = status

    refunded_total = Decimal('0')
    refunded_qty = 0
    _first_row = True
    _txn_refunded = {}   # REF-3: chek pk -> shu chekда jami qaytarilgan naqd
    try:
        with transaction.atomic():
            for p in parsed:
                # REF-1: qatorni QULFLAB qayta tekshiramiz — bir vaqtда ikki
                # refund bir xil qatorga o'tmasin (pos_exchange kabi).
                # DIQQAT: 'transaction' NULLABLE FK — uni select_related qilsak
                # LEFT OUTER JOIN bo'ladi va Postgres select_for_update (FOR UPDATE)
                # ни nullable outer join'ga qo'llay olmaydi (SQLite bunga e'tibor
                # bermaydi — shu bois test'да chiqmagan, prod'да 500 bergan).
                # variant/branch — NOT NULL (inner join), FOR UPDATE ular bilan
                # ishlaydi. transaction'ni keyin kerak bo'lganда lazy o'qiymiz.
                sale = (Sale.objects.select_for_update()
                        .select_related('variant', 'branch')
                        .filter(pk=p['sid'], branch=branch).first())
                if not sale:
                    raise _RefundAbort({'ok': False, 'error': f"sale {p['sid']} topilmadi"})
                already = sale.returns.aggregate(s=Sum('quantity'))['s'] or 0
                remaining = sale.quantity - already
                if p['qty'] > remaining:
                    raise _RefundAbort({'ok': False,
                        'error': f"{p['code']}: faqat {remaining} dona qaytarish mumkin"})
                stock = (BranchStock.objects.select_for_update()
                         .filter(variant=sale.variant, branch=sale.branch).first())
                # R2: ochiq narxli tovar sotuvда zaxira KAMAYTIRILMAGAN — shuning
                # uchun qaytarishда ham tiklamaymiz (aks holda yo'qdan zaxira).
                if stock and not sale.variant.product.is_open_price:
                    stock.stock_count = F('stock_count') + p['qty']
                    stock.save(update_fields=['stock_count'])
                # REF-3: qaytariladigan HAQIQIY naqд. Dona O'Z narxida (order_discount
                # taqsimlanmaydi). AMMO bir chek bo'yicha jami qaytarish chek
                # TO'LOVIdan oshmaydi — shu bois butun chek qaytганда chegirма
                # oxirgi qaytarishга singib, jami = to'langan summа bo'ladi
                # (egasining qoidasi: qisman → dona narxi, to'liq → chek jamisi).
                _per_unit = (sale.net_line_total() / sale.quantity
                             if sale.quantity > 0 else Decimal(sale.sale_price))
                _item_value = Decimal(p['qty']) * _per_unit
                _txn = sale.transaction
                if _txn is not None and Decimal(_txn.order_discount or 0) > 0:
                    _paid = Decimal(_txn.total)
                    if _txn.pk not in _txn_refunded:
                        _txn_refunded[_txn.pk] = sum(
                            (r.effective_cash_refund for r in
                             Return.objects.filter(sale__transaction_id=_txn.pk)),
                            Decimal('0'))
                    _prior = _txn_refunded[_txn.pk]
                    _cap = _paid - _prior if _paid > _prior else Decimal('0')
                    _cash = min(_item_value, _cap)
                    _txn_refunded[_txn.pk] = _prior + _cash
                else:
                    _cash = _item_value
                _rk = dict(sale=sale, shift=open_shift, quantity=p['qty'],
                           reason=p['reason'], refunded_by=request.user,
                           refund_cash=_cash)
                if _idem and _first_row:
                    _rk['idempotency_key'] = _idem   # S3: partiyaning bir martalik kaliti
                    _first_row = False
                try:
                    ret = Return.objects.create(**_rk)
                except IntegrityError:
                    # Poyga: boshqa so'rov ayni kalitни yozib ulgurdi — takror.
                    raise _RefundAbort({'ok': True, 'refunded_qty': 0,
                        'refunded_total': 0, 'duplicate': True}, status=200)
                refunded_qty += p['qty']
                refunded_total += _cash
    except _RefundAbort as e:
        return JsonResponse(e.payload, status=e.status)
    return JsonResponse({
        'ok': True,
        'refunded_qty': refunded_qty,
        'refunded_total': float(refunded_total),
    })


@login_required
def pos_exchange(request):
    """POST /pos/exchange/ — ALMASHTIRISH: eski tovar(lar) qaytadi, yangi
    tovar(lar) beriladi, faqat NARX FARQI hisoblanadi.

    JSON body:
      { returns:  [{sale_id, qty}],                 # qaytadigan eski tovar(lar)
        new_lines:[{stock_id, qty, sale_price}],    # beriladigan yangi tovar(lar)
        payment_method: 'cash'|'card'|'transfer',   # yangi qimmat bo'lsa — farq uchun
        reason: '...' }

    Pul modeli (kassa ANIQ to'g'ri qolishi uchun):
      old_total = qaytgan tovar qiymati,  new_total = yangi tovar qiymati
      credit = min(old_total, new_total)  -> yangi chekka order_discount (trade-in krediti)
      extra  = new_total - old_total
        extra > 0  -> mijoz farqni to'laydi (tanlangan usul), yangi chek net = extra
        extra < 0  -> do'kon farqni NAQD qaytaradi (Return.cash_refunded = |extra|)
        extra == 0 -> pul harakati yo'q
    Yangi chek net'i (= max(0,extra)) tanlangan usul bo'yicha kassaga tushadi;
    almashtirish qaytarishlari kassaga faqat cash_refunded miqdorida ta'sir qiladi.
    """
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST only'}, status=405)
    branch = _user_branch_or_403(request)
    if branch is None:
        return JsonResponse({'ok': False, 'error': 'no branch'}, status=403)

    open_shift = _open_shift_for(branch)
    if not open_shift:
        return JsonResponse({'ok': False,
            'error': "Smen ochilmagan. Almashtirish faqat ochiq smen davomida amalga oshiriladi."},
            status=400)

    try:
        data = _json.loads(request.body.decode('utf-8'))
    except ValueError:
        return JsonResponse({'ok': False, 'error': 'bad JSON'}, status=400)

    from decimal import Decimal, InvalidOperation

    def _m(v):
        try:
            d = Decimal(str(v if v not in (None, '') else 0))
        except (InvalidOperation, ValueError, TypeError):
            d = Decimal('0')
        return d if d >= 0 else Decimal('0')

    ret_items = data.get('returns') or []
    new_items = data.get('new_lines') or []
    if not ret_items:
        return JsonResponse({'ok': False, 'error': 'qaytariladigan tovar tanlanmagan'}, status=400)
    if not new_items:
        return JsonResponse({'ok': False, 'error': 'yangi tovar tanlanmagan'}, status=400)

    method = (data.get('payment_method') or 'cash').strip()
    if method not in ('cash', 'card', 'transfer'):
        method = 'cash'
    reason = (data.get('reason') or '').strip()[:200]

    parsed_new = []
    for ln in new_items:
        try:
            sid = int(ln['stock_id']); qty = int(ln['qty']); price = _m(ln['sale_price'])
        except (KeyError, ValueError, TypeError):
            return JsonResponse({'ok': False, 'error': "yangi qator noto'g'ri"}, status=400)
        if qty <= 0:
            return JsonResponse({'ok': False, 'error': "yangi qator: soni noto'g'ri"}, status=400)
        if price > Decimal('9999999999.99'):
            return JsonResponse({'ok': False,
                'error': "Narx juda katta. Iltimos, to'g'ri narx kiriting."}, status=400)
        parsed_new.append({'sid': sid, 'qty': qty, 'price': price})

    parsed_ret = []
    for it in ret_items:
        try:
            sid = int(it['sale_id']); qty = int(it['qty'])
        except (KeyError, ValueError, TypeError):
            return JsonResponse({'ok': False, 'error': "qaytish qatori noto'g'ri"}, status=400)
        if qty > 0:
            parsed_ret.append({'sale_id': sid, 'qty': qty})
    if not parsed_ret:
        return JsonResponse({'ok': False, 'error': "qaytish soni noto'g'ri"}, status=400)

    class _Abort(Exception):
        def __init__(self, payload, status=400):
            self.payload = payload; self.status = status

    try:
        with transaction.atomic():
            # ---- eski tovar(lar)ni tekshirib qiymatlash ----
            old_total = Decimal('0')
            ret_ctx = []
            for r in parsed_ret:
                sale = (Sale.objects.select_for_update()
                        .select_related('variant__product', 'branch')
                        .filter(pk=r['sale_id'], branch=branch).first())
                if not sale:
                    raise _Abort({'ok': False, 'error': f"sotuv {r['sale_id']} topilmadi"})
                already = sale.returns.aggregate(s=Sum('quantity'))['s'] or 0
                remaining = sale.quantity - already
                if r['qty'] > remaining:
                    raise _Abort({'ok': False, 'error':
                        f"{sale.variant.product.code}: faqat {remaining} dona qaytarish mumkin"})
                # Trade-in qiymati = tovarning O'Z qatori narxi (order_discount
                # taqsimlanmaydi — qaytarish bilan bir xil siyosat, egasi qarori).
                per_unit = (sale.net_line_total() / sale.quantity) if sale.quantity > 0 else sale.sale_price
                amt = Decimal(r['qty']) * Decimal(per_unit)
                old_total += amt
                ret_ctx.append((sale, r['qty']))

            # ---- yangi tovar(lar)ni tekshirib qiymatlash (qoldiqni bloklaymiz) ----
            sids = sorted({l['sid'] for l in parsed_new})
            locked = {s.pk: s for s in BranchStock.objects.select_for_update()
                      .select_related('variant__product', 'branch')
                      .filter(pk__in=sids, branch=branch)}
            new_total = Decimal('0')
            for ln in parsed_new:
                stock = locked.get(ln['sid'])
                if not stock:
                    raise _Abort({'ok': False, 'error': f"yangi tovar {ln['sid']} topilmadi"})
                if (not stock.variant.product.is_open_price
                        and ln['qty'] > stock.stock_count):
                    raise _Abort({'ok': False, 'error':
                        f"{stock.variant.product.code} {stock.variant.size}/{stock.variant.color}: "
                        f"omborda faqat {stock.stock_count} ta bor"})
                new_total += Decimal(ln['qty']) * Decimal(ln['price'])

            credit = min(old_total, new_total)
            extra = new_total - old_total
            cash_back = (old_total - new_total) if old_total > new_total else Decimal('0')
            new_pm = method if extra > 0 else 'cash'

            txn = SaleTransaction.objects.create(
                branch=branch, sold_by=request.user,
                payment_method=new_pm, payment_breakdown=[],
                note=(reason or 'Almashtirish')[:200],
                order_discount=credit,
                # DISC-1: bu CHEGIRMA EMAS — mijoz eski tovar bilan to'lagan.
                # Alohida maydonда saqlanadi, shunda hisobotlar uni kassir
                # bergan chegirma sifatida ko'rsatmaydi.
                exchange_credit=credit,
                discount_reason='Almashtirish: eski tovar hisobiga',
                shift=open_shift,
            )
            for ln in parsed_new:
                stock = locked[ln['sid']]
                if not stock.variant.product.is_open_price:
                    stock.stock_count = F('stock_count') - ln['qty']
                    stock.save()
                Sale.objects.create(
                    transaction=txn, variant=stock.variant, branch=stock.branch,
                    quantity=ln['qty'], sale_price=ln['price'],
                    cost_at_sale=stock.cost_price, line_discount=Decimal('0'),
                    sold_by=request.user,
                )

            # eski tovar(lar): omborga qaytarish + almashtirish Return'i.
            # cash_back butun almashtirishga bitta — birinchi qatorga yozamiz.
            cb_left = cash_back
            for sale, qty in ret_ctx:
                stock = BranchStock.objects.filter(
                    variant=sale.variant, branch=sale.branch).first()
                # R2: ochiq narxli tovar zaxira yuritmaydi — tiklamaymiz.
                if stock and not sale.variant.product.is_open_price:
                    stock.stock_count = F('stock_count') + qty
                    stock.save()
                this_cb = cb_left if cb_left > 0 else Decimal('0')
                cb_left = Decimal('0')
                Return.objects.create(
                    sale=sale, shift=open_shift, quantity=qty,
                    reason=(reason or f'Almashtirish → chek #{txn.pk}')[:200],
                    refunded_by=request.user,
                    is_exchange=True, cash_refunded=this_cb,
                )
    except _Abort as e:
        return JsonResponse(e.payload, status=e.status)

    return JsonResponse({
        'ok': True,
        'txn_id': txn.pk,
        'receipt_url': f'/transaction/{txn.public_id}/?autoprint=1',
        'old_total': float(old_total),
        'new_total': float(new_total),
        'extra': float(extra),
        'cash_back': float(cash_back),
        'method': new_pm,
    })


@login_required
def pos_parked_delete(request, pk):
    """POST /pos/parked/<pk>/delete/ — saqlangan savatni o'chirish."""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST only'}, status=405)
    branch = _user_branch_or_403(request)
    if branch is None:
        return JsonResponse({'ok': False, 'error': 'no branch'}, status=403)
    deleted, _ = ParkedSale.objects.filter(pk=pk, branch=branch).delete()
    return JsonResponse({'ok': bool(deleted)})


# ---------- SALE ----------

@login_required
def sale_create(request, stock_id):
    stock = get_object_or_404(BranchStock.objects.select_related('variant__product', 'branch'),
                              pk=stock_id)
    # Sotuvchilar faqat o'z filialida sotishi mumkin
    if not request.user.is_admin():
        if request.user.branch_id != stock.branch_id:
            return HttpResponseForbidden(
                "<h3>Bu filialda sotishga ruxsat yo'q.</h3>"
            )

    if request.method == 'POST':
        form = SaleForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            qty = cd['quantity']
            # MON-21: smensiz sotuv YO'Q — aks holda shift=None bilan yozilib,
            # Z-hisobotга tushmaydi.
            open_shift = _open_shift_for(stock.branch)
            if not open_shift:
                messages.error(request, "Smen ochilmagan. Avval smen oching.")
                return redirect('sale_create', stock_id=stock.id)
            if qty > stock.stock_count:
                messages.error(request,
                    f"Omborda yetarli emas. Mavjud: {stock.stock_count}, so'rov: {qty}.")
                return redirect('sale_create', stock_id=stock.id)
            with transaction.atomic():
                stock.stock_count = F('stock_count') - qty
                stock.save()
                # Wrap every sale in a SaleTransaction so the OFD/fiscal
                # submission flow is the same for single-item and cart sales.
                txn = SaleTransaction.objects.create(
                    branch=stock.branch,
                    sold_by=request.user,
                    payment_method=SaleTransaction.PaymentMethod.CASH,
                    shift=open_shift,   # MON-21
                    note=cd.get('note') or '',
                )
                Sale.objects.create(
                    transaction=txn,
                    variant=stock.variant, branch=stock.branch,
                    quantity=qty, sale_price=cd['sale_price'],
                    cost_at_sale=stock.cost_price,
                    note=cd.get('note') or '', sold_by=request.user,
                )

            # Best-effort fiscal submission (no-op when no OFD configured).
            from .fiscal import submit_for_transaction
            submit_for_transaction(txn)

            messages.success(request,
                f"Sotildi: {qty} dona × {cd['sale_price']} so'm ({stock.branch.name}).")
            return redirect('lookup')
    else:
        prefill_price = stock.sale_price if stock.sale_price > 0 else \
            stock.variant.product.default_sale_price
        form = SaleForm(initial={'sale_price': prefill_price})
    return render(request, 'inventory/sale_form.html', {
        'form': form, 'stock': stock,
    })


# ---------- BRANCHES ----------

@admin_required
def branch_list(request):
    """Filiallar ro'yxati: 30 kunlik tushum + P&L + xodimlar."""
    since_30d = timezone.now() - timedelta(days=30)
    rev_expr = ExpressionWrapper(
        F('quantity') * F('sale_price') - F('line_discount'),
        output_field=DecimalField(max_digits=14, decimal_places=2)
    )
    cost_expr = ExpressionWrapper(
        F('quantity') * F('cost_at_sale'),
        output_field=DecimalField(max_digits=14, decimal_places=2)
    )

    branches = list(Branch.objects.annotate(
        stock_total=Coalesce(Sum('stocks__stock_count'), 0),
        staff_count=Count('staff', distinct=True),
        stock_value=Coalesce(Sum(
            ExpressionWrapper(F('stocks__stock_count') * F('stocks__cost_price'),
                              output_field=DecimalField(max_digits=14, decimal_places=2))
        ), 0, output_field=DecimalField(max_digits=14, decimal_places=2)),
    ).order_by('-is_active', 'name'))

    # 30-day stats per branch
    _sales_30d = Sale.objects.filter(sold_at__gte=since_30d)
    # SAL-6: chek chegirmasi filialga ANIQ tushadi (bitta chek — bitta filial),
    # shuning uchun to'g'ridan-to'g'ri ayiriladi. Aks holda filial tushumi ham,
    # P&L (gross/net) ham chegirma miqdorida oshib ko'rinardi.
    _, _odisc_by_branch = _order_discount_share(_sales_30d, 'branch_id')
    # RET-1: qaytarish filialga aniq tegishli (Return -> Sale -> branch).
    _, _, _ret_by_branch = _returns_adjustment(_sales_30d, 'branch_id')
    stats_map = {}
    for row in (_sales_30d
                .values('branch_id')
                .annotate(rev=Sum(rev_expr), cost=Sum(cost_expr),
                          txns=Count('transaction', distinct=True),
                          qty=Sum('quantity'))):
        stats_map[row['branch_id']] = row

    for br in branches:
        s = stats_map.get(br.id, {})
        _rr, _cc = _ret_by_branch.get(br.id) or (Decimal('0'), Decimal('0'))
        rev = _net_rev(s.get('rev'), _odisc_by_branch.get(br.id)) - float(_rr)
        cost = float(_dec(s.get('cost') or 0) - _cc)
        period_fraction = 30 / 30.0  # whole 30-day window
        fixed = float((br.monthly_rent or 0) + (br.monthly_other_costs or 0)) * period_fraction
        gross = rev - cost
        net = gross - fixed
        br.m_revenue = rev
        br.m_cost = cost
        br.m_gross = gross
        br.m_fixed = fixed
        br.m_net = net
        br.m_margin = (net / rev * 100) if rev else 0
        br.m_txns = s.get('txns') or 0
        br.m_qty = s.get('qty') or 0

    total_branches = sum(1 for b in branches if b.is_active)
    total_revenue = sum(float(b.m_revenue) for b in branches)
    total_stock_value = sum(float(b.stock_value or 0) for b in branches)

    return render(request, 'inventory/branch_list.html', {
        'branches': branches,
        'total_branches': total_branches,
        'total_revenue': total_revenue,
        'total_stock_value': total_stock_value,
    })


@admin_required
def branch_create(request):
    if request.method == 'POST':
        form = BranchForm(request.POST)
        if form.is_valid():
            br = form.save()
            messages.success(request, f"Filial yaratildi: {br.name}")
            return redirect('branch_list')
    else:
        form = BranchForm()
    return render(request, 'inventory/branch_form.html', {
        'form': form, 'title': "Yangi filial",
    })


@admin_required
def branch_edit(request, pk):
    branch = get_object_or_404(Branch, pk=pk)
    if request.method == 'POST':
        form = BranchForm(request.POST, instance=branch)
        if form.is_valid():
            form.save()
            messages.success(request, 'Filial yangilandi.')
            return redirect('branch_list')
    else:
        form = BranchForm(instance=branch)
    return render(request, 'inventory/branch_form.html', {
        'form': form, 'title': f'Tahrirlash — {branch.name}', 'branch': branch,
    })


# ---------- USERS ----------

@admin_required
def user_list(request):
    """Foydalanuvchilar: 30 kunlik performans + oxirgi kirish."""
    role_filter = request.GET.get('role') or ''
    branch_id = request.GET.get('branch') or ''
    only_active = request.GET.get('active') == '1'

    users = User.objects.select_related('branch')
    if role_filter:
        users = users.filter(role=role_filter)
    if branch_id:
        try:
            users = users.filter(branch_id=int(branch_id))
        except ValueError:
            branch_id = ''
    if only_active:
        users = users.filter(is_active=True)

    users = list(users.order_by('-is_active', 'username'))

    # 30-kunlik sotuv stat'lari
    since = timezone.now() - timedelta(days=30)
    rev_expr = ExpressionWrapper(
        F('quantity') * F('sale_price') - F('line_discount'),
        output_field=DecimalField(max_digits=14, decimal_places=2)
    )
    _sales_30d = Sale.objects.filter(sold_at__gte=since)
    # SAL-6: chek chegirmasi sotuvchiga ANIQ tushadi (bitta chek — bitta
    # sotuvchi). Bu muhim: s_commission aynan shu summadan hisoblanadi, ya'ni
    # ilgari kassir chegirma qilib bergan puldan ham foiz olardi.
    _, _odisc_by_seller = _order_discount_share(_sales_30d, 'sold_by_id')
    # RET-1: qaytarilgan tovardan komissiya to'lanmasin.
    _, _, _ret_by_seller = _returns_adjustment(_sales_30d, 'sold_by_id')
    stats = (_sales_30d
             .values('sold_by_id')
             .annotate(
                 revenue=Sum(rev_expr),
                 qty=Sum('quantity'),
                 n=Count('transaction', distinct=True),
             ))
    stat_map = {s['sold_by_id']: s for s in stats}
    for u in users:
        s = stat_map.get(u.id, {})
        _rr, _ = _ret_by_seller.get(u.id) or (Decimal('0'), Decimal('0'))
        u.s_revenue = _net_rev(s.get('revenue'), _odisc_by_seller.get(u.id)) - float(_rr)
        u.s_qty = s.get('qty') or 0
        u.s_txns = s.get('n') or 0
        pct = float(u.commission_percent or 0)
        u.s_commission = u.s_revenue * pct / 100 if pct else 0

    return render(request, 'inventory/user_list.html', {
        'users': users,
        'role_filter': role_filter,
        'branch_id': branch_id,
        'only_active': only_active,
        'branches': Branch.objects.filter(is_active=True).order_by('name'),
        'role_choices': User.Role.choices,
    })


@admin_required
def user_create(request):
    if request.method == 'POST':
        form = UserCreateForm(request.POST)
        if form.is_valid():
            user = form.save()
            # H7 fix: audit account creation (password set is implicit)
            AuditLog.objects.create(
                user=request.user,
                username_snapshot=request.user.username,
                action=AuditLog.Action.CREATE,
                model_name='User',
                object_id=str(user.pk),
                object_repr=f'{user.username} ({user.get_role_display()})',
            )
            messages.success(request, f"Foydalanuvchi yaratildi: {user.username}")
            return redirect('user_list')
    else:
        form = UserCreateForm()
    return render(request, 'inventory/user_form.html', {
        'form': form, 'title': "Yangi foydalanuvchi", 'is_create': True,
    })


@admin_required
def user_edit(request, pk):
    target = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=target)
        if form.is_valid():
            new_password = (form.cleaned_data.get('new_password') or '').strip()
            user = form.save()
            # H7 fix: explicit audit row when password changes — never log the
            # password itself, only the fact that it changed (and by whom)
            if new_password:
                AuditLog.objects.create(
                    user=request.user,
                    username_snapshot=request.user.username,
                    action=AuditLog.Action.UPDATE,
                    model_name='UserPassword',
                    object_id=str(user.pk),
                    object_repr=f'Parol o\'zgartirildi: {user.username}',
                )
            messages.success(request, 'Foydalanuvchi yangilandi.')
            return redirect('user_list')
    else:
        form = UserEditForm(instance=target)
    return render(request, 'inventory/user_form.html', {
        'form': form, 'title': f'Tahrirlash — {target.username}',
        'target_user': target, 'is_create': False,
    })


# ---------- CATEGORIES ----------

@admin_required
def category_list(request):
    if request.method == 'POST':
        action = request.POST.get('action') or 'create'
        if action == 'delete':
            try:
                pk = int(request.POST.get('pk') or 0)
            except (TypeError, ValueError):
                pk = 0
            cat = Category.objects.filter(pk=pk).first()
            if cat:
                n_products = cat.products.count()
                move_to = (request.POST.get('move_to') or '').strip()
                if n_products > 0:
                    # Mahsulotlarni avval tanlangan joyga ko'chiramiz
                    if move_to == 'none':
                        cat.products.update(category=None)
                        messages.info(request,
                            f"{n_products} ta mahsulot kategoriyasiz qoldi.")
                    elif move_to.isdigit() and Category.objects.filter(
                            pk=int(move_to)).exclude(pk=cat.pk).exists():
                        target = Category.objects.get(pk=int(move_to))
                        cat.products.update(category=target)
                        messages.info(request,
                            f"{n_products} ta mahsulot \"{target.name}\" ga ko'chirildi.")
                    else:
                        messages.error(request,
                            f"\"{cat.name}\" o'chirilmadi: unda {n_products} ta "
                            "mahsulot bor — ular ko'chiriladigan joyni tanlang.")
                        return redirect('category_list')
                name = cat.name
                cat.delete()
                AuditLog.objects.create(
                    user=request.user,
                    username_snapshot=request.user.username,
                    action=AuditLog.Action.DELETE,
                    model_name='Category', object_id=str(pk),
                    object_repr=name)
                messages.success(request, f"\"{name}\" o'chirildi.")
            return redirect('category_list')
        if action == 'edit':
            try:
                pk = int(request.POST.get('pk') or 0)
            except (TypeError, ValueError):
                pk = 0
            cat = Category.objects.filter(pk=pk).first()
            if cat:
                new_name = (request.POST.get('name') or '').strip()[:120]
                new_prefix = (request.POST.get('prefix') or '').strip()[:6]
                if not new_name:
                    messages.error(request, "Nom kerak.")
                else:
                    cat.name = new_name
                    if new_prefix:
                        cat.prefix = new_prefix.upper()
                    grp = (request.POST.get('group') or '').strip()
                    cat.group_id = int(grp) if grp.isdigit() else None
                    cat.save()
                    messages.success(request, f"\"{cat.name}\" yangilandi.")
            return redirect('category_list')
        # create
        form = CategoryForm(request.POST)
        if form.is_valid():
            cat = form.save()
            messages.success(request, f"\"{cat.name}\" qo'shildi (prefiks: {cat.prefix}).")
            return redirect('category_list')
    else:
        form = CategoryForm()

    # Filter + search
    q = (request.GET.get('q') or '').strip()
    group_slug = (request.GET.get('group') or '').strip()  # men|women|kids|home|none
    categories = Category.objects.select_related('group')
    if q:
        categories = categories.filter(name__icontains=q)
    if group_slug == 'none':
        categories = categories.filter(group__isnull=True)
    elif group_slug:
        categories = categories.filter(group__slug=group_slug)

    # Bo'lim tugmalari uchun: har bo'limда nechta kategoriya + bo'limsizlar soni
    _gc = dict(Category.objects.filter(group__isnull=False)
               .values_list('group__slug').annotate(n=Count('pk')))
    groups = list(Group.objects.all())
    for g in groups:
        g.n_cats = _gc.get(g.slug, 0)
    unassigned_count = Category.objects.filter(group__isnull=True).count()

    # 30-kunlik sotuvlar va revenue agregati
    since_30d = timezone.now() - timedelta(days=30)
    rev_expr = ExpressionWrapper(
        F('quantity') * F('sale_price') - F('line_discount'),
        output_field=DecimalField(max_digits=14, decimal_places=2)
    )
    cost_expr = ExpressionWrapper(
        F('quantity') * F('cost_at_sale'),
        output_field=DecimalField(max_digits=14, decimal_places=2)
    )

    # SAL-6: KATEGORIYA — qator darajasidagi o'lchov. Bitta chek bir nechta
    # kategoriyani qamraydi va egasining qoidasi bo'yicha chek chegirmasi
    # ayrim tovarga tegishli emas, shuning uchun kategoriyalar O'Z narxida
    # qoladi va chegirma ALOHIDA ko'rsatiladi: kategoriyalar + chegirma = JAMI.
    _sales_30d = Sale.objects.filter(sold_at__gte=since_30d)
    _odisc_total, _ = _order_discount_share(_sales_30d)
    _, _, _disc_split = _order_discount_share(_sales_30d, split=True)
    # RET-1: qaytarish kategoriyaga ANIQ tegishli (Return -> Sale -> mahsulot).
    _, _, _ret_by_cat = _returns_adjustment(
        _sales_30d, 'variant__product__category_id')
    cat_stats = {}
    for row in (_sales_30d
                .values('variant__product__category_id')
                .annotate(rev=Sum(rev_expr), cost=Sum(cost_expr),
                          qty=Sum('quantity'))):
        cat_stats[row['variant__product__category_id']] = {
            'rev': float(row['rev'] or 0),
            'cost': float(row['cost'] or 0),
            'qty': row['qty'] or 0,
        }

    # Per-category stock + value
    stock_stats = {}
    for row in (BranchStock.objects
                .values('variant__product__category_id')
                .annotate(stock=Sum('stock_count'),
                          value=Sum(ExpressionWrapper(
                              F('stock_count') * F('cost_price'),
                              output_field=DecimalField(max_digits=14, decimal_places=2))))):
        stock_stats[row['variant__product__category_id']] = {
            'stock': row['stock'] or 0,
            'value': float(row['value'] or 0),
        }

    categories = list(categories.annotate(product_count=Count('products')).order_by('name'))
    for c in categories:
        s = cat_stats.get(c.id, {})
        st = stock_stats.get(c.id, {})
        _rr, _cc = _ret_by_cat.get(c.id) or (Decimal('0'), Decimal('0'))
        c.s_rev = float(_dec(s.get('rev', 0)) - _rr)
        c.s_cost = float(_dec(s.get('cost', 0)) - _cc)
        c.s_qty = s.get('qty', 0)
        c.s_profit = c.s_rev - c.s_cost
        c.s_margin = (c.s_profit / c.s_rev * 100) if c.s_rev else 0
        c.s_stock = st.get('stock', 0)
        c.s_value = st.get('value', 0)

    # Top-level stats
    total_categories = len(categories)
    no_category_count = Product.objects.filter(category__isnull=True).count()
    top_cat = max(categories, key=lambda c: c.s_rev, default=None)
    cats_rev_total = sum(c.s_rev for c in categories)
    order_discount_total = float(_odisc_total)
    net_rev_total = cats_rev_total - order_discount_total

    return render(request, 'inventory/category_list.html', {
        'categories': categories, 'form': form,
        'cats_rev_total': cats_rev_total,
        'order_discount_total': order_discount_total,
        'discount_split': _disc_split,   # DISC-1
        'net_rev_total': net_rev_total,
        'q': q,
        'groups': groups,
        'group_slug': group_slug,
        'unassigned_count': unassigned_count,
        'total_categories': total_categories,
        'no_category_count': no_category_count,
        'top_cat': top_cat,
    })


# ---------- SALES LIST ----------

# DISC-7: chegirma sababi endi RO'YXATDAN tanlanadi. Erkin matn amalда
# ishlamadi — auditда yig'ilgani: "s", "c", "raz", "3000", "1500". Kassir
# maydonni tezroq yopish uchun bitta belgi teradi va sabab ma'nosini
# yo'qotadi. Ro'yxat guruhlanadigan ma'lumot beradi.
DISCOUNT_REASONS = (
    "Ko'p tovar oldi",
    "Yaxlitlash",
    "Nuqson / kamchilik",
    "Doimiy mijoz",
    "Rahbar ruxsati",
)


def _valid_discount_reason(reason):
    """DISC-7. Ro'yxatdagi sabab yoki hech bo'lmasa mazmunli erkin matn.

    Server ESKI mijozni ham qabul qiladi. Sabab: bu PWA — service worker
    keshida qolgan eski sahifa erkin matn yuboradi, qat'iy ro'yxat esa
    kassani ish o'rtasida to'xtatib qo'yardi. Shu bois faqat ma'nosizini
    rad etamiz: bitta-ikkita belgi ("s", "c") va faqat raqam ("3000" — bu
    sabab emas, kassir summani qayta terган).
    """
    head = (reason or '').split(';')[0].strip()
    if head in DISCOUNT_REASONS:
        return True
    if head.lower().startswith('boshqa'):
        head = head.split(':', 1)[-1].strip()
    if len(head) < 3:
        return False
    if head.replace(' ', '').replace("'", '').isdigit():
        return False
    return True


# DISC-6: "yaxlitlash" — kassir chekni butun songa tushirish uchun olib
# tashlagan mayda qoldiq (2 000 so'mlik chekда 500 so'm). Bu chegirma emas,
# maydalik: hisobotда uni ulgurji chegirma bilan bir songa qo'shish egani
# chalg'itadi. Auditда ko'rilgani: 308 chekdan 264 tasi shu xil, medianasi
# 1 000 so'm, hammasi birga jamining atigi 20%i.
ROUNDING_MAX = Decimal('5000')     # bundan kattasi yaxlitlash emas
ROUNDING_STEP = Decimal('1000')    # mijoz to'lagani shu songa karrali bo'lsa


def _is_rounding(manual, gross, paid, reason=''):
    """DISC-6. Mayda summa VA chegirma jamini BUTUN songa keltirgan bo'lsa.

    Uch shart ham kerak:

      mayda            — 1 705 000 -> 1 535 000 (170 000) yaxlitlash emas,
                         bu ulgurji chegirma;
      jami butun bo'ldi — 19 500 -> 19 000;
      jami butun EMAS EDI — 100 000 lik chekда 3 000 chegirma ham "butun"
                         qoldiradi (97 000), lekin u yaxlitlash emas: summa
                         allaqachon butun edi, kassir ataylab chegirma bergan.
                         Uchinchi shartsiz karta shu xil cheklarni yaxlitlash
                         deb yashirib qo'yardi — chegirmani KAM ko'rsatish esa
                         ko'p ko'rsatishdан battar.

    DISC-7 dan keyin kassir sababni O'ZI "Yaxlitlash" deb belgilaydi va taxmin
    kerak bo'lmaydi. Taxmin faqat eski cheklar uchun qoladi. Belgilangan sabab
    ham mayda summa cheklovidan o'tadi: "Yaxlitlash" deb 50 000 yozilsa, u
    kartada QO'LDA bo'lib ko'rinishi kerak.
    """
    if manual <= 0 or manual >= ROUNDING_MAX:
        return False
    if (reason or '').strip().lower().startswith('yaxlitlash'):
        return True
    if paid <= 0 or gross <= 0:
        return False
    return paid % ROUNDING_STEP == 0 and gross % ROUNDING_STEP != 0


def _order_discount_share(sale_qs, *group_fields, split=False):
    """Filtrlangan sotuv qatorlariga to'g'ri keladigan CHEK chegirmasi ulushi.

    SAL-1. Sale qatorlarida faqat `line_discount` bor. CHEK bo'yicha umumiy
    chegirma (`order_discount`) SaleTransaction'da turadi, shuning uchun
    qatorlar ustidan Sum() olgan HAR BIR sahifa mijoz TO'LAGANIDAN ko'p
    ko'rsatardi:

        3×100 000, chek chegirmasi 60 000 (mijoz 240 000 to'lagan)
          qatorlar Sum'i         300 000
          chek jamisi (t.total)  240 000   ← to'g'risi

    QAYSI o'lchov bo'yicha ayirish mumkin — muhim farq bor:

      CHEK darajasidagi o'lchovlar (filial, sotuvchi, mijoz, kun): bitta chek
      AYNAN bitta filialga/sotuvchiga/kunga tegishli, shuning uchun chegirma
      shu guruhga to'liq va aniq tushadi — to'g'ridan-to'g'ri ayiriladi.

      QATOR darajasidagi o'lchovlar (mahsulot, kategoriya, guruh): bitta chek
      bir nechta mahsulotni qamraydi va egasining qoidasi bo'yicha (REF-3,
      net_line_total'ga qarang) chek chegirmasi ALOHIDA tovarga tegishli
      EMAS. Shuning uchun bunday ro'yxatlarda qatorlar O'Z narxida qoladi va
      chegirma ALOHIDA qator sifatida ko'rsatiladi:
          mahsulotlar 300 000 + "Chek chegirmasi −60 000" = JAMI 240 000

    Filtr chekning bir qismini tanlasa — faqat o'sha qismning ulushi olinadi
    (bitta mahsulot bo'yicha filtr butun chek chegirmasini yeb qo'ymasin).
    Faqat `order_discount > 0` bo'lgan cheklar so'raladi — chegirmasiz
    sahifada qo'shimcha yuk amalda yo'q.

    Qaytaradi: (jami_ulush, {kalit: ulush}). `group_fields` bitta bo'lsa kalit
    — skalyar, bir nechta bo'lsa — tuple, umuman berilmasa — bo'sh dict.

    `split=True` bo'lsa uchinchi qiymat ham qaytadi (DISC-1/DISC-6):
    {'promo', 'rounding', 'manual', 'exchange', 'total'} — chek chegirmasining
    qismlari.
    Ular bitta songa yig'ilgan holда hisobotlar egasi sozlagan AKSIYAni kassir
    o'zi bergan chegirmadan ajrata olmasdi, almashtirish krediti esa umuman
    chegirma bo'lmasa ham "chegirma" bo'lib ko'rinardi.
    """
    rev = F('quantity') * F('sale_price') - F('line_discount')
    # DISC-1: uchala qismni BIR so'rovда olamiz (qo'shimcha so'rov yo'q).
    disc_rows = list(SaleTransaction.objects
                     .filter(id__in=sale_qs.order_by().values('transaction_id'),
                             order_discount__gt=0)
                     .values_list('id', 'order_discount',
                                  'promo_discount', 'exchange_credit',
                                  'discount_reason'))
    if not disc_rows:
        empty = {'promo': Decimal('0'), 'rounding': Decimal('0'),
                 'manual': Decimal('0'), 'exchange': Decimal('0'),
                 'total': Decimal('0')}
        return (Decimal('0'), {}, empty) if split else (Decimal('0'), {})
    disc = {r[0]: _dec(r[1]) for r in disc_rows}
    ids = list(disc)
    # Chekning TO'LIQ summasi (filtrdan qat'i nazar) — maxraj.
    whole = {r['transaction_id']: _dec(r['s'] or 0)
             for r in (Sale.objects.filter(transaction_id__in=ids)
                       .order_by().values('transaction_id').annotate(s=Sum(rev)))}
    parts = {}
    for _id, _tot, _promo, _exch, _reason in disc_rows:
        _tot, _promo, _exch = _dec(_tot), _dec(_promo or 0), _dec(_exch or 0)
        _manual = _tot - _promo - _exch
        if _manual < 0:
            _manual = Decimal('0')
        # DISC-6: mayda qoldiqni AJRATAMIZ. Bitta songa yig'ilganда karta
        # 1 000 so'mlik yaxlitlash bilan 186 500 so'mlik ulgurji chegirmani
        # o'rtachalab, egaga "biz bunchalik chegirma bermaganmiz" degan
        # tuyg'u berardi. Ular boshqa-boshqa hodisa — alohida ko'rsatiladi.
        _round = Decimal('0')
        _gross = whole.get(_id) or Decimal('0')
        if _is_rounding(_manual, _gross, _gross - _tot, _reason):
            _round, _manual = _manual, Decimal('0')
        parts[_id] = (_promo, _round, _manual, _exch)
    fields = ['transaction_id'] + list(group_fields)
    picked = (sale_qs.filter(transaction_id__in=ids)
              .order_by().values(*fields).annotate(s=Sum(rev)))
    total = Decimal('0')
    by_group = {}
    acc = {'promo': Decimal('0'), 'rounding': Decimal('0'),
           'manual': Decimal('0'), 'exchange': Decimal('0'),
           'total': Decimal('0')}
    for r in picked:
        tid = r['transaction_id']
        w = whole.get(tid) or Decimal('0')
        if w <= 0:
            continue
        frac = _dec(r['s'] or 0) / w        # chekning tanlangan ulushi
        share = _dec(disc[tid]) * frac
        total += share
        if split:
            _p, _r, _m, _e = parts[tid]
            acc['promo'] += _p * frac
            acc['rounding'] += _r * frac
            acc['manual'] += _m * frac
            acc['exchange'] += _e * frac
            acc['total'] += share
        if group_fields:
            key = (r[group_fields[0]] if len(group_fields) == 1
                   else tuple(r[g] for g in group_fields))
            by_group[key] = by_group.get(key, Decimal('0')) + share
    return (total, by_group, acc) if split else (total, by_group)


def _returns_adjustment(sale_qs, *group_fields):
    """Qaytarilgan tovarlar uchun TUSHUM va TANNARX tuzatmasi (RET-1).

    Muammo: `/sales/` dan boshqa HECH BIR sahifa qaytarishni hisobga
    olmasdi — na tushumда, na tannarxда. 3 dona sotilib 1 dona qaytsa,
    hamma joyда 3 donaning tushumi ham, tannarxi ham turaverardi.

    Ikki tuzatma kerak, va ular BOSHQA-BOSHQA:

      TUSHUM  — kassaдан HAQIQATDA chiqqan pul (effective_cash_refund).
        Almashtirishда naqd chiqmaydi (0), demak tushum kamaymaydi —
        to'g'ri, chunki mijoz pulni oldingi chekда to'lagan va u
        do'kondа qoladi.

      TANNARX — tovar OMBORGA QAYTGAN bo'lsagina qaytariladi. Qaytgan
        tovar endi sotilgan emas, u zaxira; tannarxi COGSда qolsa
        ikki marta sanaladi. `pos_refund` ochiq narxli (is_open_price)
        tovarni omborga TIKLAMAYDI — demak ularning tannarxi COGSда
        qolishi KERAK (tovar ham ketdi, pul ham qaytdi: haqiqiy zarar).

    Shu ikki qoida almashtirishni ham, oddiy qaytarishni ham bir xil
    to'g'ri hisoblaydi:
        almashtirish : tushum −0,      tannarx −45 000  -> foyda to'g'ri
        oddiy qaytish: tushum −80 500, tannarx −45 000  -> foyda 0
        ochiq narxli : tushum −80 500, tannarx −0       -> foyda −45 000

    O'lchovlar bo'yicha guruhlash ANIQ: har Return bitta Sale qatoriga
    tegishli, u qatorда esa bitta filial/sotuvchi/mahsulot/kategoriya bor.
    Shu bois chek chegirmasidan farqli o'laroq bu yerда taqsimlash yo'q.

    Qaytaradi: (tushum_krediti, tannarx_krediti, {kalit: (tushum, tannarx)})
    """
    rets = (Return.objects.filter(sale__in=sale_qs.order_by().values('pk'))
            .select_related('sale__variant__product__category__group',
                            'sale__transaction')
            .prefetch_related('sale__transaction__lines'))
    rev_credit = Decimal('0')
    cost_credit = Decimal('0')
    by_group = {}
    for r in rets:
        sale = r.sale
        _rev = _dec(r.effective_cash_refund or 0)
        # Omborga qaytmagan tovarning tannarxi COGSда qoladi.
        _cost = (Decimal('0') if sale.variant.product.is_open_price
                 else _dec(r.quantity) * _dec(sale.cost_at_sale))
        rev_credit += _rev
        cost_credit += _cost
        if group_fields:
            key = (_ret_group_value(sale, group_fields[0]) if len(group_fields) == 1
                   else tuple(_ret_group_value(sale, g) for g in group_fields))
            cur = by_group.get(key) or (Decimal('0'), Decimal('0'))
            by_group[key] = (cur[0] + _rev, cur[1] + _cost)
    return rev_credit, cost_credit, by_group


def _ret_group_value(sale, field):
    """`_returns_adjustment` uchun guruh kalitini Sale qatoridan oladi."""
    if field == 'branch_id':
        return sale.branch_id
    if field == 'sold_by_id':
        return sale.sold_by_id
    if field == 'sold_at__date':
        return timezone.localtime(sale.sold_at).date()
    if field == 'variant__product_id':
        return sale.variant.product_id
    if field == 'variant__product__category_id':
        return sale.variant.product.category_id
    if field == 'variant__product__category__name':
        cat = sale.variant.product.category
        return cat.name if cat else None
    if field == 'variant__product__category__group__name':
        cat = sale.variant.product.category
        return cat.group.name if (cat and cat.group_id) else None
    raise ValueError(f'_returns_adjustment: qo\'llab-quvvatlanmagan o\'lchov {field}')


def _net_rev(gross, share):
    """Brutto qator summasidan chek chegirmasi ulushini ayiradi (float)."""
    return float(_dec(gross or 0) - _dec(share or 0))


@admin_required
def sales_list(request):
    """Sotuvlar ro'yxati: filterlar + kunlik jami + CSV export."""
    # Filterlar
    q = (request.GET.get('q') or '').strip()
    date_from_raw = request.GET.get('date_from') or ''
    date_to_raw = request.GET.get('date_to') or ''
    branch_id = request.GET.get('branch') or ''
    seller_id = request.GET.get('seller') or ''
    payment_method = request.GET.get('payment_method') or ''
    returned = request.GET.get('returned') or ''   # '', 'yes', 'no'
    # DISC-3: chegirma bo'yicha filtr. Egasi "kim qancha chegirma berdi?"
    # savolini bermoqда — javob berish uchun chegirmali cheklarni ajratib
    # ko'ra olish kerak. Qiymatlar: '' | any | none | manual | promo | exchange
    discount = (request.GET.get('discount') or '').strip()
    if discount not in ('', 'any', 'none', 'manual', 'promo', 'exchange'):
        discount = ''
    export = request.GET.get('export') == 'csv'
    # Ko'rinish: 'checks' — bir qator = butun chek (standart); 'items' — bir
    # qator = bitta mahsulot (mahsulotni qidirish uchun qulay).
    view_mode = (request.GET.get('view') or 'checks').strip()
    if view_mode not in ('checks', 'items'):
        view_mode = 'checks'

    # Standart: oxirgi 1 oy (sana filtri berilmagan bo'lsa). Foydalanuvchi
    # kengaytirmoqchi bo'lsa — "Sanadan" ni o'zgartiradi.
    if not date_from_raw and not date_to_raw:
        date_from_raw = (timezone.localdate() - timedelta(days=30)).strftime('%Y-%m-%d')

    # DIQQAT: _returned annotatsiyasini BAZAVIY qs'ga QO'YMAYMIZ — u returns'га
    # JOIN qilib qatorlarni ko'paytiradi va aggregate (jami) IKKI barobar
    # sanardi. Annotatsiya faqat ko'rsatiladigan 300 qatorга qo'shiladi (pastда).
    qs = Sale.objects.select_related(
        'variant__product', 'branch', 'sold_by', 'transaction')

    # Sana filtri — sold_at'ni date()'ga O'RAMAYMIZ (aks holда sale_soldat_idx
    # indeksi ishlamay jadval to'liq skanerlanardi). Aware datetime oralig'i.
    tz = timezone.get_current_timezone()
    try:
        if date_from_raw:
            df = datetime.strptime(date_from_raw, '%Y-%m-%d').date()
            qs = qs.filter(sold_at__gte=timezone.make_aware(
                datetime.combine(df, datetime.min.time()), tz))
    except ValueError:
        date_from_raw = ''
    try:
        if date_to_raw:
            dt = datetime.strptime(date_to_raw, '%Y-%m-%d').date()
            qs = qs.filter(sold_at__lt=timezone.make_aware(
                datetime.combine(dt + timedelta(days=1), datetime.min.time()), tz))
    except ValueError:
        date_to_raw = ''
    if branch_id:
        try:
            qs = qs.filter(branch_id=int(branch_id))
        except ValueError:
            branch_id = ''
    if seller_id:
        try:
            qs = qs.filter(sold_by_id=int(seller_id))
        except ValueError:
            seller_id = ''
    if payment_method:
        qs = qs.filter(transaction__payment_method=payment_method)
    # Qaytarilgan filtri — Exists() bilan (JOIN fanout emas, aggregate to'g'ri).
    _has_ret = Return.objects.filter(sale=OuterRef('pk'))
    if returned == 'yes':
        qs = qs.filter(Exists(_has_ret))    # qisman yoki to'liq qaytarilgan
    elif returned == 'no':
        qs = qs.filter(~Exists(_has_ret))   # umuman qaytarilmagan
    # DISC-3: chegirma filtri. DIQQAT — annotatsiya QO'YMAYMIZ: `transaction`
    # oldinga qarab FK (1:1 join), lekin bazaviy qs'ga annotate qo'shilsa
    # pastдagi values().annotate() (kunlik jami) o'ralmay qolib fanout xavfi
    # tug'ilardi (yuqoridagi izohga qarang). Shu bois F() to'g'ridan-to'g'ri
    # filtr ichида ishlatiladi:
    #   qo'lda = order_discount − promo_discount − exchange_credit
    if discount == 'any':
        qs = qs.filter(Q(transaction__order_discount__gt=0) | Q(line_discount__gt=0))
    elif discount == 'none':
        # Chek'siz eski sotuvlar ham "chegirmasiz" hisoblanadi.
        qs = qs.filter(line_discount=0).filter(
            Q(transaction__isnull=True) | Q(transaction__order_discount=0))
    elif discount == 'manual':
        qs = qs.filter(transaction__order_discount__gt=(
            F('transaction__promo_discount') + F('transaction__exchange_credit')))
    elif discount == 'promo':
        qs = qs.filter(transaction__promo_discount__gt=0)
    elif discount == 'exchange':
        qs = qs.filter(transaction__exchange_credit__gt=0)
    if q:
        cond = (
            Q(variant__product__name__icontains=q)
            | Q(variant__product__code__icontains=q)
            | Q(variant__barcode__icontains=q)
            | Q(transaction__customer_name__icontains=q)
            | Q(transaction__customer_phone__icontains=q)
        )
        # Chek raqami: "174" ham, "#174" ham ishlaydi.
        # Agar shunday raqamli chek BOR bo'lsa — faqat o'shani ko'rsatamiz.
        # (Aks holda raqam shtrix-kod ichida ham uchrab, begona cheklar
        #  qo'shilib kelardi.)
        _num = q.lstrip('#').strip()
        _txn = None
        # len<=12: ulkan raqam int() (>4300 xona → ValueError) yoki bigint
        # overflow (DataError) bilan 500 bermasin.
        if _num.isdigit() and len(_num) <= 12:
            _txn = SaleTransaction.objects.filter(pk=int(_num)).first()
        if _txn is not None:
            qs = qs.filter(transaction_id=_txn.pk)
        else:
            qs = qs.filter(cond)

    qs = qs.order_by('-sold_at')

    if export:
        resp = HttpResponse(content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = 'attachment; filename="sales.csv"'
        resp.write('﻿')  # BOM for Excel
        w = csv.writer(resp)
        w.writerow(['Sana', 'Vaqt', 'Kod', 'Mahsulot', 'Variant', 'Filial',
                    'Sotuvchi', 'Soni', 'Narx', 'Chegirma', 'Jami',
                    'To\'lov turi', 'Mijoz'])
        for s in qs[:10000]:
            t = s.transaction
            w.writerow([_csv_safe(c) for c in [   # SEC-20
                s.sold_at.strftime('%Y-%m-%d'),
                s.sold_at.strftime('%H:%M'),
                s.variant.product.code,
                s.variant.product.name,
                f'{s.variant.size}/{s.variant.color}',
                s.branch.name,
                s.sold_by.username,
                s.quantity,
                float(s.sale_price),
                float(s.line_discount or 0),
                float(s.total),
                t.get_payment_method_display() if t else '',
                t.customer_name if t else '',
            ]])
        return resp

    from decimal import Decimal
    # Jami — BUTUN filtr bo'yicha DB'da (faqat 300 emas). aggregate() Django 6'да
    # annotatsiyali so'rovni SUBQUERY'ga o'raydi (fanout xavfsiz), AMMO pastдаgi
    # KUNLIK values().annotate() o'ralmaydi — shu bois returns'ни bazaviy qs'ga
    # umuman qo'shmaymiz, alohida so'rovда olamiz. 'txns' — HAQIQIY chek soni.
    _rev_expr = F('quantity') * F('sale_price') - F('line_discount')
    agg = qs.aggregate(
        total=Sum(_rev_expr),
        qty=Sum('quantity'),
        txns=Count('transaction_id', distinct=True),
    )
    # SAL-1: chek chegirmasi ulushi ayiriladi — JAMI mijoz TO'LAGANI bo'lsin,
    # ya'ni CHEK ko'rinishidagi qatorlar yig'indisi va Z-hisobot bilan bir xil.
    _odisc_total, _odisc_by_day, _disc_split = _order_discount_share(
        qs, 'sold_at__date', split=True)
    gross_total = _dec(agg['total'] or 0)          # chegirmagacha
    order_discount_total = _odisc_total
    total = gross_total - _odisc_total
    qty_total = agg['qty'] or 0
    txn_count = agg['txns'] or 0
    line_count = qs.count()   # "Topilgan qatorlar" — jami satr (300 cheklovsiz)

    # ---- Qaytarishlar: ALOHIDA so'rov (sales qs'ga JOIN qilmaymiz — fanout yo'q).
    # DIQQAT: refund summasini SQL'da QAYTA YOZMAYMIZ. Ilgari bu yerda formula
    # ORM ifodasi sifatida takrorlangan edi va order_discount taqsimotini
    # HISOBGA OLMASDI — chegirmali chek qaytarilganda bu sahifa Z-hisobotdan
    # KO'PROQ ko'rsatardi (300 000 vs 240 000). Endi ikkala joy ham yagona
    # manbani — Return.effective_cash_refund'ni — ishlatadi (ARCH-6 darsi).
    ret_qs = (Return.objects
              .filter(sale__in=qs)
              .select_related('sale__transaction')
              .prefetch_related('sale__transaction__lines'))

    returned_total = Decimal('0')
    returned_qty = 0
    returned_count = 0
    _ret_by_sale = {}    # sale_id -> {'qty', 'money'}
    ret_by_day = {}      # date    -> {'qty', 'money'}
    for _r in ret_qs:
        _m = _r.effective_cash_refund or Decimal('0')
        returned_total += _m
        returned_qty += _r.quantity
        returned_count += 1
        _e = _ret_by_sale.setdefault(_r.sale_id, {'qty': 0, 'money': Decimal('0')})
        _e['qty'] += _r.quantity
        _e['money'] += _m
        _d = timezone.localtime(_r.sale.sold_at).date()
        _dd = ret_by_day.setdefault(_d, {'qty': 0, 'money': Decimal('0')})
        _dd['qty'] += _r.quantity
        _dd['money'] += _m
    net_total = total - returned_total

    # ---- SAL-2: Z-hisobot bilan SOLISHTIRISH uchun ikkinchi ko'rsatkich.
    # Yuqoridagi `returned_total` — SHU DAVRDA SOTILGAN tovarlardan qaytgani
    # (qaytarish qachon bo'lganidan qat'i nazar). Z-hisobot esa — SHU SMENDA
    # KASSADAN CHIQQAN pul. Kecha sotilib bugun qaytarilgan tovar birinchisiga
    # KECHA, ikkinchisiga BUGUN tushadi. Ikkalasi ham to'g'ri, lekin ikkalasi
    # "Qaytarilgan" deb nomlangani uchun sahifa Z-hisobotga zid ko'rinardi.
    # Endi davr ichida AMALGA OSHIRILGAN qaytarishlar ham chiqadi: bu son
    # o'sha davr smenlarining Z-hisobotlari yig'indisiga TENG.
    # DIQQAT: bu son FAQAT sana+filial oynasi uchun ma'noli. Sotuvchi/to'lov
    # turi/qidiruv filtri yoqilganда yuqoridagi `returned_total` filtrga
    # bo'ysunadi, kassadan chiqqan pulni esa sotuvchi bo'yicha bo'lib
    # bo'lmaydi (qaytarishni BOSHQA kassir rasmiylashtirgan bo'lishi mumkin).
    # Shu bois tor filtrда umuman ko'rsatmaymiz — noto'g'ri taqqoslashдan
    # ko'ra yo'qligi yaxshi.
    _comparable = not (seller_id or payment_method or q or returned or discount)
    period_ret_qs = (Return.objects
                     .select_related('sale__transaction')
                     .prefetch_related('sale__transaction__lines')) if _comparable else Return.objects.none()
    if date_from_raw:
        period_ret_qs = period_ret_qs.filter(refunded_at__gte=timezone.make_aware(
            datetime.combine(datetime.strptime(date_from_raw, '%Y-%m-%d').date(),
                             datetime.min.time()), tz))
    if date_to_raw:
        period_ret_qs = period_ret_qs.filter(refunded_at__lt=timezone.make_aware(
            datetime.combine(datetime.strptime(date_to_raw, '%Y-%m-%d').date()
                             + timedelta(days=1), datetime.min.time()), tz))
    if branch_id:
        period_ret_qs = period_ret_qs.filter(sale__branch_id=int(branch_id))
    period_returned_total = Decimal('0')
    period_returned_qty = 0
    period_returned_count = 0
    for _r in period_ret_qs:
        period_returned_total += _r.effective_cash_refund or Decimal('0')
        period_returned_qty += _r.quantity
        period_returned_count += 1
    # Farq bo'lsa — sahifada izoh chiqadi (buzuqlik emas, ta'rif farqi).
    refunds_span_periods = (_comparable
                            and period_returned_total != returned_total)

    # Ro'yxat — FAQAT ko'rsatish uchun 300 qator. Qaytarishlarни alohida so'rovда
    # olib har satrга biriktiramiz (dead _returned annotatsiyasi olib tashlandi).
    sales = list(qs[:300])
    shown_capped = line_count > 300
    for s in sales:
        r = _ret_by_sale.get(s.pk)
        # DIQQAT: Sale.returned_qty — read-only @property (u _returned'ni o'qiydi).
        # Unga to'g'ridan-to'g'ri yozib bo'lmaydi (Attribute: no setter → 500).
        # Shu bois annotatsiya maydoni _returned'ni to'ldiramiz — property uni oladi.
        s._returned = r['qty'] if r else 0
        s.returned_money = r['money'] if r else 0
        s.net_total = s.total - s.returned_money

    # Kunlik jami — DB'da (+ qaytarish alohida so'rov, sana bo'yicha).
    daily_list = []
    for row in (qs.values('sold_at__date')
                  .annotate(total=Sum(_rev_expr), qty=Sum('quantity'),
                            count=Count('transaction_id', distinct=True))
                  .order_by('-sold_at__date')[:60]):
        _d = row['sold_at__date']
        _rmoney = float((ret_by_day.get(_d) or {}).get('money') or 0)
        # SAL-1: kunlik jami ham chek chegirmasidan tozalanadi.
        _rtotal = float(_dec(row['total'] or 0) - (_odisc_by_day.get(_d) or Decimal('0')))
        daily_list.append({
            'date': _d,
            'total': _rtotal,
            'qty': row['qty'] or 0,
            'count': row['count'] or 0,
            'returned': _rmoney,
            'net': _rtotal - _rmoney,
        })

    # ---- CHEK KO'RINISHI: bir qator = butun chek. Filtrlangan qs'даgi
    # tranzaksiyalarни guruhlab, har biriга chek jamisi (t.total = to'langan),
    # dona soni, va (agar bo'lsa) qaytarilган summani biriktiramiz. Eng so'nggi
    # 300 chek ko'rsatiladi. (Chek'siz eski sotuvlar faqat 'items' ko'rinishда.)
    checks = []
    if view_mode == 'checks':
        _txn_order = list(qs.filter(transaction_id__isnull=False)
                          .values('transaction_id')
                          .annotate(_last=Max('sold_at'))
                          .order_by('-_last')[:300])
        _ids = [r['transaction_id'] for r in _txn_order]
        _tmap = {t.pk: t for t in SaleTransaction.objects
                 .filter(pk__in=_ids)
                 .select_related('branch', 'sold_by')
                 .prefetch_related('lines__variant__product')}
        for r in _txn_order:
            t = _tmap.get(r['transaction_id'])
            if t is None:
                continue
            _lines = list(t.lines.all())
            _rm = Decimal('0')
            _rq = 0
            for ln in _lines:
                _e = _ret_by_sale.get(ln.id)
                if _e:
                    _rm += _e['money']
                    _rq += _e['qty']
            _first = _lines[0].variant.product.name if _lines else ''
            _paid = t.total
            checks.append({
                'id': t.pk,
                'public_id': t.public_id,
                'sold_at': r['_last'],
                'branch': t.branch,
                'sold_by': t.sold_by,
                'payment': t.get_payment_method_display(),
                'n_items': t.item_count,
                'n_lines': len(_lines),
                'first_product': _first,
                'customer': t.customer_name,
                'total': _paid,
                'returned_money': _rm,
                'returned_qty': _rq,
                'net_total': _paid - _rm,
            })

    # Ko'rinish tugmalari uchun: joriy filtrlar (view/export'siz) querystring.
    _gc = request.GET.copy()
    _gc.pop('view', None)
    _gc.pop('export', None)
    qs_base = _gc.urlencode()

    return render(request, 'inventory/sales_list.html', {
        'view_mode': view_mode,
        'qs_base': qs_base,
        'checks': checks,
        'check_count': txn_count,
        'checks_capped': txn_count > 300,
        'sales': sales,
        'total': total,
        'qty_total': qty_total,
        'txn_count': txn_count,
        'line_count': line_count,
        'shown_capped': shown_capped,
        'returned_total': returned_total,
        'returned_qty': returned_qty,
        'returned_count': returned_count,
        'net_total': net_total,
        'gross_total': gross_total,
        'order_discount_total': order_discount_total,
        'discount_split': _disc_split,   # DISC-1
        'period_returned_total': period_returned_total,
        'period_returned_qty': period_returned_qty,
        'period_returned_count': period_returned_count,
        'refunds_span_periods': refunds_span_periods,
        'net_qty': (qty_total or 0) - (returned_qty or 0),
        'daily_list': daily_list,
        # Filter state
        'q': q,
        'date_from': date_from_raw,
        'date_to': date_to_raw,
        'branch_id': branch_id,
        'seller_id': seller_id,
        'payment_method': payment_method,
        'returned': returned,
        'discount': discount,
        # Choices
        'branches': Branch.objects.filter(is_active=True).order_by('name'),
        'sellers': User.objects.filter(is_active=True).order_by('username'),
        'payment_methods': SaleTransaction.PaymentMethod.choices,
    })


# ---------- REPORTS ----------

def _resolve_period(period, date_from, date_to):
    """Davr bo'yicha (start_dt, end_dt) qaytaradi. end_dt = davrdan keyingi kun."""
    today = timezone.localdate()
    if period == 'today':
        start = today
        end = today + timedelta(days=1)
    elif period == 'yesterday':
        start = today - timedelta(days=1)
        end = today
    elif period == 'week':
        start = today - timedelta(days=7)
        end = today + timedelta(days=1)
    elif period == 'month':
        start = today - timedelta(days=30)
        end = today + timedelta(days=1)
    elif period == 'this_month':
        start = today.replace(day=1)
        end = today + timedelta(days=1)
    elif period == 'custom':
        start = date_from or today
        end = (date_to or today) + timedelta(days=1)
    else:
        start = today - timedelta(days=7)
        end = today + timedelta(days=1)
    tz = timezone.get_current_timezone()
    start_dt = datetime.combine(start, datetime.min.time()).replace(tzinfo=tz)
    end_dt = datetime.combine(end, datetime.min.time()).replace(tzinfo=tz)
    return start, end - timedelta(days=1), start_dt, end_dt


PIVOT_DIM_CONFIG = {
    # Dimensions that need an annotation to project a value out of sold_at
    'date':     {'label': 'Sana',       'expr': TruncDate('sold_at'),     'field': 'pv_date'},
    'week':     {'label': 'Hafta',      'expr': TruncWeek('sold_at'),     'field': 'pv_week'},
    'month':    {'label': 'Oy',         'expr': TruncMonth('sold_at'),    'field': 'pv_month'},
    'weekday':  {'label': 'Hafta kuni', 'expr': ExtractWeekDay('sold_at'),'field': 'pv_wd'},
    # Direct relational fields
    'branch':   {'label': 'Filial',       'field': 'branch__name'},
    'category': {'label': 'Kategoriya',   'field': 'variant__product__category__name'},
    'product':  {'label': 'Mahsulot',     'field': 'variant__product__name'},
    'seller':   {'label': 'Sotuvchi',     'field': 'sold_by__username'},
    'payment':  {'label': "To'lov turi",  'field': 'transaction__payment_method'},
}

# Auto-hierarchy: coarser categories first when user picks multiple row dims.
# Lower number = closer to the root of the visual tree.
PIVOT_DIM_ORDER = {
    'branch':   10,
    'seller':   20,
    'category': 30,
    'product':  40,
    'payment':  50,
    'month':    60,
    'week':     70,
    'date':     80,
    'weekday':  90,
}

# Django ExtractWeekDay: Sunday=1, Monday=2, ..., Saturday=7
PIVOT_WEEKDAY_NAMES = {1: 'Yakshanba', 2: 'Dushanba', 3: 'Seshanba',
                       4: 'Chorshanba', 5: 'Payshanba',
                       6: 'Juma',      7: 'Shanba'}


def _pivot_format_key(dim, val):
    """Render a row/column key for display in the pivot table."""
    if val is None:
        return '—'
    if dim == 'weekday':
        return PIVOT_WEEKDAY_NAMES.get(int(val), str(val))
    if dim in ('date', 'week'):
        return val.strftime('%Y-%m-%d') if hasattr(val, 'strftime') else str(val)
    if dim == 'month':
        return val.strftime('%Y-%m') if hasattr(val, 'strftime') else str(val)
    if dim == 'payment':
        labels = dict(SaleTransaction._meta.get_field('payment_method').choices)
        return labels.get(val, val)
    return str(val)


def _sort_dim_value(dim, v):
    """Sort key for distinct row/col values within a dimension."""
    if v is None:
        return (1, '')
    if dim == 'weekday':
        return (0, int(v))
    if dim in ('date', 'week', 'month') and hasattr(v, 'isoformat'):
        return (0, v.isoformat())
    return (0, str(v))


def _build_pivot(sale_qs, rows_dims, cols_dim, metric):
    """Aggregate sale_qs by N row dimensions and (optional) 1 col dimension.

    rows_dims: list[str] — already sorted by auto-hierarchy (coarsest first).
    Returns:
      mode '1d'  → items: [{keys, value, pct}, ...]  + total + headers
      mode '2d'  → headers (row + col), matrix rows w/ keys+cells+total,
                   col_totals, grand_total
    """
    rev_expr = ExpressionWrapper(
        F('quantity') * F('sale_price') - F('line_discount'),
        output_field=DecimalField(max_digits=14, decimal_places=2))
    profit_expr = ExpressionWrapper(
        (F('sale_price') - F('cost_at_sale')) * F('quantity') - F('line_discount'),
        output_field=DecimalField(max_digits=14, decimal_places=2))
    annot = {
        'revenue': Sum(rev_expr),
        'qty':     Sum('quantity'),
        'count':   Count('transaction_id', distinct=True),
        'profit':  Sum(profit_expr),
    }

    extra = {}
    group_fields = []
    all_dims = list(rows_dims) + ([cols_dim] if cols_dim else [])
    for dim in all_dims:
        if not dim or dim not in PIVOT_DIM_CONFIG:
            continue
        cfg = PIVOT_DIM_CONFIG[dim]
        if 'expr' in cfg:
            extra[cfg['field']] = cfg['expr']
        group_fields.append(cfg['field'])

    qs = sale_qs.annotate(**extra) if extra else sale_qs
    agg = list(qs.values(*group_fields).annotate(**annot).order_by())

    def m(r): return r.get(metric) or 0

    row_fields = [PIVOT_DIM_CONFIG[d]['field'] for d in rows_dims]
    row_labels = [PIVOT_DIM_CONFIG[d]['label'] for d in rows_dims]

    if not cols_dim:
        # 1-D group (possibly N-level nested rows)
        bucket = {}  # tuple-of-keys → metric value
        for r in agg:
            key = tuple(r[f] for f in row_fields)
            bucket[key] = bucket.get(key, 0) + m(r)

        # Sort by row_dims order (parent first), within each level use _sort_dim_value
        items_keys = sorted(bucket.keys(),
                            key=lambda k: tuple(_sort_dim_value(rows_dims[i], k[i])
                                                for i in range(len(k))))
        total = sum(bucket.values())
        items = []
        prev_keys = [None] * len(row_fields)
        for k in items_keys:
            display = []
            for i, raw in enumerate(k):
                fmt = _pivot_format_key(rows_dims[i], raw)
                # Hide repeated parent values from adjacent rows for clarity
                if prev_keys[i] == raw and i < len(k) - 1:
                    display.append('')
                else:
                    display.append(fmt)
            prev_keys = list(k)
            value = bucket[k]
            items.append({'keys': display, 'value': value,
                          'pct': (value / total * 100) if total else 0})
        return {
            'mode': '1d',
            'row_labels': row_labels,
            'items':      items,
            'total':      total,
        }

    # 2-D pivot with N-level rows
    cols_field = PIVOT_DIM_CONFIG[cols_dim]['field']

    # Aggregate: row_key_tuple → col_key → value
    bucket = {}
    col_set = set()
    for r in agg:
        rk = tuple(r[f] for f in row_fields)
        ck = r[cols_field]
        col_set.add(ck)
        bucket.setdefault(rk, {})
        bucket[rk][ck] = bucket[rk].get(ck, 0) + m(r)

    col_keys_raw = sorted(col_set, key=lambda v: _sort_dim_value(cols_dim, v))
    row_keys = sorted(bucket.keys(),
                      key=lambda k: tuple(_sort_dim_value(rows_dims[i], k[i])
                                          for i in range(len(k))))

    matrix = []
    col_totals = [0] * len(col_keys_raw)
    grand_total = 0
    prev_keys = [None] * len(row_fields)
    for rk in row_keys:
        row_cells = []
        row_total = 0
        for ci, ck in enumerate(col_keys_raw):
            v = bucket[rk].get(ck, 0)
            row_cells.append(v)
            row_total += v
            col_totals[ci] += v
        display = []
        for i, raw in enumerate(rk):
            fmt = _pivot_format_key(rows_dims[i], raw)
            if prev_keys[i] == raw and i < len(rk) - 1:
                display.append('')
            else:
                display.append(fmt)
        prev_keys = list(rk)
        matrix.append({
            'keys':  display,
            'cells': row_cells,
            'total': row_total,
        })
        grand_total += row_total

    return {
        'mode': '2d',
        'row_labels': row_labels,
        'cols_label': PIVOT_DIM_CONFIG[cols_dim]['label'],
        'col_keys':   [_pivot_format_key(cols_dim, v) for v in col_keys_raw],
        'matrix':     matrix,
        'col_totals': col_totals,
        'grand_total': grand_total,
    }


@admin_required
def reports(request):
    form = ReportForm(request.GET or None, initial={'period': 'week', 'report_type': 'sales'})
    rows = []
    headers = []
    title = ''
    summary = {}
    pivot = None

    if form.is_valid():
        rtype = form.cleaned_data['report_type']
        period = form.cleaned_data['period']
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')
        branch = form.cleaned_data.get('branch')

        d_start, d_end, dt_start, dt_end = _resolve_period(period, date_from, date_to)

        if rtype == 'sales':
            title = "Sotuvlar hisoboti"
            qs = Sale.objects.filter(sold_at__gte=dt_start, sold_at__lt=dt_end) \
                .select_related('variant__product', 'branch', 'sold_by').order_by('-sold_at')
            if branch:
                qs = qs.filter(branch=branch)
            headers = ['Sana', 'Filial', 'Kod', 'Mahsulot', "O'lcham", 'Rang',
                       'Soni', "Narx (so'm)", "Jami (so'm)", 'Sotuvchi', 'Izoh']
            total_qty = 0
            total_rev = 0
            for s in qs:
                rows.append([
                    timezone.localtime(s.sold_at).strftime('%Y-%m-%d %H:%M'),
                    s.branch.name, s.variant.product.code, s.variant.product.name,
                    s.variant.size, s.variant.color, s.quantity,
                    s.sale_price, s.total, s.sold_by.username, s.note,
                ])
                total_qty += s.quantity
                total_rev += s.total
            # SAL-6: qatorlar O'Z narxida qoladi (chek chegirmasi ayrim
            # tovarga tegishli emas — egasining qoidasi), lekin XULOSA
            # yopilishi SHART: qatorlar − chek chegirmasi = sof daromad.
            _odisc_rep, _ = _order_discount_share(qs)
            _rret, _, _ = _returns_adjustment(qs)
            summary = {'Jami sotilgan dona': total_qty,
                       "Qatorlar bo'yicha (so'm)": total_rev,
                       "Chek chegirmasi (so'm)": -_dec(_odisc_rep),
                       "Qaytarilgan (so'm)": -_rret,
                       "Jami daromad (so'm)": (_dec(total_rev)
                                               - _dec(_odisc_rep) - _rret),
                       'Sotuvlar soni': qs.count()}

        elif rtype == 'intakes':
            title = "Qabullar hisoboti"
            qs = Intake.objects.filter(received_at__gte=dt_start, received_at__lt=dt_end) \
                .select_related('variant__product', 'branch', 'received_by') \
                .order_by('-received_at')
            if branch:
                qs = qs.filter(branch=branch)
            headers = ['Sana', 'Filial', 'Kod', 'Mahsulot', "O'lcham", 'Rang',
                       'Soni', "Tannarx (so'm)", "Jami xarajat (so'm)",
                       'Yetkazib beruvchi', 'Qabul qildi', 'Izoh']
            total_qty = 0
            total_cost = 0
            for i in qs:
                rows.append([
                    timezone.localtime(i.received_at).strftime('%Y-%m-%d %H:%M'),
                    i.branch.name, i.variant.product.code, i.variant.product.name,
                    i.variant.size, i.variant.color, i.quantity,
                    i.cost_per_unit, i.total_cost, i.supplier,
                    i.received_by.username, i.note,
                ])
                total_qty += i.quantity
                total_cost += i.total_cost
            summary = {'Jami qabul qilingan dona': total_qty,
                       "Jami xarajat (so'm)": total_cost,
                       'Qabullar soni': qs.count()}

        elif rtype == 'inventory':
            title = "Joriy ombor holati"
            qs = BranchStock.objects.filter(stock_count__gt=0) \
                .select_related('variant__product', 'branch') \
                .order_by('branch__name', 'variant__product__name')
            if branch:
                qs = qs.filter(branch=branch)
            headers = ['Filial', 'Kod', 'Mahsulot', "O'lcham", 'Rang',
                       'Soni', "Tannarx (so'm)", "Jami qiymat (so'm)"]
            total_qty = 0
            total_val = 0
            for s in qs:
                val = s.stock_count * s.cost_price
                rows.append([
                    s.branch.name, s.variant.product.code, s.variant.product.name,
                    s.variant.size, s.variant.color,
                    s.stock_count, s.cost_price, val,
                ])
                total_qty += s.stock_count
                total_val += val
            summary = {'Jami dona': total_qty, "Jami qiymat (so'm)": total_val,
                       'Variantlar soni': qs.count()}

        elif rtype == 'by_product':
            title = "Mahsulotlar bo'yicha sotuv xulosasi"
            sale_qs = Sale.objects.filter(sold_at__gte=dt_start, sold_at__lt=dt_end)
            if branch:
                sale_qs = sale_qs.filter(branch=branch)
            agg = sale_qs.values(
                'variant__product__code', 'variant__product__name'
            ).annotate(
                qty=Sum('quantity'),
                # SAL-6: line_discount umuman ayirilmasdi — bu shubhasiz xato.
                revenue=Sum(ExpressionWrapper(
                    F('quantity') * F('sale_price') - F('line_discount'),
                    output_field=DecimalField(max_digits=14, decimal_places=2))),
                count=Count('id'),
            ).order_by('-revenue')
            headers = ['Kod', 'Mahsulot', 'Sotuvlar soni', 'Sotilgan dona', "Daromad (so'm)"]
            total_qty = 0
            total_rev = 0
            for a in agg:
                rows.append([
                    a['variant__product__code'], a['variant__product__name'],
                    a['count'], a['qty'], a['revenue'],
                ])
                total_qty += a['qty'] or 0
                total_rev += a['revenue'] or 0
            _odisc_rep, _ = _order_discount_share(sale_qs)
            _rret, _, _ = _returns_adjustment(sale_qs)
            summary = {'Jami sotilgan dona': total_qty,
                       "Mahsulotlar bo'yicha (so'm)": total_rev,
                       "Chek chegirmasi (so'm)": -_dec(_odisc_rep),
                       "Qaytarilgan (so'm)": -_rret,
                       "Jami daromad (so'm)": (_dec(total_rev)
                                               - _dec(_odisc_rep) - _rret),
                       'Mahsulotlar soni': len(agg)}

        elif rtype == 'pivot':
            # Multi-select rows: pick whatever user checked, default to branch.
            picked_rows = form.cleaned_data.get('pivot_rows') or []
            if not picked_rows:
                picked_rows = ['branch']
            # Auto-hierarchy: sort by the canonical parent-first order.
            rows_dims = sorted(
                [d for d in picked_rows if d in PIVOT_DIM_ORDER],
                key=lambda d: PIVOT_DIM_ORDER[d],
            )
            cols_dim = form.cleaned_data.get('pivot_cols') or ''
            metric   = form.cleaned_data.get('pivot_metric') or 'revenue'

            metric_label = dict(PIVOT_METRIC_CHOICES).get(metric, metric)
            row_labels   = [PIVOT_DIM_CONFIG[d]['label'] for d in rows_dims]
            cols_label   = PIVOT_DIM_CONFIG[cols_dim]['label'] if cols_dim else ''

            title = (f"Pivot — {metric_label}: "
                     + ' > '.join(row_labels)
                     + (f" × {cols_label}" if cols_label else ''))

            sale_qs = Sale.objects.filter(sold_at__gte=dt_start, sold_at__lt=dt_end)
            if branch:
                sale_qs = sale_qs.filter(branch=branch)

            pivot = _build_pivot(sale_qs, rows_dims, cols_dim, metric)
            pivot['metric'] = metric
            pivot['metric_label'] = metric_label
            pivot['is_currency'] = metric in ('revenue', 'profit')
            pivot['rows_dims']  = rows_dims  # what the picker should re-check

            # CSV / PDF flat representation
            if pivot['mode'] == '1d':
                headers = row_labels + [metric_label, '% Ulush']
                rows = []
                for it in pivot['items']:
                    rows.append(list(it['keys']) + [it['value'], f"{it['pct']:.1f}%"])
                summary = {f"Jami {metric_label.lower()}": pivot['total']}
                if metric in ('revenue', 'profit'):
                    _odisc_rep, _ = _order_discount_share(sale_qs)
                    summary["Chek chegirmasi (so'm)"] = -_dec(_odisc_rep)
                    summary["Sof (so'm)"] = _dec(pivot['total']) - _dec(_odisc_rep)
            else:
                headers = row_labels + list(pivot['col_keys']) + ['Jami']
                rows = []
                for r in pivot['matrix']:
                    rows.append(list(r['keys']) + list(r['cells']) + [r['total']])
                rows.append(['Ustun jami'] + [''] * (len(row_labels) - 1)
                            + list(pivot['col_totals']) + [pivot['grand_total']])
                summary = {f"Jami {metric_label.lower()}": pivot['grand_total']}
                if metric in ('revenue', 'profit'):
                    _odisc_rep, _ = _order_discount_share(sale_qs)
                    summary["Chek chegirmasi (so'm)"] = -_dec(_odisc_rep)
                    summary["Sof (so'm)"] = (_dec(pivot['grand_total'])
                                             - _dec(_odisc_rep))

        elif rtype == 'deadstock':
            title = "O'lik zaxira — davrda sotilmagan tovarlar"
            stock_qs = BranchStock.objects.filter(stock_count__gt=0).exclude(
                variant__product__is_open_price=True).select_related('variant__product')
            if branch:
                stock_qs = stock_qs.filter(branch=branch)
            smap = {}
            for st in stock_qs:
                pid = st.variant.product_id
                d = smap.setdefault(pid, {'qty': 0, 'val': 0,
                    'code': st.variant.product.code, 'name': st.variant.product.name})
                d['qty'] += st.stock_count
                d['val'] += st.stock_count * st.cost_price
            sold_qs = Sale.objects.filter(sold_at__gte=dt_start, sold_at__lt=dt_end)
            if branch:
                sold_qs = sold_qs.filter(branch=branch)
            sold_map = {r['variant__product']: r['q'] for r in
                        sold_qs.values('variant__product').annotate(q=Sum('quantity'))}
            headers = ['Kod', 'Mahsulot', 'Ombordagi soni',
                       "Band pul (so'm)", 'Davrda sotilgan']
            dead = []
            for pid, d in smap.items():
                if sold_map.get(pid, 0) == 0:
                    dead.append((d['val'], d['code'], d['name'], d['qty']))
            dead.sort(reverse=True)
            total_val = 0
            for val, code, name, qty in dead:
                rows.append([code, name, qty, val, 0])
                total_val += val
            summary = {"O'lik zaxira mahsulotlari": len(dead),
                       "Band bo'lgan pul (so'm)": total_val}

        elif rtype == 'reorder':
            title = "Qayta buyurtma tavsiyasi (14 kunlik)"
            days = max(1, (d_end - d_start).days) if d_start and d_end else 30
            sold_qs = Sale.objects.filter(sold_at__gte=dt_start, sold_at__lt=dt_end)
            if branch:
                sold_qs = sold_qs.filter(branch=branch)
            sold_map = {r['variant__product']: r['q'] for r in
                        sold_qs.values('variant__product').annotate(q=Sum('quantity'))}
            stock_qs = BranchStock.objects.exclude(
                variant__product__is_open_price=True).select_related('variant__product')
            if branch:
                stock_qs = stock_qs.filter(branch=branch)
            smap = {}
            for st in stock_qs:
                pid = st.variant.product_id
                d = smap.setdefault(pid, {'qty': 0,
                    'code': st.variant.product.code, 'name': st.variant.product.name})
                d['qty'] += st.stock_count
            headers = ['Kod', 'Mahsulot', 'Ombordagi soni', "Kunlik o'rtacha",
                       'Qolgan kun', 'Tavsiya (dona)']
            urgent = []
            for pid, d in smap.items():
                sold = sold_map.get(pid, 0)
                daily = sold / days if sold else 0
                if daily <= 0:
                    continue
                days_left = d['qty'] / daily
                if days_left < 14:
                    need = max(0, round(daily * 14) - d['qty'])
                    urgent.append((days_left, d['code'], d['name'], d['qty'],
                                   round(daily, 1), round(days_left), need))
            urgent.sort()
            for dl, code, name, qty, daily, dleft, need in urgent:
                rows.append([code, name, qty, daily, dleft, need])
            summary = {"Qayta buyurtma kerak bo'lgan": len(urgent),
                       "Davr (kun)": days}

        elif rtype == 'margin':
            title = "Kategoriya bo'yicha marja"
            sale_qs = Sale.objects.filter(sold_at__gte=dt_start, sold_at__lt=dt_end)
            if branch:
                sale_qs = sale_qs.filter(branch=branch)
            agg = sale_qs.values('variant__product__category__name').annotate(
                qty=Sum('quantity'),
                rev=Sum(ExpressionWrapper(F('quantity') * F('sale_price'),
                        output_field=DecimalField(max_digits=14, decimal_places=2))),
                cost=Sum(ExpressionWrapper(F('quantity') * F('cost_at_sale'),
                        output_field=DecimalField(max_digits=14, decimal_places=2))),
            ).order_by('-rev')
            headers = ['Kategoriya', 'Sotilgan dona', "Daromad (so'm)",
                       "Tannarx (so'm)", "Foyda (so'm)", 'Marja %']
            t_rev = 0; t_cost = 0; t_qty = 0
            for a in agg:
                rev = a['rev'] or 0; cost = a['cost'] or 0; profit = rev - cost
                margin = (profit / rev * 100) if rev else 0
                rows.append([a['variant__product__category__name'] or '—',
                             a['qty'] or 0, rev, cost, profit, round(margin, 1)])
                t_rev += rev; t_cost += cost; t_qty += a['qty'] or 0
            _odisc_rep, _ = _order_discount_share(sale_qs)
            # RET-1: qaytarish tushumdan ham, tannarxdan ham chiqadi.
            _rret, _cret, _ = _returns_adjustment(sale_qs)
            _net_rep = _dec(t_rev) - _dec(_odisc_rep) - _rret
            t_cost = _dec(t_cost) - _cret
            summary = {"Kategoriyalar bo'yicha (so'm)": t_rev,
                       "Chek chegirmasi (so'm)": -_dec(_odisc_rep),
                       "Qaytarilgan (so'm)": -_rret,
                       "Jami daromad (so'm)": _net_rep,
                       "Jami foyda (so'm)": _net_rep - _dec(t_cost),
                       'Umumiy marja %': (round(float((_net_rep - _dec(t_cost))
                                                      / _net_rep * 100), 1)
                                          if _net_rep else 0)}

        elif rtype == 'payouts':
            title = "Kassa chiqimlari (naqd)"
            qs = (CashPayout.objects
                  .filter(created_at__gte=dt_start, created_at__lt=dt_end)
                  .select_related('branch', 'created_by')
                  .order_by('-created_at'))
            if branch:
                qs = qs.filter(branch=branch)
            headers = ['Sana', 'Filial', 'Kategoriya', "Summa (so'm)",
                       'Sotuvchi', 'Smen', 'Izoh']
            total = 0
            by_cat = {}
            for p in qs:
                amt = int(round(p.amount))
                rows.append([
                    timezone.localtime(p.created_at).strftime('%Y-%m-%d %H:%M'),
                    p.branch.name if p.branch else '—',
                    p.get_category_display(),
                    amt,
                    p.created_by.username if p.created_by else '—',
                    f'#{p.shift_id}',
                    p.note,
                ])
                total += amt
                by_cat[p.get_category_display()] = by_cat.get(p.get_category_display(), 0) + amt
            summary = {"Jami olingan (so'm)": total, 'Chiqimlar soni': len(rows)}
            for cat, amt in sorted(by_cat.items(), key=lambda kv: -kv[1]):
                summary[f"{cat} (so'm)"] = amt

        if request.GET.get('export') == 'csv':
            return _csv_response(title, headers, rows, summary,
                                 d_start, d_end, branch)
        if request.GET.get('export') == 'pdf':
            return _pdf_response(title, headers, rows, summary,
                                 d_start, d_end, branch)

        return render(request, 'inventory/reports.html', {
            'form': form, 'rows': rows, 'headers': headers, 'title': title,
            'summary': summary, 'd_start': d_start, 'd_end': d_end,
            'branch': branch, 'pivot': pivot,
        })

    return render(request, 'inventory/reports.html', {
        'form': form, 'rows': None, 'headers': [], 'title': '',
        'pivot': None,
    })


# ---------- CART / MULTI-ITEM SALE ----------

CART_KEY = 'cart'  # session dict: {str(stock_id): qty}


def _get_cart(request):
    return request.session.get(CART_KEY, {}) or {}


def _save_cart(request, cart):
    # drop zero/negative entries
    cart = {k: int(v) for k, v in cart.items() if int(v) > 0}
    request.session[CART_KEY] = cart
    request.session.modified = True
    return cart


def _cart_lines(cart):
    """Resolve cart {stock_id: qty} → list of {stock, qty, available, ok}."""
    if not cart:
        return []
    stock_ids = [int(sid) for sid in cart.keys() if str(sid).isdigit()]
    stocks = {bs.id: bs for bs in BranchStock.objects.filter(id__in=stock_ids)
              .select_related('variant__product', 'branch')}
    lines = []
    for sid, qty in cart.items():
        stock = stocks.get(int(sid))
        if not stock:
            continue
        lines.append({
            'stock': stock,
            'qty': int(qty),
            'available': stock.stock_count,
            'ok': int(qty) <= stock.stock_count,
            'subtotal': int(qty) * (stock.sale_price or stock.variant.product.default_sale_price),
        })
    return lines


def cart_count(request):
    """Helper used by base.html navbar to show cart badge."""
    cart = _get_cart(request)
    return sum(int(v) for v in cart.values())


@login_required
def cart_add(request, stock_id):
    """POST /cart/add/<stock_id>/  with qty (default 1)."""
    stock = get_object_or_404(BranchStock.objects.select_related('branch'), pk=stock_id)
    if not request.user.is_admin() and request.user.branch_id != stock.branch_id:
        return HttpResponseForbidden("Bu filialda sotishga ruxsat yo'q.")

    try:
        qty = max(1, int(request.POST.get('qty') or 1))
    except ValueError:
        qty = 1
    cart = _get_cart(request)
    new_qty = int(cart.get(str(stock_id), 0)) + qty
    if new_qty > stock.stock_count:
        new_qty = stock.stock_count
        messages.warning(request,
            f"Omborda faqat {stock.stock_count} dona bor.")
    cart[str(stock_id)] = new_qty
    _save_cart(request, cart)
    messages.success(request,
        f"Savatga qo'shildi: {stock.variant.product.code} × {qty} dona.")
    next_url = request.POST.get('next') or request.META.get('HTTP_REFERER') or 'lookup'
    return redirect(next_url if next_url.startswith('/') else 'lookup')


@login_required
def cart_view(request):
    cart = _get_cart(request)
    lines = _cart_lines(cart)
    # If user is a seller, only their branch's items
    if not request.user.is_admin():
        lines = [l for l in lines if l['stock'].branch_id == request.user.branch_id]
    total = sum(l['subtotal'] for l in lines)
    return render(request, 'inventory/cart.html', {
        'lines': lines, 'total': total,
        'payment_methods': SaleTransaction.PaymentMethod.choices,
    })


@login_required
def cart_update(request):
    """POST: update qty of an item or remove it."""
    cart = _get_cart(request)
    stock_id = request.POST.get('stock_id')
    action = request.POST.get('action')
    if stock_id in cart:
        if action == 'remove':
            del cart[stock_id]
        else:
            try:
                qty = int(request.POST.get('qty') or 0)
            except ValueError:
                qty = 0
            if qty <= 0:
                del cart[stock_id]
            else:
                stock = BranchStock.objects.filter(pk=stock_id).first()
                if stock:
                    cart[stock_id] = min(qty, stock.stock_count)
    _save_cart(request, cart)
    return redirect('cart_view')


@login_required
def cart_clear(request):
    request.session[CART_KEY] = {}
    request.session.modified = True
    messages.info(request, "Savat bo'shatildi.")
    return redirect('lookup')


@login_required
def checkout(request):
    cart = _get_cart(request)
    lines = _cart_lines(cart)
    if not request.user.is_admin():
        lines = [l for l in lines if l['stock'].branch_id == request.user.branch_id]
    if not lines:
        messages.warning(request, "Savat bo'sh.")
        return redirect('lookup')

    # All lines must be from same branch
    branches = {l['stock'].branch_id for l in lines}
    if len(branches) > 1:
        messages.error(request,
            "Bitta chekda turli filiallardan tovar bo'lmaydi. Ortiqcha tovarlarni olib tashlang.")
        return redirect('cart_view')
    if any(not l['ok'] for l in lines):
        messages.error(request, "Ba'zi tovarlar yetarli emas — savatni tekshiring.")
        return redirect('cart_view')

    if request.method == 'POST':
        branch = lines[0]['stock'].branch
        # MON-21: smensiz sotuv YO'Q (pos_checkout kabi) — aks holda shift=None
        # bilan yozilib, Z-hisobotга tushmaydi.
        open_shift = _open_shift_for(branch)
        if not open_shift:
            messages.error(request, "Smen ochilmagan. Avval smen oching.")
            return redirect('cart_view')
        # MON-20: to'lov turini normallashtiramiz + tekshiramiz (xom qiymat
        # DB constraint'ga urилиб 500 bermasin, va soxta bucket'га tushmasin).
        _pm = _norm_pay_method(request.POST.get('payment_method') or 'cash')
        if _pm not in ('cash', 'card', 'transfer'):
            messages.error(request, "To'lov turi noto'g'ri.")
            return redirect('cart_view')
        with transaction.atomic():
            txn = SaleTransaction.objects.create(
                branch=branch,
                sold_by=request.user,
                payment_method=_pm,
                shift=open_shift,
                customer_name=request.POST.get('customer_name') or '',
                customer_phone=request.POST.get('customer_phone') or '',
                note=request.POST.get('note') or '',
            )
            for line in lines:
                stock = line['stock']
                qty = line['qty']
                stock.stock_count = F('stock_count') - qty
                stock.save()
                Sale.objects.create(
                    transaction=txn,
                    variant=stock.variant, branch=stock.branch,
                    quantity=qty,
                    sale_price=stock.sale_price or stock.variant.product.default_sale_price,
                    cost_at_sale=stock.cost_price,
                    sold_by=request.user,
                )
            request.session[CART_KEY] = {}
            request.session.modified = True

        from .fiscal import submit_for_transaction
        submit_for_transaction(txn)

        messages.success(request,
            f"Sotuv yakunlandi: {len(lines)} ta mahsulot.")
        return redirect('transaction_detail', pk=txn.pk)

    return render(request, 'inventory/checkout.html', {
        'lines': lines, 'total': sum(l['subtotal'] for l in lines),
        'payment_methods': SaleTransaction.PaymentMethod.choices,
    })


@login_required
def transaction_detail(request, token):
    txn = get_object_or_404(
        SaleTransaction.objects.select_related('branch', 'sold_by')
            .prefetch_related('lines__variant__product'),
        public_id=token,   # SEC-6: ketma-ket PK emas, tasodifiy token
    )
    if not request.user.is_admin() and request.user.branch_id != txn.branch_id:
        return HttpResponseForbidden("Bu chekni ko'rishga ruxsat yo'q.")
    return render(request, 'inventory/transaction_detail.html', {'txn': txn})


# ---------- RETURNS ----------

@login_required
def return_create(request, sale_id):
    sale = get_object_or_404(Sale.objects.select_related('variant__product', 'branch'),
                             pk=sale_id)
    if not request.user.is_admin() and request.user.branch_id != sale.branch_id:
        return HttpResponseForbidden()

    max_returnable = sale.quantity - sale.returned_qty
    if max_returnable <= 0:
        messages.warning(request, "Bu sotuv allaqachon to'liq qaytarilgan.")
        return redirect('sales_list')

    # Require open shift for cash reconciliation
    open_shift = _open_shift_for(sale.branch)
    if request.method == 'POST' and not open_shift:
        messages.error(request,
            "Smen ochilmagan. Qaytarish faqat ochiq smen davomida amalga oshiriladi.")
        return redirect('return_create', sale_id=sale.id)

    if request.method == 'POST':
        try:
            qty = int(request.POST.get('quantity') or 0)
        except ValueError:
            qty = 0
        reason = (request.POST.get('reason') or '').strip()
        if qty < 1 or qty > max_returnable:
            messages.error(request,
                f"Qaytarish miqdori 1 dan {max_returnable} gacha bo'lishi kerak.")
            return redirect('return_create', sale_id=sale.id)
        with transaction.atomic():
            stock = BranchStock.objects.filter(
                variant=sale.variant, branch=sale.branch
            ).first()
            # R2: ochiq narxli tovar zaxira yuritmaydi — tiklamaymiz.
            if stock and not sale.variant.product.is_open_price:
                stock.stock_count = F('stock_count') + qty
                stock.save()
            Return.objects.create(
                shift=open_shift,
                sale=sale, quantity=qty, reason=reason,
                refunded_by=request.user,
            )
        messages.success(request,
            f"Qaytarildi: {qty} dona × {sale.sale_price} so'm.")
        return redirect('sales_list')

    return render(request, 'inventory/return_form.html', {
        'sale': sale, 'max_returnable': max_returnable,
    })


# ---------- AUDIT LOG ----------

@admin_required
def audit_list(request):
    logs = AuditLog.objects.select_related('user').all()
    # Filters
    action = request.GET.get('action') or ''
    user_filter = request.GET.get('user') or ''
    model = request.GET.get('model') or ''
    q = (request.GET.get('q') or '').strip()
    date_from = request.GET.get('date_from') or ''
    date_to = request.GET.get('date_to') or ''
    export = request.GET.get('export') == 'csv'

    if action:
        logs = logs.filter(action=action)
    if user_filter:
        logs = logs.filter(username_snapshot=user_filter)
    if model:
        logs = logs.filter(model_name=model)
    if q:
        logs = logs.filter(
            Q(username_snapshot__icontains=q) |
            Q(object_repr__icontains=q) |
            Q(model_name__icontains=q) |
            Q(object_id=q)
        )
    try:
        if date_from:
            logs = logs.filter(created_at__date__gte=datetime.strptime(date_from, '%Y-%m-%d').date())
    except ValueError:
        date_from = ''
    try:
        if date_to:
            logs = logs.filter(created_at__date__lte=datetime.strptime(date_to, '%Y-%m-%d').date())
    except ValueError:
        date_to = ''

    logs = logs.order_by('-created_at')

    if export:
        resp = HttpResponse(content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = 'attachment; filename="audit_log.csv"'
        resp.write('﻿')
        w = csv.writer(resp)
        w.writerow(['Sana', 'Vaqt', 'Foydalanuvchi', 'Amal', 'Model',
                    'Obyekt ID', 'Obyekt', 'IP'])
        for l in logs[:10000]:
            w.writerow([_csv_safe(c) for c in [   # SEC-20
                l.created_at.strftime('%Y-%m-%d'),
                l.created_at.strftime('%H:%M:%S'),
                l.username_snapshot or '',
                l.get_action_display(),
                l.model_name or '',
                l.object_id or '',
                l.object_repr or '',
                l.ip or '',   # SEC-15: model maydoni `ip` (ip_address emas — 500 berardi)
            ]])
        return resp

    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(logs, 50)
    page = paginator.get_page(request.GET.get('page'))

    actions = AuditLog.Action.choices
    users = (AuditLog.objects.values_list('username_snapshot', flat=True)
             .distinct().order_by('username_snapshot'))
    users = [u for u in users if u]
    models_list = (AuditLog.objects.values_list('model_name', flat=True)
                   .distinct().order_by('model_name'))
    models_list = [m for m in models_list if m]

    return render(request, 'inventory/audit_list.html', {
        'page': page, 'actions': actions, 'users': users,
        'models_list': models_list,
        'f_action': action, 'f_user': user_filter, 'f_model': model, 'q': q,
        'date_from': date_from, 'date_to': date_to,
    })


# ---------- CUSTOMERS ----------

@admin_required
def customer_list(request):
    """Mijozlar ro'yxati: qidiruv + segment filter + CSV export."""
    q = (request.GET.get('q') or '').strip()
    segment = request.GET.get('segment') or ''
    export = request.GET.get('export') == 'csv'

    customers = Customer.objects.all()
    if q:
        customers = customers.filter(Q(name__icontains=q) | Q(phone__icontains=q))

    revenue_expr = ExpressionWrapper(
        F('transactions__lines__quantity') * F('transactions__lines__sale_price'),
        output_field=DecimalField(max_digits=14, decimal_places=2),
    )
    customers = customers.annotate(
        txn_count=Count('transactions', distinct=True),
        total_spent=Coalesce(Sum(revenue_expr), 0,
                             output_field=DecimalField(max_digits=14, decimal_places=2)),
        last_visit=Max('transactions__sold_at'),
    ).order_by('-total_spent', 'name')

    # Segment thresholds (sums in UZS)
    VIP_THRESHOLD = 5_000_000   # 5M+ → VIP
    REGULAR_THRESHOLD = 500_000  # 500k+ → Regular

    if segment == 'vip':
        customers = customers.filter(total_spent__gte=VIP_THRESHOLD)
    elif segment == 'regular':
        customers = customers.filter(total_spent__gte=REGULAR_THRESHOLD,
                                     total_spent__lt=VIP_THRESHOLD)
    elif segment == 'new':
        customers = customers.filter(total_spent__lt=REGULAR_THRESHOLD)

    customers = customers[:300]

    if export:
        resp = HttpResponse(content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = 'attachment; filename="customers.csv"'
        resp.write('﻿')
        w = csv.writer(resp)
        w.writerow(['Ism', 'Telefon', 'Tag', 'Cheklar', 'Jami sarf',
                    "O'rtacha chek", 'Oxirgi tashrif'])
        for c in customers:
            avg = (c.total_spent / c.txn_count) if c.txn_count else 0
            w.writerow([_csv_safe(c2) for c2 in [   # SEC-20
                c.name or '', c.phone or '', c.tags or '',
                c.txn_count or 0,
                float(c.total_spent or 0),
                float(avg),
                c.last_visit.strftime('%Y-%m-%d %H:%M') if c.last_visit else '',
            ]])
        return resp

    customers = list(customers)
    # Compute average ticket and segment label for display
    for c in customers:
        c.avg_ticket = (float(c.total_spent) / c.txn_count) if c.txn_count else 0
        spent = float(c.total_spent or 0)
        if spent >= VIP_THRESHOLD:
            c.segment_label = 'VIP'
            c.segment_class = 'bg-warning text-dark'
        elif spent >= REGULAR_THRESHOLD:
            c.segment_label = 'Doimiy'
            c.segment_class = 'bg-success'
        else:
            c.segment_label = 'Yangi'
            c.segment_class = 'bg-secondary'

    # Aggregate stats for header
    total_count = Customer.objects.count()
    total_revenue = Customer.objects.annotate(
        s=Coalesce(Sum(revenue_expr), 0,
                   output_field=DecimalField(max_digits=14, decimal_places=2))
    ).aggregate(t=Sum('s'))['t'] or 0

    return render(request, 'inventory/customer_list.html', {
        'customers': customers,
        'q': q, 'segment': segment,
        'total_count': total_count,
        'total_revenue': total_revenue,
    })


@admin_required
def customer_detail(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    txns = (customer.transactions.select_related('branch', 'sold_by')
            .prefetch_related('lines__variant__product')
            .order_by('-sold_at')[:50])
    revenue_expr = ExpressionWrapper(
        F('lines__quantity') * F('lines__sale_price') - F('lines__line_discount'),
        output_field=DecimalField(max_digits=14, decimal_places=2),
    )
    stats = customer.transactions.aggregate(
        n=Count('id', distinct=True),
        total=Coalesce(Sum(revenue_expr), 0,
                       output_field=DecimalField(max_digits=14, decimal_places=2)),
        first_visit=Min('sold_at'),
        last_visit=Max('sold_at'),
        odisc=Coalesce(Sum('order_discount'), 0,
                       output_field=DecimalField(max_digits=14, decimal_places=2)),
        # DISC-1: mijoz — CHEK darajasidagi o'lchov, shuning uchun uch qismni
        # to'g'ridan-to'g'ri chekdan yig'amiz (taqsimlash kerak emas).
        opromo=Coalesce(Sum('promo_discount'), 0,
                        output_field=DecimalField(max_digits=14, decimal_places=2)),
        oexch=Coalesce(Sum('exchange_credit'), 0,
                       output_field=DecimalField(max_digits=14, decimal_places=2)),
    )
    # SAL-6: mijoz HAQIQATDA to'lagani. Ilgari na qator chegirmasi, na chek
    # chegirmasi ayirilardi — "VIP / Doimiy / Yangi" segmenti ham shu shishgan
    # summaga qarab qo'yilardi. Mijoz — CHEK darajasidagi o'lchov, shuning
    # uchun order_discount to'g'ridan-to'g'ri ayiriladi.
    total = float(_dec(stats['total'] or 0) - _dec(stats['odisc'] or 0))
    order_discount_total = float(stats['odisc'] or 0)
    _promo = _dec(stats['opromo'] or 0)
    _exch = _dec(stats['oexch'] or 0)
    _manual = _dec(stats['odisc'] or 0) - _promo - _exch
    _disc_split = {'promo': _promo, 'exchange': _exch,
                   'manual': _manual if _manual > 0 else Decimal('0'),
                   'total': _dec(stats['odisc'] or 0)}
    n_txns = stats['n'] or 0
    avg_ticket = total / n_txns if n_txns else 0

    # Segment
    if total >= 5_000_000:
        segment = ('VIP', 'bg-warning text-dark')
    elif total >= 500_000:
        segment = ('Doimiy', 'bg-success')
    else:
        segment = ('Yangi', 'bg-secondary')

    # Top 5 favorite products (with image)
    fav_qs = (Sale.objects
              .filter(transaction__customer=customer)
              .values('variant__product_id',
                      'variant__product__code',
                      'variant__product__name')
              .annotate(qty=Sum('quantity'),
                        revenue=Sum(F('quantity') * F('sale_price')
                                    - F('line_discount'),
                                    output_field=DecimalField(max_digits=14,
                                                              decimal_places=2)))
              .order_by('-qty')[:5])
    favorites = list(fav_qs)
    # Attach product image URL where available
    fav_ids = [f['variant__product_id'] for f in favorites]
    prod_imgs = {p.id: (p.image.url if p.image else '')
                 for p in Product.objects.filter(id__in=fav_ids)}
    for f in favorites:
        f['image_url'] = prod_imgs.get(f['variant__product_id'], '')

    # 6-month spend chart (by month)
    from collections import OrderedDict
    from datetime import date as _date
    chart = OrderedDict()
    today = timezone.localdate()
    # Last 6 months
    cur = today.replace(day=1)
    months = []
    for _ in range(6):
        months.append(cur)
        # Previous month
        if cur.month == 1:
            cur = cur.replace(year=cur.year - 1, month=12)
        else:
            cur = cur.replace(month=cur.month - 1)
    months.reverse()
    for m in months:
        chart[m.strftime('%Y-%m')] = 0
    # Sum per month
    for t in customer.transactions.all():
        key = t.sold_at.astimezone(timezone.get_current_timezone()).strftime('%Y-%m')
        if key in chart:
            chart[key] += float(t.total)
    chart_labels = [m.strftime('%b %Y') for m in months]
    chart_data = list(chart.values())

    return render(request, 'inventory/customer_detail.html', {
        'order_discount_total': order_discount_total,
        'discount_split': _disc_split,   # DISC-1
        'customer': customer, 'txns': txns,
        'stats': stats,
        'total_spent': total,
        'avg_ticket': avg_ticket,
        'n_txns': n_txns,
        'first_visit': stats.get('first_visit'),
        'last_visit': stats.get('last_visit'),
        'segment_label': segment[0],
        'segment_class': segment[1],
        'favorites': favorites,
        'chart_labels': chart_labels,
        'chart_data': chart_data,
    })


@login_required
def pos_customer_lookup(request):
    """JSON: GET /pos/customer/?phone=998... → returns matches."""
    phone = (request.GET.get('phone') or '').strip()
    if len(phone) < 3:
        return JsonResponse({'matches': []})
    cleaned = ''.join(c for c in phone if c.isdigit() or c == '+')
    if not cleaned:
        return JsonResponse({'matches': []})
    qs = Customer.objects.filter(phone__icontains=cleaned)[:5]
    return JsonResponse({'matches': [
        {'id': c.pk, 'name': c.name, 'phone': c.phone} for c in qs
    ]})


# ---------- PWA ----------

def _serve_static_file(rel_path, content_type):
    from django.conf import settings
    import os
    path = os.path.join(settings.BASE_DIR, 'static', rel_path)
    if not os.path.exists(path):
        return HttpResponse(status=404)
    # OPS-11: fayl huquqi buzuq bo'lsa (masalan deploy'дан keyin sw.js o'qilmasa)
    # 500 emas, 404 qaytaramiz — brauzer eski SW'ни saqlaydi, sayt ochiq qoladi.
    try:
        with open(path, 'rb') as f:
            data = f.read()
    except OSError:
        return HttpResponse(status=404)
    response = HttpResponse(data, content_type=content_type)
    response['Cache-Control'] = 'public, max-age=3600'
    return response


def manifest(request):
    """PWA manifest served at /manifest.webmanifest"""
    return _serve_static_file('manifest.webmanifest', 'application/manifest+json')


def service_worker(request):
    """PWA service worker — must be served at root scope."""
    response = _serve_static_file('sw.js', 'application/javascript')
    response['Service-Worker-Allowed'] = '/'
    return response


# ---------- QR / ETIKETKA ----------

@login_required
def qr_image(request, code):
    """Mahsulot kodi uchun QR PNG"""
    norm = normalize_code(code)
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=8, border=2,
    )
    qr.add_data(norm)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, 'PNG')
    response = HttpResponse(buf.getvalue(), content_type='image/png')
    response['Cache-Control'] = 'public, max-age=3600'
    return response


# Do'kon nomi — peshtaxta narx etiketkasida chiqadi
PRICE_LABEL_STORE_NAME = 'Koreys Bozor'
PRICE_LABEL_SIZES = {'30x20': (30, 20), '58x40': (58, 40)}


@login_required
def price_labels(request):
    """Peshtaxta (narx) etiketkasi: mahsulot nomi + narx + do'kon nomi.

    Barcode YO'Q — bu sotuvchi skanerlaydigan emas, tokchaga qo'yiladigan
    narx yorlig'i. Ikki qog'oz o'lchami: 30×20 mm va 58×40 mm.
    Manba: ?codes=KOD1,KOD2  yoki  ?code=KOD  yoki  ?ids=<variant id'lar>.
    """
    size = request.GET.get('size') or '58x40'
    if size not in PRICE_LABEL_SIZES:
        size = '58x40'
    w, h = PRICE_LABEL_SIZES[size]

    try:
        copies = max(1, min(200, int(request.GET.get('copies') or 1)))
    except ValueError:
        copies = 1
    price_param = request.GET.get('price')

    # --- Turlarni yig'ish: codes= / code= (mahsulotning HAMMA turi) yoki
    #     ids= (aniq turlar). Har turning o'z narxi bo'lgani uchun etiketka
    #     MAHSULOTGA emas, TURGA chiqariladi.
    variants = []
    seen = set()

    def _add(v):
        if v and v.pk not in seen:
            seen.add(v.pk)
            variants.append(v)

    raw_codes = (request.GET.get('codes') or request.GET.get('code') or '')
    codes = [c.strip() for c in raw_codes.split(',') if c.strip()]
    products = []
    if codes:
        norm = [normalize_code(c.upper()) for c in codes]
        pmap = {p.code: p for p in Product.objects.filter(code__in=norm)}
        for c in norm:
            p = pmap.get(c)
            if not p:
                continue
            products.append(p)
            for v in (ProductVariant.objects.filter(product=p)
                      .select_related('product')
                      .order_by('color', 'size')):
                _add(v)

    ids = [int(i) for i in (request.GET.get('ids') or '').split(',')
           if i.strip().isdigit()]
    if ids:
        vmap = {v.pk: v for v in ProductVariant.objects.filter(pk__in=ids)
                .select_related('product')}
        for i in ids:
            _add(vmap.get(i))

    # --- Har TURNING o'z sotuv narxi ---
    branch = getattr(request.user, 'branch', None)
    labels = []
    for v in variants:
        if price_param:
            price = price_param
        else:
            qs = BranchStock.objects.filter(variant=v, sale_price__gt=0)
            st = (qs.filter(branch=branch).first() if branch else None) or qs.first()
            price = st.sale_price if st else v.product.default_sale_price
        try:
            price_str = f"{int(round(float(price))):,}"
        except (ValueError, TypeError):
            price_str = str(price)
        vtext = ' / '.join(x for x in (v.color, v.size) if x)
        labels.append({'variant': v, 'product': v.product, 'vtext': vtext,
                       'price': price, 'price_str': price_str})

    render_labels = []
    for lb in labels:
        for _ in range(copies):
            render_labels.append(lb)

    # Yo'lak (toolbar) havolalari uchun manbani saqlaymiz
    if ids:
        src_qs = 'ids=' + ','.join(str(i) for i in ids)
    else:
        src_qs = 'codes=' + ','.join(p.code for p in products)

    # "Orqaga" — kelib chiqgan sahifaga (faqat ichki yo'l; ochiq-redirect emas)
    back_raw = (request.GET.get('back') or '').strip()
    if back_raw.startswith('/') and not back_raw.startswith('//'):
        back_url = back_raw
    else:
        back_url = reverse('product_list')

    return render(request, 'inventory/price_labels.html', {
        'labels': render_labels, 'uniq_labels': labels,
        'copies': copies, 'size': size,
        'w': w, 'h': h, 'n_products': len(labels),
        'codes': ','.join(p.code for p in products),
        'src_qs': src_qs, 'back_url': back_url,
        'store_name': PRICE_LABEL_STORE_NAME,
    })


# ---------- INSIGHTS / ANALYTICS ----------

def _insights_context(request):
    """Insights view va export uchun barcha hisoblar."""
    period = request.GET.get('period', 'month')
    branch_id = request.GET.get('branch') or ''

    today = timezone.localdate()
    end_date = today   # davr oxirgi (kiritilган) kuni; odatda bugun
    if period == 'week':
        d_start = today - timedelta(days=7)
        days = 7
    elif period == 'today':
        d_start = today
        days = 1
    elif period == 'quarter':
        d_start = today - timedelta(days=90)
        days = 90
    elif period == 'this_month':
        # Kalendar oy: shu oyning 1-kunidan BUGUNГА qadar (rolling 30 kun EMAS).
        d_start = today.replace(day=1)
        days = (today - d_start).days + 1
    elif period == 'last_month':
        # To'liq o'tgan oy: 1-kunidan oxirgi kunigача.
        end_date = today.replace(day=1) - timedelta(days=1)   # o'tgan oy oxiri
        d_start = end_date.replace(day=1)                      # o'tgan oy boshi
        days = (end_date - d_start).days + 1
    else:
        period = 'month'
        d_start = today - timedelta(days=30)
        days = 30

    _PERIOD_LABELS = {'today': 'Bugun', 'week': '7 kun', 'month': '30 kun',
                      'quarter': '90 kun', 'this_month': 'Bu oy',
                      'last_month': "O'tgan oy"}
    period_label = _PERIOD_LABELS.get(period, period)

    tz = timezone.get_current_timezone()
    dt_start = datetime.combine(d_start, datetime.min.time()).replace(tzinfo=tz)
    dt_end = datetime.combine(end_date + timedelta(days=1), datetime.min.time()).replace(tzinfo=tz)
    # Oldingi davr (taqqoslash uchun)
    prev_d_start = d_start - timedelta(days=days)
    prev_dt_start = datetime.combine(prev_d_start, datetime.min.time()).replace(tzinfo=tz)
    prev_dt_end = dt_start

    sales = Sale.objects.filter(sold_at__gte=dt_start, sold_at__lt=dt_end)
    branches_all = Branch.objects.filter(is_active=True)
    selected_branch = None
    if branch_id:
        try:
            selected_branch = Branch.objects.get(pk=int(branch_id))
            sales = sales.filter(branch=selected_branch)
        except (Branch.DoesNotExist, ValueError):
            pass

    # Net revenue = qty*price - line_discount. Whole-order discount is
    # attached to SaleTransaction so we subtract it at the txn-id aggregate.
    revenue_expr = ExpressionWrapper(
        F('quantity') * F('sale_price') - F('line_discount'),
        output_field=DecimalField(max_digits=14, decimal_places=2)
    )

    totals = sales.aggregate(
        revenue=Sum(revenue_expr),
        qty=Sum('quantity'),
        sales_count=Count('id'),
    )

    # Subtract whole-order discounts (stored on SaleTransaction)
    txn_qs = SaleTransaction.objects.filter(
        sold_at__gte=dt_start, sold_at__lt=dt_end
    )
    if branch_id and selected_branch:
        txn_qs = txn_qs.filter(branch=selected_branch)
    order_disc = txn_qs.aggregate(s=Sum('order_discount'))['s'] or 0

    # SAL-6: sarlavhadagi raqam allaqachon sof edi, lekin PASTDAGI barcha
    # bo'linmalar brutto qolgan edi — sahifa o'zi bilan ziddiyatda:
    # "Tushum 240 000", ostida "Filial B — 300 000".
    # CHEK darajasidagi o'lchovlar (filial, sotuvchi) uchun ulush aniq
    # tushadi — ayiramiz. QATOR darajasidagilar (mahsulot, kategoriya,
    # guruh) O'Z narxida qoladi va chegirma alohida ko'rsatiladi (egasining
    # qoidasi: chek chegirmasi ayrim tovarga tegishli emas).
    _, _od_by_branch = _order_discount_share(sales, 'branch_id')
    _, _od_by_seller = _order_discount_share(sales, 'sold_by_id')
    _, _, _disc_split = _order_discount_share(sales, split=True)

    # RET-1: qaytarish tushumdan ham, tannarxdan ham chiqadi. Ilgari
    # /insights/ qaytarishni UMUMAN ko'rmasdi — 3 dona sotilib 1 dona
    # qaytsa ham uchalasining tushumi va tannarxi turaverardi.
    _ret_rev, _ret_cost, _ret_by_prod = _returns_adjustment(
        sales, 'variant__product_id')
    _, _, _ret_by_branch2 = _returns_adjustment(sales, 'branch_id')
    _, _, _ret_by_seller2 = _returns_adjustment(sales, 'sold_by_id')
    _, _, _ret_by_catname = _returns_adjustment(
        sales, 'variant__product__category__name')
    _, _, _ret_by_groupname = _returns_adjustment(
        sales, 'variant__product__category__group__name')

    revenue = (totals['revenue'] or 0) - order_disc - _ret_rev
    qty = totals['qty'] or 0
    sales_count = totals['sales_count'] or 0
    discount_total = order_disc + (sales.aggregate(s=Sum('line_discount'))['s'] or 0)

    # Foyda hisoblash: revenue - cost. cost = sum(quantity * variant.branch_stock.cost_price)
    # Sale modelida cost_price yo'q, lekin oxirgi BranchStock dan oladi (taxmin).
    # Aniqroq bo'lishi uchun har sotuv pay'tida cost_price'ni Sale'da saqlash kerak edi;
    # hozircha BranchStock.cost_price'dan foydalanamiz.
    # Cost is snapshotted on the Sale itself (cost_at_sale), so we can aggregate in SQL.
    cost_expr = ExpressionWrapper(
        F('quantity') * F('cost_at_sale'),
        output_field=DecimalField(max_digits=14, decimal_places=2)
    )
    total_cost = _dec(sales.aggregate(c=Sum(cost_expr))['c'] or 0) - _ret_cost

    profit_by_product_qs = sales.values(
        'variant__product_id', 'variant__product__code', 'variant__product__name',
    ).annotate(
        revenue=Sum(revenue_expr),
        cost=Sum(cost_expr),
        qty=Sum('quantity'),
    )
    profit_by_product = {}
    for row in profit_by_product_qs:
        _rr, _cc = (_ret_by_prod.get(row['variant__product_id'])
                    or (Decimal('0'), Decimal('0')))
        _rev = _dec(row['revenue'] or 0) - _rr
        _cost = _dec(row['cost'] or 0) - _cc
        profit_by_product[row['variant__product_id']] = {
            'code': row['variant__product__code'],
            'name': row['variant__product__name'],
            'revenue': _rev,
            'cost': _cost,
            'profit': _rev - _cost,
            'qty': row['qty'] or 0,
        }

    profit = revenue - total_cost
    margin = (profit / revenue * 100) if revenue else 0

    # TOP foyda keltirgan mahsulotlar (revenue emas, balki real foyda)
    top_profit = sorted(profit_by_product.values(),
                        key=lambda x: x['profit'], reverse=True)[:10]
    for p in top_profit:
        p['margin'] = (p['profit'] / p['revenue'] * 100) if p['revenue'] else 0

    # O'rtacha sotuv summasi va eng katta sotuv
    avg_sale_value = (revenue / sales_count) if sales_count else 0
    largest_sale = sales.annotate(
        # SAL-6: line_discount ayirilmasdi — chegirmali qator "eng katta
        # sotuv" bo'lib chiqib qolishi mumkin edi.
        line=ExpressionWrapper(F('quantity') * F('sale_price')
                               - F('line_discount'),
                               output_field=DecimalField(max_digits=14, decimal_places=2))
    ).order_by('-line').first()

    # O'tgan davr taqqoslash (week-over-week / period-over-period)
    prev_sales = Sale.objects.filter(sold_at__gte=prev_dt_start, sold_at__lt=prev_dt_end)
    if branch_id and selected_branch:
        prev_sales = prev_sales.filter(branch=selected_branch)
    prev_totals = prev_sales.aggregate(
        revenue=Sum(revenue_expr), qty=Sum('quantity'), n=Count('id'),
    )
    prev_revenue = prev_totals['revenue'] or 0
    prev_qty = prev_totals['qty'] or 0
    prev_n = prev_totals['n'] or 0

    def growth(curr, prev):
        if not prev:
            return None
        return float((curr - prev) / prev * 100)

    rev_growth = growth(float(revenue), float(prev_revenue))
    qty_growth = growth(qty, prev_qty)
    sales_growth = growth(sales_count, prev_n)

    # Filiallar taqqoslash (current period) + P&L (qat'iy xarajatlardan keyin)
    period_fraction = days / 30.0
    branch_compare = []
    for br in branches_all:
        b_sales = sales.filter(branch=br)
        b_agg = b_sales.aggregate(
            rev=Sum(revenue_expr), cost=Sum(cost_expr),
            qty=Sum('quantity'), n=Count('id'),
        )
        _rr, _cc = _ret_by_branch2.get(br.id) or (Decimal('0'), Decimal('0'))
        b_rev = _net_rev(b_agg['rev'], _od_by_branch.get(br.id)) - float(_rr)
        b_cost = float(_dec(b_agg['cost'] or 0) - _cc)
        b_profit = b_rev - b_cost
        b_n = b_agg['n'] or 0
        b_qty = b_agg['qty'] or 0
        b_stock = BranchStock.objects.filter(branch=br).aggregate(
            s=Sum('stock_count'))['s'] or 0
        rent_period = float(br.monthly_rent or 0) * period_fraction
        other_period = float(br.monthly_other_costs or 0) * period_fraction
        fixed_period = rent_period + other_period
        net_profit = b_profit - fixed_period
        branch_compare.append({
            'branch': br,
            'revenue': b_rev,
            'cost': b_cost,
            'profit': b_profit,  # gross (foyda tannarxdan keyin)
            'fixed_period': fixed_period,
            'rent_period': rent_period,
            'other_period': other_period,
            'net_profit': net_profit,
            'net_margin': (net_profit / b_rev * 100) if b_rev else 0,
            'margin': (b_profit / b_rev * 100) if b_rev else 0,
            'sales_count': b_n,
            'qty': b_qty,
            'avg_sale': (b_rev / b_n) if b_n else 0,
            'stock': b_stock,
        })

    # Tovar aylanmasi: kunlik o'rtacha sotuv va omborda qancha kun yetadi
    # Top sotilgan mahsulotlar uchun
    turnover = []
    for pid, pdata in sorted(profit_by_product.items(),
                             key=lambda x: x[1]['qty'], reverse=True)[:10]:
        daily_avg = pdata['qty'] / days if days else 0
        current_stock = BranchStock.objects.filter(
            variant__product_id=pid
        ).aggregate(s=Sum('stock_count'))['s'] or 0
        days_left = (current_stock / daily_avg) if daily_avg else None
        turnover.append({
            'code': pdata['code'], 'name': pdata['name'],
            'daily_avg': daily_avg, 'stock': current_stock,
            'days_left': days_left,
        })

    # TOP 10 mahsulot (qty va daromad bo'yicha)
    top_products = list(sales.values(
        'variant__product__id', 'variant__product__code',
        'variant__product__name',
    ).annotate(
        qty=Sum('quantity'),
        revenue=Sum(revenue_expr),
        n_sales=Count('id'),
    ).order_by('-revenue')[:10])
    for _tp in top_products:      # RET-1
        _rr, _ = (_ret_by_prod.get(_tp['variant__product__id'])
                  or (Decimal('0'), Decimal('0')))
        _tp['revenue'] = _dec(_tp['revenue'] or 0) - _rr
    top_products.sort(key=lambda r: -r['revenue'])

    # SLOW MOVERS — bu davrda umuman sotilmagan mahsulotlar
    sold_product_ids = sales.values_list('variant__product_id', flat=True).distinct()
    slow_products = list(
        Product.objects.exclude(id__in=sold_product_ids)
        .select_related('category')
        .annotate(stock=Sum('variants__branch_stocks__stock_count'))[:15]
    )

    # TOP sotuvchilar + komissiya
    top_sellers = list(sales.values(
        'sold_by__id', 'sold_by__username',
        'sold_by__first_name', 'sold_by__last_name',
        'sold_by__branch__name',
        'sold_by__commission_percent',
    ).annotate(
        revenue=Sum(revenue_expr),
        qty=Sum('quantity'),
        n_sales=Count('id'),
    ).order_by('-revenue')[:10])
    for s in top_sellers:
        pct = float(s.get('sold_by__commission_percent') or 0)
        # SOTUVCHI — chek darajasidagi o'lchov. Komissiya ham shu sof
        # summadan hisoblanadi (aks holda kassir chegirma qilib bergan
        # puldan ham foiz olardi).
        _rr, _ = (_ret_by_seller2.get(s.get('sold_by__id'))
                  or (Decimal('0'), Decimal('0')))
        rev = (_net_rev(s.get('revenue'), _od_by_seller.get(s.get('sold_by__id')))
               - float(_rr))
        s['revenue'] = rev
        s['commission_percent'] = pct
        s['commission'] = rev * pct / 100 if pct else 0

    # Filiallar bo'yicha (donut/bar uchun)
    by_branch = list(sales.values(
        'branch__id', 'branch__name'
    ).annotate(
        revenue=Sum(revenue_expr),
        qty=Sum('quantity'),
        n_sales=Count('id'),
    ).order_by('-revenue'))
    for _b in by_branch:      # FILIAL — chek darajasidagi o'lchov
        _rr, _ = (_ret_by_branch2.get(_b['branch__id'])
                  or (Decimal('0'), Decimal('0')))
        _b['revenue'] = (_net_rev(_b['revenue'],
                                  _od_by_branch.get(_b['branch__id'])) - float(_rr))
    by_branch.sort(key=lambda r: -r['revenue'])

    # Kategoriya bo'yicha
    by_category = list(sales.values(
        'variant__product__category__name',
    ).annotate(
        revenue=Sum(revenue_expr),
        qty=Sum('quantity'),
    ).order_by('-revenue'))
    for _r in by_category:        # RET-1
        _rr, _ = (_ret_by_catname.get(_r['variant__product__category__name'])
                  or (Decimal('0'), Decimal('0')))
        _r['revenue'] = _dec(_r['revenue'] or 0) - _rr
    by_category.sort(key=lambda r: -r['revenue'])

    # Bo'lim (4 katta guruh) bo'yicha
    by_group = list(sales.values(
        'variant__product__category__group__name',
    ).annotate(
        revenue=Sum(revenue_expr),
        qty=Sum('quantity'),
    ).order_by('-revenue'))
    for r in by_group:            # RET-1 (nomlar o'zgartirilishidan OLDIN)
        _rr, _ = (_ret_by_groupname.get(r['variant__product__category__group__name'])
                  or (Decimal('0'), Decimal('0')))
        r['revenue'] = _dec(r['revenue'] or 0) - _rr
        if not r['variant__product__category__group__name']:
            r['variant__product__category__group__name'] = "Bo'limsiz"
    by_group.sort(key=lambda r: -r['revenue'])

    # Ierarxiya daraxti: Bo'lim -> Kategoriya (daromad bo'yicha)
    _tree_rows = list(sales.values(
        'variant__product__category__group__name',
        'variant__product__category__group__sort_order',
        'variant__product__category__id',
        'variant__product__category__name',
    ).annotate(revenue=Sum(revenue_expr), qty=Sum('quantity')))
    _grand = sum(float(r['revenue'] or 0) for r in _tree_rows) or 1
    _tree = {}
    for r in _tree_rows:
        gname = r['variant__product__category__group__name'] or "Bo'limsiz"
        gorder = r['variant__product__category__group__sort_order']
        gorder = 99 if gorder is None else gorder
        node = _tree.setdefault(gname, {
            'name': gname, 'order': gorder, 'revenue': 0.0, 'qty': 0, 'cats': []})
        _rr, _ = (_ret_by_catname.get(r['variant__product__category__name'])
                  or (Decimal('0'), Decimal('0')))
        rev = float(_dec(r['revenue'] or 0) - _rr)   # RET-1
        node['revenue'] += rev
        node['qty'] += r['qty'] or 0
        node['cats'].append({
            'id': r['variant__product__category__id'],
            'name': r['variant__product__category__name'] or 'Kategoriyasiz',
            'revenue': rev, 'qty': r['qty'] or 0})
    group_tree = sorted(_tree.values(), key=lambda x: (x['order'], -x['revenue']))
    for node in group_tree:
        node['share'] = node['revenue'] / _grand * 100
        node['cats'].sort(key=lambda c: -c['revenue'])
        cmax = max((c['revenue'] for c in node['cats']), default=0) or 1
        for c in node['cats']:
            c['bar'] = c['revenue'] / cmax * 100
    _color_tree(group_tree)

    # Sotuv trendi (kunlar bo'yicha)
    daily_map = {}
    for i in range(days):
        d = d_start + timedelta(days=i)
        daily_map[d.isoformat()] = {'qty': 0, 'revenue': 0}
    for s in sales:
        day_key = timezone.localtime(s.sold_at).date().isoformat()
        if day_key in daily_map:
            daily_map[day_key]['qty'] += s.quantity
            daily_map[day_key]['revenue'] += float(s.quantity * s.sale_price)
    daily_labels = list(daily_map.keys())
    daily_revenue = [daily_map[k]['revenue'] for k in daily_labels]
    daily_qty = [daily_map[k]['qty'] for k in daily_labels]

    # Soatlar bo'yicha (0-23)
    hour_buckets = [0] * 24
    for s in sales:
        h = timezone.localtime(s.sold_at).hour
        hour_buckets[h] += s.quantity

    # Hafta kunlari bo'yicha
    weekday_labels = ['Du', 'Se', 'Ch', 'Pa', 'Ju', 'Sh', 'Ya']
    weekday_buckets = [0] * 7
    for s in sales:
        wd = timezone.localtime(s.sold_at).weekday()  # 0=Monday
        weekday_buckets[wd] += s.quantity

    # Ombor risk: sof aksiya, narx
    out_of_stock = list(
        BranchStock.objects.filter(stock_count=0)
        .select_related('variant__product', 'branch')[:20]
    )

    # Qaytarilishlar: davr ichida eng ko'p qaytarilgan mahsulotlar.
    # Sifat muammosi indikatori — agar % returned high bo'lsa, mahsulot
    # tekshirilishi kerak (nuqson, o'lcham xato, va h.k.).
    returns_qs = Return.objects.filter(
        refunded_at__gte=dt_start, refunded_at__lt=dt_end,
    )
    if branch_id and selected_branch:
        returns_qs = returns_qs.filter(sale__branch=selected_branch)
    returns_total_qty = returns_qs.aggregate(s=Sum('quantity'))['s'] or 0
    returns_total_amount = returns_qs.aggregate(
        s=Sum(ExpressionWrapper(
            F('quantity') * F('sale__sale_price'),
            output_field=DecimalField(max_digits=14, decimal_places=2),
        ))
    )['s'] or 0
    return_rate = (returns_total_qty / qty * 100) if qty else 0

    top_returns_qs = (returns_qs
        .values('sale__variant__product_id',
                'sale__variant__product__code',
                'sale__variant__product__name')
        .annotate(qty_returned=Sum('quantity'))
        .order_by('-qty_returned')[:10])
    top_returns = []
    for row in top_returns_qs:
        pid = row['sale__variant__product_id']
        sold = profit_by_product.get(pid, {}).get('qty', 0)
        returned = row['qty_returned']
        rate = (returned / sold * 100) if sold else 0
        # Top reasons for this product
        reasons = (returns_qs.filter(sale__variant__product_id=pid)
                   .exclude(reason='')
                   .values('reason')
                   .annotate(n=Count('id'))
                   .order_by('-n')[:3])
        top_returns.append({
            'code': row['sale__variant__product__code'],
            'name': row['sale__variant__product__name'],
            'returned': returned,
            'sold': sold,
            'rate': rate,
            'reasons': [r['reason'] for r in reasons],
        })

    # Foyda yuqori mahsulotlar (markup bo'yicha eng yuqori)
    high_margin = list(
        Product.objects.order_by('-markup_percent')[:8]
    )

    context = {
        'period': period,
        'period_label': period_label,
        'd_start': d_start,
        'd_end': end_date,
        'days': days,
        'discount_total': discount_total,
        # SAL-6: MAHSULOT/KATEGORIYA/GURUH ro'yxatlari O'Z narxida — bu son
        # ular ostiga "Chek chegirmasi −X" qatori bo'lib tushadi, shunda
        # ro'yxat + chegirma = sarlavhadagi "Tushum".
        'order_discount_total': float(order_disc),
        'discount_split': _disc_split,   # DISC-1
        'selected_branch': selected_branch,
        'branches_all': branches_all,
        'revenue': revenue,
        'qty': qty,
        'sales_count': sales_count,
        'total_cost': total_cost,
        'profit': profit,
        'margin': margin,
        'avg_sale_value': avg_sale_value,
        'largest_sale': largest_sale,
        'prev_revenue': prev_revenue,
        'prev_qty': prev_qty,
        'prev_n': prev_n,
        'rev_growth': rev_growth,
        'qty_growth': qty_growth,
        'sales_growth': sales_growth,
        'top_products': top_products,
        'top_profit': top_profit,
        'slow_products': slow_products,
        'top_sellers': top_sellers,
        'by_branch': by_branch,
        'by_category': by_category,
        'by_group': by_group,
        'group_tree': group_tree,
        'branch_compare': branch_compare,
        'turnover': turnover,
        'daily_labels': daily_labels,
        'daily_revenue': daily_revenue,
        'daily_qty': daily_qty,
        'hour_buckets': hour_buckets,
        'weekday_labels': weekday_labels,
        'weekday_buckets': weekday_buckets,
        'out_of_stock': out_of_stock,
        'high_margin': high_margin,
        'returns_total_qty': returns_total_qty,
        'returns_total_amount': returns_total_amount,
        'return_rate': return_rate,
        'top_returns': top_returns,
    }

    # ===== TIER D — Advanced BI =====

    # D1: Sales forecast — 90-day daily revenue → linear regression → next 30 days
    fc_start = today - timedelta(days=90)
    fc_dt_start = datetime.combine(fc_start, datetime.min.time()).replace(tzinfo=tz)
    fc_sales = Sale.objects.filter(sold_at__gte=fc_dt_start, sold_at__lt=dt_end)
    if branch_id and selected_branch:
        fc_sales = fc_sales.filter(branch=selected_branch)
    daily_rev = {}
    for i in range(91):  # 0..90 days back
        d = today - timedelta(days=90 - i)
        daily_rev[d.isoformat()] = 0.0
    for row in fc_sales.values('sold_at', 'quantity', 'sale_price', 'line_discount'):
        d = row['sold_at'].astimezone(tz).date().isoformat()
        if d in daily_rev:
            daily_rev[d] += (float(row['quantity']) * float(row['sale_price'])
                             - float(row['line_discount'] or 0))
    # Simple linear regression: y = a + b*x where x is day index (0..90)
    xs = list(range(len(daily_rev)))
    ys = list(daily_rev.values())
    n = len(xs)
    if n > 1 and sum(ys) > 0:
        sx = sum(xs); sy = sum(ys)
        sxy = sum(x * y for x, y in zip(xs, ys))
        sxx = sum(x * x for x in xs)
        denom = (n * sxx - sx * sx)
        if denom != 0:
            slope = (n * sxy - sx * sy) / denom
            intercept = (sy - slope * sx) / n
        else:
            slope = 0; intercept = sy / n if n else 0
    else:
        slope = 0; intercept = 0
    # Build 90-actual + 30-predicted series
    actual_labels = []
    actual_values = []
    forecast_labels = []
    forecast_values = []
    for i, (date_iso, v) in enumerate(daily_rev.items()):
        actual_labels.append(date_iso[5:])
        actual_values.append(round(v))
    for j in range(1, 31):  # next 30 days
        future_date = today + timedelta(days=j)
        forecast_labels.append(future_date.isoformat()[5:])
        x_future = (n - 1) + j  # extending the index
        predicted = max(0, intercept + slope * x_future)
        forecast_values.append(round(predicted))
    total_forecast_30d = sum(forecast_values)
    context['fc_actual_labels'] = actual_labels
    context['fc_actual_values'] = actual_values
    context['fc_forecast_labels'] = forecast_labels
    context['fc_forecast_values'] = forecast_values
    context['fc_slope'] = slope
    context['fc_total_30d'] = total_forecast_30d
    context['fc_trend_direction'] = 'up' if slope > 0 else ('down' if slope < 0 else 'flat')

    # D2: Customer cohort retention (6 months)
    from collections import defaultdict
    cohort_first = {}  # customer_id -> first purchase month (YYYY-MM)
    cohort_activity = defaultdict(set)  # (cohort_month, activity_month) -> customers
    for row in (SaleTransaction.objects
                .filter(customer__isnull=False,
                        sold_at__gte=today - timedelta(days=180))
                .values('customer_id', 'sold_at')):
        cid = row['customer_id']
        m = row['sold_at'].astimezone(tz).strftime('%Y-%m')
        if cid not in cohort_first or m < cohort_first[cid]:
            cohort_first[cid] = m
        cohort_activity[(cohort_first.get(cid, m), m)].add(cid)
    # We need to re-pass since cohort_first might have updated
    cohort_first = {}
    for row in (SaleTransaction.objects
                .filter(customer__isnull=False,
                        sold_at__gte=today - timedelta(days=180))
                .order_by('sold_at')
                .values('customer_id', 'sold_at')):
        cid = row['customer_id']
        if cid not in cohort_first:
            cohort_first[cid] = row['sold_at'].astimezone(tz).strftime('%Y-%m')
    # Build cohort matrix
    cohort_activity = defaultdict(set)
    for row in (SaleTransaction.objects
                .filter(customer__isnull=False,
                        sold_at__gte=today - timedelta(days=180))
                .values('customer_id', 'sold_at')):
        cid = row['customer_id']
        if cid not in cohort_first:
            continue
        m = row['sold_at'].astimezone(tz).strftime('%Y-%m')
        cohort_activity[(cohort_first[cid], m)].add(cid)
    # List of 6 months back
    cohort_months = []
    cur_month_d = today.replace(day=1)
    for _ in range(6):
        cohort_months.append(cur_month_d.strftime('%Y-%m'))
        if cur_month_d.month == 1:
            cur_month_d = cur_month_d.replace(year=cur_month_d.year - 1, month=12)
        else:
            cur_month_d = cur_month_d.replace(month=cur_month_d.month - 1)
    cohort_months.reverse()  # oldest first
    cohort_rows = []
    for cohort in cohort_months:
        cohort_size = sum(1 for cid, m in cohort_first.items() if m == cohort)
        if cohort_size == 0:
            cohort_rows.append({'cohort': cohort, 'size': 0, 'cells': []})
            continue
        cells = []
        for activity in cohort_months:
            if activity < cohort:
                cells.append(None)
                continue
            active = len(cohort_activity.get((cohort, activity), set()))
            pct = (active / cohort_size * 100) if cohort_size else 0
            cells.append({'count': active, 'pct': round(pct, 1)})
        cohort_rows.append({'cohort': cohort, 'size': cohort_size, 'cells': cells})
    context['cohort_months'] = cohort_months
    context['cohort_rows'] = cohort_rows

    # ===== BI EXTENSIONS =====

    # Hour × Weekday heatmap (revenue grid 7×24)
    heatmap = [[0 for _ in range(24)] for _ in range(7)]  # 0=Mon
    for sold_at, qty, price, disc in sales.values_list('sold_at', 'quantity', 'sale_price', 'line_discount'):
        local_dt = timezone.localtime(sold_at)
        wd = local_dt.weekday()
        h = local_dt.hour
        heatmap[wd][h] += float(qty) * float(price) - float(disc or 0)
    flat_heatmap = []
    for wd in range(7):
        for h in range(24):
            flat_heatmap.append({'x': h, 'y': wd, 'v': heatmap[wd][h]})
    max_heat = max((max(row) for row in heatmap), default=0) or 1
    context['heatmap'] = heatmap
    context['heatmap_max'] = max_heat
    context['heatmap_hour_labels'] = list(range(24))
    context['heatmap_day_labels'] = ['Du', 'Se', 'Ch', 'Pa', 'Ju', 'Sh', 'Ya']

    # Year-over-year: same date range last year
    last_year_start = dt_start.replace(year=dt_start.year - 1)
    last_year_end = dt_end.replace(year=dt_end.year - 1)
    ly_sales = Sale.objects.filter(sold_at__gte=last_year_start, sold_at__lt=last_year_end)
    if branch_id and selected_branch:
        ly_sales = ly_sales.filter(branch=selected_branch)
    ly_agg = ly_sales.aggregate(
        revenue=Sum(revenue_expr), qty=Sum('quantity'),
        txns=Count('transaction', distinct=True),
    )
    ly_revenue = float(ly_agg['revenue'] or 0)
    ly_qty = ly_agg['qty'] or 0
    ly_txns = ly_agg['txns'] or 0
    def _pct(now, prev):
        if not prev: return None
        return (float(now) - float(prev)) / float(prev) * 100
    context['ly_revenue'] = ly_revenue
    context['ly_qty'] = ly_qty
    context['ly_txns'] = ly_txns
    context['yoy_revenue_delta'] = _pct(revenue, ly_revenue)
    context['yoy_qty_delta'] = _pct(qty, ly_qty)
    context['yoy_txns_delta'] = _pct(sales_count, ly_txns)

    # ABC analysis: sort by revenue, classify by cumulative %.
    # L10 fix: thresholds configurable via query string (?abc_a=80&abc_b=95)
    try:
        abc_a_threshold = float(request.GET.get('abc_a') or 80)
        abc_b_threshold = float(request.GET.get('abc_b') or 95)
    except (TypeError, ValueError):
        abc_a_threshold, abc_b_threshold = 80, 95
    abc_a_threshold = max(1, min(99, abc_a_threshold))
    abc_b_threshold = max(abc_a_threshold + 1, min(99.5, abc_b_threshold))

    sorted_products = sorted(profit_by_product.values(),
                             key=lambda x: float(x['revenue'] or 0), reverse=True)
    total_rev = sum(float(p['revenue'] or 0) for p in sorted_products) or 1
    abc_summary = {'A': 0, 'B': 0, 'C': 0}
    abc_count = {'A': 0, 'B': 0, 'C': 0}
    cumulative = 0
    for p in sorted_products:
        pct = float(p['revenue'] or 0) / total_rev * 100
        cumulative += pct
        if cumulative <= abc_a_threshold:
            cls = 'A'
        elif cumulative <= abc_b_threshold:
            cls = 'B'
        else:
            cls = 'C'
        p['abc'] = cls
        abc_summary[cls] += float(p['revenue'] or 0)
        abc_count[cls] += 1
    context['abc_a_threshold'] = abc_a_threshold
    context['abc_b_threshold'] = abc_b_threshold
    context['abc_top_a'] = [p for p in sorted_products if p['abc'] == 'A'][:8]
    context['abc_summary'] = abc_summary
    context['abc_count'] = abc_count

    # Dead stock: products with zero sales in last 90 days but stock > 0
    cutoff_90d = timezone.now() - timedelta(days=90)
    recently_sold_ids = set(Sale.objects.filter(sold_at__gte=cutoff_90d)
                            .values_list('variant__product_id', flat=True).distinct())
    has_stock = (BranchStock.objects.filter(stock_count__gt=0)
                 .values_list('variant__product_id', flat=True).distinct())
    dead_ids = set(has_stock) - recently_sold_ids
    dead_products = list(Product.objects
                         .filter(id__in=list(dead_ids))
                         .annotate(stock=Sum('variants__branch_stocks__stock_count'),
                                   value=Sum(ExpressionWrapper(
                                       F('variants__branch_stocks__stock_count')
                                       * F('variants__branch_stocks__cost_price'),
                                       output_field=DecimalField(max_digits=14, decimal_places=2))))
                         .order_by('-value')[:10])
    context['dead_products'] = dead_products
    context['dead_count'] = len(dead_ids)

    # Margin alerts: products where last sale was at/below cost
    margin_alerts = []
    bad_sales = (sales.filter(sale_price__lte=F('cost_at_sale'))
                 .select_related('variant__product', 'branch')
                 .order_by('-sold_at')[:5])
    for s in bad_sales:
        margin_alerts.append({
            'product': s.variant.product,
            'branch': s.branch.name,
            'date': s.sold_at,
            'sale_price': float(s.sale_price),
            'cost': float(s.cost_at_sale),
            'loss': (float(s.cost_at_sale) - float(s.sale_price)) * s.quantity,
        })
    context['margin_alerts'] = margin_alerts

    return context


@admin_required
def insights(request):
    ctx = _insights_context(request)
    fmt = request.GET.get('export', '').lower()
    if fmt == 'csv':
        return _insights_csv(ctx)
    if fmt == 'pdf':
        return _insights_pdf(ctx)
    return render(request, 'inventory/insights.html', ctx)


def _insights_csv(ctx):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    fn = f"Biznes_tahlili_{ctx['d_start']}_{ctx['d_end']}.csv"
    response['Content-Disposition'] = f'attachment; filename="{fn}"'
    response.write('﻿')
    w = csv.writer(response)
    branch_str = ctx['selected_branch'].name if ctx['selected_branch'] else 'Barcha filiallar'
    w.writerow([f"yurit — Biznes tahlili"])
    w.writerow([f"Davr: {ctx['d_start']} — {ctx['d_end']}  ({ctx['days']} kun)"])
    w.writerow([f"Filial: {branch_str}"])
    w.writerow([])

    w.writerow(['ASOSIY KO\'RSATKICHLAR'])
    w.writerow(["Daromad (so'm)", ctx['revenue']])
    w.writerow(["Tannarx (so'm)", ctx['total_cost']])
    w.writerow(["Sof foyda (so'm)", ctx['profit']])
    w.writerow(['Foyda marjasi (%)', f"{ctx['margin']:.2f}"])
    w.writerow(['Sotuvlar soni', ctx['sales_count']])
    w.writerow(['Sotilgan dona', ctx['qty']])
    w.writerow(["O'rtacha sotuv summasi (so'm)", round(ctx['avg_sale_value'])])
    if ctx['largest_sale']:
        w.writerow(["Eng katta bitta sotuv (so'm)", ctx['largest_sale'].line])
    w.writerow([])

    w.writerow(["OLDINGI DAVR BILAN TAQQOSLASH"])
    w.writerow(['Ko\'rsatkich', 'Hozir', 'Oldin', "O'sish %"])
    g = lambda v: f"{v:+.1f}%" if v is not None else '—'
    w.writerow(['Daromad', ctx['revenue'], ctx['prev_revenue'], g(ctx['rev_growth'])])
    w.writerow(['Dona', ctx['qty'], ctx['prev_qty'], g(ctx['qty_growth'])])
    w.writerow(['Sotuvlar', ctx['sales_count'], ctx['prev_n'], g(ctx['sales_growth'])])
    w.writerow([])

    w.writerow([f"TOP-10 MAHSULOT (DAROMAD BO'YICHA)"])
    w.writerow(['Kod', 'Mahsulot', 'Dona', 'Sotuvlar', "Daromad (so'm)"])
    for p in ctx['top_products']:
        w.writerow([_csv_safe(c) for c in [   # SEC-20
                    p['variant__product__code'], p['variant__product__name'],
                    p['qty'], p['n_sales'], p['revenue']]])
    w.writerow([])

    w.writerow([f"TOP-10 MAHSULOT (SOF FOYDA BO'YICHA)"])
    w.writerow(['Kod', 'Mahsulot', 'Dona', "Daromad", "Tannarx", "Foyda", 'Marja %'])
    for p in ctx['top_profit']:
        w.writerow([_csv_safe(c) for c in [   # SEC-20
                    p['code'], p['name'], p['qty'], p['revenue'],
                    p['cost'], p['profit'], f"{p['margin']:.1f}"]])
    w.writerow([])

    w.writerow(["SOTILMAYOTGAN MAHSULOTLAR"])
    w.writerow(['Kod', 'Mahsulot', 'Kategoriya', 'Omborda dona'])
    for p in ctx['slow_products']:
        w.writerow([_csv_safe(c) for c in [   # SEC-20
                    p.code, p.name,
                    p.category.name if p.category else '—',
                    p.stock or 0]])
    w.writerow([])

    w.writerow(["FILIALLAR TAQQOSLASH"])
    w.writerow(['Filial', "Daromad", "Tannarx", "Foyda", 'Marja %',
                'Sotuvlar', 'Dona', "O'rtacha sotuv", 'Omborda dona'])
    for b in ctx['branch_compare']:
        w.writerow([_csv_safe(c) for c in [   # SEC-20
                    b['branch'].name, b['revenue'], b['cost'], b['profit'],
                    f"{b['margin']:.1f}", b['sales_count'], b['qty'],
                    round(b['avg_sale']), b['stock']]])
    w.writerow([])

    w.writerow(["TOP-10 SOTUVCHILAR"])
    w.writerow(['Username', 'Ism', 'Filial', 'Sotuvlar', 'Dona', "Daromad (so'm)"])
    for s in ctx['top_sellers']:
        w.writerow([_csv_safe(c) for c in [   # SEC-20
            s['sold_by__username'],
            f"{s['sold_by__first_name'] or ''} {s['sold_by__last_name'] or ''}".strip(),
            s['sold_by__branch__name'] or '—',
            s['n_sales'], s['qty'], s['revenue'],
        ]])
    w.writerow([])

    w.writerow(["TOVAR AYLANMASI (Top sotilganlar)"])
    w.writerow(['Kod', 'Mahsulot', 'Kunlik o\'rtacha (dona)', 'Hozir omborda', 'Necha kun yetadi'])
    for t in ctx['turnover']:
        dleft = f"{t['days_left']:.0f}" if t['days_left'] is not None else '—'
        w.writerow([_csv_safe(c) for c in [   # SEC-20
                    t['code'], t['name'], f"{t['daily_avg']:.2f}", t['stock'], dleft]])
    w.writerow([])

    w.writerow(["OMBORDA TUGAGAN"])
    w.writerow(['Filial', 'Kod', "O'lcham", 'Rang'])
    for o in ctx['out_of_stock']:
        w.writerow([_csv_safe(c) for c in [   # SEC-20
                    o.branch.name, o.variant.product.code,
                    o.variant.size, o.variant.color]])
    return response


def _insights_pdf(ctx):
    response = HttpResponse(content_type='application/pdf')
    fn = f"Biznes_tahlili_{ctx['d_start']}_{ctx['d_end']}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{fn}"'
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=15*mm, bottomMargin=15*mm,
        title="Biznes tahlili",
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle('H1', parent=styles['Title'], fontSize=18, spaceAfter=2, alignment=0)
    h2 = ParagraphStyle('H2', parent=styles['Heading3'], fontSize=12,
                        spaceBefore=8, spaceAfter=4, textColor=colors.HexColor('#0d6efd'))
    meta = ParagraphStyle('M', parent=styles['Normal'], fontSize=9,
                          textColor=colors.HexColor('#666'))

    branch_str = ctx['selected_branch'].name if ctx['selected_branch'] else 'Barcha filiallar'
    elements = [
        Paragraph("yurit — Biznes tahlili", h1),
        Paragraph(f"Davr: <b>{ctx['d_start']} — {ctx['d_end']}</b> ({ctx['days']} kun) &nbsp; | &nbsp; "
                  f"Filial: <b>{branch_str}</b> &nbsp; | &nbsp; "
                  f"Yaratildi: {timezone.localtime().strftime('%Y-%m-%d %H:%M')}", meta),
        Spacer(1, 4*mm),
    ]

    def header_style(t):
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#bbb')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.white, colors.HexColor('#f4f6fa')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))
        return t

    # KPI table
    elements.append(Paragraph("Asosiy ko'rsatkichlar", h2))
    kpi = [
        ['Ko\'rsatkich', 'Qiymat', 'Oldingi davr', "O'sish %"],
        ["Daromad (so'm)", _fmt(ctx['revenue']), _fmt(ctx['prev_revenue']),
         f"{ctx['rev_growth']:+.1f}%" if ctx['rev_growth'] is not None else '—'],
        ["Tannarx (so'm)", _fmt(ctx['total_cost']), '—', '—'],
        ["Sof foyda (so'm)", _fmt(ctx['profit']), '—', '—'],
        ['Foyda marjasi (%)', f"{ctx['margin']:.2f}%", '—', '—'],
        ['Sotuvlar', _fmt(ctx['sales_count']), _fmt(ctx['prev_n']),
         f"{ctx['sales_growth']:+.1f}%" if ctx['sales_growth'] is not None else '—'],
        ['Dona', _fmt(ctx['qty']), _fmt(ctx['prev_qty']),
         f"{ctx['qty_growth']:+.1f}%" if ctx['qty_growth'] is not None else '—'],
        ["O'rtacha sotuv (so'm)", _fmt(ctx['avg_sale_value']), '—', '—'],
    ]
    elements.append(header_style(Table(kpi, repeatRows=1,
                                       colWidths=[55*mm, 40*mm, 40*mm, 35*mm])))

    # Filial taqqoslash
    if ctx['branch_compare']:
        elements.append(Paragraph("Filiallar taqqoslash", h2))
        data = [['Filial', "Daromad", "Foyda", 'Marja', 'Sotuvlar', 'Dona', 'Omborda']]
        for b in ctx['branch_compare']:
            data.append([
                b['branch'].name, _fmt(b['revenue']), _fmt(b['profit']),
                f"{b['margin']:.1f}%", _fmt(b['sales_count']),
                _fmt(b['qty']), _fmt(b['stock']),
            ])
        elements.append(header_style(Table(data, repeatRows=1)))

    # Top by revenue
    elements.append(Paragraph("TOP-10 mahsulot (daromad bo'yicha)", h2))
    data = [['#', 'Kod', 'Mahsulot', 'Dona', 'Sotuvlar', "Daromad"]]
    for i, p in enumerate(ctx['top_products'], 1):
        data.append([str(i), p['variant__product__code'],
                     p['variant__product__name'][:40],
                     _fmt(p['qty']), _fmt(p['n_sales']), _fmt(p['revenue'])])
    elements.append(header_style(Table(data, repeatRows=1)))

    # Top by profit
    elements.append(Paragraph("TOP-10 mahsulot (sof foyda bo'yicha)", h2))
    data = [['#', 'Kod', 'Mahsulot', "Daromad", "Foyda", 'Marja']]
    for i, p in enumerate(ctx['top_profit'], 1):
        data.append([str(i), p['code'], p['name'][:40],
                     _fmt(p['revenue']), _fmt(p['profit']),
                     f"{p['margin']:.1f}%"])
    elements.append(header_style(Table(data, repeatRows=1)))

    # Slow movers
    if ctx['slow_products']:
        elements.append(Paragraph("Sotilmayotgan mahsulotlar", h2))
        data = [['Kod', 'Mahsulot', 'Kategoriya', 'Omborda']]
        for p in ctx['slow_products']:
            data.append([p.code, p.name[:40],
                         p.category.name if p.category else '—',
                         _fmt(p.stock or 0)])
        elements.append(header_style(Table(data, repeatRows=1)))

    # Top sellers
    elements.append(Paragraph("Top-10 sotuvchilar", h2))
    data = [['#', 'Username', 'Filial', 'Sotuvlar', 'Dona', "Daromad"]]
    for i, s in enumerate(ctx['top_sellers'], 1):
        data.append([str(i), s['sold_by__username'],
                     s['sold_by__branch__name'] or '—',
                     _fmt(s['n_sales']), _fmt(s['qty']), _fmt(s['revenue'])])
    elements.append(header_style(Table(data, repeatRows=1)))

    # Turnover
    if ctx['turnover']:
        elements.append(Paragraph("Tovar aylanmasi", h2))
        data = [['Kod', 'Mahsulot', 'Kunlik o\'rtacha', 'Omborda', 'Necha kun yetadi']]
        for t in ctx['turnover']:
            dleft = f"{t['days_left']:.0f} kun" if t['days_left'] is not None else '—'
            data.append([t['code'], t['name'][:40],
                         f"{t['daily_avg']:.2f}", _fmt(t['stock']), dleft])
        elements.append(header_style(Table(data, repeatRows=1)))

    def _footer(canvas, doc_obj):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#999'))
        canvas.drawRightString(doc_obj.pagesize[0] - 15*mm, 8*mm,
                               f"yurit — bet {doc_obj.page}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)
    response.write(buf.getvalue())
    buf.close()
    return response


def _csv_response(title, headers, rows, summary, d_start, d_end, branch):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    fn = f"{title.replace(' ', '_')}_{d_start}_{d_end}.csv"
    response['Content-Disposition'] = f'attachment; filename="{fn}"'
    response.write('﻿')  # BOM for Excel
    writer = csv.writer(response)
    writer.writerow([title])
    writer.writerow([f"Davr: {d_start} — {d_end}"])
    if branch:
        writer.writerow([f"Filial: {branch.name}"])
    else:
        writer.writerow(["Filial: Barchasi"])
    writer.writerow([])
    writer.writerow(headers)
    for r in rows:
        writer.writerow([_csv_safe(c) for c in r])   # SEC-20
    writer.writerow([])
    writer.writerow(['Xulosa:'])
    for k, v in summary.items():
        writer.writerow([_csv_safe(k), _csv_safe(v)])   # SEC-20
    return response


def _fmt(v):
    """PDF jadval uchun qiymatni formatlash."""
    if v is None:
        return ''
    if hasattr(v, 'quantize'):  # Decimal
        return f"{v:,.0f}".replace(',', ' ')
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return f"{v:,}".replace(',', ' ')
    return str(v)


def _pdf_response(title, headers, rows, summary, d_start, d_end, branch):
    response = HttpResponse(content_type='application/pdf')
    fn = f"{title.replace(' ', '_').replace(chr(39), '')}_{d_start}_{d_end}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{fn}"'

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=12*mm, rightMargin=12*mm,
        topMargin=12*mm, bottomMargin=12*mm,
        title=title,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleU', parent=styles['Title'], fontSize=16, spaceAfter=4, alignment=0,
    )
    meta_style = ParagraphStyle(
        'MetaU', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#555'),
    )

    elements = []
    elements.append(Paragraph(title, title_style))
    branch_str = branch.name if branch else "Barcha filiallar"
    elements.append(Paragraph(
        f"Davr: <b>{d_start} — {d_end}</b> &nbsp; | &nbsp; "
        f"Filial: <b>{branch_str}</b> &nbsp; | &nbsp; "
        f"Yaratildi: {timezone.localtime().strftime('%Y-%m-%d %H:%M')}",
        meta_style,
    ))
    elements.append(Spacer(1, 6*mm))

    data = [list(headers)]
    for row in rows:
        data.append([_fmt(c) for c in row])
    if len(data) == 1:
        data.append(['—'] + [''] * (len(headers) - 1))

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#bbbbbb')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [colors.white, colors.HexColor('#f4f6fa')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(table)

    if summary:
        elements.append(Spacer(1, 8*mm))
        elements.append(Paragraph("<b>Xulosa:</b>", styles['Heading4']))
        summary_data = [[k, _fmt(v)] for k, v in summary.items()]
        st = Table(summary_data, colWidths=[100*mm, 60*mm])
        st.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#bbbbbb')),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f4f6fa')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(st)

    def _page_footer(canvas, doc_obj):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#999'))
        canvas.drawRightString(
            doc_obj.pagesize[0] - 12*mm, 8*mm,
            f"yurit Ombor Boshqaruv  -  bet {doc_obj.page}"
        )
        canvas.restoreState()

    doc.build(elements, onFirstPage=_page_footer, onLaterPages=_page_footer)
    response.write(buf.getvalue())
    buf.close()
    return response


# ---------- MIJOZ SO'ROVLARI (customer product requests / demand log) ----------

def _request_branch(request):
    """Best-effort branch for a new request. Sellers → their branch;
    admins → POS-selected/assigned branch, else first active branch."""
    b = getattr(request.user, 'branch', None)
    if b:
        return b
    if request.user.is_admin():
        return Branch.objects.filter(is_active=True).order_by('name').first()
    return None


@login_required
def product_requests(request):
    """Talab jurnali: mijozlar so'ragan, lekin bizda yo'q mahsulotlar.

    Bir xil nom bo'yicha guruhlanadi — eng ko'p so'ralganlar tepada.
    """
    Status = ProductRequest.Status

    # --- Pending, grouped by name (case-insensitive via smart_title on input) ---
    pending = list(
        ProductRequest.objects.filter(status=Status.NEW)
        .select_related('requested_by')
        .order_by('-created_at')
    )
    groups = {}
    for r in pending:
        key = r.name.strip().lower()
        g = groups.get(key)
        if not g:
            g = groups[key] = {
                'name': r.name, 'count': 0,
                'phones': [], 'notes': [],
                'first': r.created_at, 'last': r.created_at,
            }
        g['count'] += 1
        g['last'] = max(g['last'], r.created_at)
        g['first'] = min(g['first'], r.created_at)
        if r.customer_phone and r.customer_phone not in g['phones']:
            g['phones'].append(r.customer_phone)
        if r.note and r.note not in g['notes']:
            g['notes'].append(r.note)
    pending_groups = sorted(
        groups.values(), key=lambda g: (-g['count'], -g['last'].timestamp())
    )

    # --- Recently resolved (brought in / dismissed) ---
    resolved_recent = list(
        ProductRequest.objects
        .filter(status__in=[Status.STOCKED, Status.DISMISSED])
        .select_related('resolved_by')
        .order_by('-resolved_at', '-created_at')[:25]
    )

    stats = {
        'pending_total': len(pending),
        'pending_items': len(pending_groups),
        'top_count': pending_groups[0]['count'] if pending_groups else 0,
    }

    return render(request, 'inventory/product_requests.html', {
        'pending_groups': pending_groups,
        'resolved_recent': resolved_recent,
        'stats': stats,
    })


@login_required
def product_request_add(request):
    """Yangi so'rov qo'shish. Sotuvchi ham, admin ham qo'sha oladi.
    `next` bo'lsa — o'sha sahifaga qaytadi (Qidiruv integratsiyasi)."""
    if request.method != 'POST':
        return redirect('product_requests')

    name = smart_title((request.POST.get('name') or '').strip())
    phone = (request.POST.get('customer_phone') or '').strip()
    note = (request.POST.get('note') or '').strip()
    nxt = (request.POST.get('next') or '').strip()

    if not name:
        messages.warning(request, "Mahsulot nomi bo'sh bo'lishi mumkin emas.")
        return redirect(nxt or 'product_requests')

    pr = ProductRequest.objects.create(
        name=name[:200],
        customer_phone=phone[:40],
        note=note[:255],
        branch=_request_branch(request),
        requested_by=request.user,
    )

    # How many times has this item been asked (still pending)?
    same = ProductRequest.objects.filter(
        name__iexact=name, status=ProductRequest.Status.NEW
    ).count()
    if same > 1:
        messages.success(
            request,
            f"\"{pr.name}\" so'rovlarga qo'shildi — bu mahsulot {same} marta so'raldi."
        )
    else:
        messages.success(request, f"\"{pr.name}\" so'rovlarga qo'shildi.")

    # Redirect back to a safe local page only
    if nxt.startswith('/'):
        return redirect(nxt)
    return redirect('product_requests')


@login_required
def product_request_resolve(request):
    """Nom bo'yicha barcha kutilayotgan so'rovlarni yopish:
    'stock' = keltirildi, 'dismiss' = rad etildi. Faqat admin."""
    if request.method != 'POST':
        return redirect('product_requests')
    if not request.user.is_admin():
        return HttpResponseForbidden("Faqat administrator uchun.")

    name = (request.POST.get('name') or '').strip()
    action = (request.POST.get('action') or '').strip()
    if not name:
        return redirect('product_requests')

    status_map = {
        'stock': ProductRequest.Status.STOCKED,
        'dismiss': ProductRequest.Status.DISMISSED,
    }
    new_status = status_map.get(action)
    if not new_status:
        messages.warning(request, "Noma'lum amal.")
        return redirect('product_requests')

    n = (ProductRequest.objects
         .filter(name__iexact=name, status=ProductRequest.Status.NEW)
         .update(status=new_status,
                 resolved_by=request.user,
                 resolved_at=timezone.now()))
    if action == 'stock':
        messages.success(request, f"\"{name}\" — keltirildi deb belgilandi ({n} ta so'rov).")
    else:
        messages.info(request, f"\"{name}\" — rad etildi ({n} ta so'rov).")
    return redirect('product_requests')


# ==================== NARXLAR (narx / marja / ulgurji) ====================

PRICE_PAGE_SIZE = 100

#: marja holati bo'yicha filtrlar — panel va ro'yxat shu kalitlarni ishlatadi
PRICE_ISSUES = [
    ('no_cost',   "Tannarx yo'q",      'Tannarx 0 — foyda hisoblanmaydi'),
    ('zero',      "Marja 0%",          'Sotuv narxi tannarxga teng'),
    ('loss',      "Zararga",           'Sotuv narxi tannarxdan past'),
    ('no_sale',   "Sotuv narxi yo'q",  'Sotuv narxi 0'),
    ('low',       "Marja past (<10%)", "Marja 10% dan kam"),
    ('no_ws',     "Ulgurji yo'q",      "Ulgurji narx qo'yilmagan"),
]


def _lock_stocks(qs):
    """STK-14: BranchStock qatorlarini JOIN'SIZ qulflaydi.

    Postgres FOR UPDATE ni outer join'ning NULLABLE tomoniga qo'llay olmaydi:

        psycopg.errors.FeatureNotSupported:
        FOR UPDATE cannot be applied to the nullable side of an outer join

    `_price_qs` 'variant__product__category' ni select_related qiladi,
    `Product.category` esa null=True — demak LEFT OUTER JOIN paydo bo'ladi va
    butun /prices/apply/ 500 beradi. SQLite select_for_update'ni UMUMAN
    e'tiborsiz qoldiradi, shuning uchun test bazasida chiqmaydi va faqat
    prodda ko'rinadi. Aynan shu xato ilgari pos_refund'да ham bo'lgan
    (REF-1 izohiga qarang) — ya'ni bu takrorlanuvchi tuzoq.

    Yechim: qulflanadigan so'rovда JOIN umuman bo'lmasin. Avval PK'larni
    o'qiymiz (bu FOR UPDATE emas, join zarar qilmaydi), keyin faqat PK
    bo'yicha qulflaymiz.
    """
    ids = list(qs.order_by().values_list('pk', flat=True))
    return BranchStock.objects.select_for_update().filter(pk__in=ids)


def _price_qs(request, params=None):
    """Filtrlangan BranchStock queryset. STK-9: `params` — GET yoki POST.
    price_apply POST bo'lgani uchun filtrlar GET'да YO'Q edi; natijada
    `_price_qs` BUTUN katalogni qaytarib, hammani qayta narxlardi."""
    from decimal import Decimal
    if params is None:
        params = request.GET
    qs = (BranchStock.objects
          .select_related('variant__product__category', 'branch')
          .order_by('variant__product__name', 'variant__color', 'variant__size'))

    branch_id = (params.get('branch') or '').strip()
    if branch_id:
        qs = qs.filter(branch_id=branch_id)
    elif getattr(request.user, 'branch_id', None) and request.user.role != 'admin':
        qs = qs.filter(branch=request.user.branch)

    q = (params.get('q') or '').strip()
    if q:
        qs = qs.filter(Q(variant__product__name__icontains=q)
                       | Q(variant__product__code__icontains=q)
                       | Q(variant__barcode__icontains=q)
                       | Q(variant__color__icontains=q)
                       | Q(variant__size__icontains=q))

    cat = (params.get('category') or '').strip()
    if cat:
        qs = qs.filter(variant__product__category_id=cat)

    issue = (params.get('issue') or '').strip()
    if issue == 'no_cost':
        qs = qs.filter(cost_price__lte=0)
    elif issue == 'zero':
        qs = qs.filter(cost_price__gt=0, sale_price=F('cost_price'))
    elif issue == 'loss':
        qs = qs.filter(cost_price__gt=0, sale_price__gt=0,
                       sale_price__lt=F('cost_price'))
    elif issue == 'no_sale':
        qs = qs.filter(sale_price__lte=0)
    elif issue == 'low':
        qs = qs.filter(cost_price__gt=0, sale_price__gt=0,
                       sale_price__lt=F('cost_price') * Decimal('1.10'))
    elif issue == 'no_ws':
        qs = qs.filter(wholesale_price__lte=0)
    return qs


@admin_required
def price_list(request):
    """Narx/marja jadvali — filtr, qatorlab tahrir, ommaviy qo'llash."""
    from decimal import Decimal
    from django.core.paginator import Paginator
    qs = _price_qs(request)

    base = BranchStock.objects.all()
    branch_id = (request.GET.get('branch') or '').strip()
    if branch_id:
        base = base.filter(branch_id=branch_id)
    counts = {
        'no_cost': base.filter(cost_price__lte=0).count(),
        'zero': base.filter(cost_price__gt=0, sale_price=F('cost_price')).count(),
        'loss': base.filter(cost_price__gt=0, sale_price__gt=0,
                            sale_price__lt=F('cost_price')).count(),
        'no_sale': base.filter(sale_price__lte=0).count(),
        'low': base.filter(cost_price__gt=0, sale_price__gt=0,
                           sale_price__lt=F('cost_price') * Decimal('1.10')).count(),
        'no_ws': base.filter(wholesale_price__lte=0).count(),
    }

    issue_cards = [(k, label, hint, counts.get(k, 0))
                   for k, label, hint in PRICE_ISSUES]

    paginator = Paginator(qs, PRICE_PAGE_SIZE)
    page = paginator.get_page(request.GET.get('page'))

    params = request.GET.copy()
    params.pop('page', None)
    return render(request, 'inventory/price_list.html', {
        'page_obj': page,
        'rows': page.object_list,
        'total_count': paginator.count,
        'categories': Category.objects.order_by('name'),
        'branches': Branch.objects.filter(is_active=True).order_by('name'),
        'issues': PRICE_ISSUES,
        'issue_cards': issue_cards,
        'cur': {
            'q': request.GET.get('q', ''),
            'category': request.GET.get('category', ''),
            'branch': branch_id,
            'issue': request.GET.get('issue', ''),
        },
        'querystring': params.urlencode(),
    })


# Eslatma: BranchStock o'zgarishini signals.py avtomatik AuditLog'ga yozadi
# ({field: [old, new]} ko'rinishida), shuning uchun bu yerda qo'shimcha yozuv
# yaratilmaydi — aks holda tarixda dublikat chiqadi.


@admin_required
def price_apply(request):
    """Qatorlab tahrir + ommaviy amal (marja qo'llash, narxni foizga o'zgartirish)."""
    from decimal import Decimal, InvalidOperation
    if request.method != 'POST':
        return redirect('price_list')
    back = request.POST.get('back') or reverse('price_list')

    def dec(v, allow_negative=False):
        """Narxlar manfiy bo'lmaydi; foiz esa manfiy bo'lishi mumkin (-5%).

        V1: BO'SH katak = "o'zgarmadi" (None), NOL emas — aks holда katakни
        tozalab qayta yozmoqchi bo'lган kassir narxni 0 ga tushirib, tovar
        tekinга sotilar edi.
        V4: 10 xonaдан katta son (max_digits=12, decimal_places=2) saqlashда
        InvalidOperation berib butun atomik blokни orqага qaytarardi — shu yerда
        rad etamiz.
        """
        s = str(v).replace(' ', '').replace(',', '.').strip()
        if s == '':
            return None
        try:
            d = Decimal(s)
        except (InvalidOperation, ValueError, TypeError):
            return None
        if not d.is_finite():          # 'inf' / 'nan' — N2
            return None
        if not allow_negative and d < 0:
            return None
        if abs(d) >= Decimal('10000000000'):   # 10 butun xona chegarasi
            return None
        return d

    mode = request.POST.get('mode') or 'rows'
    changed = 0

    skipped = 0
    if mode == 'rows':
        # Jadvalda qo'lda o'zgartirilgan kataklar
        ids = request.POST.getlist('row_id')
        with transaction.atomic():
            for i, sid in enumerate(ids):
                stock = BranchStock.objects.select_for_update().filter(pk=sid).first()
                if not stock:
                    continue
                get = lambda n: (request.POST.getlist(n)[i]
                                 if i < len(request.POST.getlist(n)) else '')
                rc, rs, rw = get('row_cost'), get('row_sale'), get('row_ws')
                new_cost = dec(rc)
                new_sale = dec(rs)
                new_ws = dec(rw)
                # E2: bo'sh EMAS, lekin rad etilган (manfiy/noto'g'ri/juda katta)
                # qiymatlarни sanaymiz — jim "muvaffaqiyat" o'rniga ogohlantiramiz.
                for _rv, _pv in ((rc, new_cost), (rs, new_sale), (rw, new_ws)):
                    if (_rv or '').strip() != '' and _pv is None:
                        skipped += 1
                ch = {}
                if new_cost is not None and new_cost != stock.cost_price:
                    ch['tannarx'] = [str(stock.cost_price), str(new_cost)]
                    stock.cost_price = new_cost
                if new_sale is not None and new_sale != stock.sale_price:
                    ch['sotuv'] = [str(stock.sale_price), str(new_sale)]
                    stock.sale_price = new_sale
                if new_ws is not None and new_ws != stock.wholesale_price:
                    ch['ulgurji'] = [str(stock.wholesale_price), str(new_ws)]
                    stock.wholesale_price = new_ws
                if ch:
                    stock.save(update_fields=['cost_price', 'sale_price',
                                              'wholesale_price'])
                    changed += 1
        messages.success(request, f"{changed} ta qator yangilandi.")
        if skipped:
            messages.warning(request,
                f"{skipped} ta qiymat qabul qilinmadi (manfiy yoki noto'g'ri) — "
                "o'sha kataklarni tekshiring.")
        return redirect(back)

    # ---- ommaviy amal ----
    op = request.POST.get('op') or ''
    pct = dec(request.POST.get('pct'), allow_negative=True)
    if pct is None:
        messages.error(request, "Foiz qiymati noto'g'ri.")
        return redirect(back)

    scope = request.POST.get('scope') or 'selected'
    if scope == 'selected':
        sel = request.POST.getlist('sel')
        if not sel:
            messages.error(request, "Hech qanday qator tanlanmagan.")
            return redirect(back)
        targets = BranchStock.objects.filter(pk__in=sel)
    else:
        # STK-9: filtrlarni POST'дан o'qiymiz (GET bo'sh). Va agar HECH QANDAY
        # filtr bo'lmasa — bu BUTUN katalogni qamrab oladi. Tasodifan hammani
        # qayta narxlamaslik uchun aniq filtr yoki tasdiq talab qilamiz.
        _has_filter = any((request.POST.get(k) or '').strip()
                          for k in ('branch', 'q', 'category', 'issue'))
        if not _has_filter and request.POST.get('confirm_all') != '1':
            messages.error(request,
                "Filtr tanlanmagan — bu BUTUN katalogга ta'sir qiladi. "
                "Avval filtrlang yoki qatorlarni belgilab tanlang.")
            return redirect(back)
        targets = _price_qs(request, request.POST)      # POST filtridagi qatorlar

    factor = Decimal('1') + pct / Decimal('100')
    q2 = Decimal('0.01')

    with transaction.atomic():
        # STK-14: select_related'siz QULFLAYMIZ — pastdagi halqa faqat narx
        # maydonlarini o'qiydi, bogʻlangan jadvallar kerak emas.
        for stock in _lock_stocks(targets):
            ch = {}
            if op == 'margin_from_cost':
                # tannarx bor -> sotuv = tannarx × (1 + marja)
                if stock.cost_price > 0:
                    new = (stock.cost_price * factor).quantize(q2)
                    if new != stock.sale_price:
                        ch['sotuv'] = [str(stock.sale_price), str(new)]
                        stock.sale_price = new
            elif op == 'cost_from_sale':
                # sotuv narxi to'g'ri -> tannarx = sotuv / (1 + marja)
                if stock.sale_price > 0 and factor > 0:
                    new = (stock.sale_price / factor).quantize(q2)
                    if new != stock.cost_price:
                        ch['tannarx'] = [str(stock.cost_price), str(new)]
                        stock.cost_price = new
            elif op == 'wholesale_from_cost':
                if stock.cost_price > 0:
                    new = (stock.cost_price * factor).quantize(q2)
                    if new != stock.wholesale_price:
                        ch['ulgurji'] = [str(stock.wholesale_price), str(new)]
                        stock.wholesale_price = new
            elif op == 'margin_add':
                # Mavjud marjaga qo'shish: 20% edi, +5 -> 25%.
                # sotuv = tannarx × (1 + (joriy_marja + pct)/100)
                if stock.cost_price > 0 and stock.sale_price > 0:
                    cur_m = (stock.sale_price / stock.cost_price
                             - Decimal('1')) * Decimal('100')
                    new_m = cur_m + pct
                    if new_m > Decimal('-100'):
                        new = (stock.cost_price
                               * (Decimal('1') + new_m / Decimal('100'))
                               ).quantize(q2)
                        if new >= 0 and new != stock.sale_price:
                            ch['sotuv'] = [str(stock.sale_price), str(new)]
                            stock.sale_price = new
            elif op == 'wholesale_margin_add':
                if stock.cost_price > 0 and stock.wholesale_price > 0:
                    cur_m = (stock.wholesale_price / stock.cost_price
                             - Decimal('1')) * Decimal('100')
                    new_m = cur_m + pct
                    if new_m > Decimal('-100'):
                        new = (stock.cost_price
                               * (Decimal('1') + new_m / Decimal('100'))
                               ).quantize(q2)
                        if new >= 0 and new != stock.wholesale_price:
                            ch['ulgurji'] = [str(stock.wholesale_price), str(new)]
                            stock.wholesale_price = new
            elif op == 'bump_sale':
                if stock.sale_price > 0:
                    new = (stock.sale_price * factor).quantize(q2)
                    if new != stock.sale_price:
                        ch['sotuv'] = [str(stock.sale_price), str(new)]
                        stock.sale_price = new
            elif op == 'bump_wholesale':
                if stock.wholesale_price > 0:
                    new = (stock.wholesale_price * factor).quantize(q2)
                    if new != stock.wholesale_price:
                        ch['ulgurji'] = [str(stock.wholesale_price), str(new)]
                        stock.wholesale_price = new
            if ch:
                stock.save(update_fields=['cost_price', 'sale_price',
                                          'wholesale_price'])
                changed += 1

    messages.success(request, f"{changed} ta narx yangilandi.")
    return redirect(back)


@admin_required
def price_history(request):
    """Narx o'zgarishlari tarixi (AuditLog'dan)."""
    from django.core.paginator import Paginator
    logs = (AuditLog.objects
            .filter(model_name='BranchStock')
            .filter(changes__has_any_keys=['cost_price', 'sale_price',
                                           'wholesale_price'])
            .select_related('user')
            .order_by('-created_at'))
    q = (request.GET.get('q') or '').strip()
    if q:
        # object_repr'da mahsulot KODI turadi — nom bo'yicha qidirish uchun
        # avval mos zaxira yozuvlarini topamiz.
        ids = list(BranchStock.objects
                   .filter(Q(variant__product__name__icontains=q)
                           | Q(variant__product__code__icontains=q))
                   .values_list('pk', flat=True)[:5000])
        logs = logs.filter(Q(object_repr__icontains=q)
                           | Q(object_id__in=[str(i) for i in ids]))
    page = Paginator(logs, 80).get_page(request.GET.get('page'))

    # object_id -> BranchStock: nomni chiroyli ko'rsatish uchun
    oids = [l.object_id for l in page.object_list if (l.object_id or '').isdigit()]
    stocks = {str(s.pk): s for s in BranchStock.objects
              .filter(pk__in=oids)
              .select_related('variant__product', 'branch')}
    for l in page.object_list:
        l.stock = stocks.get(l.object_id)

    return render(request, 'inventory/price_history.html', {
        'page_obj': page, 'logs': page.object_list, 'q': q,
    })


# ---------- AKSIYALAR (Promotion) ----------

@admin_required
def promotion_list(request):
    """Aksiyalar ro'yxati. Model va POS mantiqi bor edi — endi boshqaruvi ham."""
    promos = (Promotion.objects
              .select_related('category')
              .prefetch_related('target_products')
              .order_by('-is_active', '-created_at'))
    now = timezone.now()
    for p in promos:
        p.is_running = (p.is_active and p.valid_from <= now
                        and (p.valid_until is None or p.valid_until >= now))
    return render(request, 'inventory/promotion_list.html', {
        'promos': promos,
        'types': Promotion.Type.choices,
        'categories': Category.objects.order_by('name'),
        'now': now,
    })


@admin_required
def promotion_save(request):
    """Aksiya yaratish yoki tahrirlash (bitta forma)."""
    from decimal import Decimal, InvalidOperation
    if request.method != 'POST':
        return redirect('promotion_list')
    pk = (request.POST.get('pk') or '').strip()
    promo = Promotion.objects.filter(pk=pk).first() if pk else Promotion()

    name = (request.POST.get('name') or '').strip()[:120]
    if not name:
        messages.error(request, "Aksiya nomini kiriting.")
        return redirect('promotion_list')

    ptype = request.POST.get('promo_type') or Promotion.Type.PERCENT_OFF
    if ptype not in dict(Promotion.Type.choices):
        ptype = Promotion.Type.PERCENT_OFF

    def num(v, default=0):
        # V7: "50%", "50 " kabi kiritishlar ham ishlasin — ilgari '%' belgisi
        # InvalidOperation berib, num() 0 qaytarар va "50%" JIM 0% aksiyaга
        # aylanardi.
        s = str(v if v is not None else '').strip()
        for ch in (' ', ' ', ' ', '%', "'"):
            s = s.replace(ch, '')
        s = s.replace(',', '.')
        try:
            d = Decimal(s or default)
            return d if d.is_finite() else Decimal(default)
        except (InvalidOperation, ValueError, TypeError):
            return Decimal(default)

    percent = num(request.POST.get('percent'))
    percent = max(Decimal('0'), min(Decimal('100'), percent))
    # V7: foizli aksiya turlari uchun foiz 0 bo'lsa — bu foydasiz "0% chegirma"
    # bo'lardi (lekin muvaffaqiyat deб ko'rsatilardi). Rad etamiz.
    if ptype in (Promotion.Type.PERCENT_OFF, Promotion.Type.NTH_PERCENT) and percent <= 0:
        messages.error(request, "Foizli aksiya uchun foiz 0 dan katta bo'lishi kerak.")
        return redirect('promotion_list')
    try:
        qty_required = max(1, int(request.POST.get('qty_required') or 1))
    except (TypeError, ValueError):
        qty_required = 1
    try:
        qty_free = max(0, int(request.POST.get('qty_free') or 0))
    except (TypeError, ValueError):
        qty_free = 0

    promo.name = name
    promo.promo_type = ptype
    promo.percent = percent
    promo.qty_required = qty_required
    promo.qty_free = qty_free
    promo.category = Category.objects.filter(
        pk=request.POST.get('category') or 0).first()
    promo.is_active = bool(request.POST.get('is_active'))

    vf = (request.POST.get('valid_from') or '').strip()
    vu = (request.POST.get('valid_until') or '').strip()
    tz = timezone.get_current_timezone()
    if vf:
        try:
            promo.valid_from = timezone.make_aware(
                datetime.strptime(vf, '%Y-%m-%d'), tz)
        except (ValueError, TypeError):
            pass
    promo.valid_until = None
    if vu:
        try:
            promo.valid_until = timezone.make_aware(
                datetime.strptime(vu, '%Y-%m-%d').replace(hour=23, minute=59), tz)
        except (ValueError, TypeError):
            pass
    # V7: tugash sanasi boshlanish sanasidан oldин bo'lmasин — aks holда aksiya
    # hech qachon ishlamaydi (lekin "saqlandi" deб ko'rsatilardi).
    if (promo.valid_from and promo.valid_until
            and promo.valid_until < promo.valid_from):
        messages.error(request,
            "Tugash sanasi boshlanish sanasidan oldin bo'lmasin.")
        return redirect('promotion_list')
    promo.save()

    AuditLog.objects.create(
        user=request.user, username_snapshot=request.user.username,
        action=(AuditLog.Action.UPDATE if pk else AuditLog.Action.CREATE),
        model_name='Promotion', object_id=str(promo.pk),
        object_repr=f'Aksiya: {promo.name}'[:300],
    )
    messages.success(request, f"\"{promo.name}\" saqlandi.")
    return redirect('promotion_list')


@admin_required
def promotion_delete(request, pk):
    if request.method != 'POST':
        return redirect('promotion_list')
    promo = get_object_or_404(Promotion, pk=pk)
    name = promo.name
    AuditLog.objects.create(
        user=request.user, username_snapshot=request.user.username,
        action=AuditLog.Action.DELETE, model_name='Promotion',
        object_id=str(pk), object_repr=f'Aksiya: {name}'[:300],
    )
    promo.delete()
    messages.success(request, f"\"{name}\" o'chirildi.")
    return redirect('promotion_list')


# ---------- TEZKOR SOTUV toifalari (POS paneli sozlamasi) ----------

@admin_required
def quick_sell_settings(request):
    """POS 'Tezkor sotuv' toifalari va narxlarini tahrirlash."""
    if request.method == 'POST':
        action = request.POST.get('action') or 'save'

        if action == 'add':
            name = (request.POST.get('new_name') or '').strip()[:60]
            if not name:
                messages.error(request, "Toifa nomini kiriting.")
            elif QuickSellItem.objects.filter(name__iexact=name).exists():
                messages.error(request, f"\"{name}\" allaqachon bor.")
            else:
                last = QuickSellItem.objects.order_by('-order').first()
                item = QuickSellItem.objects.create(
                    name=name,
                    prices=_parse_prices(request.POST.get('new_prices')),
                    icon=(request.POST.get('new_icon') or 'bi-bag').strip()[:40],
                    order=(last.order + 10) if last else 10,
                )
                _quick_sell_sync(item)
                messages.success(request, f"\"{name}\" qo'shildi.")
            return redirect('quick_sell_settings')

        if action == 'delete':
            item = QuickSellItem.objects.filter(pk=request.POST.get('pk')).first()
            if item:
                name = item.name
                item.delete()
                messages.success(request, f"\"{name}\" o'chirildi.")
            return redirect('quick_sell_settings')

        # ---- qatorlab saqlash ----
        n = 0
        for item in QuickSellItem.objects.all():
            pk = str(item.pk)
            name = (request.POST.get(f'name_{pk}') or '').strip()[:60]
            if not name:
                continue
            prices = _parse_prices(request.POST.get(f'prices_{pk}'))
            icon = (request.POST.get(f'icon_{pk}') or 'bi-bag').strip()[:40]
            try:
                order = int(request.POST.get(f'order_{pk}') or item.order)
            except (TypeError, ValueError):
                order = item.order
            active = bool(request.POST.get(f'active_{pk}'))
            if (name != item.name or prices != item.price_list
                    or icon != item.icon or order != item.order
                    or active != item.is_active):
                item.name = name
                item.prices = prices
                item.icon = icon
                item.order = order
                item.is_active = active
                item.save()
                _quick_sell_sync(item)      # yangi narxga tur + zaxira
                n += 1
        messages.success(request, f"{n} ta toifa yangilandi.")
        return redirect('quick_sell_settings')

    items = list(QuickSellItem.objects.select_related('product'))
    for it in items:
        _quick_sell_sync(it)
    return render(request, 'inventory/quick_sell_settings.html', {
        'items': items,
        'category_name': QUICK_SELL_CATEGORY,
    })


def _parse_prices(text):
    """"2500, 3000; 5000" -> [2500, 3000, 5000]"""
    out = []
    for chunk in re.split(r'[,;\s]+', (text or '').strip()):
        chunk = re.sub(r'[^0-9.]', '', chunk)
        if not chunk:
            continue
        try:
            v = int(float(chunk))
        except ValueError:
            continue
        if v > 0 and v not in out:
            out.append(v)
    return sorted(out)



PRICE_TAG_SEP = ' · '


# Sof narx belgisi: "115 000" yoki "115 000 (2)".
# MUHIM: faqat MING-AJRATGICHLI (probelli) sonlar narx deb qaraladi. Ilgari
# HAR QANDAY raqam ("1273", "3027", "545") ham narx deb belgilanardi — bu
# raqamli RANG/KOD'larni bir "oila"ga qo'shib yuborardi va qabulда qoldiq
# NOTO'G'RI turga tushib, kassada "omborda yo'q" chiqardi. Tizim yaratadigan
# narx belgisi doim probelli (masalan 115 000), shuning uchun probelni talab
# qilamiz — "1273" kabi kodlar endi o'z rangi bo'lib qoladi.
_PRICE_TAG_RE = re.compile(r'^\d{1,3}(\s\d{3})+(\s*\(\d+\))?$')


def _base_color(color):
    """Narx belgisi olib tashlangan asosiy rang.

    "Qora · 115 000" -> "Qora"
    "115 000"        -> ""      (asosiy rang bo'sh bo'lgan holat — ajratuvchi
                                 yo'q, shuning uchun alohida tekshiramiz;
                                 aks holda har qabulda yangi tur ochilib
                                 ketardi)
    """
    c = (color or '').strip()
    if c in ('—', '-', '–'):
        c = ''            # "rang yo'q" belgisi — bo'sh deb qaraymiz
    if PRICE_TAG_SEP in c:
        b = c.split(PRICE_TAG_SEP)[0].strip()
        return '' if b in ('—', '-', '–') else b
    if _PRICE_TAG_RE.match(c):
        return ''
    return c


def _price_tag(sale_price):
    return f"{int(sale_price):,}".replace(',', ' ')


def resolve_row_price_color(size, color, sale_price, seen):
    """Bitta qabulda bir xil o'lcham ikki xil narxda kelsa — ikkinchi qatorga
    narx belgisi qo'shamiz, shunda u alohida tur bo'lib o'z kodini oladi.

    Masalan "Maria" mahsuloti, 50-o'lcham:
        6 dona  — 10 000 so'm  ->  "50 / —"        (o'z kodi)
        3 dona  — 12 000 so'm  ->  "50 / 12 000"   (boshqa kod)
    Ikkalasi ham 50-o'lcham, lekin kassada kod -> tur -> narx bo'lgani uchun
    bitta turga ikki narx sig'maydi.

    Narxi ham, o'lchami ham bir xil bo'lsa — bu haqiqiy takror, xato qaytaramiz.

    seen: {(o'lcham, rang): narx} — chaqiruvchi tomonda saqlanadi.
    Qaytaradi: (rang, takror_mi)
    """
    from decimal import Decimal
    try:
        price = Decimal(str(sale_price or 0))
    except Exception:
        price = Decimal('0')

    key = (size, color)
    if key not in seen:
        seen[key] = price
        return color, False
    if seen[key] == price:
        return color, True          # bir xil o'lcham + bir xil narx = takror

    base = _base_color(color)
    tag = _price_tag(price)
    n = 1
    while True:
        suffix = tag if n == 1 else f"{tag} ({n})"
        new_color = (f"{base}{PRICE_TAG_SEP}{suffix}" if base else suffix)[:50]
        k = (size, new_color)
        if k not in seen:
            seen[k] = price
            return new_color, False
        if seen[k] == price:
            return new_color, True
        n += 1


def resolve_price_variant(variant, branch, sale_price):
    """Sotuv narxi omborodagidan FARQ qilsa — alohida tur (o'z shtrix-kodi bilan).

    Nega: shtrix-kod turga tegishli, kassa esa kod -> tur -> narx bo'yicha
    ishlaydi. Ya'ni bir turga ikki xil narx sig'maydi — oxirgi qabul
    avvalgisini bosib ketadi va etiketkalar noto'g'ri narx bilan chiqadi.

    Qoida:
      1. Shu narxdagi "birodar" tur allaqachon bo'lsa — o'shani ishlatamiz
         (har qabulda yangi tur yaratilib ketmasin).
      2. Bazaviy tur bo'sh bo'lsa (ombor 0) — narxni bemalol yangilaymiz.
      3. Aks holda yangi tur ochamiz: rang "... · 115 000" ko'rinishida
         belgilanadi va yangi EAN-13 beriladi.

    Qaytaradi: aslida ishlatilishi kerak bo'lgan ProductVariant.
    """
    from decimal import Decimal
    if variant is None or branch is None:
        return variant
    try:
        sale = Decimal(str(sale_price or 0))
    except Exception:
        return variant
    if sale <= 0:
        return variant  # narx ko'rsatilmagan — eski xatti-harakat

    base_color = _base_color(variant.color)
    sibs = [v for v in ProductVariant.objects.filter(
                product=variant.product, size=variant.size)
            if _base_color(v.color) == base_color]

    stocks = {st.variant_id: st for st in BranchStock.objects.filter(
        variant__in=sibs, branch=branch)}

    # 1) Ayni shu narxdagi tur bormi?
    for v in sibs:
        st = stocks.get(v.pk)
        if st is not None and st.sale_price == sale:
            return v

    # 2) Kelgan turning o'zi bo'sh bo'lsa — o'shani ishlatamiz
    own = stocks.get(variant.pk)
    if own is None or own.stock_count <= 0:
        return variant

    # 3) Yangi tur — o'z kodi va o'z narxi bilan
    tag = _price_tag(sale)
    color = f"{base_color}{PRICE_TAG_SEP}{tag}" if base_color else tag
    color = color[:50]
    n = 1
    while ProductVariant.objects.filter(product=variant.product,
                                        size=variant.size, color=color).exists():
        n += 1
        color = f"{color[:44]} ({n})"

    new_v = ProductVariant.objects.create(
        product=variant.product, size=variant.size, color=color)
    code = gen_internal_ean13(new_v.pk)
    k = 0
    while ProductVariant.objects.filter(barcode=code).exclude(pk=new_v.pk).exists():
        k += 1
        code = gen_internal_ean13(new_v.pk + k * 100000)
    new_v.barcode = code
    new_v.save(update_fields=['barcode'])
    return new_v


@admin_required
def variant_split_batch(request):
    """Bir turdagi zaxiraning bir qismini ALOHIDA turga (o'z shtrix-kodi va
    o'z narxi bilan) ajratib olish.

    Nega shunday: shtrix-kod TURGA tegishli va kassada kod -> tur -> narx
    bo'lib ishlaydi. Demak "boshqa kod + boshqa narx" degani — alohida tur.
    Masalan 155 dona ichidan 28 tasi boshqa narxda kelgan bo'lsa, o'shani
    ajratamiz: qolgan 127 eski kodda qoladi, 28 tasi yangi kod oladi.
    """
    from decimal import Decimal, InvalidOperation
    if request.method != 'POST':
        return redirect('product_list')

    src = get_object_or_404(ProductVariant.objects.select_related('product'),
                            pk=request.POST.get('variant') or 0)
    branch = Branch.objects.filter(pk=request.POST.get('branch') or 0).first() \
        or getattr(request.user, 'branch', None)
    back = request.POST.get('back') or reverse('product_detail',
                                               args=[src.product.code])
    if branch is None:
        messages.error(request, "Filial topilmadi.")
        return redirect(back)

    def dec(v, default='0'):
        try:
            return Decimal(str(v).replace(' ', '').replace(',', '.') or default)
        except (InvalidOperation, ValueError, TypeError):
            return Decimal(default)

    try:
        qty = int(float(request.POST.get('qty') or 0))
    except (TypeError, ValueError):
        qty = 0
    cost = dec(request.POST.get('cost'))
    sale = dec(request.POST.get('sale'))
    if qty <= 0:
        messages.error(request, "Ajratiladigan miqdorni kiriting.")
        return redirect(back)
    if cost < 0 or sale < 0:
        messages.error(request, "Narx manfiy bo'lmaydi.")
        return redirect(back)

    src_stock = BranchStock.objects.filter(variant=src, branch=branch).first()
    if src_stock is None or src_stock.stock_count < qty:
        have = src_stock.stock_count if src_stock else 0
        messages.error(request, f"Omborda faqat {have} dona bor.")
        return redirect(back)

    # Yangi tur nomi: narx bilan farqlanadi (unique_together: product+size+color)
    tag = (request.POST.get('tag') or '').strip()[:40] or f"{int(sale or cost):,}".replace(',', ' ')
    base_color = (src.color or '').strip()
    color = f"{base_color} · {tag}".strip(' ·') if base_color else tag
    color = color[:50]
    n = 1
    while ProductVariant.objects.filter(product=src.product, size=src.size,
                                        color=color).exists():
        n += 1
        color = f"{color[:44]} ({n})"

    with transaction.atomic():
        new_v = ProductVariant.objects.create(
            product=src.product, size=src.size, color=color)
        # Yangi, betakror shtrix-kod
        code = gen_internal_ean13(new_v.pk)
        while ProductVariant.objects.filter(barcode=code).exclude(pk=new_v.pk).exists():
            code = gen_internal_ean13(new_v.pk + 100000)
        new_v.barcode = code
        new_v.save(update_fields=['barcode'])

        BranchStock.objects.create(
            variant=new_v, branch=branch, stock_count=qty,
            cost_price=cost or src_stock.cost_price,
            sale_price=sale or src_stock.sale_price,
            wholesale_price=src_stock.wholesale_price,
        )
        src_stock.stock_count = F('stock_count') - qty
        src_stock.save(update_fields=['stock_count'])

        AuditLog.objects.create(
            user=request.user, username_snapshot=request.user.username,
            action=AuditLog.Action.CREATE, model_name='ProductVariant',
            object_id=str(new_v.pk),
            object_repr=f"Partiya ajratildi: {src.product.name} — {qty} dona"[:300],
            changes={'ajratildi': [str(src.pk), str(new_v.pk)],
                     'miqdor': ['0', str(qty)],
                     'shtrix_kod': ['', code]},
        )

    messages.success(
        request,
        f"{qty} dona alohida turga ajratildi — yangi shtrix-kod {code}, "
        f"sotuv narxi {sale or src_stock.sale_price} so'm.")
    return redirect(back)


@admin_required
def intake_mixed(request):
    """Aralash qabul — bitta sahifada butun yuk.

    "Kiyim/poyabzal" (o'lchamlar setkasi) va "Turlar jadvali" (qatorlab)
    birlashtirilgan: har mahsulot uchun qulay usulni tanlaysiz, hammasi
    bitta qabulga yoziladi.
    """
    branches = Branch.objects.filter(is_active=True).order_by('name')
    return render(request, 'inventory/intake_mixed.html', {
        'branches': branches,
        'categories': Category.objects.all().order_by('name'),
        'user_branch': getattr(request.user, 'branch', None),
        'suppliers': Supplier.objects.filter(is_active=True).order_by('name')[:200],
    })


@admin_required
def intake_mixed_save(request):
    """POST JSON — bir nechta mahsulotni bitta qabulda saqlaydi.

    Kutilgan shakl:
      {branch, supplier, invoice_number, note,
       products: [{code|name, category, brand, cost, marja, price,
                   lines: [{size, color, qty, cost, price, barcode}]}]}

    Ikkala usul (setka / qatorlab) ham AYNI "lines" shaklini yuboradi —
    farq faqat ko'rinishda, shuning uchun bu yerda bitta yo'l yetadi.
    """
    from decimal import Decimal, InvalidOperation
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Faqat POST'}, status=405)
    try:
        data = _json.loads(request.body.decode('utf-8'))
    except (ValueError, UnicodeDecodeError):
        return JsonResponse({'ok': False, 'error': "So'rov noto'g'ri"}, status=400)

    def dec(v, default='0'):
        try:
            return Decimal(str(v).replace(' ', '').replace(',', '.') or default)
        except (InvalidOperation, ValueError, TypeError):
            return Decimal(default)

    branch = Branch.objects.filter(pk=data.get('branch') or 0,
                                   is_active=True).first()
    if branch is None:
        return JsonResponse({'ok': False, 'error': 'Filial tanlanmagan'}, status=400)

    products_in = data.get('products') or []
    if not products_in:
        return JsonResponse({'ok': False, 'error': "Mahsulot qo'shilmagan"}, status=400)

    supplier_text = (data.get('supplier') or '').strip()[:200]
    supplier_obj = (Supplier.objects.filter(name__iexact=supplier_text).first()
                    if supplier_text else None)

    created_variant_ids = []
    total_qty = 0
    new_products = 0

    try:
        with transaction.atomic():
            session = IntakeSession.objects.create(
                branch=branch,
                supplier=supplier_obj,
                supplier_text='' if supplier_obj else supplier_text,
                received_by=request.user,
                invoice_number=(data.get('invoice_number') or '').strip()[:80],
                note=(data.get('note') or '').strip() or 'Aralash qabul',
            )

            for pi, p in enumerate(products_in, start=1):
                lines = [ln for ln in (p.get('lines') or [])
                         if int(float(ln.get('qty') or 0)) > 0]
                if not lines:
                    continue

                code = (p.get('code') or '').strip()
                product = Product.objects.filter(code__iexact=code).first() if code else None
                if product is None:
                    name = smart_title((p.get('name') or '').strip())
                    if not name:
                        raise ValueError(f"{pi}-mahsulot: nom kiritilmagan")
                    category = Category.objects.filter(pk=p.get('category') or 0).first()
                    if category is None:
                        raise ValueError(f"{pi}-mahsulot ({name}): kategoriya tanlanmagan")
                    brand = smart_title((p.get('brand') or '').strip())
                    product = Product.objects.filter(name__iexact=name,
                                                     category=category).first()
                    if product is None:
                        product = Product.objects.create(
                            name=name, brand=brand, category=category)
                        new_products += 1

                # Mahsulot darajasidagi narx — qatorda ko'rsatilmasa shu ishlatiladi
                p_cost = dec(p.get('cost'))
                p_price = dec(p.get('price'))
                p_marja = dec(p.get('marja'))
                if p_price <= 0 and p_cost > 0 and p_marja > 0:
                    p_price = (p_cost * (Decimal('1') + p_marja / Decimal('100'))
                               ).quantize(Decimal('0.01'))
                if p_cost <= 0 and p_price > 0 and p_marja > 0:
                    p_cost = (p_price / (Decimal('1') + p_marja / Decimal('100'))
                              ).quantize(Decimal('0.01'))

                for ln in lines:
                    qty = int(float(ln.get('qty') or 0))
                    size = (ln.get('size') or '').strip()[:30] or '—'
                    color = (ln.get('color') or '').strip()[:50]
                    cost = dec(ln.get('cost')) if str(ln.get('cost') or '').strip() else p_cost
                    price = dec(ln.get('price')) if str(ln.get('price') or '').strip() else p_price
                    if price <= 0:
                        raise ValueError(f"{product.name} ({size}): sotuv narxi yo'q")

                    variant, _ = ProductVariant.objects.get_or_create(
                        product=product, size=size, color=color)
                    bc = (ln.get('barcode') or '').strip()
                    if bc and not variant.barcode:
                        if not ProductVariant.objects.filter(barcode=bc).exclude(
                                pk=variant.pk).exists():
                            variant.barcode = bc
                            variant.save(update_fields=['barcode'])
                    if not variant.barcode:
                        c = gen_internal_ean13(variant.pk)
                        k = 0
                        while ProductVariant.objects.filter(barcode=c).exclude(
                                pk=variant.pk).exists():
                            k += 1
                            c = gen_internal_ean13(variant.pk + k * 100000)
                        variant.barcode = c
                        variant.save(update_fields=['barcode'])

                    # Sotuv narxi omborodagidan farq qilsa — alohida tur
                    variant = resolve_price_variant(variant, branch, price)
                    if not variant.barcode:
                        c = gen_internal_ean13(variant.pk)
                        variant.barcode = c
                        variant.save(update_fields=['barcode'])

                    stock, _ = BranchStock.objects.get_or_create(
                        variant=variant, branch=branch,
                        defaults={'cost_price': cost, 'sale_price': price})
                    if cost > 0:  # STK-8 weighted-average
                        stock.cost_price = weighted_cost(
                            stock.stock_count if isinstance(stock.stock_count, int) else 0,
                            stock.cost_price, qty, cost)
                    stock.stock_count = F('stock_count') + qty
                    stock.sale_price = price
                    stock.save(update_fields=['stock_count', 'cost_price', 'sale_price'])

                    Intake.objects.create(
                        session=session, supplier_ref=supplier_obj,
                        variant=variant, branch=branch,
                        quantity=qty, cost_per_unit=cost, sale_price=price,
                        supplier=supplier_text, note='Aralash qabul',
                        received_by=request.user)

                    created_variant_ids.append(variant.pk)
                    total_qty += qty

                if product.default_sale_price == 0 and p_price > 0:
                    product.default_sale_price = p_price
                    product.save(update_fields=['default_sale_price'])

            if not created_variant_ids:
                raise ValueError("Hech qanday miqdor kiritilmagan")
    except ValueError as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)

    try:
        from .notifications import notify_intake_session
        notify_intake_session(session)
    except Exception:
        pass  # xabar yuborilmasa ham qabul saqlanadi

    ids = ','.join(str(i) for i in dict.fromkeys(created_variant_ids))
    return JsonResponse({
        'ok': True,
        'session_id': session.pk,
        'total_qty': total_qty,
        'variants': len(set(created_variant_ids)),
        'new_products': new_products,
        'labels_url': f"{reverse('variant_labels')}?ids={ids}&copies=stock",
        'session_url': reverse('intake_session_detail', args=[session.pk]),
    })


@admin_required
def product_image(request, code):
    """Mahsulot rasmini yuklash / o'chirish.

    Nakladnoy rasmlaridagi kabi: telefondan suratga olish ham, fayl tanlash
    ham ishlaydi (shablonda `capture` atributi bor).
    """
    product = get_object_or_404(Product, code__iexact=code)
    back = reverse('product_detail', args=[product.code])
    if request.method != 'POST':
        return redirect(back)

    if request.POST.get('action') == 'delete':
        if product.image:
            product.image.delete(save=False)
            product.image = None
            product.save(update_fields=['image'])
            messages.success(request, "Rasm o'chirildi.")
        return redirect(back)

    f = request.FILES.get('image')
    if not f:
        messages.error(request, "Rasm tanlanmadi.")
        return redirect(back)
    if f.size > 12 * 1024 * 1024:
        messages.error(request, "Rasm juda katta (12 MB dan oshmasin).")
        return redirect(back)
    ctype = (getattr(f, 'content_type', '') or '').lower()
    if not ctype.startswith('image/'):
        messages.error(request, "Faqat rasm fayli yuklanadi.")
        return redirect(back)

    old = product.image.name if product.image else None
    product.image = f
    product.save(update_fields=['image'])
    if old and old != product.image.name:
        # eskisini tashlab yubormaymiz — joy egallab qolmasin
        try:
            product.image.storage.delete(old)
        except Exception:
            pass
    messages.success(request, "Rasm saqlandi.")
    return redirect(back)


@admin_required
def product_export(request):
    """Filtrlangan mahsulot ro'yxatini Excel'ga chiqarish.

    Sahifadagi qidiruv/kategoriya/zaxira filtrlari va saralash aynan
    saqlanadi — ekranda ko'rgan ro'yxatning o'zi tushadi.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    rows = (BranchStock.objects
            .select_related('variant', 'variant__product',
                            'variant__product__category', 'branch')
            .exclude(variant__product__is_open_price=True))

    q = (request.GET.get('q') or '').strip()
    if q:
        rows = rows.filter(
            Q(variant__product__name__icontains=q) |
            Q(variant__product__code__icontains=q) |
            Q(variant__product__brand__icontains=q) |
            Q(variant__barcode__icontains=q) |
            Q(variant__color__icontains=q) |
            Q(variant__size__icontains=q))
    cat = request.GET.get('category')
    if cat and str(cat).isdigit():
        rows = rows.filter(variant__product__category_id=int(cat))
    sf = request.GET.get('stock')
    if sf == 'zero':
        rows = rows.filter(stock_count=0)
    elif sf == 'low':
        rows = rows.filter(stock_count__gt=0, stock_count__lte=3)
    elif sf == 'in_stock':
        rows = rows.filter(stock_count__gt=0)

    rows = rows.order_by('variant__product__name', 'variant__size', 'variant__color')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Mahsulotlar'
    head = ['Kod', 'Nomi', 'Brend', 'Kategoriya', 'Rang', "O'lcham",
            'Shtrix-kod', 'Filial', 'Ombor', 'Tannarx', 'Sotuv narxi',
            'Marja %', 'Zaxira qiymati']
    ws.append(head)
    hf = Font(bold=True, color='FFFFFF')
    fill = PatternFill('solid', fgColor='4F46E5')
    for i in range(1, len(head) + 1):
        cl = ws.cell(row=1, column=i)
        cl.font = hf; cl.fill = fill
        cl.alignment = Alignment(horizontal='center')

    n = 0
    for st in rows.iterator(chunk_size=500):
        v = st.variant; p = v.product
        cost = float(st.cost_price or 0)
        sale = float(st.sale_price or 0)
        marja = round((sale - cost) / cost * 100, 1) if cost > 0 else None
        ws.append([_csv_safe(p.code), _csv_safe(p.name), _csv_safe(p.brand or ''),
                   _csv_safe(p.category.name if p.category else ''),
                   _csv_safe(v.color or ''), _csv_safe(v.size or ''),
                   _csv_safe(v.barcode or ''),
                   _csv_safe(st.branch.name), st.stock_count, cost, sale, marja,
                   round(st.stock_count * cost, 2)])   # SEC-20
        n += 1

    widths = [11, 34, 14, 16, 12, 10, 16, 16, 8, 12, 12, 9, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(head))}{n + 1}'

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    stamp = timezone.localtime().strftime('%Y-%m-%d_%H%M')
    resp['Content-Disposition'] = f'attachment; filename="mahsulotlar_{stamp}.xlsx"'
    wb.save(resp)
    return resp


# Har bo'lim uchun alohida, aniq (professional) rang
GROUP_COLORS = {
    'Erkaklar':            '#2563EB',   # ko'k
    'Ayollar':             '#DB2777',   # pushti
    'Bolalar':             '#F59E0B',   # amber
    'Parfumeriya va uy':   '#0D9488',   # teal
    "Bo'limsiz":           '#6B7280',   # kulrang
}


def _hex_rgba(hexstr, alpha):
    h = hexstr.lstrip('#')
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return f'rgba({r},{g},{b},{alpha:.2f})'


def _color_tree(nodes):
    """Daraxtga rang beradi: har bo'lim o'z rangi, kategoriyalar esa o'sha
    rangning quyuqdan ochiqqa tuslari (kattadan kichraygan sari ochroq)."""
    for node in nodes:
        base = GROUP_COLORS.get(node['name'], '#6B7280')
        node['color'] = base
        n = max(len(node['cats']), 1)
        for i, c in enumerate(node['cats']):
            alpha = max(0.42, 1.0 - (i / n) * 0.6)
            c['color'] = _hex_rgba(base, alpha)


@admin_required
def warehouse(request):
    """Ombor — to'liq nazorat va tahlil.

    Hammasi AGREGAT so'rovlar bilan hisoblanadi: 1475 zaxira yozuvi va
    ~19 000 dona bo'lgani uchun har qator bo'yicha alohida so'rov yubormaymiz.
    """
    from decimal import Decimal
    branches = list(Branch.objects.filter(is_active=True).order_by('name'))
    branch_id = request.GET.get('branch') or ''
    branch = next((b for b in branches if str(b.pk) == str(branch_id)), None)

    stock = BranchStock.objects.select_related(
        'variant__product__category', 'branch')
    if branch:
        stock = stock.filter(branch=branch)
    stock = stock.exclude(variant__product__is_open_price=True)

    money = lambda expr: ExpressionWrapper(
        expr, output_field=DecimalField(max_digits=18, decimal_places=2))
    COST = money(F('stock_count') * F('cost_price'))
    RETAIL = money(F('stock_count') * F('sale_price'))

    # ---------- Umumiy ko'rsatkichlar ----------
    agg = stock.aggregate(
        units=Sum('stock_count'),
        cost_val=Sum(COST),
        retail_val=Sum(RETAIL),
        rows=Count('pk'),
        skus=Count('variant', distinct=True),
        products=Count('variant__product', distinct=True),
        zero=Count('pk', filter=Q(stock_count=0)),
        low=Count('pk', filter=Q(stock_count__gt=0, stock_count__lte=3)),
    )
    cost_val = float(agg['cost_val'] or 0)
    retail_val = float(agg['retail_val'] or 0)

    # ---------- Sotuv tezligi (90 kun) ----------
    d90 = timezone.now() - timedelta(days=90)
    sales_q = Sale.objects.filter(sold_at__gte=d90)
    if branch:
        sales_q = sales_q.filter(branch=branch)
    # SAL-6: bu yerда na qator chegirmasi, na chek chegirmasi ayirilardi —
    # 90 kunlik "sotilgan qiymat" ikki karra shishgan bo'lishi mumkin edi.
    # line_discount endi ayiriladi (bu shubhasiz xato edi); chek chegirmasi
    # esa MAHSULOT — qator darajasidagi o'lchov bo'lgani uchun alohida
    # ko'rsatiladi, qatorlarga taqsimlanmaydi.
    sold_map = {r['variant__product_id']: r for r in (
        sales_q.values('variant__product_id')
        .annotate(qty=Sum('quantity'),
                  revenue=Sum(money(F('quantity') * F('sale_price')
                                    - F('line_discount'))),
                  last=Max('sold_at')))}
    _odisc_90d, _ = _order_discount_share(sales_q)
    _, _, _disc_split = _order_discount_share(sales_q, split=True)
    order_discount_90d = float(_odisc_90d)

    # ---------- Kategoriya matritsasi ----------
    cat_rows = list(stock.values(
        'variant__product__category__id', 'variant__product__category__name'
    ).annotate(
        units=Sum('stock_count'),
        cost_val=Sum(COST),
        retail_val=Sum(RETAIL),
        products=Count('variant__product', distinct=True),
    ).order_by('-cost_val'))
    matrix = []
    for r in cat_rows:
        cv = float(r['cost_val'] or 0)
        rv = float(r['retail_val'] or 0)
        matrix.append({
            'id': r['variant__product__category__id'],
            'name': r['variant__product__category__name'] or 'Kategoriyasiz',
            'units': r['units'] or 0,
            'products': r['products'] or 0,
            'cost_val': cv,
            'retail_val': rv,
            'margin': ((rv - cv) / cv * 100) if cv > 0 else None,
            'share': (cv / cost_val * 100) if cost_val > 0 else 0,
        })

    # ---------- Bo'lim (4 katta guruh) bo'yicha ----------
    grp_rows = list(stock.values(
        'variant__product__category__group__id',
        'variant__product__category__group__name'
    ).annotate(
        units=Sum('stock_count'),
        cost_val=Sum(COST),
        retail_val=Sum(RETAIL),
        products=Count('variant__product', distinct=True),
    ).order_by('-cost_val'))
    groups_matrix = []
    for r in grp_rows:
        cv = float(r['cost_val'] or 0)
        rv = float(r['retail_val'] or 0)
        groups_matrix.append({
            'name': r['variant__product__category__group__name'] or "Bo'limsiz",
            'units': r['units'] or 0,
            'products': r['products'] or 0,
            'cost_val': cv,
            'retail_val': rv,
            'margin': ((rv - cv) / cv * 100) if cv > 0 else None,
            'share': (cv / cost_val * 100) if cost_val > 0 else 0,
        })

    # ---------- Ierarxiya daraxti: Bo'lim -> Kategoriya -> mahsulotlar ----------
    tree_rows = list(stock.values(
        'variant__product__category__group__id',
        'variant__product__category__group__name',
        'variant__product__category__group__sort_order',
        'variant__product__category__id',
        'variant__product__category__name',
    ).annotate(
        units=Sum('stock_count'),
        cost_val=Sum(COST),
        products=Count('variant__product', distinct=True),
    ))
    _tree = {}
    for r in tree_rows:
        gname = r['variant__product__category__group__name'] or "Bo'limsiz"
        gorder = r['variant__product__category__group__sort_order'] or 99
        node = _tree.setdefault(gname, {
            'name': gname, 'order': gorder, 'cost_val': 0.0,
            'products': 0, 'units': 0, 'cats': []})
        cv = float(r['cost_val'] or 0)
        node['cost_val'] += cv
        node['products'] += r['products'] or 0
        node['units'] += r['units'] or 0
        node['cats'].append({
            'id': r['variant__product__category__id'],
            'name': r['variant__product__category__name'] or 'Kategoriyasiz',
            'cost_val': cv,
            'products': r['products'] or 0,
            'units': r['units'] or 0,
        })
    group_tree = sorted(_tree.values(), key=lambda x: (x['order'], -x['cost_val']))
    for node in group_tree:
        node['share'] = (node['cost_val'] / cost_val * 100) if cost_val > 0 else 0
        node['cats'].sort(key=lambda c: -c['cost_val'])
        cmax = max((c['cost_val'] for c in node['cats']), default=0) or 1
        for c in node['cats']:
            c['bar'] = c['cost_val'] / cmax * 100
    _color_tree(group_tree)

    # ---------- ABC tahlili (90 kunlik tushum bo'yicha) ----------
    prod_stock = {r['variant__product_id']: r for r in stock.values(
        'variant__product_id').annotate(
            units=Sum('stock_count'), cost_val=Sum(COST))}
    ranked = []
    for pid, srow in prod_stock.items():
        s = sold_map.get(pid) or {}
        ranked.append({
            'pid': pid,
            'revenue': float(s.get('revenue') or 0),
            'qty': s.get('qty') or 0,
            'last': s.get('last'),
            'units': srow['units'] or 0,
            'cost_val': float(srow['cost_val'] or 0),
        })
    ranked.sort(key=lambda x: -x['revenue'])
    total_rev = sum(x['revenue'] for x in ranked) or 1.0
    run = 0.0
    abc = {'A': {'n': 0, 'val': 0.0, 'rev': 0.0},
           'B': {'n': 0, 'val': 0.0, 'rev': 0.0},
           'C': {'n': 0, 'val': 0.0, 'rev': 0.0}}
    for x in ranked:
        run += x['revenue']
        pct = run / total_rev * 100
        cls = 'A' if pct <= 80 and x['revenue'] > 0 else ('B' if pct <= 95 and x['revenue'] > 0 else 'C')
        x['abc'] = cls
        abc[cls]['n'] += 1
        abc[cls]['val'] += x['cost_val']
        abc[cls]['rev'] += x['revenue']

    # ---------- O'lik zaxira ----------
    names = {p.pk: p for p in Product.objects.filter(
        pk__in=[x['pid'] for x in ranked]).select_related('category')}
    now = timezone.now()
    dead = []
    for x in ranked:
        if x['units'] <= 0 or x['qty'] > 0:
            continue
        p = names.get(x['pid'])
        if p is None:
            continue
        last = x['last']
        dead.append({
            'product': p,
            'units': x['units'],
            'cost_val': x['cost_val'],
            'days': (now - last).days if last else None,
        })
    dead.sort(key=lambda d: -d['cost_val'])
    dead_val = sum(d['cost_val'] for d in dead)

    # ---------- Buyurtma kerak (sotuv tezligiga qarab) ----------
    reorder = []
    for x in ranked:
        if x['qty'] <= 0:
            continue
        per_day = x['qty'] / 90.0
        days_left = (x['units'] / per_day) if per_day > 0 else None
        if days_left is not None and days_left <= 21:
            p = names.get(x['pid'])
            if p is None:
                continue
            reorder.append({
                'product': p,
                'units': x['units'],
                'per_day': per_day,
                'days_left': days_left,
                'suggest': max(1, int(per_day * 30 - x['units'])),
            })
    reorder.sort(key=lambda r: r['days_left'])

    # ---------- Kirim / chiqim harakati (12 hafta) ----------
    d84 = timezone.now() - timedelta(days=84)
    ink_q = Intake.objects.filter(received_at__gte=d84)
    sal_q = Sale.objects.filter(sold_at__gte=d84)
    if branch:
        ink_q = ink_q.filter(branch=branch)
        sal_q = sal_q.filter(branch=branch)
    from django.db.models.functions import TruncWeek
    in_map = {r['w'].date(): r['q'] for r in ink_q.annotate(w=TruncWeek('received_at'))
              .values('w').annotate(q=Sum('quantity')) if r['w']}
    out_map = {r['w'].date(): r['q'] for r in sal_q.annotate(w=TruncWeek('sold_at'))
               .values('w').annotate(q=Sum('quantity')) if r['w']}
    weeks, flow_in, flow_out = [], [], []
    start = (timezone.localdate() - timedelta(days=83))
    start -= timedelta(days=start.weekday())
    for i in range(12):
        w = start + timedelta(weeks=i)
        weeks.append(w.strftime('%d.%m'))
        flow_in.append(int(in_map.get(w, 0) or 0))
        flow_out.append(int(out_map.get(w, 0) or 0))

    # Tuzatish jadvali — eng ko'p pul turgan yozuvlar
    adjust_rows = list(stock.annotate(val=COST).order_by('-val')[:40])

    _abc_total = sum(v['val'] for v in abc.values()) or 0.0
    abc_rows = [{'cls': c,
                 'n': abc[c]['n'],
                 'val': abc[c]['val'],
                 'rev': abc[c]['rev'],
                 'share': (abc[c]['val'] / _abc_total * 100) if _abc_total else 0}
                for c in ('A', 'B', 'C')]

    return render(request, 'inventory/warehouse.html', {
        'order_discount_90d': order_discount_90d,
        'branches': branches,
        'sel_branch': branch,
        'kpi': {
            'units': agg['units'] or 0,
            'cost_val': cost_val,
            'retail_val': retail_val,
            'profit': retail_val - cost_val,
            'margin': ((retail_val - cost_val) / cost_val * 100) if cost_val > 0 else None,
            'skus': agg['skus'] or 0,
            'products': agg['products'] or 0,
            'zero': agg['zero'] or 0,
            'low': agg['low'] or 0,
            'dead_n': len(dead),
            'dead_val': dead_val,
            'dead_share': (dead_val / cost_val * 100) if cost_val > 0 else 0,
        },
        'adjust_rows': adjust_rows,
        'matrix': matrix,
        'matrix_max': max([m['cost_val'] for m in matrix], default=0),
        'groups_matrix': groups_matrix,
        'groups_max': max([g['cost_val'] for g in groups_matrix], default=0),
        'group_tree': group_tree,
        'abc_rows': abc_rows,
        'dead': dead[:40],
        'dead_total': len(dead),
        'reorder': reorder[:40],
        'reorder_total': len(reorder),
        'weeks': weeks,
        'flow_in': flow_in,
        'flow_out': flow_out,
    })


@admin_required
def warehouse_adjust(request):
    """Zaxirani tuzatish — sabab bilan va tarixga yozilib.

    Farq Intake sifatida yoziladi (musbat yoki manfiy), shuning uchun
    qabul tarixida ham, hisobotlarda ham ko'rinadi.
    """
    from decimal import Decimal
    if request.method != 'POST':
        return redirect('warehouse')
    back = request.POST.get('back') or reverse('warehouse')
    st = BranchStock.objects.filter(pk=request.POST.get('stock') or 0).select_related(
        'variant__product', 'branch').first()
    if st is None:
        messages.error(request, "Zaxira yozuvi topilmadi.")
        return redirect(back)
    try:
        new_count = int(float(request.POST.get('count') or 0))
    except (TypeError, ValueError):
        messages.error(request, "Miqdor noto'g'ri.")
        return redirect(back)
    if new_count < 0:
        messages.error(request, "Miqdor manfiy bo'lmaydi.")
        return redirect(back)

    reason = (request.POST.get('reason') or '').strip()[:60] or "Qo'lda tuzatish"

    with transaction.atomic():
        # STK-9: qatorni QULFLAB, HOZIRGI (yangi) qoldiqni o'qiymiz — delta va
        # Intake yozuvi eskirgan o'qishга emas, jonli qiymatga nisbatan.
        st = (BranchStock.objects.select_for_update()
              .select_related('variant__product', 'branch').filter(pk=st.pk).first())
        if st is None:
            messages.error(request, "Zaxira yozuvi topilmadi.")
            return redirect(back)
        old = st.stock_count
        delta = new_count - old
        if delta == 0:
            messages.info(request, "O'zgarish yo'q.")
            return redirect(back)
        st.stock_count = new_count
        st.save(update_fields=['stock_count'])
        Intake.objects.create(
            variant=st.variant, branch=st.branch,
            quantity=delta, cost_per_unit=st.cost_price,
            sale_price=st.sale_price,
            is_return=delta < 0,
            return_reason=reason if delta < 0 else '',
            note=f"Ombor tuzatishi: {reason}",
            received_by=request.user)
        AuditLog.objects.create(
            user=request.user, username_snapshot=request.user.username,
            action=AuditLog.Action.UPDATE, model_name='BranchStock',
            object_id=str(st.pk),
            object_repr=f"{st.variant.product.name} {st.variant.size}"[:300],
            changes={'stock_count': [str(old), str(new_count)],
                     'sabab': ['', reason]},
        )
    messages.success(
        request,
        f"{st.variant.product.name} — {old} → {new_count} dona ({reason}).")
    return redirect(back)


# ===========================================================================
#  ONLAYN DO'KON (ochiq sayt)
#  Mijozlar uchun: katalog, savat, buyurtma. Kirish talab qilinmaydi.
#  Narx ko'rsatiladi, ombor qoldig'i KO'RSATILMAYDI.
# ===========================================================================

SHOP_NAME = 'Koreys Bozor'
SHOP_CITY = 'Urganch shahar'
SHOP_PHONE = '+998 90 000 00 00'
SHOP_TELEGRAM = 'https://t.me/koreysbozor'
CART_KEY = 'shop_cart'


def _shop_branch():
    return Branch.objects.filter(is_active=True).order_by('pk').first()


def _shop_stock_qs():
    """Saytda ko'rsatiladigan tovarlar.

    Narxsizlar (sotuv = 0) va ochiq narxli xizmat yozuvlari chiqmaydi —
    aks holda saytda "0 so'm" bo'lib ko'rinardi.
    """
    return (BranchStock.objects
            .filter(branch=_shop_branch(), stock_count__gt=0, sale_price__gt=0)
            .exclude(variant__product__is_open_price=True)
            .select_related('variant', 'variant__product',
                            'variant__product__category'))


def _shop_groups():
    """4 bo'lim + har birida nechta tovar (faqat omborda bori). Bo'sh
    bo'limlar ko'rsatilmaydi."""
    counts = dict(
        _shop_stock_qs()
        .filter(variant__product__category__group__isnull=False)
        .values_list('variant__product__category__group__slug')
        .annotate(n=Count('variant__product', distinct=True)))
    out = []
    for g in Group.objects.all():
        n = counts.get(g.slug, 0)
        if n:
            out.append({'slug': g.slug, 'name': g.name, 'n': n})
    return out


def _shop_ctx(request, **extra):
    cart = request.session.get(CART_KEY) or {}
    ctx = {
        'shop_name': SHOP_NAME, 'shop_city': SHOP_CITY,
        'shop_phone': SHOP_PHONE, 'shop_telegram': SHOP_TELEGRAM,
        'cart_count': sum(int(v) for v in cart.values()),
        'nav_groups': _shop_groups(),   # sayt menyusi — har sahifada
    }
    ctx.update(extra)
    return ctx


def shop_home(request):
    """Bosh sahifa — do'kon haqida + mashhur bo'limlar."""
    stock = _shop_stock_qs()
    cats = (stock.values('variant__product__category__id',
                         'variant__product__category__name')
            .annotate(n=Count('variant__product', distinct=True))
            .order_by('-n')[:10])
    categories = [{'id': c['variant__product__category__id'],
                   'name': c['variant__product__category__name'] or 'Boshqa',
                   'n': c['n']} for c in cats]

    # So'nggi qo'shilganlar
    latest_ids = list(stock.order_by('-variant__product__created_at')
                      .values_list('variant__product_id', flat=True)[:60])
    seen, pids = set(), []
    for pid in latest_ids:
        if pid not in seen:
            seen.add(pid); pids.append(pid)
        if len(pids) >= 8:
            break
    products = _shop_cards(pids)
    # Sayt pastidagi raqamlar — haqiqiy assortimentdan
    stat_products = stock.values('variant__product').distinct().count()
    stat_categories = (stock.values('variant__product__category')
                       .distinct().count())
    return render(request, 'shop/home.html',
                  _shop_ctx(request, categories=categories, products=products,
                            stat_products=stat_products,
                            stat_categories=stat_categories))


def _shop_cards(pids):
    """Mahsulot kartochkalari: nom, eng arzon narx, rasm."""
    if not pids:
        return []
    rows = (_shop_stock_qs().filter(variant__product_id__in=pids)
            .values('variant__product_id')
            .annotate(price_min=Min('sale_price'), price_max=Max('sale_price')))
    price_map = {r['variant__product_id']: r for r in rows}
    prods = Product.objects.filter(pk__in=pids).select_related('category')
    out = []
    for p in prods:
        r = price_map.get(p.pk)
        if not r:
            continue
        out.append({'p': p, 'price_min': r['price_min'], 'price_max': r['price_max']})
    # pids tartibini saqlaymiz
    order = {pid: i for i, pid in enumerate(pids)}
    out.sort(key=lambda x: order.get(x['p'].pk, 999))
    return out


def shop_catalog(request):
    """Katalog — qidiruv, kategoriya, sahifalash."""
    from django.core.paginator import Paginator
    stock = _shop_stock_qs()
    q = (request.GET.get('q') or '').strip()
    cat = request.GET.get('category') or ''
    group = (request.GET.get('group') or '').strip()   # men|women|kids|home
    if q:
        stock = stock.filter(
            Q(variant__product__name__icontains=q) |
            Q(variant__product__brand__icontains=q) |
            Q(variant__product__category__name__icontains=q))
    if group:
        stock = stock.filter(variant__product__category__group__slug=group)
    if cat.isdigit():
        stock = stock.filter(variant__product__category_id=int(cat))

    pids = list(dict.fromkeys(
        stock.order_by('variant__product__name')
        .values_list('variant__product_id', flat=True)))
    paginator = Paginator(pids, 24)
    page = paginator.get_page(request.GET.get('page') or 1)
    products = _shop_cards(list(page.object_list))

    # Kategoriya chiplari: bo'lim tanlangan bo'lsa faqat o'sha bo'limnikini
    categories = (Category.objects.filter(
        products__variants__branch_stocks__in=_shop_stock_qs())
        .annotate(n=Count('products', distinct=True))
        .order_by('name').distinct())
    if group:
        categories = categories.filter(group__slug=group)

    # Bo'lim nomi (sarlavha uchun)
    group_name = ''
    if group:
        g = Group.objects.filter(slug=group).first()
        group_name = g.name if g else ''

    return render(request, 'shop/catalog.html', _shop_ctx(
        request, products=products, page_obj=page, paginator=paginator,
        q=q, category_id=cat, categories=categories,
        group_slug=group, group_name=group_name))


def shop_product(request, code):
    product = get_object_or_404(Product, code=normalize_code(code))
    rows = _shop_stock_qs().filter(variant__product=product)
    if not rows.exists():
        raise Http404
    variants = [{'v': r.variant, 'price': r.sale_price, 'stock_id': r.pk}
                for r in rows.order_by('variant__size', 'variant__color')]
    return render(request, 'shop/product.html', _shop_ctx(
        request, product=product, variants=variants,
        price_min=min(v['price'] for v in variants),
        price_max=max(v['price'] for v in variants)))


# ---------- Savat (sessiyada saqlanadi) ----------

def _cart_items(request):
    """Savatdagi qatorlar. Yo'q bo'lib qolgan tovarlar jimgina tushib qoladi."""
    cart = request.session.get(CART_KEY) or {}
    if not cart:
        return [], 0
    rows = {str(r.pk): r for r in _shop_stock_qs().filter(pk__in=list(cart.keys()))}
    items, total = [], 0
    changed = False
    for sid, qty in list(cart.items()):
        st = rows.get(str(sid))
        if st is None:
            cart.pop(sid, None); changed = True
            continue
        qty = max(1, min(int(qty), st.stock_count))
        if str(cart[sid]) != str(qty):
            cart[sid] = qty; changed = True
        line = st.sale_price * qty
        total += line
        items.append({'stock': st, 'variant': st.variant,
                      'product': st.variant.product,
                      'qty': qty, 'price': st.sale_price, 'total': line})
    if changed:
        request.session[CART_KEY] = cart
        request.session.modified = True
    return items, total


@require_POST
def shop_cart_add(request):
    sid = (request.POST.get('stock') or '').strip()
    try:
        qty = max(1, int(request.POST.get('qty') or 1))
    except (TypeError, ValueError):
        qty = 1
    st = _shop_stock_qs().filter(pk=sid or 0).first()
    if st is None:
        messages.error(request, 'Bu tovar hozir mavjud emas.')
        return redirect('shop_catalog')
    cart = request.session.get(CART_KEY) or {}
    cart[str(st.pk)] = min(int(cart.get(str(st.pk), 0)) + qty, st.stock_count)
    request.session[CART_KEY] = cart
    request.session.modified = True
    messages.success(request, f'"{st.variant.product.name}" savatga qo\'shildi.')
    return redirect(request.POST.get('next') or 'shop_cart')


@require_POST
def shop_cart_update(request):
    cart = request.session.get(CART_KEY) or {}
    sid = (request.POST.get('stock') or '').strip()
    action = request.POST.get('action')
    if sid in cart:
        if action == 'remove':
            cart.pop(sid)
        else:
            try:
                q = int(request.POST.get('qty') or 1)
            except (TypeError, ValueError):
                q = 1
            if q <= 0:
                cart.pop(sid)
            else:
                cart[sid] = q
    request.session[CART_KEY] = cart
    request.session.modified = True
    return redirect('shop_cart')


def shop_cart(request):
    items, total = _cart_items(request)
    return render(request, 'shop/cart.html',
                  _shop_ctx(request, items=items, total=total))


def shop_checkout(request):
    items, total = _cart_items(request)
    if not items:
        return redirect('shop_cart')

    # Onlayn to'lov faqat merchant kalitlari sozlangan bo'lsa ko'rinadi
    online_ready = bool(getattr(settings, 'CLICK_MERCHANT_ID', '') or
                        getattr(settings, 'PAYME_MERCHANT_ID', ''))

    if request.method == 'POST':
        name = (request.POST.get('name') or '').strip()[:120]
        phone = (request.POST.get('phone') or '').strip()[:32]
        address = (request.POST.get('address') or '').strip()
        note = (request.POST.get('note') or '').strip()
        pay = request.POST.get('payment') or WebOrder.Payment.ON_DELIVERY
        if pay == WebOrder.Payment.ONLINE and not online_ready:
            pay = WebOrder.Payment.ON_DELIVERY
        if not name or not phone:
            messages.error(request, 'Ism va telefon raqamini kiriting.')
        else:
            with transaction.atomic():
                order = WebOrder.objects.create(
                    branch=_shop_branch(), customer_name=name,
                    customer_phone=phone, address=address, note=note,
                    payment_method=pay, total=total)
                for it in items:
                    WebOrderLine.objects.create(
                        order=order, variant=it['variant'],
                        quantity=it['qty'], price=it['price'])
            request.session[CART_KEY] = {}
            request.session.modified = True
            try:
                from .notifications import notify_web_order
                notify_web_order(order)
            except Exception:
                pass  # xabar ketmasa ham buyurtma saqlanadi
            return redirect('shop_order_done', pk=order.pk)

    return render(request, 'shop/checkout.html', _shop_ctx(
        request, items=items, total=total, online_ready=online_ready))


def shop_order_done(request, pk):
    order = get_object_or_404(WebOrder.objects.prefetch_related(
        'lines__variant__product'), pk=pk)
    return render(request, 'shop/done.html', _shop_ctx(request, order=order))


# ---------- Xodimlar uchun: kelgan buyurtmalar ----------

@admin_required
def web_orders(request):
    status = request.GET.get('status') or ''
    qs = (WebOrder.objects.select_related('branch', 'handled_by')
          .prefetch_related('lines__variant__product'))
    if status in dict(WebOrder.Status.choices):
        qs = qs.filter(status=status)
    counts = {s: WebOrder.objects.filter(status=s).count()
              for s, _ in WebOrder.Status.choices}
    return render(request, 'inventory/web_orders.html', {
        'orders': qs[:100], 'status': status, 'counts': counts,
        'statuses': WebOrder.Status.choices,
    })


@admin_required
@require_POST
def web_order_status(request, pk):
    order = get_object_or_404(WebOrder, pk=pk)
    new = request.POST.get('status')
    if new in dict(WebOrder.Status.choices):
        order.status = new
        order.handled_by = request.user
        order.handled_at = timezone.now()
        order.save(update_fields=['status', 'handled_by', 'handled_at'])
        messages.success(request, f'#{order.pk} — {order.get_status_display()}.')
    return redirect(request.POST.get('back') or 'web_orders')
